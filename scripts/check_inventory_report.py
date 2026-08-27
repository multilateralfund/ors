r"""
Use this script to validate Inventory report output agains provided comments / errors.

Example:
    python check_inventory_report.py \
        ~/Downloads/2026.08 Inventory\ report.xlsx \
        ~/Downloads/Validation\(Aug12,\ 2026\)-Comments.xlsx \
        -o invalid_inventory_report.md

"""

import argparse
import os
import sys
import django
from operator import itemgetter

from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime


# Setup required to use ORM outside Django commands.
sys.path.append(os.getcwd())
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "multilateralfund.settings")
django.setup()


from core.models import Project
from core.api.export.projects_inventory_report import tz_naive


def formatted_date(d):
    if d and isinstance(d, datetime):
        return d.strftime("%b-%y")
    return d


def compute_mya_completion_date(meta_project):
    result = None

    if meta_project:
        result = tz_naive(meta_project.end_date)

        projects = Project.objects.really_all().filter(
            version__gte=3, meta_project=meta_project, project_end_date__isnull=False
        )

        for project in projects:
            project_end_date = tz_naive(project.project_end_date)

            if result is None or project_end_date > result:
                result = project_end_date

    return result


def _get_category(project):
    category = project.meta_project.type if project.meta_project else project.category
    options = {
        "Individual": "IND",
        "Multi-year agreement": "MYA",
    }
    result = options.get(category, category)

    if result != "MYA" and project.cluster and project.cluster.category == "MYA":
        result = "MYA"

    return result


def fetch_project_info(project_id):
    project = Project.objects.get(pk=project_id)

    final_version = project.final_version

    agreement_date = tz_naive(final_version.date_per_agreement)
    project_end_date = tz_naive(final_version.project_end_date)

    meta_project = project.meta_project

    mya_end_date = tz_naive(meta_project.end_date) if meta_project else None
    mya_completion_date = compute_mya_completion_date(meta_project)

    extended_date = (
        tz_naive(meta_project.extended_date_of_completion) if meta_project else None
    )

    return {
        "Date per agreement": formatted_date(agreement_date),
        "Project end date": formatted_date(project_end_date),
        "End date (MYA)": formatted_date(mya_end_date),
        "Computed End date (MYA)": formatted_date(mya_completion_date),
        "Extended date (MYA)": formatted_date(extended_date),
        "Category": _get_category(project),
        "Status": project.status.code,
        "Adjustment": project.adjustment,
        "Project type": project.project_type.code,
        "Metacode": project.metacode,
    }


def read_comments(path):
    wb = load_workbook(path, read_only=True)
    sheet = wb.active

    result = {}

    current_header = ""

    for row in sheet.rows:
        row_data = [c.value for c in row]

        if any(row_data) and len(row_data) > 1:
            col_code = row_data[0]
            col_incorrect = row_data[1]
            col_correct = row_data[2]

            if col_code and not (col_incorrect or col_correct):
                continue

            if col_code.strip().lower() in ["legacy code", "code"]:
                current_header = col_incorrect.replace("Incorrect", "").strip()
                continue

            result.setdefault(current_header, []).append(
                (col_code, col_incorrect, col_correct)
            )

    return result


def read_inventory(path):
    wb = load_workbook(path, read_only=True)
    sheet = wb.active

    result = {}

    # sheet.rows is a generator, need a single reference so we can
    # build the header and then proceed to read the rows.
    _rows = sheet.rows
    header = [c.value for c in next(_rows)]

    for row in _rows:
        row_data = [c.value for c in row]
        project_data = dict(zip(header, row_data))

        result.setdefault(project_data["Code"], []).append(project_data)

        legacy_code = project_data["Legacy Code"]

        if legacy_code:
            result.setdefault(legacy_code, []).append(project_data)

    return result


def report_sums(inventory_data, comments_data):
    sum_inventory = 0
    for code, data in inventory_data.items():
        if len(data) > 1:
            print(f"Duplicate entries found for {code}: {[d["id"] for d in data]}")
        sum_inventory += len(data)

    print("Inventory total:", sum_inventory)
    print("Comments total:", sum(len(x) for x in comments_data.values()))


def check_value_differs(correct_value, current_value):

    if isinstance(correct_value, datetime) and isinstance(current_value, datetime):
        return (correct_value.year, correct_value.month) != (
            current_value.year,
            current_value.month,
        )

    if correct_value == "NOT EMPTY" and current_value:
        return False

    return current_value != correct_value


def validate_data(inventory_data, comments_data):
    result = {}

    count = 0
    for header, validated_projects in comments_data.items():
        print(f"Parsing header: {header}")
        for legacy_code, incorrect_value, correct_value in validated_projects:

            if legacy_code not in inventory_data:
                count += 1
                print(f"{count}. {legacy_code} missing from inventory!")
                continue

            matched_projects = inventory_data[legacy_code]
            invalid_projects = []

            for project in matched_projects:
                current_value = project[header]

                value_differs = check_value_differs(correct_value, current_value)
                if value_differs:
                    invalid_projects.append((project["id"], project))

            if len(invalid_projects) == len(matched_projects):
                count += 1
                for p_id, project in invalid_projects:
                    print(
                        f"{count}. [{p_id}] {legacy_code} {header}: {current_value} => {correct_value}"
                    )
                    result.setdefault(header, []).append(
                        {
                            "current": current_value,
                            "correct": correct_value,
                            "project": project,
                        }
                    )

    return result


def build_report(invalid_data):
    base_header = ["id", "code", "legacy code"]
    info_header = [
        "Metacode",
        "Date per agreement",
        "Project end date",
        "End date (MYA)",
        "Computed End date (MYA)",
        "Extended date (MYA)",
        "Category",
        "Status",
        "Project type",
    ]

    tables = {}

    for group, invalid in invalid_data.items():
        group_header = base_header.copy()
        group_header.extend([f"Incorrect {group}", f"Correct {group}"])
        group_header.extend(info_header)

        group_table = [group_header]

        for inv in invalid:
            project = inv["project"]
            info = fetch_project_info(project["id"])
            row = [project["id"], project["Code"], project["Legacy Code"]]
            row.extend([formatted_date(inv["current"]), formatted_date(inv["correct"])])
            row.extend([info[name] for name in info_header])
            group_table.append(row)

        tables[group] = group_table

    sort_on = itemgetter((base_header + info_header).index("Metacode") + 2)

    for group, table in tables.items():
        tables[group] = [table[0]] + sorted(table[1:], key=sort_on)

    return tables


def render_md(report):
    result = []

    for group, table in report.items():
        result.append(f"# {group}\n\n")
        header = table[0]
        result.append(f"| {' | '.join(header)} |")
        result.append(f"{'| -' * len(header)} |")
        for row in table[1:]:
            result.append(f"| {' | '.join([str(x) for x in row])} |")

        result.append("\n\n")

    return "\n".join(result)


def write_xls(report, xls_path):
    wb = Workbook(write_only=True)

    for group, table in report.items():
        sheet = wb.create_sheet(group)
        sheet.title = group

        for row in table:
            sheet.append(row)

    wb.save(xls_path)


def main(inventory_path, comments_paths: list[Path], md_output=None, xls_output=None):
    comments_data = {}

    for idx, comments_path in enumerate(comments_paths):
        comment_data = read_comments(comments_path)
        for comment_group, flagged_projects in comment_data.items():
            existing_comment_data = comments_data.setdefault(comment_group, [])
            existing_mapping = {c: correct for c, _, correct in existing_comment_data}

            for code, _, correct in flagged_projects:
                if code in existing_mapping and existing_mapping[code] != correct:
                    print(
                        f"Conflicting data for {code}: {comments_paths[idx - 1]} {existing_mapping[code]} != {correct} {comments_path}"
                    )

            existing_comment_data.extend(flagged_projects)

    inventory_data = read_inventory(inventory_path)

    report_sums(inventory_data, comments_data)
    invalid_data = validate_data(inventory_data, comments_data)

    report = build_report(invalid_data)
    rendered_report = render_md(report)

    print(rendered_report)

    if md_output is not None:
        with open(md_output, "w") as md_out_file:
            md_out_file.write(rendered_report)

    if xls_output is not None:
        write_xls(report, xls_output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Validate inventory report against provided comments.",
    )
    parser.add_argument("inventory", help="Inventory report XLSX.")
    parser.add_argument("comments", nargs="+", help="Comments XLSX.")
    parser.add_argument("-o", "--md-output", help="Output MD file.")
    parser.add_argument("--xls-output", help="Output XLSX file.")

    args = parser.parse_args()

    for arg in ["inventory", "comments"]:
        arg_value = getattr(args, arg)
        file_names = arg_value if isinstance(arg_value, list) else [arg_value]
        for file_name in file_names:
            arg_path = Path(file_name)

            print(f"{arg.title()}: {arg_path}")

            if not arg_path.exists():
                print(f"Error: {arg_path} not found!")

    inventory_path = Path(args.inventory)
    comments_paths = [Path(p) for p in args.comments]

    main(inventory_path, comments_paths, args.md_output, args.xls_output)
