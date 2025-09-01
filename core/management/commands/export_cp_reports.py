from django.core.management import BaseCommand
from openpyxl import Workbook

from core.models.adm import AdmColumn
from core.models.country_programme import CPReport, CPReportFormatColumn


class Command(BaseCommand):
    help = "Export country programme reports"

    def export_section_a_b(self, ws, cp_report, columns, section):
        cp_records = cp_report.cprecords.filter(section=section)
        for cp_record in cp_records:
            row = [
                cp_report.country.name,
                cp_report.year,
                cp_record.get_chemical_name(),
            ]
            usages = {
                str(cp_usage.usage_id): cp_usage.quantity
                for cp_usage in cp_record.record_usages.all()
            }
            for column in columns:
                if str(column.usage_id) in usages:
                    row.append(usages[str(column.usage_id)])
                else:
                    row.append(0)

            row += [
                cp_record.imports or 0,
                cp_record.exports or 0,
                cp_record.production or 0,
                cp_record.manufacturing_blends or 0,
                cp_record.import_quotas or 0,
                cp_record.export_quotas or 0,
                (
                    cp_record.banned_date.strftime("%d/%m/%Y")
                    if cp_record.banned_date
                    else ""
                ),
                cp_record.remarks,
            ]
            ws.append(row)

    def export_section_c(self, ws, cp_report):
        cp_prices = cp_report.prices.all()
        for cp_price in cp_prices:
            ws.append(
                [
                    cp_report.country.name,
                    cp_report.year,
                    cp_price.get_chemical_name(),
                    cp_price.previous_year_price,
                    cp_price.current_year_price,
                    cp_price.is_fob,
                    cp_price.is_retail,
                    cp_price.remarks,
                ],
            )

    def export_section_d(self, ws, cp_report):
        cp_generations = cp_report.cpgenerations.all()
        for cp_generation in cp_generations:
            ws.append(
                [
                    cp_report.country.name,
                    cp_report.year,
                    "HFC-23",
                    cp_generation.all_uses,
                    cp_generation.feedstock,
                    # cp_generation.destruction,  # deprecated, removed in frontend
                    cp_generation.other_uses_quantity,
                    cp_generation.other_uses_remarks,
                ],
            )

    def export_section_e(self, ws, cp_report):
        cp_emissions = cp_report.cpemissions.all()
        for cp_emission in cp_emissions:
            ws.append(
                [
                    cp_report.country.name,
                    cp_report.year,
                    cp_emission.facility,
                    cp_emission.total,
                    cp_emission.all_uses,
                    cp_emission.stored_at_start_of_year,
                    cp_emission.feedstock_gc,
                    cp_emission.destruction,
                    cp_emission.feedstock_wpc,
                    cp_emission.destruction_wpc,
                    cp_emission.generated_emissions,
                    cp_emission.stored_at_end_of_year,
                    cp_emission.remarks,
                ],
            )

    def export_adm_b_c(self, ws, cp_report, columns, section):
        adm = cp_report.adm_records.select_related("row").filter(section=section)
        for obj in adm:
            if obj.row.text == "N/A":
                continue
            row_text = (
                f"{obj.row.index} {obj.row.text}" if obj.row.index else obj.row.text
            )
            row = [cp_report.country.name, cp_report.year, row_text]
            for column in columns:
                if column == obj.column:
                    row.append(obj.value_text)
                else:
                    row.append("")
            ws.append(row)

    def export_adm_d(self, ws, cp_report):
        adm_d = cp_report.adm_records.select_related("value_choice").filter(section="D")
        for obj in adm_d:
            row_text = (
                f"{obj.row.index} {obj.row.text}" if obj.row.index else obj.row.text
            )
            ws.append(
                [
                    cp_report.country.name,
                    cp_report.year,
                    row_text,
                    obj.value_choice.value if obj.value_choice else "",
                    obj.value_text,
                ],
            )

    def handle(self, *args, **kwargs):
        wb = Workbook()
        del wb["Sheet"]
        ws1 = wb.create_sheet()
        ws2 = wb.create_sheet()
        ws3 = wb.create_sheet()
        ws4 = wb.create_sheet()
        ws5 = wb.create_sheet()
        ws6 = wb.create_sheet()
        ws7 = wb.create_sheet()
        ws8 = wb.create_sheet()
        ws9 = wb.create_sheet()

        ws1.title = "Section A"
        ws2.title = "Section B"
        ws3.title = "Section C"
        ws4.title = "Section D"
        ws5.title = "Section E"
        ws6.title = "Section F"
        ws7.title = "ADM B"
        ws8.title = "ADM C"
        ws9.title = "ADM D"

        report_fields = ["Country", "Year"]
        record_fields = [
            "Import",
            "Export",
            "Production",
            "Manufacturing of blends",
            "Import quotas",
            "Export quotas",
            "Date ban commenced",
            "Remarks",
        ]

        # Section A
        section_a_columns = (
            CPReportFormatColumn.objects.select_related("usage")
            .filter(section="A")
            .exclude(usage__full_name="Methyl bromide")
            .distinct("usage_id")
        )
        ws1.append(
            report_fields
            + ["Substance"]
            + [column.usage.full_name for column in section_a_columns]
            + record_fields
        )

        # Section B
        section_b_columns = (
            CPReportFormatColumn.objects.select_related("usage")
            .filter(section="B")
            .exclude(usage__full_name="Refrigeration")
            .exclude(usage__full_name="Refrigeration Manufacturing")
            .distinct("usage_id")
        )
        ws2.append(
            report_fields
            + ["Substance"]
            + [column.usage.full_name for column in section_b_columns]
            + record_fields
        )

        # Section C
        ws3.append(
            report_fields
            + [
                "Substance",
                "Previous year price",
                "Current year price",
                "FOB Price",
                "Retail Price",
                "Remarks",
            ],
        )

        # Section D
        ws4.append(
            report_fields
            + [
                "Substance",
                "Total production for all uses",
                "Production for feedstock uses within your country",
                # "Captured for destruction",  # destruction removed in frontend
                (
                    "Production for exempted essential, critical, "
                    "high-ambient-temperature or other uses within your country "
                    "- Quantity"
                ),
                (
                    "Production for exempted essential, critical, "
                    "high-ambient-temperature or other uses within your country "
                    "- Decision / type of use or remarks"
                ),
            ],
        )

        # Section E
        ws5.append(
            report_fields
            + [
                "Facility name",
                "Total amount generated (tonnes)",
                "Amount stored at the beginning of the year (tonnes)",
                "Amount generated and captured (tonnes) - For uses excluding feedstocks",
                "Amount generated and captured (tonnes) - For feedstock use in your country",
                "Amount generated and captured (tonnes) - For destruction",
                "Amount used for feedstock without prior capture (tonnes)",
                "Amount destroyed in the facility without prior capture (tonnes)",
                "Amount of generated emissions (tonnes)",
                "Amount stored at the end of the year (tonnes)",
                "Remarks",
            ],
        )

        # Section F
        ws6.append(report_fields + ["Comment"])

        # ADM B
        adm_b_columns = (
            AdmColumn.objects.filter(section="B")
            .exclude(type="parent")
            .order_by("sort_order")
        )
        ws7.append(
            report_fields
            + ["Type of action"]
            + [column.name for column in adm_b_columns]
        )

        # ADM C
        adm_c_columns = AdmColumn.objects.filter(section="C").order_by("sort_order")
        ws8.append(
            report_fields
            + ["Description"]
            + [column.display_name for column in adm_c_columns]
        )

        # ADM D
        ws9.append(report_fields + ["Question", "Answer choice", "Answer text"])

        # New CP report format
        for cp_report in (
            CPReport.objects.select_related("country")
            .filter(year__gt=2018)
            .order_by("country", "-year")
        ):
            # Section A
            self.export_section_a_b(ws1, cp_report, section_a_columns, "A")
            # Section B
            self.export_section_a_b(ws2, cp_report, section_b_columns, "B")
            # Section C
            self.export_section_c(ws3, cp_report)
            # Section D
            self.export_section_d(ws4, cp_report)
            # Section E
            self.export_section_e(ws5, cp_report)
            # Section F
            ws6.append([cp_report.country.name, cp_report.year, cp_report.comment])

        # Old format
        for cp_report in (
            CPReport.objects.select_related("country")
            .filter(year__lte=2018)
            .order_by("country", "-year")
        ):
            # Section A
            self.export_section_a_b(ws1, cp_report, section_a_columns, "A")
            # Section C
            self.export_section_c(ws3, cp_report)
            # ADM B
            self.export_adm_b_c(ws7, cp_report, adm_b_columns, "B")
            # ADM C
            self.export_adm_b_c(ws8, cp_report, adm_c_columns, "C")
            # ADM D
            self.export_adm_d(ws9, cp_report)

        wb.save("Country_programme_export.xlsx")
