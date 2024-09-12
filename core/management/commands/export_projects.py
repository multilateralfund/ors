from django.core.management import BaseCommand
from openpyxl import Workbook

from core.models.project import Project, ProjectFund, ProjectOdsOdp


class Command(BaseCommand):
    help = "Export projects"

    def export_projects(self, ws):
        projects = Project.objects.select_related(
            "country",
            "agency",
            "subsector__sector",
            "project_type",
            "status",
            "cluster",
            "approval_meeting",
            "meeting_transf",
            "meta_project",
        ).filter(code__isnull=False)

        for p in projects:
            ws.append(
                [
                    p.meta_project.type if p.meta_project else "",
                    p.meta_project.code if p.meta_project else "",
                    p.country.name,
                    p.country.iso3,
                    p.agency.name,
                    p.agency.code,
                    p.national_agency,
                    ", ".join(p.coop_agencies.all()),
                    p.legacy_code,
                    p.code,
                    p.serial_number_legacy,
                    p.serial_number,
                    p.additional_funding,
                    p.mya_code,
                    p.title,
                    p.description,
                    p.excom_provision,
                    p.project_type.name,
                    p.project_type.code,
                    p.project_type_legacy,
                    p.cluster.name if p.cluster else "",
                    p.cluster.code if p.cluster else "",
                    p.status.name,
                    p.status.code,
                    p.approval_meeting.number if p.approval_meeting else "",
                    p.meeting_transf.number if p.meeting_transf else "",
                    p.decision.number if p.decision else "",
                    p.project_duration,
                    p.stage,
                    p.tranche,
                    p.compliance,
                    p.sector.name if p.sector else "",
                    p.sector.code if p.sector else "",
                    p.sector_legacy,
                    p.subsector.name if p.subsector else "",
                    p.subsector.code if p.subsector else "",
                    p.subsector_legacy,
                    p.mya_subsector,
                    p.substance_type,
                    p.impact,
                    p.impact_production,
                    p.substance_phasedout,
                    p.fund_disbursed,
                    p.fund_disbursed_psc,
                    p.capital_cost,
                    p.operating_cost,
                    p.contingency_cost,
                    p.effectiveness_cost,
                    p.total_fund_transferred,
                    p.total_psc_transferred,
                    p.total_fund_approved,
                    p.total_psc_cost,
                    p.total_grant,
                    p.date_approved,
                    p.date_completion,
                    p.date_actual,
                    p.date_per_agreement,
                    p.remarks,
                    p.umbrella_project,
                    p.loan,
                    p.intersessional_approval,
                    p.retroactive_finance,
                    p.withdrawn,
                    p.incomplete,
                    p.issue,
                    p.issue_description,
                    p.application,
                    p.products_manufactured,
                    p.plan,
                    p.technology,
                    p.impact_co2mt,
                    p.impact_prod_co2mt,
                    p.ods_phasedout_co2mt,
                    p.hcfc_stage,
                    p.date_comp_revised,
                    p.date_per_decision,
                    p.local_ownership,
                    p.export_to,
                    p.submission_category,
                    p.submission_number,
                    p.programme_officer,
                    p.funds_allocated,
                    p.support_cost_psc,
                    p.project_cost,
                    p.date_received,
                    p.revision_number,
                    p.date_of_revision,
                    p.agency_remarks,
                    p.submission_comments,
                    p.reviewed_mfs,
                    p.correspondance_no,
                    p.plus,
                    p.source_file,
                ],
            )

    def export_project_fund(self, ws):
        project_funds = ProjectFund.objects.select_related("project", "meeting").all()
        for p in project_funds:
            ws.append(
                [
                    p.project.code,
                    p.amount,
                    p.support_psc,
                    p.meeting.number if p.meeting else "",
                    p.interest,
                    p.date,
                    p.fund_type,
                    p.sort_order,
                ],
            )

    def export_project_ods_odp(self, ws):
        projects_ods_odp = ProjectOdsOdp.objects.select_related(
            "project", "ods_substance", "ods_blend"
        ).all()
        for p in projects_ods_odp:
            ws.append(
                [
                    p.project.code,
                    p.ods_substance.name if p.ods_substance else "",
                    p.ods_blend.name if p.ods_blend else "",
                    p.ods_display_name,
                    p.odp,
                    p.ods_replacement,
                    p.co2_mt,
                    p.ods_type,
                    p.sort_order,
                ],
            )

    def handle(self, *args, **kwargs):
        wb = Workbook()
        del wb["Sheet"]
        ws1 = wb.create_sheet()
        ws2 = wb.create_sheet()
        ws3 = wb.create_sheet()
        ws1.title = "Projects"
        ws2.title = "ProjectFounds"
        ws3.title = "ProjectOdsOdp"

        # Projects
        ws1.append(
            [
                "Meta project type",
                "Meta project code",
                "Country name",
                "country iso3",
                "Agency name",
                "Agency code",
                "National agency",
                "Coop agencies",
                "Legacy code",
                "Project code",
                "Serial number legacy",
                "Serial number",
                "Additional funding",
                "Mya code",
                "Title",
                "Description",
                "Excom provision",
                "Project type name",
                "Project type code",
                "Project type legacy",
                "Cluster name",
                "Cluster code",
                "Status name",
                "Status code",
                "Approval meeting number",
                "Meeting transf number",
                "Decision number",
                "Project duration",
                "Stage",
                "Tranche",
                "Compliance",
                "Sector name",
                "Sector code",
                "Sector legacy",
                "Subsector name",
                "Subsector code",
                "Subsector legacy",
                "Mya subsector",
                "Substance type",
                "Impact",
                "Impact production",
                "Substance phasedout",
                "Fund disbursed",
                "Fund disbursed psc",
                "Capital cost",
                "Operating cost",
                "Contingency cost",
                "Effectiveness cost",
                "Total fund transferred",
                "Total psc transferred",
                "Total fund approved",
                "Total psc cost",
                "Total grant",
                "Date approved",
                "Date completion",
                "Date actual",
                "Date per agreement",
                "Remarks",
                "Umbrella project",
                "Loan",
                "Intersessional approval",
                "Retroactive finance",
                "Withdrawn",
                "Incomplete",
                "Issue",
                "Issue description",
                "Application",
                "Products manufactured",
                "Plan",
                "Technology",
                "Impact co2mt",
                "Impact prod co2mt",
                "Ods phasedout co2mt",
                "HCFC stage",
                "Date comp revised",
                "Date per decision",
                "Local ownership",
                "Export to",
                "Submission category",
                "Submission number",
                "Programme officer",
                "Funds allocated",
                "Support cost psc",
                "Project cost",
                "Date received",
                "Revision number",
                "Date of revision",
                "Agency remarks",
                "Submission comments",
                "Reviewed mfs",
                "Correspondance no",
                "Plus",
                "Source file",
            ],
        )
        self.export_projects(ws1)

        # ProjectFounds
        ws2.append(
            [
                "Project code",
                "Amount",
                "Support psc",
                "Meeting number",
                "Interest",
                "Date",
                "Fund type",
                "Sort order",
            ],
        )
        self.export_project_fund(ws2)

        # ProjectOdsOdp
        ws3.append(
            [
                "Project code",
                "Ods substance",
                "Ods blend",
                "Ods display name",
                "Odp",
                "Ods replacement",
                "CO₂ mt",
                "Ods type",
                "Sort order",
            ],
        )
        self.export_project_ods_odp(ws3)

        wb.save("Projects_export.xlsx")
