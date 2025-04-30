# TODO: split the file into multiple files
# pylint: disable=C0302,R0914

import json
import urllib
from base64 import b64decode
from datetime import datetime
from decimal import Decimal
from itertools import zip_longest

import openpyxl
from django.conf import settings
from django.core.files.base import ContentFile
from django.db import models, transaction
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, generics, mixins, status, views, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

from core.api.export.replenishment import (
    EMPTY_ROW,
    ScaleOfAssessmentTemplateWriter,
    StatusOfTheFundTemplateWriter,
    StatisticsTemplateWriter,
    StatusOfContributionsSummaryTemplateWriter,
    StatusOfContributionsTriennialTemplateWriter,
    StatusOfContributionsAnnualTemplateWriter,
    ConsolidatedInputDataWriter,
)
from core.api.filters.replenishment import (
    InvoiceFilter,
    PaymentFilter,
    ReplenishmentFilter,
    ScaleOfAssessmentFilter,
)
from core.api.permissions import IsUserAllowedReplenishment
from core.api.serializers import (
    CountrySerializer,
    InvoiceSerializer,
    InvoiceCreateSerializer,
    PaymentSerializer,
    PaymentCreateSerializer,
    ReplenishmentSerializer,
    ScaleOfAssessmentSerializer,
    DisputedContributionCreateSerializer,
    DisputedContributionReadSerializer,
    ScaleOfAssessmentExcelExportSerializer,
    EmptyInvoiceSerializer,
    ExternalAllocationSerializer,
    ExternalIncomeAnnualSerializer,
    StatusOfTheFundFileSerializer,
)
from core.api.utils import workbook_response, validate_files
from core.api.views.utils import (
    TriennialStatusOfContributionsAggregator,
    AnnualStatusOfContributionsAggregator,
    add_period_status_of_contributions_response_worksheet,
    SummaryStatusOfContributionsAggregator,
    add_summary_status_of_contributions_response_worksheet,
    add_statistics_status_of_contributions_response_worksheet,
    StatisticsStatusOfContributionsAggregator,
    get_as_of_date,
    get_budget_years,
)
from core.models import (
    Agency,
    AnnualContributionStatus,
    Country,
    DisputedContribution,
    ExternalIncomeAnnual,
    ExternalAllocation,
    FermGainLoss,
    Invoice,
    InvoiceFile,
    Payment,
    PaymentFile,
    Replenishment,
    ScaleOfAssessment,
    ScaleOfAssessmentVersion,
    TriennialContributionStatus,
    StatusOfTheFundFile,
)

EXPORT_RESOURCES_DIR = settings.ROOT_DIR / "api" / "export" / "templates"


class ReplenishmentCountriesViewSet(viewsets.GenericViewSet, mixins.ListModelMixin):
    """
    Viewset for all the countries that are available for a replenishment.
    """

    serializer_class = CountrySerializer

    def get_queryset(self):
        user = self.request.user
        queryset = Country.objects.all()
        if user.user_type in (
            user.UserType.COUNTRY_USER,
            user.UserType.COUNTRY_SUBMITTER,
        ):
            queryset = queryset.filter(id=user.country_id)
        return queryset.order_by("name")


class ReplenishmentCountriesSOAViewSet(viewsets.GenericViewSet, mixins.ListModelMixin):
    """
    Viewset that returns all SOA countries.

    There is no need for special security here.
    """

    serializer_class = CountrySerializer

    def get_queryset(self):
        # `contributions` is the related name for the SoA -> Country FK
        countries_qs = Country.objects.annotate(
            contributions_count=models.Count(models.F("contributions"))
        ).filter(contributions_count__gt=0)

        return countries_qs.order_by("name")


class ReplenishmentAsOfDateViewSet(viewsets.GenericViewSet, mixins.ListModelMixin):
    permission_classes = [IsUserAllowedReplenishment]

    def list(self, request, *args, **kwargs):
        self.check_permissions(request)

        as_of_date = get_as_of_date().strftime("%d %B %Y")
        data = {"as_of_date": as_of_date}

        return Response(status=status.HTTP_200_OK, data=data)


class ReplenishmentBudgetYearsViewSet(viewsets.GenericViewSet, mixins.ListModelMixin):
    permission_classes = [IsUserAllowedReplenishment]

    def list(self, request, *args, **kwargs):
        self.check_permissions(request)

        return Response(status=status.HTTP_200_OK, data=get_budget_years())


class ReplenishmentViewSet(
    viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin
):
    """
    Viewset for all replenishments that are available.
    """

    model = Replenishment
    serializer_class = ReplenishmentSerializer
    permission_classes = [IsUserAllowedReplenishment]
    filterset_class = ReplenishmentFilter

    def get_queryset(self):
        return Replenishment.objects.prefetch_related(
            models.Prefetch(
                "scales_of_assessment_versions",
                queryset=ScaleOfAssessmentVersion.objects.order_by("-version"),
            )
        ).order_by("-start_year")

    @transaction.atomic
    def perform_create(self, serializer):
        previous_replenishment = Replenishment.objects.order_by("-start_year").first()
        try:
            previous_replenishment_final_version = (
                previous_replenishment.scales_of_assessment_versions.get(is_final=True)
            )
        except ScaleOfAssessmentVersion.DoesNotExist as exc:
            raise ValidationError(
                {
                    "non_field_errors": "The current replenishment hasn't been finalized yet."
                }
            ) from exc
        except ScaleOfAssessmentVersion.MultipleObjectsReturned as exc:
            raise ValidationError(
                {
                    "non_field_errors": "There are multiple final versions for the previous replenishment."
                }
            ) from exc

        new_replenishment = serializer.save(
            amount=0,
            start_year=previous_replenishment.start_year + 3,
            end_year=previous_replenishment.end_year + 3,
        )

        # Create draft version
        new_version = ScaleOfAssessmentVersion.objects.create(
            replenishment=new_replenishment
        )

        # Copy scales of assessment from previous replenishment
        new_scales_of_assessment = []
        for (
            previous_scale_of_assessment
        ) in previous_replenishment_final_version.scales_of_assessment.values(
            "country", "currency"
        ):
            new_scale_of_assessment = ScaleOfAssessment(
                version=new_version,
                country_id=previous_scale_of_assessment["country"],
                currency=previous_scale_of_assessment["currency"],
            )
            new_scales_of_assessment.append(new_scale_of_assessment)

        ScaleOfAssessment.objects.bulk_create(new_scales_of_assessment)


class ScaleOfAssessmentViewSet(
    viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin
):
    """
    Viewset for all scales of assessment.
    """

    model = ScaleOfAssessment
    filterset_class = ScaleOfAssessmentFilter
    serializer_class = ScaleOfAssessmentSerializer
    permission_classes = [IsUserAllowedReplenishment]

    def get_queryset(self):
        return (
            ScaleOfAssessment.objects.select_related(
                "country", "version__replenishment"
            )
            .prefetch_related(
                models.Prefetch(
                    "version__replenishment__scales_of_assessment_versions",
                    queryset=ScaleOfAssessmentVersion.objects.order_by("-version"),
                )
            )
            .order_by("country__name")
        )

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        # pylint: disable=too-many-locals
        # pylint: disable=too-many-statements
        input_data = request.data

        try:
            replenishment = Replenishment.objects.get(id=input_data["replenishment_id"])
        except Replenishment.DoesNotExist as exc:
            raise ValidationError(
                {"replenishment_id": "Replenishment does not exist."}
            ) from exc

        if input_data.get("amount"):
            replenishment.amount = input_data["amount"]
            replenishment.save()

        should_create_new_version = input_data.get("createNewVersion")
        final = input_data.get("final") or False
        meeting_number = input_data.get("meeting") or ""
        decision_number = input_data.get("decision") or ""
        comment = input_data.get("comment") or ""
        decision_pdf = input_data.get("decision_pdf") or {}
        currency_date_range_start = input_data.get("currency_date_range_start") or ""
        currency_date_range_end = input_data.get("currency_date_range_end") or ""

        previous_version = replenishment.scales_of_assessment_versions.order_by(
            "-version"
        ).first()

        if previous_version.is_final:
            raise ValidationError(
                {
                    "non_field_errors": "The current replenishment has already been finalized."
                }
            )

        if decision_pdf.get("data") and decision_pdf.get("filename"):
            decision_file = ContentFile(
                b64decode(decision_pdf["data"]), name=decision_pdf["filename"]
            )
        else:
            decision_file = None

        if should_create_new_version or final:
            # Marking as final always creates a new version that is marked as final
            version = ScaleOfAssessmentVersion.objects.create(
                replenishment=replenishment,
                is_final=final,
                meeting_number=meeting_number,
                decision_number=decision_number,
                version=previous_version.version + 1,
                comment=comment,
                decision_pdf=decision_file,
                currency_date_range_start=currency_date_range_start,
                currency_date_range_end=currency_date_range_end,
            )
        else:
            version = previous_version
            version.decision_number = decision_number
            version.meeting_number = meeting_number
            version.is_final = final
            version.comment = comment
            version.decision_pdf = decision_file
            version.currency_date_range_start = currency_date_range_start
            version.currency_date_range_end = currency_date_range_end

            version.save()

        # Delete all scales of assessment if updating the latest version
        if not should_create_new_version:
            version.scales_of_assessment.all().delete()

        serializer = ScaleOfAssessmentSerializer(data=input_data["data"], many=True)
        serializer.is_valid(raise_exception=True)
        scales_of_assessment = [
            ScaleOfAssessment(version=version, **scale_of_assessment)
            for scale_of_assessment in serializer.validated_data
        ]

        ScaleOfAssessment.objects.bulk_create(scales_of_assessment)

        if final:
            # Delete all previous status of contributions data, if any
            AnnualContributionStatus.objects.filter(
                year__gte=replenishment.start_year, year__lte=replenishment.end_year
            ).delete()
            TriennialContributionStatus.objects.filter(
                start_year__gte=replenishment.start_year,
                end_year__lte=replenishment.end_year,
            ).delete()
            annual_contributions = []
            triennial_contributions = []
            # Create Status of Contributions data
            for scale_of_assessment in scales_of_assessment:
                annual_contributions.extend(
                    [
                        AnnualContributionStatus(
                            year=replenishment.start_year,
                            country=scale_of_assessment.country,
                            agreed_contributions=(
                                scale_of_assessment.amount / Decimal("3")
                            ),
                            outstanding_contributions=(
                                scale_of_assessment.amount / Decimal("3")
                            ),
                        ),
                        AnnualContributionStatus(
                            year=replenishment.start_year + 1,
                            country=scale_of_assessment.country,
                            agreed_contributions=(
                                scale_of_assessment.amount / Decimal("3")
                            ),
                            outstanding_contributions=(
                                scale_of_assessment.amount / Decimal("3")
                            ),
                        ),
                        AnnualContributionStatus(
                            year=replenishment.start_year + 2,
                            country=scale_of_assessment.country,
                            agreed_contributions=(
                                scale_of_assessment.amount / Decimal("3")
                            ),
                            outstanding_contributions=(
                                scale_of_assessment.amount / Decimal("3")
                            ),
                        ),
                    ]
                )
                triennial_contributions.append(
                    TriennialContributionStatus(
                        start_year=replenishment.start_year,
                        end_year=replenishment.end_year,
                        country=scale_of_assessment.country,
                        # Populating agreed/outstanding contributions in full;
                        # dealing with current values will happen via DB view.
                        agreed_contributions=scale_of_assessment.amount,
                        outstanding_contributions=scale_of_assessment.amount,
                    )
                )

            AnnualContributionStatus.objects.bulk_create(annual_contributions)
            TriennialContributionStatus.objects.bulk_create(triennial_contributions)

        headers = self.get_success_headers(serializer.data)
        return Response(
            {},
            status=(
                status.HTTP_201_CREATED
                if should_create_new_version or final
                else status.HTTP_200_OK
            ),
            headers=headers,
        )

    @action(methods=["GET"], detail=False, url_path="export")
    def export(self, request, *args, **kwargs):
        if "start_year" not in request.query_params:
            raise ValidationError({"start_year": "This query parameter is required."})
        if "version" not in request.query_params:
            raise ValidationError({"version": "This query parameter is required."})

        start_year = int(request.query_params["start_year"])
        queryset = self.filter_queryset(self.get_queryset())

        # Open template file
        wb = openpyxl.load_workbook(
            filename=EXPORT_RESOURCES_DIR / "Scale of Assesement.xlsx"
        )
        # Delete non-useful external links from these files; they are just exports
        # pylint: disable=protected-access
        wb._external_links = []
        ws = wb.active

        data = [
            {"no": i + 1, **soa}
            for i, soa in enumerate(
                ScaleOfAssessmentExcelExportSerializer(queryset, many=True).data
            )
        ]
        data_count = len(data)

        ScaleOfAssessmentTemplateWriter(
            ws,
            data,
            data_count,
            start_year,
        ).write()

        return workbook_response(
            f"Scales of Assessment {start_year} - {start_year + 2}",
            wb,
        )


class ReplenishmentScaleOfAssessmentVersionFileDownloadView(generics.RetrieveAPIView):
    permission_classes = [IsUserAllowedReplenishment]
    queryset = ScaleOfAssessmentVersion.objects.all()
    lookup_field = "id"

    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.decision_pdf is None:
            return Response(status=status.HTTP_404_NOT_FOUND)

        response = HttpResponse(
            obj.decision_pdf.file.read(), content_type="application/octet-stream"
        )
        file_name = urllib.parse.quote(obj.decision_pdf_name)
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        return response


class AnnualStatusOfContributionsView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        year = kwargs["year"]

        self.check_permissions(request)
        agg = AnnualStatusOfContributionsAggregator(year)

        data = {}
        soc_qs = agg.get_status_of_contributions_qs()
        data["status_of_contributions"] = [
            {
                "country": CountrySerializer(country).data,
                "agreed_contributions": country.agreed_contributions,
                "cash_payments": country.cash_payments,
                "bilateral_assistance": country.bilateral_assistance,
                "promissory_notes": country.promissory_notes,
                "outstanding_contributions": country.outstanding_contributions,
                "gain_loss": country.gain_loss,
            }
            for country in soc_qs
        ]

        ceit_countries_qs = agg.get_ceit_countries()
        data["ceit"] = agg.get_ceit_data(ceit_countries_qs)
        data["ceit_countries"] = CountrySerializer(ceit_countries_qs, many=True).data

        data["total"] = agg.get_total()

        disputed_contribution_amount = agg.get_disputed_contribution_amount()

        data["total"]["agreed_contributions_with_disputed"] = (
            data["total"]["agreed_contributions"] + disputed_contribution_amount
        )
        data["total"]["outstanding_contributions_with_disputed"] = (
            data["total"]["outstanding_contributions"] + disputed_contribution_amount
        )
        data["disputed_contributions"] = disputed_contribution_amount
        data["disputed_contributions_per_country"] = DisputedContributionReadSerializer(
            DisputedContribution.objects.filter(
                year=year, country__isnull=False
            ).select_related("country"),
            many=True,
        ).data

        return Response(data)


class AnnualStatusOfContributionsExportView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        year = kwargs["year"]

        self.check_permissions(request)
        agg = AnnualStatusOfContributionsAggregator(year)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        add_period_status_of_contributions_response_worksheet(
            wb, agg, f"YR{year}", year
        )

        return workbook_response(f"Status of Contributions {year}", wb)


class StatusOfContributionsExportView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]
    SUMMARY_WORKSHEET_NAME = "Summary Status of Contributions"
    TRIENIAL_WORKSHEET_NAME = "2024-26 Contributions"

    def get(self, request, *args, **kwargs):

        self.check_permissions(request)

        years = request.query_params.get("years")
        triennial_start_years = request.query_params.get("triennials")

        as_of_date = get_as_of_date()

        agg = SummaryStatusOfContributionsAggregator()
        soc_qs = agg.get_status_of_contributions_qs()
        summary_data = [
            {
                "no": index + 1,
                "country": country.name,
                "agreed_contributions": country.agreed_contributions,
                "cash_payments": country.cash_payments,
                "bilateral_assistance": country.bilateral_assistance,
                "promissory_notes": country.promissory_notes,
                "outstanding_contributions": country.outstanding_contributions,
                "gain_loss": country.gain_loss,
            }
            for index, country in enumerate(soc_qs)
        ]
        data_count = len(summary_data)
        disputed_contributions = agg.get_disputed_contribution_amount()

        # Open template file
        wb = openpyxl.load_workbook(
            filename=EXPORT_RESOURCES_DIR / "ContributionsFormatted.xlsx"
        )
        # Delete non-useful external links from these files; they are just exports
        # pylint: disable=protected-access
        wb._external_links = []

        ws = wb[self.SUMMARY_WORKSHEET_NAME]

        StatusOfContributionsSummaryTemplateWriter(
            ws,
            summary_data,
            data_count,
            None,
            as_of_date=as_of_date,
            disputed_contributions=disputed_contributions,
        ).write()
        sheet_names = [self.SUMMARY_WORKSHEET_NAME]

        triennial_ws = wb[self.TRIENIAL_WORKSHEET_NAME]
        # Now add triennials
        triennial_start_years = (
            [] if not triennial_start_years else triennial_start_years.split(",")
        )
        for triennial_start_year in triennial_start_years:
            start_year = int(triennial_start_year)
            end_year = start_year + 2
            agg = TriennialStatusOfContributionsAggregator(start_year, end_year)

            soc_qs = agg.get_status_of_contributions_qs()
            triennial_data = [
                {
                    "no": index + 1,
                    "country": country.name,
                    "agreed_contributions": country.agreed_contributions,
                    "cash_payments": country.cash_payments,
                    "bilateral_assistance": country.bilateral_assistance,
                    "promissory_notes": country.promissory_notes,
                    "outstanding_contributions": country.outstanding_contributions,
                    "gain_loss": country.gain_loss,
                }
                for index, country in enumerate(soc_qs)
            ]
            disputed_contributions = agg.get_disputed_contribution_amount()
            ceit_countries_qs = agg.get_ceit_countries()
            ceit_data = agg.get_ceit_data(ceit_countries_qs)

            ws = wb.copy_worksheet(triennial_ws)
            StatusOfContributionsTriennialTemplateWriter(
                ws,
                triennial_data,
                len(triennial_data),
                start_year,
                as_of_date=as_of_date,
                disputed_contributions=disputed_contributions,
                ceit_data=ceit_data,
            ).write()
            sheet_name = f"{start_year}-{end_year} Contributions"
            # Save sheet with the updated title
            ws.title = sheet_name
            # Make sure sheet doesn't get deleted in the end
            sheet_names.append(sheet_name)

        # Now add years
        years = [] if not years else years.split(",")
        for year in years:
            year = int(year)
            # We can use the same writer
            agg = AnnualStatusOfContributionsAggregator(year)

            soc_qs = agg.get_status_of_contributions_qs()
            annual_data = [
                {
                    "no": index + 1,
                    "country": country.name,
                    "agreed_contributions": country.agreed_contributions,
                    "cash_payments": country.cash_payments,
                    "bilateral_assistance": country.bilateral_assistance,
                    "promissory_notes": country.promissory_notes,
                    "outstanding_contributions": country.outstanding_contributions,
                    "gain_loss": country.gain_loss,
                }
                for index, country in enumerate(soc_qs)
            ]
            disputed_contributions = agg.get_disputed_contribution_amount()
            ceit_countries_qs = agg.get_ceit_countries()
            ceit_data = agg.get_ceit_data(ceit_countries_qs)

            ws = wb.copy_worksheet(triennial_ws)
            StatusOfContributionsAnnualTemplateWriter(
                ws,
                annual_data,
                len(annual_data),
                year,
                as_of_date=as_of_date,
                disputed_contributions=disputed_contributions,
                ceit_data=ceit_data,
            ).write()
            sheet_name = f"{year} Contributions"
            # Save sheet with the updated title
            ws.title = sheet_name
            # Make sure sheet doesn't get deleted in the end
            sheet_names.append(sheet_name)

        # Delete all other worksheets
        for name in wb.sheetnames:
            if name not in sheet_names:
                wb.remove(wb[name])

        return workbook_response("Status of Contributions", wb)


class TriennialStatusOfContributionsView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        start_year = kwargs["start_year"]
        end_year = kwargs["end_year"]

        self.check_permissions(request)
        agg = TriennialStatusOfContributionsAggregator(start_year, end_year)

        data = {}
        soc_qs = agg.get_status_of_contributions_qs()
        data["status_of_contributions"] = [
            {
                "country": CountrySerializer(country).data,
                "agreed_contributions": country.agreed_contributions,
                "cash_payments": country.cash_payments,
                "bilateral_assistance": country.bilateral_assistance,
                "promissory_notes": country.promissory_notes,
                "outstanding_contributions": country.outstanding_contributions,
                "gain_loss": country.gain_loss,
            }
            for country in soc_qs
        ]

        ceit_countries_qs = agg.get_ceit_countries()
        data["ceit"] = agg.get_ceit_data(ceit_countries_qs)
        data["ceit_countries"] = CountrySerializer(ceit_countries_qs, many=True).data

        data["total"] = agg.get_total()

        disputed_contribution_amount = agg.get_disputed_contribution_amount()
        data["total"]["agreed_contributions_with_disputed"] = (
            data["total"]["agreed_contributions"] + disputed_contribution_amount
        )
        data["total"]["outstanding_contributions_with_disputed"] = (
            data["total"]["outstanding_contributions"] + disputed_contribution_amount
        )
        data["disputed_contributions"] = disputed_contribution_amount
        data["disputed_contributions_per_country"] = DisputedContributionReadSerializer(
            DisputedContribution.objects.filter(
                year__gte=start_year, year__lte=end_year, country__isnull=False
            ).select_related("country"),
            many=True,
        ).data

        return Response(data)


class TriennialStatusOfContributionsExportView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        start_year = kwargs["start_year"]
        end_year = kwargs["end_year"]

        self.check_permissions(request)
        agg = TriennialStatusOfContributionsAggregator(start_year, end_year)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        add_period_status_of_contributions_response_worksheet(
            wb, agg, f"YR{start_year}_{str(end_year)[2:]}", f"{start_year}-{end_year}"
        )

        return workbook_response(f"Status of Contributions {start_year}-{end_year}", wb)


class SummaryStatusOfContributionsView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        self.check_permissions(request)
        agg = SummaryStatusOfContributionsAggregator()

        data = {}
        soc_qs = agg.get_status_of_contributions_qs()
        data["status_of_contributions"] = [
            {
                "country": CountrySerializer(country).data,
                "agreed_contributions": country.agreed_contributions,
                "cash_payments": country.cash_payments,
                "bilateral_assistance": country.bilateral_assistance,
                "promissory_notes": country.promissory_notes,
                "outstanding_contributions": country.outstanding_contributions,
                "gain_loss": country.gain_loss,
            }
            for country in soc_qs
        ]

        data["ceit"] = agg.get_ceit_data()
        data["ceit_countries"] = CountrySerializer(
            Country.objects.filter(ceit_statuses__is_ceit=True).distinct(), many=True
        ).data

        current_year = datetime.now().year
        data_current_year = TriennialContributionStatus.objects.filter(
            end_year__lt=current_year
        ).aggregate(
            bilateral_assistance=models.Sum("bilateral_assistance", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
            cash_payments=models.Sum("cash_payments", default=0),
            agreed_contributions=models.Sum("agreed_contributions", default=0),
        )

        if current_year % 3 == 2:
            # Current year is the start year of a triennial period
            current_triennial_data = AnnualContributionStatus.objects.filter(
                year=current_year
            ).aggregate(
                bilateral_assistance=models.Sum("bilateral_assistance", default=0),
                promissory_notes=models.Sum("promissory_notes", default=0),
                cash_payments=models.Sum("cash_payments", default=0),
                agreed_contributions=models.Sum("agreed_contributions", default=0),
            )
        elif current_year % 3 == 0:
            # Current year is in the middle of a triennial period
            current_triennial_data = AnnualContributionStatus.objects.filter(
                year__in=[current_year - 1, current_year]
            ).aggregate(
                bilateral_assistance=models.Sum("bilateral_assistance", default=0),
                promissory_notes=models.Sum("promissory_notes", default=0),
                cash_payments=models.Sum("cash_payments", default=0),
                agreed_contributions=models.Sum("agreed_contributions", default=0),
            )
        else:
            # Current year is the end year of a triennial period
            current_triennial_data = TriennialContributionStatus.objects.filter(
                end_year=current_year
            ).aggregate(
                bilateral_assistance=models.Sum("bilateral_assistance", default=0),
                promissory_notes=models.Sum("promissory_notes", default=0),
                cash_payments=models.Sum("cash_payments", default=0),
                agreed_contributions=models.Sum("agreed_contributions", default=0),
            )

        data_current_year["bilateral_assistance"] += current_triennial_data[
            "bilateral_assistance"
        ]
        data_current_year["promissory_notes"] += current_triennial_data[
            "promissory_notes"
        ]
        data_current_year["cash_payments"] += current_triennial_data["cash_payments"]
        data_current_year["agreed_contributions"] += current_triennial_data[
            "agreed_contributions"
        ]
        data["percentage_total_paid_current_year"] = (
            (
                data_current_year["cash_payments"]
                + data_current_year["bilateral_assistance"]
                + data_current_year["promissory_notes"]
            )
            / data_current_year["agreed_contributions"]
            * Decimal("100")
        )
        data["total"] = agg.get_total()
        data["total"]["gain_loss"] = agg.get_gain_loss()
        data["total"]["interest_earned"] = agg.get_interest_earned()

        disputed_contribution_amount = agg.get_disputed_contribution_amount()
        data["total"]["agreed_contributions_with_disputed"] = (
            data["total"]["agreed_contributions"] + disputed_contribution_amount
        )
        data["total"]["outstanding_contributions_with_disputed"] = (
            data["total"]["outstanding_contributions"] + disputed_contribution_amount
        )
        data["disputed_contributions"] = disputed_contribution_amount
        data["disputed_contributions_per_country"] = DisputedContributionReadSerializer(
            DisputedContribution.objects.filter(country__isnull=False).select_related(
                "country"
            ),
            many=True,
        ).data

        return Response(data)


class SummaryStatusOfContributionsExportView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        self.check_permissions(request)
        agg = SummaryStatusOfContributionsAggregator()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        add_summary_status_of_contributions_response_worksheet(wb, agg)

        current_year = datetime.now().year

        return workbook_response(f"Summary Status of Contributions {current_year}", wb)


class StatisticsStatusOfContributionsView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        self.check_permissions(request)

        current_year = datetime.now().year
        statistics_agg = StatisticsStatusOfContributionsAggregator()
        soc_data = statistics_agg.get_soc_data()
        external_income = statistics_agg.get_external_income_data()

        response = []
        # Annotating soc data with external income would make more queries
        # (separate for interest and miscellaneous income) and would be less
        # readable.
        for soc, income in zip_longest(soc_data, external_income):
            total_payments = (
                soc["cash_payments_sum"]
                + soc["bilateral_assistance_sum"]
                + soc["promissory_notes_sum"]
            )
            response.append(
                {
                    "start_year": soc["start_year"],
                    "end_year": soc["end_year"],
                    "agreed_contributions": soc["agreed_contributions_sum"],
                    "cash_payments": soc["cash_payments_sum"],
                    "bilateral_assistance": soc["bilateral_assistance_sum"],
                    "promissory_notes": soc["promissory_notes_sum"],
                    "total_payments": total_payments,
                    "disputed_contributions": soc["disputed_contributions"],
                    "outstanding_contributions": soc["outstanding_contributions_sum"],
                    "payment_pledge_percentage": (
                        total_payments
                        / soc["agreed_contributions_sum"]
                        * Decimal("100")
                    ),
                    "interest_earned": income["interest_earned"],
                    "miscellaneous_income": income["miscellaneous_income"],
                    "total_income": (
                        total_payments
                        + income["interest_earned"]
                        + income["miscellaneous_income"]
                    ),
                    "percentage_outstanding_agreed": (
                        soc["outstanding_contributions_sum"]
                        / soc["agreed_contributions_sum"]
                        * Decimal("100")
                    ),
                    "outstanding_ceit": soc["outstanding_ceit"],
                    "percentage_outstanding_ceit": (
                        soc["outstanding_ceit"]
                        / soc["agreed_contributions_sum"]
                        * Decimal("100")
                    ),
                }
            )

        summary_agg = SummaryStatusOfContributionsAggregator()
        # Interest earned and miscellaneous income are kept in separate models
        external_income_interest = ExternalIncomeAnnual.objects.aggregate(
            interest_earned=models.Sum("interest_earned", default=0),
            miscellaneous_income=models.Sum("miscellaneous_income", default=0),
        )
        totals = {
            "start_year": soc_data[0]["start_year"],
            "end_year": current_year,
            **summary_agg.get_total(),
            "disputed_contributions": summary_agg.get_disputed_contribution_amount(),
            "interest_earned": external_income_interest["interest_earned"],
            "miscellaneous_income": external_income_interest["miscellaneous_income"],
            "outstanding_ceit": summary_agg.get_ceit_data()[
                "outstanding_contributions"
            ],
        }

        totals["total_payments"] = (
            totals["cash_payments"]
            + totals["bilateral_assistance"]
            + totals["promissory_notes"]
        )
        totals["payment_pledge_percentage"] = (
            totals["total_payments"] / totals["agreed_contributions"] * Decimal("100")
        )
        totals["total_income"] = (
            totals["total_payments"]
            + totals["interest_earned"]
            + totals["miscellaneous_income"]
        )
        totals["percentage_outstanding_agreed"] = (
            totals["outstanding_contributions"]
            / totals["agreed_contributions"]
            * Decimal("100")
        )
        totals["percentage_outstanding_ceit"] = (
            totals["outstanding_ceit"] / totals["agreed_contributions"] * Decimal("100")
        )
        response.append(totals)

        return Response(response)


class StatisticsExportView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        self.check_permissions(request)

        statistics_agg = StatisticsStatusOfContributionsAggregator()
        soc_data = statistics_agg.get_soc_data()
        external_income = statistics_agg.get_external_income_data()

        statistics_data = []
        for soc, income in zip_longest(soc_data, external_income):
            total_payments = (
                soc["cash_payments_sum"]
                + soc["bilateral_assistance_sum"]
                + soc["promissory_notes_sum"]
            )
            statistics_data.append(
                {
                    0: f"{soc['start_year']}-{soc['end_year']}",
                    1: soc["agreed_contributions_sum"],
                    2: soc["cash_payments_sum"],
                    3: soc["bilateral_assistance_sum"],
                    4: soc["promissory_notes_sum"],
                    5: total_payments,
                    6: soc["disputed_contributions"],
                    7: soc["outstanding_contributions_sum"],
                    8: (total_payments / soc["agreed_contributions_sum"]),
                    # One empty row in between
                    10: income["interest_earned"],
                    # One empty row in between
                    12: income["miscellaneous_income"],
                    # One empty row in between
                    14: (
                        total_payments
                        + income["interest_earned"]
                        + income["miscellaneous_income"]
                    ),
                    # One empty row in between
                    16: f"{soc['start_year']}-{soc['end_year']}",
                    17: soc["agreed_contributions_sum"],
                    18: total_payments,
                    19: (total_payments / soc["agreed_contributions_sum"]),
                    # Total income
                    20: (
                        total_payments
                        + income["interest_earned"]
                        + income["miscellaneous_income"]
                    ),
                    21: soc["outstanding_contributions_sum"],
                    22: (
                        soc["outstanding_contributions_sum"]
                        / soc["agreed_contributions_sum"]
                    ),
                    23: soc["outstanding_ceit"],
                    24: (soc["outstanding_ceit"] / soc["agreed_contributions_sum"]),
                }
            )

        WORKSHEET_NAME = "Statistics"
        # Open template file
        wb = openpyxl.load_workbook(
            filename=EXPORT_RESOURCES_DIR / "ContributionsFormatted.xlsx"
        )
        # Delete non-useful external links from these files; they are just exports
        # pylint: disable=protected-access
        wb._external_links = []

        ws = wb[WORKSHEET_NAME]

        data_count = len(statistics_data)

        StatisticsTemplateWriter(
            ws,
            statistics_data,
            data_count,
            None,
            as_of_date=get_as_of_date(),
        ).write()

        # Delete all other worksheets
        for name in wb.sheetnames:
            if name != WORKSHEET_NAME:
                wb.remove(wb[name])

        return workbook_response(WORKSHEET_NAME, wb)


class StatisticsStatusOfContributionsExportView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        self.check_permissions(request)

        current_year = datetime.now().year
        periods = (
            TriennialContributionStatus.objects.values("start_year", "end_year")
            .distinct()
            .order_by("start_year")
        )
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        add_statistics_status_of_contributions_response_worksheet(wb, periods)

        summary_agg = SummaryStatusOfContributionsAggregator()
        add_summary_status_of_contributions_response_worksheet(wb, summary_agg)

        # Most likely too many queries, but it's readable
        # An alternative would be to do more Python manipulations with
        # queries that try to gather more data at once, it's probably
        # in the same order of magnitude performance-wise
        for period in reversed(periods):
            triennial_agg = TriennialStatusOfContributionsAggregator(
                period["start_year"], period["end_year"]
            )
            add_period_status_of_contributions_response_worksheet(
                wb,
                triennial_agg,
                f"YR{period['start_year']}_{str(period['end_year'])[2:]}",
                f"{period['start_year']}-{period['end_year']}",
            )

            for year in range(period["end_year"], period["start_year"] - 1, -1):
                if year > current_year:
                    continue

                annual_agg = AnnualStatusOfContributionsAggregator(year)
                add_period_status_of_contributions_response_worksheet(
                    wb, annual_agg, f"YR{year}", year
                )

        return workbook_response(f"All Status Of Contributions {current_year}", wb)


class DisputedContributionViewSet(
    viewsets.GenericViewSet,
    mixins.CreateModelMixin,
    mixins.DestroyModelMixin,
):
    model = DisputedContribution
    permission_classes = [IsUserAllowedReplenishment]

    def get_serializer_class(self):
        if self.request.method in ["POST", "PUT"]:
            return DisputedContributionCreateSerializer
        return DisputedContributionReadSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = DisputedContribution.objects.all()
        if user.user_type in (
            user.UserType.COUNTRY_USER,
            user.UserType.COUNTRY_SUBMITTER,
        ):
            queryset = queryset.filter(country_id=user.country_id)

        return queryset.select_related("country")

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """
        Override base DRF create to update annual/triennial agreed contributions.
        """
        amount = request.data.get("amount")
        country_id = request.data.get("country")
        year = request.data.get("year")

        ret = super().create(request, *args, **kwargs)

        try:
            annual_contribution = AnnualContributionStatus.objects.get(
                country_id=int(country_id),
                year=int(year),
            )
            triennial_contribution = TriennialContributionStatus.objects.get(
                country_id=int(country_id),
                start_year__lte=int(year),
                end_year__gte=int(year),
            )
        except AnnualContributionStatus.DoesNotExist:
            return ret
        except TriennialContributionStatus.DoesNotExist:
            return ret

        annual_contribution.agreed_contributions -= Decimal(amount)
        # It may seem wrong to substract from outstanding_contributions as well,
        # but because outstanding_contributions is not calculated on the fly, and
        # its value depends on agreed_contributions, we need to update it as well.
        annual_contribution.outstanding_contributions -= Decimal(amount)
        annual_contribution.save(
            update_fields=["agreed_contributions", "outstanding_contributions"]
        )

        triennial_contribution.agreed_contributions -= Decimal(amount)
        triennial_contribution.outstanding_contributions -= Decimal(amount)
        triennial_contribution.save(
            update_fields=["agreed_contributions", "outstanding_contributions"]
        )

        return ret


class BilateralAssistanceViewSet(
    viewsets.GenericViewSet,
    mixins.CreateModelMixin,
):
    model = AnnualContributionStatus
    permission_classes = [IsUserAllowedReplenishment]

    def get_queryset(self):
        user = self.request.user
        queryset = AnnualContributionStatus.objects.all()
        if user.user_type in (
            user.UserType.COUNTRY_USER,
            user.UserType.COUNTRY_SUBMITTER,
        ):
            queryset = queryset.filter(country_id=user.country_id)

        return queryset.select_related("country")

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """
        This viewset actually only updates the bilateral_assistance-related
        fields on Annual/Triennial contribution statuses.
        """
        # pylint: disable=too-many-locals
        input_data = request.data
        amount = input_data.get("amount")
        meeting_id = input_data.get("meeting_id")
        decision = input_data.get("decision_number", "")
        comment = input_data.get("comment", "")
        if amount is None:
            raise ValidationError(
                {"amount": "Bilateral assistance amount needs to be provided."}
            )
        if meeting_id is None:
            raise ValidationError(
                {"meeting": "Bilateral assistance meeting needs to be provided."}
            )

        try:
            annual_contribution = AnnualContributionStatus.objects.get(
                country_id=input_data["country_id"],
                year=input_data["year"],
            )
            triennial_contribution = TriennialContributionStatus.objects.get(
                country_id=input_data["country_id"],
                start_year__lte=input_data["year"],
                end_year__gte=input_data["year"],
            )
        except AnnualContributionStatus.DoesNotExist as exc:
            raise ValidationError(
                {
                    "year": (
                        "Annual Contribution Status for this year/country "
                        "does not exist."
                    )
                }
            ) from exc
        except TriennialContributionStatus.DoesNotExist as exc:
            raise ValidationError(
                {
                    "year": (
                        "Triennial Contribution Status for this year/country "
                        "does not exist."
                    )
                }
            ) from exc

        annual_contribution.bilateral_assistance += Decimal(amount)
        annual_contribution.outstanding_contributions -= Decimal(amount)
        annual_contribution.bilateral_assistance_meeting_id = meeting_id
        annual_contribution.bilateral_assistance_decision_number = decision
        annual_contribution.bilateral_assistance_comment = comment
        annual_contribution.save(
            update_fields=[
                "bilateral_assistance",
                "outstanding_contributions",
                "bilateral_assistance_meeting_id",
                "bilateral_assistance_decision_number",
                "bilateral_assistance_comment",
            ]
        )

        triennial_contribution.bilateral_assistance += Decimal(amount)
        triennial_contribution.outstanding_contributions -= Decimal(amount)
        triennial_contribution.bilateral_assistance_meeting_id = meeting_id
        triennial_contribution.bilateral_assistance_decision_number = decision
        triennial_contribution.bilateral_assistance_comment = comment
        triennial_contribution.save(
            update_fields=[
                "bilateral_assistance",
                "outstanding_contributions",
                "bilateral_assistance_meeting_id",
                "bilateral_assistance_decision_number",
                "bilateral_assistance_comment",
            ]
        )

        return Response(status=status.HTTP_200_OK, data=[])


class ReplenishmentDashboardView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        self.check_permissions(request)

        current_year = datetime.now().year
        latest_closed_triennial = (
            Replenishment.objects.filter(end_year__lt=current_year)
            .order_by("-start_year")
            .first()
        )
        income_interest = ExternalIncomeAnnual.objects.aggregate(
            interest_earned=models.Sum("interest_earned", default=0),
            miscellaneous_income=models.Sum("miscellaneous_income", default=0),
        )
        allocations = ExternalAllocation.objects.aggregate(
            undp=models.Sum("undp", default=0),
            unep=models.Sum("unep", default=0),
            unido=models.Sum("unido", default=0),
            world_bank=models.Sum("world_bank", default=0),
            staff_contracts=models.Sum("staff_contracts", default=0),
            treasury_fees=models.Sum("treasury_fees", default=0),
            monitoring_fees=models.Sum("monitoring_fees", default=0),
            technical_audit=models.Sum("technical_audit", default=0),
            information_strategy=models.Sum("information_strategy", default=0),
        )

        computed_summary_data = TriennialContributionStatus.objects.aggregate(
            cash_payments=models.Sum("cash_payments", default=0),
            bilateral_assistance=models.Sum("bilateral_assistance", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
        )
        computed_party_data = (
            Country.objects.filter(
                triennial_contributions_status__isnull=False,
            )
            .prefetch_related("triennial_contributions_status")
            .annotate(
                outstanding_contributions_sum=models.Sum(
                    "triennial_contributions_status__outstanding_contributions",
                    default=0,
                )
            )
            .aggregate(
                parties_paid_in_advance_count=models.Count(
                    "id",
                    filter=models.Q(outstanding_contributions_sum__lt=0),
                ),
                parties_paid_count=models.Count(
                    "id",
                    filter=models.Q(outstanding_contributions_sum=0),
                ),
                parties_have_to_pay_count=models.Count(
                    "id",
                    filter=models.Q(outstanding_contributions_sum__gt=0),
                ),
            )
        )

        gain_loss = FermGainLoss.objects.aggregate(
            total=models.Sum("amount", default=0)
        )["total"]

        computed_summary_data_latest_closed_triennial = (
            TriennialContributionStatus.objects.filter(
                start_year=latest_closed_triennial.start_year,
                end_year=latest_closed_triennial.end_year,
            ).aggregate(
                agreed_contributions=models.Sum("agreed_contributions", default=0),
                cash_payments=models.Sum("cash_payments", default=0),
                bilateral_assistance=models.Sum("bilateral_assistance", default=0),
                promissory_notes=models.Sum("promissory_notes", default=0),
            )
        )
        payment_pledge_percentage_latest_closed_triennial = (
            (
                computed_summary_data_latest_closed_triennial["cash_payments"]
                + computed_summary_data_latest_closed_triennial["bilateral_assistance"]
                + computed_summary_data_latest_closed_triennial["promissory_notes"]
            )
            / computed_summary_data_latest_closed_triennial["agreed_contributions"]
            * Decimal("100")
        )

        pledges = (
            TriennialContributionStatus.objects.values("start_year", "end_year")
            .annotate(
                outstanding_pledges=models.Sum("outstanding_contributions", default=0),
                agreed_pledges=models.Sum("agreed_contributions", default=0),
                total_payments=models.Sum("cash_payments", default=0),
            )
            .order_by("start_year")
        )

        as_of_date = get_as_of_date()

        external_income = ExternalIncomeAnnual.objects.values(
            "year",
            "triennial_start_year",
            "interest_earned",
            "miscellaneous_income",
        ).order_by("-triennial_start_year", "-year")

        agencies = list(
            Agency.objects.filter(is_for_replenishment=True)
            .values("name", "agency_type")
            .order_by("agency_type", "name")
        )

        data = {
            "as_of_date": as_of_date.strftime("%d %B %Y"),
            "overview": {
                "payment_pledge_percentage": payment_pledge_percentage_latest_closed_triennial,
                "gain_loss": gain_loss,
                "parties_paid_in_advance_count": computed_party_data[
                    "parties_paid_in_advance_count"
                ],
                "parties_paid_count": computed_party_data["parties_paid_count"],
                "parties_have_to_pay_count": computed_party_data[
                    "parties_have_to_pay_count"
                ],
            },
            "income": {
                "cash_payments": computed_summary_data["cash_payments"],
                "bilateral_assistance": computed_summary_data["bilateral_assistance"],
                "interest_earned": income_interest["interest_earned"],
                "promissory_notes": computed_summary_data["promissory_notes"],
                "miscellaneous_income": income_interest["miscellaneous_income"],
            },
            "allocations": {
                **allocations,
                "bilateral_assistance": computed_summary_data["bilateral_assistance"],
                "gain_loss": gain_loss,
            },
            "charts": {
                "outstanding_pledges": [
                    {
                        "start_year": pledge["start_year"],
                        "end_year": pledge["end_year"],
                        "outstanding_pledges": pledge["outstanding_pledges"],
                    }
                    for pledge in pledges
                    if pledge["end_year"] < current_year
                ],
                "pledged_contributions": [
                    {
                        "start_year": pledge["start_year"],
                        "end_year": pledge["end_year"],
                        "agreed_pledges": pledge["agreed_pledges"],
                    }
                    for pledge in pledges
                ],
                "payments": [
                    {
                        "start_year": pledge["start_year"],
                        "end_year": pledge["end_year"],
                        "total_payments": pledge["total_payments"],
                    }
                    for pledge in pledges
                ],
            },
            "external_income": external_income,
            "agencies": agencies,
        }

        data["overview"]["balance"] = sum(data["income"].values()) - sum(
            data["allocations"].values()
        )

        return Response(data)


class ReplenishmentDashboardExportView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get_status(self):
        income_interest = ExternalIncomeAnnual.objects.aggregate(
            interest_earned=models.Sum("interest_earned", default=0),
            miscellaneous_income=models.Sum("miscellaneous_income", default=0),
        )
        allocations = ExternalAllocation.objects.aggregate(
            undp=models.Sum("undp", default=0),
            unep=models.Sum("unep", default=0),
            unido=models.Sum("unido", default=0),
            world_bank=models.Sum("world_bank", default=0),
            staff_contracts=models.Sum("staff_contracts", default=0),
            treasury_fees=models.Sum("treasury_fees", default=0),
            monitoring_fees=models.Sum("monitoring_fees", default=0),
            technical_audit=models.Sum("technical_audit", default=0),
            information_strategy=models.Sum("information_strategy", default=0),
        )

        computed_summary_data = TriennialContributionStatus.objects.aggregate(
            cash_payments=models.Sum("cash_payments", default=0),
            bilateral_assistance=models.Sum("bilateral_assistance", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
        )
        gain_loss = FermGainLoss.objects.aggregate(
            total=models.Sum("amount", default=0)
        )["total"]

        total_income = sum(
            [
                computed_summary_data["cash_payments"],
                computed_summary_data["promissory_notes"],
                computed_summary_data["bilateral_assistance"],
                income_interest["interest_earned"],
                income_interest["miscellaneous_income"],
            ]
        )

        total_allocations_agencies = sum(
            [
                allocations["undp"],
                allocations["unep"],
                allocations["unido"],
                allocations["world_bank"],
            ]
        )

        total_provisions = sum(
            [
                total_allocations_agencies,
                allocations["staff_contracts"],
                allocations["treasury_fees"],
                allocations["monitoring_fees"],
                allocations["technical_audit"],
                allocations["information_strategy"],
                computed_summary_data["bilateral_assistance"],
                gain_loss,
            ]
        )

        promissory_notes = computed_summary_data["promissory_notes"]
        promissory_notes = (
            promissory_notes if abs(promissory_notes) > Decimal(5) else Decimal(0)
        )

        balance = total_income - total_provisions

        data = [
            (
                None,
                None,
                computed_summary_data["cash_payments"],
            ),
            (
                None,
                None,
                promissory_notes,
            ),
            (
                None,
                None,
                computed_summary_data["bilateral_assistance"],
            ),
            (
                None,
                None,
                income_interest["interest_earned"],
            ),
            (
                None,
                None,
                income_interest["miscellaneous_income"],
            ),
            EMPTY_ROW,
            (
                None,
                None,
                total_income,
            ),
            EMPTY_ROW,
            EMPTY_ROW,
            (None, allocations["undp"], None),
            (None, allocations["unep"], None),
            (None, allocations["unido"], None),
            (None, allocations["world_bank"], None),
            (None, "-", None),
            (None, "-", None),
            (
                None,
                None,
                total_allocations_agencies,
            ),
            EMPTY_ROW,
            EMPTY_ROW,
            (
                None,
                None,
                allocations["staff_contracts"],
            ),
            (None, None, allocations["treasury_fees"]),
            (
                None,
                None,
                allocations["monitoring_fees"],
            ),
            (None, None, allocations["technical_audit"]),
            (None, None, None),
            (
                None,
                None,
                allocations["information_strategy"],
            ),
            (
                None,
                None,
                computed_summary_data["bilateral_assistance"],
            ),
            EMPTY_ROW,
            (None, None, gain_loss),
            EMPTY_ROW,
            (
                None,
                None,
                total_provisions,
            ),
            None,
            (None, None, balance),
        ]
        return data

    def get(self, request, *args, **kwargs):
        WORKSHEET_NAME = "Status of the Fund"
        # Open template file
        wb = openpyxl.load_workbook(
            filename=EXPORT_RESOURCES_DIR / "ContributionsFormatted.xlsx"
        )
        # Delete non-useful external links from these files; they are just exports
        # pylint: disable=protected-access
        wb._external_links = []

        ws = wb[WORKSHEET_NAME]

        status_data = self.get_status()
        data_count = len(status_data)

        as_of_date = get_as_of_date()

        StatusOfTheFundTemplateWriter(
            ws,
            status_data,
            data_count,
            None,
            as_of_date=as_of_date,
        ).write()

        # Delete all other worksheets
        for name in wb.sheetnames:
            if name != WORKSHEET_NAME:
                wb.remove(wb[name])

        return workbook_response(WORKSHEET_NAME, wb)


class ReplenishmentExternalAllocationViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """
    Viewset for all the ExternalAllocation.
    """

    model = ExternalAllocation

    permission_classes = [IsUserAllowedReplenishment]
    serializer_class = ExternalAllocationSerializer
    ordering_fields = ["year"]

    def get_queryset(self):
        return ExternalAllocation.objects.all()

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        agency_name = request.data.get("agency_name")
        if agency_name is not None:
            return super().create(request, *args, **kwargs)

        # If no agency name is used, it's a special case of using the endpoint
        data_to_create = []

        meeting_id = request.data.get("meeting_id")
        decision_number = request.data.get("decision_number", "")
        comment = request.data.get("comment", "")

        for key in request.data.keys():
            params = {}
            if key.startswith("staff_contracts_"):
                params = {
                    "year": int(key.replace("staff_contracts_", "")),
                    "staff_contracts": Decimal(request.data.get(key)),
                }
            elif key.startswith("treasury_fees_"):
                params = {
                    "year": int(key.replace("treasury_fees_", "")),
                    "treasury_fees": Decimal(request.data.get(key)),
                }
            elif key.startswith("monitoring_fees_"):
                params = {
                    "year": int(key.replace("monitoring_fees_", "")),
                    "monitoring_fees": Decimal(request.data.get(key)),
                }
            else:
                # Non-relevant key
                continue

            data_to_create.append(
                {
                    "meeting_id": meeting_id,
                    "decision_number": decision_number,
                    "comment": comment,
                    **params,
                }
            )
        serializer = ExternalAllocationSerializer(data=data_to_create, many=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()

        headers = self.get_success_headers(serializer.data)
        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )


class ReplenishmentExternalIncomeAnnualViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """
    Viewset for the ExternalIncomeAnnual.
    """

    model = ExternalIncomeAnnual

    permission_classes = [IsUserAllowedReplenishment]
    serializer_class = ExternalIncomeAnnualSerializer
    ordering_fields = ["-year", "quarter"]

    def get_queryset(self):
        return ExternalIncomeAnnual.objects.all()


class ReplenishmentInvoiceViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """
    Viewset for all the invoices.
    """

    model = Invoice

    permission_classes = [IsUserAllowedReplenishment]
    serializer_class = InvoiceSerializer
    filterset_class = InvoiceFilter
    filter_backends = [
        DjangoFilterBackend,
        filters.OrderingFilter,
        filters.SearchFilter,
    ]
    ordering_fields = [
        "date_of_issuance",
    ]
    search_fields = [
        "number",
        "country__name",
        "year",
        "currency",
    ]

    def get_queryset(self):
        user = self.request.user
        queryset = Invoice.objects.all()
        if user.user_type in (
            user.UserType.COUNTRY_USER,
            user.UserType.COUNTRY_SUBMITTER,
        ):
            queryset = queryset.filter(country_id=user.country_id)

        ordering_params = self.request.query_params.get("ordering", "")
        if ordering_params:
            queryset = queryset.order_by(ordering_params)
        return queryset.select_related("country").prefetch_related(
            "invoice_files",
        )

    def get_serializer_class(self):
        if self.request.method in ["POST", "PUT"]:
            return InvoiceCreateSerializer
        return InvoiceSerializer

    def list(self, request, *args, **kwargs):
        """
        Return Scale of Assessment joined with the invoices to show countries
        that have not paid yet.
        """

        year_min = request.query_params.get("year_min")
        year_max = request.query_params.get("year_max")
        if (year_min is None or year_max is None) and request.query_params.get(
            "hide_no_invoice"
        ) != "true":
            raise ValueError(
                "year_min and year_max parameters are mandatory "
                "when querying for not issued invoices"
            )

        invoice_qs = self.filter_queryset(self.get_queryset())
        invoice_data = InvoiceSerializer(invoice_qs, many=True).data
        # pylint: disable=too-many-boolean-expressions
        if (
            "search" in request.query_params
            or "country_id" in request.query_params
            or "date_of_issuance" in request.query_params.get("ordering", "")
            or request.query_params.get("hide_no_invoice") == "true"
            or (
                request.query_params.get("status", None) is not None
                and request.query_params.get("status") != "not_issued"
            )
            or request.user.user_type == request.user.UserType.COUNTRY_USER
        ):
            # If filtered, we should not send the empty invoices
            return Response(
                invoice_data,
                status=status.HTTP_200_OK,
            )

        # Once we got here we certainly have year_min and year_max set
        year_min = int(year_min)
        year_max = int(year_max)
        countries_without_invoices_data = []
        for year in range(year_min, year_max + 1):
            countries_with_invoices = [
                invoice["country"]["id"]
                for invoice in invoice_data
                if invoice["year"] == year
            ]
            countries_without_invoices = (
                ScaleOfAssessment.objects.filter(
                    version__is_final=True,
                    version__replenishment__start_year__lte=year,
                    version__replenishment__end_year__gte=year,
                )
                .exclude(country_id__in=countries_with_invoices)
                .select_related("country")
                .order_by("country__name")
            )
            countries_without_invoices_data.extend(
                EmptyInvoiceSerializer(
                    countries_without_invoices, many=True, context={"year": year}
                ).data
            )

        if request.query_params.get("status") == "not_issued":
            data = countries_without_invoices_data
        else:
            data = [
                *invoice_data,
                *countries_without_invoices_data,
            ]

        ordering_params = request.query_params.get("ordering", "")
        if "country" in ordering_params:
            data = sorted(
                data,
                key=lambda x: x["country"]["name"],
                reverse="-country" in ordering_params,
            )

        return Response(
            data,
            status=status.HTTP_200_OK,
        )

    def _parse_invoice_new_files(self, request):
        number_of_files = int(request.data.get("nr_new_files", 0))
        files = [
            {
                "type": request.data.get(f"files[{file_no}][type]"),
                "contents": request.data.get(f"files[{file_no}][file]"),
            }
            for file_no in range(number_of_files)
        ]
        return files

    def _create_new_invoice_files(self, invoice, files_list):
        invoice_files = []
        for invoice_file in files_list:
            invoice_files.append(
                InvoiceFile(
                    invoice=invoice,
                    filename=invoice_file["contents"].name,
                    file=invoice_file["contents"],
                )
            )
        InvoiceFile.objects.bulk_create(invoice_files)

    def _parse_is_ferm_flag(self, request):
        is_ferm = request.data.get("is_ferm", None)
        if is_ferm == "true":
            return True
        return False

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        files = self._parse_invoice_new_files(request)

        # request.data is not mutable and we need to perform some boolean-string magic
        # for the is_ferm field, because we receive it from a forn.
        request_data = request.data.copy()

        is_ferm = self._parse_is_ferm_flag(request)
        request_data["is_ferm"] = is_ferm

        serializer = InvoiceCreateSerializer(data=request_data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # First create the actual invoice
        self.perform_create(serializer)
        invoice = serializer.instance

        # Now create the files for this Invoice
        self._create_new_invoice_files(invoice, files)

        # And finally set the ScaleOfAssessment if all needed fields are specified
        if is_ferm is not None and invoice.year:
            ScaleOfAssessment.objects.filter(
                version__replenishment__start_year__lte=invoice.year,
                version__replenishment__end_year__gte=invoice.year,
                version__is_final=True,
                country=invoice.country,
            ).update(opted_for_ferm=is_ferm)

        headers = self.get_success_headers(serializer.data)
        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        current_obj = self.get_object()

        # request.data is not mutable and we need to perform some boolean-string magic
        # for the is_ferm field, because we receive it from a forn.
        request_data = request.data.copy()

        is_ferm = self._parse_is_ferm_flag(request)
        request_data["is_ferm"] = is_ferm

        new_files = self._parse_invoice_new_files(request)
        files_to_delete = json.loads(request.data.get("deleted_files", "[]"))

        # First perform the update for the Invoice fields
        serializer = InvoiceCreateSerializer(current_obj, data=request_data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        self.perform_update(serializer)

        # Now create the new files for this Invoice
        self._create_new_invoice_files(current_obj, new_files)

        # And delete the ones that need to be deleted
        current_obj.invoice_files.filter(id__in=files_to_delete).delete()

        # And finally set the ScaleOfAssessment if all needed fields are specified
        if is_ferm is not None and current_obj.year:
            ScaleOfAssessment.objects.filter(
                version__replenishment__start_year__lte=current_obj.year,
                version__replenishment__end_year__gte=current_obj.year,
                version__is_final=True,
                country=current_obj.country,
            ).update(opted_for_ferm=is_ferm)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_200_OK, headers=headers)


class ReplenishmentInvoiceFileDownloadView(generics.RetrieveAPIView):
    permission_classes = [IsUserAllowedReplenishment]
    lookup_field = "id"

    def get_queryset(self):
        user = self.request.user
        queryset = InvoiceFile.objects.all()
        if user.user_type in (
            user.UserType.COUNTRY_USER,
            user.UserType.COUNTRY_SUBMITTER,
        ):
            queryset = queryset.filter(invoice__country_id=user.country_id)

        return queryset

    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        response = HttpResponse(
            obj.file.read(), content_type="application/octet-stream"
        )
        file_name = urllib.parse.quote(obj.filename)
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        return response


class ReplenishmentPaymentViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """
    Viewset for all the payments.
    """

    model = Payment

    permission_classes = [IsUserAllowedReplenishment]
    serializer_class = PaymentSerializer
    filterset_class = PaymentFilter
    filter_backends = [
        DjangoFilterBackend,
        filters.OrderingFilter,
        filters.SearchFilter,
    ]
    ordering_fields = [
        "amount_assessed",
        "country",
        "date",
    ]
    search_fields = [
        "payment_for_years",
        "country__name",
        "comment",
        "invoice__number",
        "date",
        "currency",
        "exchange_rate",
        "ferm_gain_or_loss",
    ]

    def get_queryset(self):
        user = self.request.user
        queryset = Payment.objects.all()
        if user.user_type in (
            user.UserType.COUNTRY_USER,
            user.UserType.COUNTRY_SUBMITTER,
        ):
            queryset = queryset.filter(country_id=user.country_id)

        return queryset.select_related("country")

    def get_serializer_class(self):
        if self.request.method in ["POST", "PUT"]:
            return PaymentCreateSerializer
        return PaymentSerializer

    def _parse_payment_new_files(self, request):
        number_of_files = int(request.data.get("nr_new_files", 0))
        files = [
            {
                "type": request.data.get(f"files[{file_no}][type]"),
                "contents": request.data.get(f"files[{file_no}][file]"),
            }
            for file_no in range(number_of_files)
        ]
        return files

    def _create_new_payment_files(self, payment, files_list):
        payment_files = []
        for payment_file in files_list:
            payment_files.append(
                PaymentFile(
                    payment=payment,
                    filename=payment_file["contents"].name,
                    file=payment_file["contents"],
                )
            )
        PaymentFile.objects.bulk_create(payment_files)

    def _parse_is_ferm_flag(self, request):
        is_ferm = request.data.get("is_ferm", None)
        if is_ferm == "true":
            return True
        return False

    def _set_scale_of_assessment_ferm(self, payment, is_ferm):
        if is_ferm is not None and payment.payment_for_years:
            years_list = [
                int(year)
                for year in payment.payment_for_years
                if year not in ["deferred", "arrears"]
            ]
            for year in years_list:
                ScaleOfAssessment.objects.filter(
                    version__replenishment__start_year__lte=year,
                    version__replenishment__end_year__gte=year,
                    version__is_final=True,
                    country=payment.country,
                ).update(opted_for_ferm=is_ferm)

    def _set_annual_triennial_contributions(self, payment, old_amount=None):
        """
        To be called when a payment is added or updated.

        Assumes that Annual/Triennial ContributionStatus objects exist for the payment's
        country & years.

        Assumes calling method/block is wrapped in transaction.atomic.

        For updates the `old_value` parameter is used; if it's set, we will substract
        the old amount from the cash payments and only then add the new amount.
        """
        years_list = []
        if payment.payment_for_years:
            years_list = [
                int(year)
                for year in payment.payment_for_years
                if year not in ["deferred", "arrears"]
            ]
        amount_to_add = payment.amount_assessed
        if old_amount is not None:
            amount_to_add -= old_amount
        AnnualContributionStatus.objects.filter(
            country=payment.country, year__in=years_list
        ).update(
            cash_payments=models.F("cash_payments") + amount_to_add,
            outstanding_contributions=models.F("outstanding_contributions")
            - amount_to_add,
        )
        for year in years_list:
            # Updating objects one by one to avoid race conditions
            # (same triennial object might need to be updated multiple times)
            TriennialContributionStatus.objects.filter(
                country=payment.country, start_year__lte=year, end_year__gte=year
            ).update(
                cash_payments=models.F("cash_payments") + amount_to_add,
                outstanding_contributions=models.F("outstanding_contributions")
                - amount_to_add,
            )

    def _unset_annual_triennial_contributions(self, payment):
        """
        To be called when a payment is deleted.

        Assumes that Annual/Triennial ContributionStatus objects exist for the payment's
        country & years.

        Assumes calling method/block is wrapped in transaction.atomic.
        """
        years_list = []
        if payment.payment_for_years:
            years_list = [
                int(year)
                for year in payment.payment_for_years
                if year not in ["deferred", "arrears"]
            ]
        AnnualContributionStatus.objects.filter(
            country=payment.country, year__in=years_list
        ).update(
            cash_payments=models.F("cash_payments") - payment.amount_assessed,
            outstanding_contributions=models.F("outstanding_contributions")
            + payment.amount_assessed,
        )
        for year in years_list:
            # Updating objects one by one to avoid race conditions
            # (same triennial object might need to be updated multiple times)
            TriennialContributionStatus.objects.filter(
                country=payment.country, start_year__lte=year, end_year__gte=year
            ).update(
                cash_payments=models.F("cash_payments") - payment.amount_assessed,
                outstanding_contributions=models.F("outstanding_contributions")
                + payment.amount_assessed,
            )

    def _set_ferm(self, payment, old_amount=None):
        """
        Updates global FERM data based on a added/updated payment.
        `old_amount` only used for updating payments so we know what to add/substract
        """
        if payment.ferm_gain_or_loss is None:
            return

        years_list = []
        if payment.payment_for_years:
            years_list = [
                int(year)
                for year in payment.payment_for_years
                if year not in ["deferred", "arrears"]
            ]
        if not years_list:
            return

        amount_to_add = payment.ferm_gain_or_loss
        if old_amount is not None:
            amount_to_add -= old_amount

        ferm_gain_loss_qs = FermGainLoss.objects.filter(
            country=payment.country, year__in=years_list
        )
        if ferm_gain_loss_qs.exists():
            ferm_gain_loss_qs.update(amount=models.F("amount") + amount_to_add)
        else:
            FermGainLoss.objects.create(
                country=payment.country,
                year=years_list[0],
                amount=amount_to_add,
            )

    def _unset_ferm(self, payment):
        """
        Updates global FERM data based on a deleted payment.
        """
        if payment.ferm_gain_or_loss is None:
            return

        years_list = []
        if payment.payment_for_years:
            years_list = [
                int(year)
                for year in payment.payment_for_years
                if year not in ["deferred", "arrears"]
            ]
        FermGainLoss.objects.filter(
            country=payment.country, year__in=years_list
        ).update(amount=models.F("amount") - payment.ferm_gain_or_loss)

    def _set_invoice_status(self, payment):
        """
        Update linked invoice when a payment is added or edited.
        """
        if payment.invoice is None:
            return
        invoice = payment.invoice
        # This relied on partially_paid and paid having the same char values in both
        # invoices and payments models.
        invoice.status = payment.status
        invoice.save(update_fields=["status"])

    def _unset_invoice_status(self, payment):
        """
        Update linked invoice when a payment is deleted.

        Called just before payment deletion.

        If other payments exist for that invoice, marking it as partially paid.
        If no other payments exist, marking it as pending.
        """
        if payment.invoice is None:
            return
        invoice = payment.invoice
        # The invoice still has this payment
        if invoice.payments.count() <= 1:
            invoice.status = invoice.InvoiceStatus.PENDING
        else:
            invoice.status = invoice.InvoiceStatus.PARTIALLY_PAID
        invoice.save(update_fields=["status"])

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        files = self._parse_payment_new_files(request)
        if file_errors := validate_files([f["contents"] for f in files]):
            return Response(file_errors, status=status.HTTP_400_BAD_REQUEST)
        # request.data is not mutable and we need to perform some boolean-string magic
        # for the is_ferm field, because we receive it from a forn.
        request_data = request.data.copy()

        is_ferm = self._parse_is_ferm_flag(request)
        request_data["is_ferm"] = is_ferm

        serializer = PaymentCreateSerializer(data=request_data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # First create the actual payment
        self.perform_create(serializer)
        payment = serializer.instance

        # Now create the files for this Payment
        self._create_new_payment_files(payment, files)

        # Update the annual/triennial contributions, FERM & invoice status
        self._set_annual_triennial_contributions(payment)
        self._set_ferm(payment)
        self._set_invoice_status(payment)

        # And finally set the ScaleOfAssessment if all needed fields are specified
        self._set_scale_of_assessment_ferm(payment, is_ferm)

        headers = self.get_success_headers(serializer.data)
        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        current_obj = self.get_object()

        previous_amount = current_obj.amount_assessed
        previous_ferm_gain_loss = current_obj.ferm_gain_or_loss

        # request.data is not mutable and we need to perform some string-bollean magic
        # for the is_ferm field, because we receive it from a forn.
        request_data = request.data.copy()
        is_ferm = self._parse_is_ferm_flag(request)
        request_data["is_ferm"] = is_ferm

        new_files = self._parse_payment_new_files(request)
        if file_errors := validate_files([f["contents"] for f in new_files]):
            return Response(file_errors, status=status.HTTP_400_BAD_REQUEST)
        files_to_delete = json.loads(request.data.get("deleted_files", "[]"))

        # First perform the update for the Payment fields
        serializer = PaymentCreateSerializer(current_obj, data=request_data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        self.perform_update(serializer)

        # Now create the new files for this Payment
        self._create_new_payment_files(current_obj, new_files)

        # And delete the ones that need to be deleted
        current_obj.payment_files.filter(id__in=files_to_delete).delete()

        # Update the annual/triennial contributions, FERM & invoice status
        self._set_annual_triennial_contributions(
            current_obj, old_amount=previous_amount
        )
        self._set_ferm(current_obj, old_amount=previous_ferm_gain_loss)
        self._set_invoice_status(current_obj)

        # And finally set the ScaleOfAssessment if all needed fields are specified
        self._set_scale_of_assessment_ferm(current_obj, is_ferm)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_200_OK, headers=headers)

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # First update the contributions, FERM & invoice status
        self._unset_annual_triennial_contributions(instance)
        self._unset_ferm(instance)
        self._unset_invoice_status(instance)

        # Then actually delete the payment
        return super().destroy(request, *args, **kwargs)


class ReplenishmentPaymentFileDownloadView(generics.RetrieveAPIView):
    permission_classes = [IsUserAllowedReplenishment]
    lookup_field = "id"

    def get_queryset(self):
        user = self.request.user
        queryset = PaymentFile.objects.all()
        if user.user_type in (
            user.UserType.COUNTRY_USER,
            user.UserType.COUNTRY_SUBMITTER,
        ):
            queryset = queryset.filter(invoice__country_id=user.country_id)

        return queryset

    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        response = HttpResponse(
            obj.file.read(), content_type="application/octet-stream"
        )
        file_name = urllib.parse.quote(obj.filename)
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        return response


class StatusOfTheFundFileViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    serializer_class = StatusOfTheFundFileSerializer
    permission_classes = [IsUserAllowedReplenishment]
    lookup_field = "id"
    queryset = StatusOfTheFundFile.objects.all()

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        file = request.data.get("file")
        if file is None:
            raise ValidationError({"file": "File contents must be uploaded."})

        file_to_save = ContentFile(
            b64decode(file.get("data")), name=file.get("filename")
        )

        if file_errors := validate_files([file_to_save]):
            return Response(file_errors, status=status.HTTP_400_BAD_REQUEST)

        serializer = StatusOfTheFundFileSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        created_obj = serializer.save()

        created_obj.file = file_to_save
        created_obj.save(update_fields=["file"])

        headers = self.get_success_headers(serializer.data)
        return Response(
            {},
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    def retrieve(self, request, *args, **kwargs):
        obj = self.get_object()
        response = HttpResponse(
            obj.file.read(), content_type="application/octet-stream"
        )
        file_name = urllib.parse.quote(obj.filename)
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        return response


class ConsolidatedInputDataExportView(views.APIView):
    permission_classes = [IsUserAllowedReplenishment]

    def get(self, request, *args, **kwargs):
        self.check_permissions(request)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ConsolidatedInputDataWriter(wb).write()

        return workbook_response("Backend Input Data", wb)
