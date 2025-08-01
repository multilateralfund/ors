# pylint: disable=C0302,R0914

from datetime import datetime

from django.db import models
from django.db.models import Q, F
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import ValidationError

from collections import defaultdict
from constance import config
from core.api.export.base import configure_sheet_print
from core.api.export.replenishment import (
    StatusOfContributionsWriter,
    StatisticsStatusOfContributionsWriter,
)
from core.models import (
    AnnualContributionStatus,
    BilateralAssistance,
    Country,
    DisputedContribution,
    ExternalIncomeAnnual,
    ExternalAllocation,
    FermGainLoss,
    Payment,
    ProjectHistory,
    TriennialContributionStatus,
    TriennialContributionView,
)
from core.models.business_plan import BusinessPlan
from core.models.country_programme import CPRecord, CPReport, CPReportFormatRow
from core.models.country_programme_archive import CPRecordArchive, CPReportArchive


BPACTIVITY_ORDERING_FIELDS = [
    "agency__name",
    "title",
    "country__iso3",
    "country__name",
    "bp_chemical_type__name",
    "project_type__code",
    "project_cluster__code",
    "sector__code",
    "subsector__code",
    "status",
    "is_multi_year",
]

SUBSTANCE_GROUP_ID_TO_CATEGORY = {
    "AI": "CFC",
    "AII": "Halon",
    "BI": "CFC",
    "BII": "CTC",
    "BIII": "TCA",
    "CI": "HCFC",
    "CII": "HBFC",
    "CIII": "Halon",
    "EI": "MBR",
    "F": "HFC",
    "uncontrolled": "Other",
    "legacy": "Legacy",
}


def log_project_history(project, request_user, description):
    ProjectHistory.objects.create(
        project=project,
        description=description,
        user=request_user,
    )


def get_country_region_dict():
    """
    Get a dictionary of country regions

    @return: dictionary of country regions
    """
    countries = Country.objects.all().select_related("parent__parent")
    country_region_dict = {}
    for country in countries:
        if country.parent_id and country.parent.parent_id:
            country_region_dict[country.id] = country.parent.parent.name
        elif country.parent_id:
            country_region_dict[country.id] = country.parent.name
        else:
            country_region_dict[country.id] = country.name

    return country_region_dict


def get_cp_report_from_request(request, cp_report_class):
    """
    Get country programme report from request query params

    @param request: request object
    @param cp_report_class: country programme report class

    @return: country programme report
    """
    cp_report_id = request.query_params.get("cp_report_id")
    country_id = request.query_params.get("country_id")
    year = request.query_params.get("year")
    cp_report = None

    try:
        user = request.user
        cp_report_qs = cp_report_class.objects.all()
        if user.has_perm("core.can_view_only_own_country") and not user.has_perm(
            "core.can_view_all_countries"
        ):
            cp_report_qs = cp_report_qs.filter(country=user.country)

        if cp_report_id:
            cp_report = cp_report_qs.get(id=cp_report_id)
        elif country_id and year:
            cp_report = cp_report_qs.get(country_id=country_id, year=year)
    except cp_report_class.DoesNotExist as e:
        raise ValidationError({"error": "Country programme report not found"}) from e

    return cp_report


def get_year_params_from_request(request):
    """
    Get min year and max year from request query params
    -> if min year and max year are not provided, set current year for both
    @param request: request object
    @return: min year and max year as integers
    """
    min_year = request.query_params.get("min_year")
    max_year = request.query_params.get("max_year")

    current_year = datetime.now().year
    if not min_year and not max_year:
        # set current year for min year and max year
        min_year = current_year
        max_year = current_year
    if not max_year:
        max_year = current_year
    if not min_year:
        min_year = 1995

    return int(min_year), int(max_year)


def get_archive_reports_final_for_year(year):
    """
    Get the max version for each archive report that does not have a final report

    @param year: year

    @return: list of archive reports (country_id, max_version)
    """
    countries_list = list(
        CPReport.objects.filter(
            year=year,
            status=CPReport.CPReportStatus.FINAL,
        )
        .values_list("country_id")
        .all()
    )

    # get the max version for each archive report that does not have a final report
    archive_reports = (
        CPReportArchive.objects.filter(
            year=year,
            status=CPReport.CPReportStatus.FINAL,
        )
        .exclude(country_id__in=countries_list)
        .values("country_id")
        .annotate(max_version=models.Max("version"))
        .values_list("country_id", "max_version")
        .all()
    )

    return archive_reports


def get_archive_reports_final_for_years(min_year, max_year):
    """
    Get the max version for each archive report that does not have a final report
    This will take into account the range of years [min_year, max_year]

    @param min_year: min year
    @param max_year: max year

    @return: list of archive reports (country_id, year, max_version)
    """
    current_reports = (
        CPReport.objects.filter(
            year__gte=min_year,
            year__lte=max_year,
            status=CPReport.CPReportStatus.FINAL,
        )
        .values_list("country_id", "year")
        .all()
    )

    # get the max version for each archive report (country&yuear)
    # that does not have a final report
    archive_reports_q = CPReportArchive.objects.filter(
        year__gte=min_year,
        year__lte=max_year,
        status=CPReport.CPReportStatus.FINAL,
    )
    # exclude the current reports
    for country, year in current_reports:
        archive_reports_q = archive_reports_q.exclude(
            country_id=country,
            year=year,
        )

    return (
        archive_reports_q.values("country_id", "year")
        .annotate(max_version=models.Max("version"))
        .values_list("country_id", "year", "max_version")
        .all()
    )


def get_final_records_for_years(min_year, max_year, filter_list=None, list_sort=True):
    """
    Get all the final records for the years in the range [min_year, max_year]
     - first get the final records for the countries that have a final report (CPReport)
     - then get the max version for each archive report that does not have a final report
     - get all the records for the archive reports
     - union the final records with the archive records in a dict
     - get the display_substance for years using the CPReportFormatRow
     - set the final list of records
        using the display_substance and the records from the dict

    @param min_year: min year
    @param max_year: max year
    @param filter_list: list of filters to apply to the records

    @return: list of records (CPRecord objects and CPRecordArchive objects)
    """
    if not filter_list:
        filter_list = []

    final_records = CPRecord.objects.get_for_years(min_year, max_year).filter(
        country_programme_report__status=CPReport.CPReportStatus.FINAL,
        *filter_list,
    )

    # get the max version for each archive report that does not have a final report
    archive_reports = get_archive_reports_final_for_years(min_year, max_year)

    # get all the records for the archive reports
    archive_records = []
    if archive_reports:
        archive_records = (
            CPRecordArchive.objects.get_for_years(min_year, max_year)
            .filter(
                *filter_list,
            )
            .filter(
                # get the records for the max version of the archive reports
                *[
                    models.Q(
                        country_programme_report__country_id=c,
                        country_programme_report__year=y,
                        country_programme_report__version=v,
                    )
                    for c, y, v in archive_reports
                ],
                _connector=models.Q.OR,
            )
        )

    # union the final records with the archive records
    reported_list = list(final_records) + list(archive_records)
    # set dict (country, year)
    existent_records = defaultdict(dict)
    for r in reported_list:
        country_year = (
            r.country_programme_report.country_id,
            r.country_programme_report.year,
        )
        chemical_key = (
            f"substance_{r.substance_id}" if r.substance else f"blend_{r.blend_id}"
        )
        existent_records[country_year][chemical_key] = r

    # get display_substance for years
    displayed_rows = {}
    for year in range(min_year, max_year + 1):
        displayed_rows[year] = (
            CPReportFormatRow.objects.get_for_year(year)
            .filter(*filter_list)
            .select_related("substance__group", "blend")
            .all()
        )

    # set the final list of records
    # if the country does not have the display_substance for the year,
    # then include a 0 value record
    final_list = []
    for country, year in existent_records:
        added_chemical_keys = set()
        for chemical_key, record in existent_records[(country, year)].items():
            added_chemical_keys.add(chemical_key)
            final_list.append(record)

        for row in displayed_rows[year]:
            chemical = row.substance or row.blend
            chemical_key = (
                f"blend_{chemical.id}" if row.blend else f"substance_{chemical.id}"
            )
            if chemical_key not in added_chemical_keys:
                cp_record_data = {
                    "country_programme_report": CPReport(
                        country_id=country, year=year, version=0
                    ),
                    "substance": chemical if row.substance else None,
                    "blend": chemical if row.blend else None,
                    "id": 0,
                }
                final_list.append(CPRecord(**cp_record_data))

    if not list_sort:
        return final_list

    # sort the final list
    final_list.sort(
        key=lambda x: (
            (
                x.country_programme_report.year,
                x.country_programme_report.country.name,
                (
                    x.substance.sort_order or float("inf")
                    if x.substance
                    else x.blend.sort_order or float("inf")
                ),
            )
        )
    )

    return final_list


def set_chemical_items_dict(
    item_cls, existing_items, section, cp_report, append_items=True
):
    """
    Set chemical records dict that are displayed for this report and section

    @param existing_items: list of existing ItemCls objects (CPRecord / CPPrices)
    @param section: str - section name
    @param cp_report: obj - CPReport object
    @param append_items: bool - if True, append new ItemCls objects to the existing items

    @return: dict of chemical records
    structure: {chemical_id: CPRecord object}
    """
    # set existing_items_dict
    existing_items_dict = {}
    for item in existing_items:
        dict_key = (
            f"blend_{item.blend_id}" if item.blend else f"substance_{item.substance_id}"
        )
        existing_items_dict[dict_key] = item

    # get all substances or blends that are displayed in all formats
    displayed_rows = (
        CPReportFormatRow.objects.get_for_year(cp_report.year)
        .filter(section=section)
        .select_related("substance__group", "blend")
        .prefetch_related(
            "substance__excluded_usages",
            "blend__excluded_usages",
            "blend__components",
        )
        .all()
    )

    # add substances or blends that are not in the existing_items_dict yet
    for row in displayed_rows:
        chemical = row.substance or row.blend
        chemical_key = (
            f"blend_{chemical.id}" if row.blend else f"substance_{chemical.id}"
        )
        if append_items and chemical_key not in existing_items_dict:
            cp_record_data = {
                "country_programme_report": cp_report,
                "substance": chemical if row.substance else None,
                "blend": chemical if row.blend else None,
                "id": 0,
            }
            if section in ["A", "B"]:
                cp_record_data["section"] = section
            existing_items_dict[chemical_key] = item_cls(**cp_record_data)

        if chemical_key in existing_items_dict:
            existing_items_dict[chemical_key].sort_order = row.sort_order

    return list(existing_items_dict.values())


def get_displayed_items(
    item_cls, cp_report, section, existing_items, with_sort=True, append_items=True
):
    """
    Returns a list of ItemCls objects for the given section and cp_report_id
    -> if there is no record for a substance or blend that is displayed in all formats
    then append a new ItemCls object to the list with the substance or blend
    and the cp_report_id

    @param item_cls: ItemCls class (CPRecord / CPPrices)
    @param cp_report_id: int - country programme report id
    @param section: str - section name
    @param existing_items: list of existing ItemCls objects
    @param with_sort: bool - if True, sort the final list
    @param append_items: bool - if True, append new ItemCls objects to the existing items

    @return: final list of ItemCls objects
    """

    # set the list of chemicals that are displayed for this report and section
    final_list = set_chemical_items_dict(
        item_cls, existing_items, section, cp_report, append_items
    )

    if with_sort:
        final_list.sort(
            key=lambda x: (
                (
                    (
                        x.substance.group.name
                        if "Other" not in x.substance.group.name_alt
                        else "zzzbbb"
                    ),  # other substances needs to be displayed last
                    getattr(x, "sort_order", float("inf")),
                    x.substance.sort_order,
                    x.substance.name,
                )
                if x.substance
                else (
                    (
                        "zzzbbb" if x.blend.is_related_preblended_polyol else "zzzaaa"
                    ),  # preblended polyols needs to be displayed last
                    getattr(x, "sort_order", float("inf")),
                    x.blend.sort_order or float("inf"),
                    x.blend.name,
                )
            )
        )

    return final_list


def get_cp_record(cp_report_id, section, cp_record_class):
    return (
        cp_record_class.objects.select_related(
            "substance__group",
            "blend",
            "country_programme_report__country",
        )
        .prefetch_related(
            "record_usages__usage",
            "substance__excluded_usages",
            "blend__excluded_usages",
            "blend__components",
        )
        .filter(country_programme_report_id=cp_report_id, section=section)
        .all()
    )


def get_displayed_records(cp_report, section, cp_record_class, append_items=True):
    """
    Returns a list of CPRecord objects for the given section and cp_report_id

    @param cp_report_id: int - country programme report id
    @param section: str - section name

    @return: list of CPRecord objects
    """

    exist_records = get_cp_record(cp_report.id, section, cp_record_class)
    final_list = get_displayed_items(
        cp_record_class, cp_report, section, exist_records, append_items=append_items
    )

    return final_list


def get_cp_prices(cp_report, cp_prices_class, append_items=True):
    exist_records = (
        cp_prices_class.objects.select_related("substance__group", "blend")
        .filter(country_programme_report=cp_report.id)
        .all()
    )
    final_list = get_displayed_items(
        cp_prices_class,
        cp_report,
        "C",
        exist_records,
        with_sort=False,
        append_items=append_items,
    )
    # set sort order for section C (we have set sort_order )
    final_list.sort(
        key=lambda x: (
            (
                (
                    x.substance.name[:4]
                    if "HCFC" in x.substance.name or "HFC" in x.substance.name
                    else "zzzBBB"
                ),  # other substances needs to be displayed last
                getattr(x, "sort_order", float("inf")) or float("inf"),
                getattr(x.substance, "sort_order", float("inf")) or float("inf"),
                x.substance.name,
            )
            if x.substance
            else (
                "zzzAAA",
                getattr(x, "sort_order", float("inf")) or float("inf"),
                getattr(x.blend, "sort_order", float("inf")) or float("inf"),
                x.blend.name,
            )
        )
    )
    return final_list


def get_business_plan_from_request(request):
    """
    Get business plan from request query params

    @param request: request object

    @return: business plan
    """
    business_plan_id = request.query_params.get("business_plan_id")
    year_start = request.query_params.get("year_start")
    year_end = request.query_params.get("year_end")
    bp_status = request.query_params.get("bp_status")
    business_plan = None

    try:
        if business_plan_id:
            business_plan = BusinessPlan.objects.get(id=business_plan_id)
        elif all([year_start, year_end, bp_status]):
            business_plan = BusinessPlan.objects.get(
                year_start=year_start,
                year_end=year_end,
                status=bp_status,
            )
        if not business_plan:
            raise BusinessPlan.DoesNotExist
    except BusinessPlan.DoesNotExist as e:
        raise ValidationError({"error": "Business plan not found"}) from e

    return business_plan


def copy_fields(obj, obj_old, fields):
    for field in fields:
        obj[f"{field}_old"] = obj_old[field] if obj_old else None


def rename_fields(obj, fields):
    """
    This is used for "old" records/activities that have now been deleted.
    """
    for field in fields:
        old_value = obj.pop(field, None)
        obj[field] = None
        obj[f"{field}_old"] = old_value


def delete_fields(obj, fields):
    for field in fields:
        obj.pop(field, None)


class SummaryStatusOfContributionsAggregator:
    """
    Aggregator for the summary status of contributions using the
    TriennialContributionView data.
    """

    # Years for which the bilateral assistance is taken from triennial data
    bilateral_triennial_years = [1994, 1995, 1996, 1997, 1998, 1999]
    bilateral_triennial_start_years = [1994, 1997]

    def get_status_of_contributions_qs(self):
        """
        @return: List of contributor countries annotated with the SoC data, similar
        to the Excel sheets YR91_XX.
        """
        return (
            Country.objects.filter(triennialcontributionview__isnull=False)
            .prefetch_related("triennialcontributionview", "ferm_gain_loss")
            .annotate(
                agreed_contributions=models.Sum(
                    "triennialcontributionview__current_agreed_contributions", default=0
                ),
                cash_payments=models.Sum(
                    "triennialcontributionview__cash_payments", default=0
                ),
                bilateral_assistance=models.Sum(
                    models.Case(
                        # For years 1994-1999, we use the values from the view
                        models.When(
                            triennialcontributionview__start_year__in=self.bilateral_triennial_start_years,
                            then=F("triennialcontributionview__bilateral_assistance"),
                        ),
                        default=0,
                        output_field=models.DecimalField(),
                    ),
                    default=0,
                )
                + models.Subquery(
                    BilateralAssistance.objects.exclude(
                        year__in=self.bilateral_triennial_years
                    )
                    .filter(
                        country=models.OuterRef("pk"),
                    )
                    .values("country__pk")
                    .annotate(total=models.Sum("amount", default=0))
                    .values("total")[:1]
                ),
                promissory_notes=models.Sum(
                    "triennialcontributionview__promissory_notes", default=0
                ),
                outstanding_contributions=models.Sum(
                    "triennialcontributionview__current_outstanding_contributions",
                    default=0,
                ),
                gain_loss=models.Subquery(
                    FermGainLoss.objects.filter(country=models.OuterRef("pk"))
                    .values("country__pk")
                    .annotate(total=models.Sum("amount", default=0))
                    .values("total")[:1]
                ),
            )
            .order_by("name")
        )

    def get_ceit_data(self):
        # Relies on the presence of `start/end_year` fields on the underlying model
        ceit_countries_triennial_filter = (
            Q(country__ceit_statuses__is_ceit=True)
            & Q(country__ceit_statuses__start_year__lte=F("start_year"))
            & (
                Q(country__ceit_statuses__end_year__gte=F("end_year"))
                | Q(country__ceit_statuses__end_year__isnull=True)
            )
        )
        # Relies on the presence of a `year` field on the underlying model
        ceit_countries_annual_filter = (
            Q(country__ceit_statuses__is_ceit=True)
            & Q(country__ceit_statuses__start_year__lte=F("year"))
            & (
                Q(country__ceit_statuses__end_year__gte=F("year"))
                | Q(country__ceit_statuses__end_year__isnull=True)
            )
        )

        ret = TriennialContributionView.objects.filter(
            ceit_countries_triennial_filter
        ).aggregate(
            agreed_contributions=models.Sum("current_agreed_contributions", default=0),
            cash_payments=models.Sum("cash_payments", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
            outstanding_contributions=models.Sum(
                "current_outstanding_contributions", default=0
            ),
        )

        # Populate gain_loss, bilateral_assistance and disputed_contributions fields;
        # these do not have a direct relationship with TriennialContributionStatus/View
        ret["gain_loss"] = FermGainLoss.objects.filter(
            ceit_countries_annual_filter
        ).aggregate(total=models.Sum("amount", default=0))["total"]

        ret["disputed_contributions"] = DisputedContribution.objects.filter(
            ceit_countries_annual_filter
        ).aggregate(total=models.Sum("amount", default=0))["total"]

        ret["bilateral_assistance"] = BilateralAssistance.objects.filter(
            ceit_countries_annual_filter
        ).aggregate(total=models.Sum("amount", default=0))["total"]

        return ret

    def get_total(self):
        ret = TriennialContributionView.objects.aggregate(
            agreed_contributions=models.Sum("current_agreed_contributions", default=0),
            cash_payments=models.Sum("cash_payments", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
            outstanding_contributions=models.Sum(
                "current_outstanding_contributions", default=0
            ),
        )
        # Taking the bliateral 1994-1996 and 1997-1999 data from the triennials
        bilateral_assistance = TriennialContributionView.objects.filter(
            start_year__in=self.bilateral_triennial_start_years
        ).aggregate(total=models.Sum("bilateral_assistance", default=0))["total"]
        # And the rest of the bilateral data from the BilateralAssistance model
        bilateral_assistance += BilateralAssistance.objects.exclude(
            year__in=self.bilateral_triennial_years
        ).aggregate(total=models.Sum("amount", default=0))["total"]
        ret["bilateral_assistance"] = bilateral_assistance
        return ret

    def get_gain_loss(self):
        return FermGainLoss.objects.aggregate(total=models.Sum("amount", default=0))[
            "total"
        ]

    def get_disputed_contribution_amount(self):
        return DisputedContribution.objects.aggregate(
            total=models.Sum("amount", default=0)
        )["total"]

    def get_interest_earned(self):
        return ExternalIncomeAnnual.objects.aggregate(
            total=models.Sum("interest_earned", default=0)
        )["total"]


class TriennialStatusOfContributionsAggregator:
    """
    Aggregator for the triennial status of contributions using the
    TriennialContributionView data.
    """

    def __init__(self, start_year, end_year):
        self.start_year = start_year
        self.end_year = end_year

    def get_status_of_contributions_qs(self):
        """
        @return: List of contributor countries annotated with the SoC data, similar
        to the Excel sheets e.g. YR1991-1993.
        """
        return (
            Country.objects.filter(
                triennialcontributionview__start_year=self.start_year,
                triennialcontributionview__end_year=self.end_year,
            )
            .prefetch_related("triennialcontributionview", "ferm_gain_loss")
            .annotate(
                agreed_contributions=models.Sum(
                    "triennialcontributionview__current_agreed_contributions", default=0
                ),
                cash_payments=models.Sum(
                    "triennialcontributionview__cash_payments", default=0
                ),
                bilateral_assistance=models.Case(
                    # For the 1994-1996 triennial
                    models.When(
                        triennialcontributionview__start_year=1994,
                        then=models.Sum(
                            "triennialcontributionview__bilateral_assistance", default=0
                        ),
                    ),
                    # For the 1997-1999 triennial
                    models.When(
                        triennialcontributionview__start_year=1997,
                        then=models.Sum(
                            "triennialcontributionview__bilateral_assistance", default=0
                        ),
                    ),
                    default=models.Subquery(
                        BilateralAssistance.objects.filter(
                            country=models.OuterRef("pk"),
                            year__gte=self.start_year,
                            year__lte=self.end_year,
                        )
                        .values("country__pk")
                        .annotate(total=models.Sum("amount", default=0))
                        .values("total")[:1]
                    ),
                    output_field=models.DecimalField(),
                ),
                promissory_notes=models.Sum(
                    "triennialcontributionview__promissory_notes", default=0
                ),
                outstanding_contributions=models.Sum(
                    "triennialcontributionview__current_outstanding_contributions",
                    default=0,
                ),
                gain_loss=models.Subquery(
                    FermGainLoss.objects.filter(
                        country=models.OuterRef("pk"),
                        year__gte=self.start_year,
                        year__lte=self.end_year,
                    )
                    .values("country__pk")
                    .annotate(total=models.Sum("amount", default=0))
                    .values("total")[:1]
                ),
            )
            .order_by("name")
        )

    def get_ceit_countries(self):
        return Country.objects.filter(
            Q(ceit_statuses__is_ceit=True)
            & Q(ceit_statuses__start_year__lte=self.start_year)
            & (
                Q(ceit_statuses__end_year__gte=self.end_year)
                | Q(ceit_statuses__end_year__isnull=True)
            ),
        )

    def get_ceit_data(self, ceit_countries_qs):
        ret = TriennialContributionView.objects.filter(
            start_year=self.start_year,
            end_year=self.end_year,
            country_id__in=ceit_countries_qs.values_list("id", flat=True),
        ).aggregate(
            agreed_contributions=models.Sum("current_agreed_contributions", default=0),
            cash_payments=models.Sum("cash_payments", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
            outstanding_contributions=models.Sum(
                "current_outstanding_contributions", default=0
            ),
        )
        # Add gain/loss & disputed & bilateral for CEIT countries
        ret["gain_loss"] = FermGainLoss.objects.filter(
            year__gte=self.start_year,
            year__lte=self.end_year,
            country_id__in=ceit_countries_qs.values_list("id", flat=True),
        ).aggregate(total=models.Sum("amount", default=0))["total"]

        ret["disputed_contributions"] = DisputedContribution.objects.filter(
            year__gte=self.start_year,
            year__lte=self.end_year,
            country_id__in=ceit_countries_qs.values_list("id", flat=True),
        ).aggregate(total=models.Sum("amount", default=0))["total"]

        ret["bilateral_assistance"] = BilateralAssistance.objects.filter(
            year__gte=self.start_year,
            year__lte=self.end_year,
            country_id__in=ceit_countries_qs.values_list("id", flat=True),
        ).aggregate(total=models.Sum("amount", default=0))["total"]

        return ret

    def get_total(self):
        ret = TriennialContributionView.objects.filter(
            start_year=self.start_year, end_year=self.end_year
        ).aggregate(
            agreed_contributions=models.Sum("current_agreed_contributions", default=0),
            cash_payments=models.Sum("cash_payments", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
            outstanding_contributions=models.Sum(
                "current_outstanding_contributions", default=0
            ),
        )
        # Adding gain/loss totals
        ret["gain_loss"] = FermGainLoss.objects.filter(
            year__gte=self.start_year, year__lte=self.end_year
        ).aggregate(total=models.Sum("amount", default=0))["total"]

        # Adding bilateral assistance totals
        if self.start_year in [1994, 1997]:
            # For the 1994-1996 and 1997-1999 triennials, use the value from
            # the TriennialContributionView
            ret["bilateral_assistance"] = TriennialContributionView.objects.filter(
                start_year=self.start_year
            ).aggregate(total=models.Sum("bilateral_assistance", default=0))["total"]
        else:
            # For all other triennials, use the BilateralAssistance model
            ret["bilateral_assistance"] = BilateralAssistance.objects.filter(
                year__gte=self.start_year, year__lte=self.end_year
            ).aggregate(total=models.Sum("amount", default=0))["total"]

        # Adding interest earned totals
        yearly_or_quarterly_data = ExternalIncomeAnnual.objects.filter(
            year__isnull=False, year__gte=self.start_year, year__lte=self.end_year
        ).aggregate(total=models.Sum("interest_earned", default=0))["total"]
        triennial_data = ExternalIncomeAnnual.objects.filter(
            triennial_start_year=self.start_year
        ).aggregate(total=models.Sum("interest_earned", default=0))["total"]
        ret["interest_earned"] = yearly_or_quarterly_data + triennial_data

        return ret

    def get_disputed_contribution_amount(self):
        return DisputedContribution.objects.filter(
            year__gte=self.start_year, year__lte=self.end_year
        ).aggregate(total=models.Sum("amount", default=0))["total"]


class AnnualStatusOfContributionsAggregator:
    """
    Aggregator for the annual status of contributions using the AnnualContributionStatus data.
    """

    def __init__(self, year):
        self.year = year

    def get_status_of_contributions_qs(self):
        """
        @return: List of contributor countries annotated with the SoC data, similar
        to the Excel sheets e.g. YR2005.
        """
        return (
            Country.objects.filter(
                annual_contributions_status__year=self.year,
            )
            .prefetch_related("annual_contributions_status")
            .annotate(
                agreed_contributions=models.Sum(
                    "annual_contributions_status__agreed_contributions", default=0
                ),
                cash_payments=models.Sum(
                    "annual_contributions_status__cash_payments", default=0
                ),
                bilateral_assistance=models.Subquery(
                    BilateralAssistance.objects.filter(
                        country=models.OuterRef("pk"),
                        year=self.year,
                    )
                    .values("country__pk")
                    .annotate(total=models.Sum("amount", default=0))
                    .values("total")[:1]
                ),
                promissory_notes=models.Sum(
                    "annual_contributions_status__promissory_notes", default=0
                ),
                outstanding_contributions=models.Sum(
                    "annual_contributions_status__outstanding_contributions", default=0
                ),
                gain_loss=models.Subquery(
                    FermGainLoss.objects.filter(
                        country=models.OuterRef("pk"),
                        year=self.year,
                    )
                    .values("country__pk")
                    .annotate(total=models.Sum("amount", default=0))
                    .values("total")[:1]
                ),
            )
            .order_by("name")
        )

    def get_ceit_countries(self):
        return Country.objects.filter(
            Q(ceit_statuses__is_ceit=True)
            & Q(ceit_statuses__start_year__lte=self.year)
            & (
                Q(ceit_statuses__end_year__gte=self.year)
                | Q(ceit_statuses__end_year__isnull=True)
            ),
        )

    def get_ceit_data(self, ceit_countries_qs):
        ceit_country_ids = ceit_countries_qs.values_list("id", flat=True)
        ret = AnnualContributionStatus.objects.filter(
            year=self.year,
            country_id__in=ceit_country_ids,
        ).aggregate(
            agreed_contributions=models.Sum("agreed_contributions", default=0),
            cash_payments=models.Sum("cash_payments", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
            outstanding_contributions=models.Sum(
                "outstanding_contributions", default=0
            ),
        )
        ret["gain_loss"] = FermGainLoss.objects.filter(
            year=self.year, country_id__in=ceit_country_ids
        ).aggregate(total=models.Sum("amount", default=0))["total"]
        ret["disputed_contributions"] = DisputedContribution.objects.filter(
            year=self.year, country_id__in=ceit_country_ids
        ).aggregate(total=models.Sum("amount", default=0))["total"]
        ret["bilateral_assistance"] = BilateralAssistance.objects.filter(
            year=self.year, country_id__in=ceit_country_ids
        ).aggregate(total=models.Sum("amount", default=0))["total"]

        return ret

    def get_total(self):
        ret = AnnualContributionStatus.objects.filter(year=self.year).aggregate(
            agreed_contributions=models.Sum("agreed_contributions", default=0),
            cash_payments=models.Sum("cash_payments", default=0),
            promissory_notes=models.Sum("promissory_notes", default=0),
            outstanding_contributions=models.Sum(
                "outstanding_contributions", default=0
            ),
        )
        ret["gain_loss"] = FermGainLoss.objects.filter(year=self.year).aggregate(
            total=models.Sum("amount", default=0)
        )["total"]
        ret["bilateral_assistance"] = BilateralAssistance.objects.filter(
            year=self.year
        ).aggregate(total=models.Sum("amount", default=0))["total"]
        ret["interest_earned"] = ExternalIncomeAnnual.objects.filter(
            year__isnull=False, year=self.year
        ).aggregate(total=models.Sum("interest_earned", default=0))["total"]

        return ret

    def get_disputed_contribution_amount(self):
        try:
            return DisputedContribution.objects.filter(year=self.year).aggregate(
                total=models.Sum("amount", default=0)
            )["total"]
        except DisputedContribution.DoesNotExist:
            return 0


class StatisticsStatusOfContributionsAggregator:
    """
    Aggregator for the statistics status of contributions using the
    TriennialContributionView data.
    """

    def get_soc_data(self):
        return (
            TriennialContributionView.objects.values("start_year", "end_year")
            .annotate(
                agreed_contributions_sum=models.Sum(
                    "current_agreed_contributions", default=0
                ),
                cash_payments_sum=models.Sum("cash_payments", default=0),
                bilateral_assistance_sum=models.Subquery(
                    BilateralAssistance.objects.filter(
                        year__gte=models.OuterRef("start_year"),
                        year__lte=models.OuterRef("end_year"),
                    )
                    # Group by replenishment start year
                    .annotate(start_year_replenishment=models.OuterRef("start_year"))
                    .values("start_year_replenishment")
                    .annotate(total=models.Sum("amount", default=0))
                    .values("total")[:1]
                ),
                promissory_notes_sum=models.Sum("promissory_notes", default=0),
                outstanding_contributions_sum=models.Sum(
                    "current_outstanding_contributions", default=0
                ),
                disputed_contributions=models.Subquery(
                    DisputedContribution.objects.filter(
                        year__gte=models.OuterRef("start_year"),
                        year__lte=models.OuterRef("end_year"),
                    )
                    # Group by replenishment start year
                    .annotate(start_year_replenishment=models.OuterRef("start_year"))
                    .values("start_year_replenishment")
                    .annotate(total=models.Sum("amount"))
                    .values("total")[:1]
                ),
                outstanding_ceit=models.Sum(
                    "current_outstanding_contributions",
                    default=0,
                    filter=Q(country__ceit_statuses__is_ceit=True)
                    & Q(country__ceit_statuses__start_year__lte=F("start_year"))
                    & (
                        Q(country__ceit_statuses__end_year__gte=F("end_year"))
                        | Q(country__ceit_statuses__end_year__isnull=True)
                    ),
                ),
            )
            .order_by("start_year")
        )

    def get_external_income_data(self):
        # ExternalIncomeAnnual can now hold triennial, annual and even quarterly data
        # This aggregator requires triennial aggregation, so we need to adjust the
        # values.
        ret = []
        triennials = list(
            TriennialContributionStatus.objects.values_list("start_year", "end_year")
            .distinct()
            .order_by("start_year")
        )
        for start_year, end_year in triennials:
            results = ExternalIncomeAnnual.objects.filter(
                models.Q(triennial_start_year=start_year)
                | (models.Q(year__gte=start_year) & models.Q(year__lte=end_year))
            ).aggregate(
                interest_earned=models.Sum("interest_earned", default=0),
                miscellaneous_income=models.Sum("miscellaneous_income", default=0),
            )
            ret.append({"start_year": start_year, "end_year": end_year, **results})
        return ret


def add_period_status_of_contributions_response_worksheet(
    wb, agg, sheet_name, sheet_period
):
    soc_qs = agg.get_status_of_contributions_qs()
    data = [
        {
            "country": country.name,
            "agreed_contributions": country.agreed_contributions,
            "cash_payments": country.cash_payments,
            "bilateral_assistance": country.bilateral_assistance,
            "promissory_notes": country.promissory_notes,
            "outstanding_contributions": country.outstanding_contributions,
        }
        for country in soc_qs
    ]
    total = agg.get_total()
    disputed_contributions_amount = agg.get_disputed_contribution_amount()

    ceit_data = agg.get_ceit_data(agg.get_ceit_countries())

    data.extend(
        [
            {
                "country": "SUB-TOTAL",
                "agreed_contributions": total["agreed_contributions"],
                "cash_payments": total["cash_payments"],
                "bilateral_assistance": total["bilateral_assistance"],
                "promissory_notes": total["promissory_notes"],
                "outstanding_contributions": total["outstanding_contributions"],
            },
            {
                "country": "Disputed contributions",
                "agreed_contributions": disputed_contributions_amount,
                "outstanding_contributions": disputed_contributions_amount,
            },
            {
                "country": "TOTAL",
                "agreed_contributions": total["agreed_contributions"]
                + disputed_contributions_amount,
                "cash_payments": total["cash_payments"],
                "bilateral_assistance": total["bilateral_assistance"],
                "promissory_notes": total["promissory_notes"],
                "outstanding_contributions": total["outstanding_contributions"]
                + disputed_contributions_amount,
            },
            {
                "country": "CEIT",
                "agreed_contributions": ceit_data["agreed_contributions"],
                "cash_payments": ceit_data["cash_payments"],
                "bilateral_assistance": ceit_data["bilateral_assistance"],
                "promissory_notes": ceit_data["promissory_notes"],
                "outstanding_contributions": ceit_data["outstanding_contributions"],
            },
        ]
    )

    ws = wb.create_sheet(sheet_name)
    configure_sheet_print(ws, "landscape")

    StatusOfContributionsWriter(ws, period=sheet_period).write(data)


def add_summary_status_of_contributions_response_worksheet(wb, agg):
    soc_qs = agg.get_status_of_contributions_qs()
    data = [
        {
            "country": country.name,
            "agreed_contributions": country.agreed_contributions,
            "cash_payments": country.cash_payments,
            "bilateral_assistance": country.bilateral_assistance,
            "promissory_notes": country.promissory_notes,
            "outstanding_contributions": country.outstanding_contributions,
            "gain_loss": country.gain_loss,
        }
        for country in soc_qs
    ]
    total = agg.get_total()
    gain_loss = agg.get_gain_loss()
    disputed_contributions_amount = agg.get_disputed_contribution_amount()
    ceit_data = agg.get_ceit_data()

    data.extend(
        [
            {
                "country": "SUB-TOTAL",
                "agreed_contributions": total["agreed_contributions"],
                "cash_payments": total["cash_payments"],
                "bilateral_assistance": total["bilateral_assistance"],
                "promissory_notes": total["promissory_notes"],
                "outstanding_contributions": total["outstanding_contributions"],
                "gain_loss": gain_loss,
            },
            {
                "country": "Disputed contributions",
                "agreed_contributions": disputed_contributions_amount,
                "outstanding_contributions": disputed_contributions_amount,
            },
            {
                "country": "TOTAL",
                "agreed_contributions": total["agreed_contributions"]
                + disputed_contributions_amount,
                "cash_payments": total["cash_payments"],
                "bilateral_assistance": total["bilateral_assistance"],
                "promissory_notes": total["promissory_notes"],
                "outstanding_contributions": total["outstanding_contributions"]
                + disputed_contributions_amount,
                "gain_loss": gain_loss,
            },
            {
                "country": "CEIT",
                "agreed_contributions": ceit_data["agreed_contributions"],
                "cash_payments": ceit_data["cash_payments"],
                "bilateral_assistance": ceit_data["bilateral_assistance"],
                "promissory_notes": ceit_data["promissory_notes"],
                "outstanding_contributions": ceit_data["outstanding_contributions"],
                "gain_loss": ceit_data["gain_loss"],
            },
        ]
    )

    current_year = datetime.now().year
    ws = wb.create_sheet(f"YR91_{str(current_year)[2:]}")
    configure_sheet_print(ws, "landscape")

    StatusOfContributionsWriter(
        ws,
        period=f"1991-{current_year}",
        extra_headers=[
            {
                "id": "gain_loss",
                "headerName": "Exchange (Gain)/Loss. NB:Negative amount = Gain",
                "column_width": 25,
            },
        ],
    ).write(data)


def add_statistics_status_of_contributions_response_worksheet(wb, periods):
    current_year = datetime.now().year
    ws = wb.create_sheet("Statistics")
    configure_sheet_print(ws, "landscape")

    headers = [
        {
            "id": "description",
            "headerName": "Description",
            "column_width": 25,
        }
    ]

    for period in periods:
        headers.append(
            {
                "id": f"{period['start_year']}-{period['end_year']}",
                "headerName": f"{period['start_year']}-{period['end_year']}",
                "column_width": 25,
            }
        )
    headers.append(
        {
            "id": "summary",
            "headerName": f"1991-{current_year}",
            "column_width": 25,
        }
    )

    agg = StatisticsStatusOfContributionsAggregator()
    soc_data = agg.get_soc_data()
    external_income_data = agg.get_external_income_data()

    columns_number = len(headers)
    last_column_letter = get_column_letter(columns_number)
    last_period_column_letter = get_column_letter(columns_number - 1)

    data = [
        {
            "description": "Pledged contributions",
            "summary": f"=SUM(B10:{last_period_column_letter}10)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": soc[
                    "agreed_contributions_sum"
                ]
                for soc in soc_data
            },
        },
        {
            "description": "Cash payments/received",
            "summary": f"=SUM(B11:{last_period_column_letter}11)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": soc["cash_payments_sum"]
                for soc in soc_data
            },
        },
        {
            "description": "Bilateral assistance",
            "summary": f"=SUM(B12:{last_period_column_letter}12)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": soc[
                    "bilateral_assistance_sum"
                ]
                for soc in soc_data
            },
        },
        {
            "description": "Promissory notes",
            "summary": f"=SUM(B13:{last_period_column_letter}13)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": soc["promissory_notes_sum"]
                for soc in soc_data
            },
        },
        {
            "description": "Total payments",
            "summary": f"=SUM(B14:{last_period_column_letter}14)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": f"=SUM({get_column_letter(i+2)}11:{get_column_letter(i+2)}13)"
                for i, soc in enumerate(soc_data)
            },
        },
        {
            "description": "Disputed contributions",
            "summary": f"=SUM(B15:{last_period_column_letter}15)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": soc["disputed_contributions"]
                for soc in soc_data
            },
        },
        {
            "description": "Outstanding pledges",
            "summary": f"=SUM(B16:{last_period_column_letter}16)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": soc[
                    "outstanding_contributions_sum"
                ]
                for soc in soc_data
            },
        },
        {
            "description": "Payments %age to pledges",
            "summary": f"={last_column_letter}14/{last_column_letter}10 * 100",
            **{
                f"{soc['start_year']}-{soc['end_year']}": (
                    f"={get_column_letter(i+2)}14" f"/{get_column_letter(i+2)}10 * 100"
                )
                for i, soc in enumerate(soc_data)
            },
        },
        {},
        {
            "description": "Interest earned",
            "summary": f"=SUM(B19:{last_period_column_letter}19)",
            **{
                f"{external_income['start_year']}-{external_income['end_year']}": external_income[
                    "interest_earned"
                ]
                for external_income in external_income_data
            },
        },
        {},
        {
            "description": "Miscellaneous income",
            "summary": f"=SUM(B21:{last_period_column_letter}21)",
            **{
                f"{external_income['start_year']}-{external_income['end_year']}": external_income[
                    "miscellaneous_income"
                ]
                for external_income in external_income_data
            },
        },
        {},
        {
            "description": "TOTAL INCOME",
            "summary": f"=SUM(B23:{last_period_column_letter}23)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": (
                    f"={get_column_letter(i+2)}14"
                    f"+{get_column_letter(i+2)}19"
                    f"+{get_column_letter(i+2)}21"
                )
                for i, soc in enumerate(soc_data)
            },
        },
        {},
        {
            "description": "Accumulated figures",
            "summary": f"1991-{current_year}",
            **{
                f"{soc['start_year']}-{soc['end_year']}": f"{soc['start_year']}-{soc['end_year']}"
                for soc in soc_data
            },
        },
        {
            "description": "Total pledges",
            "summary": f"={last_column_letter}10",
            **{
                f"{soc['start_year']}-{soc['end_year']}": f"={get_column_letter(i+2)}10"
                for i, soc in enumerate(soc_data)
            },
        },
        {
            "description": "Total payments",
            "summary": f"={last_column_letter}14",
            **{
                f"{soc['start_year']}-{soc['end_year']}": f"={get_column_letter(i+2)}14"
                for i, soc in enumerate(soc_data)
            },
        },
        {
            "description": "Payments %age to pledges",
            "summary": f"={last_column_letter}17",
            **{
                f"{soc['start_year']}-{soc['end_year']}": f"={get_column_letter(i+2)}17"
                for i, soc in enumerate(soc_data)
            },
        },
        {
            "description": "Total income",
            "summary": f"={last_column_letter}23",
            **{
                f"{soc['start_year']}-{soc['end_year']}": f"={get_column_letter(i + 2)}23"
                for i, soc in enumerate(soc_data)
            },
        },
        {
            "description": "Total outstanding contributions",
            "summary": f"={last_column_letter}26-{last_column_letter}27",
            **{
                f"{soc['start_year']}-{soc['end_year']}": f"={get_column_letter(i+2)}26-{get_column_letter(i+2)}27"
                for i, soc in enumerate(soc_data)
            },
        },
        {
            "description": "As % to total pledges",
            "summary": f"={last_column_letter}30/{last_column_letter}26 * 100",
            **{
                f"{soc['start_year']}-{soc['end_year']}": (
                    f"={get_column_letter(i+2)}30" f"/{get_column_letter(i+2)}26 * 100"
                )
                for i, soc in enumerate(soc_data)
            },
        },
        {
            "description": "Outstanding contributions for certain Countries with Economies in Transition (CEITs)",
            "summary": f"=SUM(B32:{last_period_column_letter}32)",
            **{
                f"{soc['start_year']}-{soc['end_year']}": soc["outstanding_ceit"]
                for soc in soc_data
            },
        },
        {
            "description": "CEITs' oustandings %age to pledges",
            "summary": f"={last_column_letter}32/{last_column_letter}26 * 100",
            **{
                f"{soc['start_year']}-{soc['end_year']}": (
                    f"={get_column_letter(i+2)}32" f"/{get_column_letter(i+2)}26 * 100"
                )
                for i, soc in enumerate(soc_data)
            },
        },
    ]

    StatisticsStatusOfContributionsWriter(
        ws,
        headers,
        f"1991-{current_year}",
    ).write(data)


def get_as_of_date():
    try:
        latest_payment = Payment.objects.latest("date")
    except Payment.DoesNotExist:
        latest_payment = None
    return (
        latest_payment.date
        if latest_payment and latest_payment.date
        else config.DEFAULT_REPLENISHMENT_AS_OF_DATE
    )


def get_budget_years():
    return {
        "secretariat_and_executive_committee": ExternalAllocation.objects.filter(
            staff_contracts__gt=0
        ).aggregate(max_year=models.Max("year"))["max_year"],
        "treasury_fees": ExternalAllocation.objects.filter(
            treasury_fees__gt=0
        ).aggregate(max_year=models.Max("year"))["max_year"],
        "monitoring_fees": ExternalAllocation.objects.filter(
            monitoring_fees__gt=0
        ).aggregate(max_year=models.Max("year"))["max_year"],
    }
