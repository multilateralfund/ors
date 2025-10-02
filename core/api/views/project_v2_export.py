import datetime
from functools import partial
from typing import Iterable

import io
import pathlib
import logging

import openpyxl
import docx

from docx.table import Table
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P

from django.http import FileResponse

from rest_framework import serializers

from core.api.serializers.meta_project import MetaProjecMyaDetailsSerializer
from core.api.serializers.project_v2 import ProjectDetailsV2Serializer
from core.api.serializers.business_plan import BPActivityExportSerializer

from core.models.user import User
from core.models.project import Project
from core.models.business_plan import BPActivity
from core.models.project_metadata import ProjectSpecificFields
from core.models.project_metadata import ProjectField

from core.api.export.base import configure_sheet_print, WriteOnlyBase
from core.api.export.business_plan import BPActivitiesWriter

from core.api.utils import workbook_response

logger = logging.getLogger(__name__)


def format_iso_date(isodate=None):
    if isodate:
        if isinstance(isodate, str):
            date = datetime.datetime.fromisoformat(isodate)
        elif isinstance(isodate, (datetime.date, datetime.datetime)):
            date = isodate
        else:
            return ""
        return date.strftime("%d/%m/%Y")
    return ""


def get_headers_identifiers():
    return [
        {
            "id": "country",
            "headerName": "Country",
        },
        {
            "id": "meeting",
            "headerName": "Meeting number",
        },
        {
            "id": "agency",
            "headerName": "Agency",
        },
        {
            "id": "cluster",
            "headerName": "Cluster",
            "method": lambda r, h: r[h["id"]]["name"],
            "column_width": WriteOnlyBase.COLUMN_WIDTH * 1.5,
        },
        {
            "id": "submission_status",
            "headerName": "Submission status",
        },
    ]


def get_headers_bp():
    return []


def get_headers_cross_cutting():
    return [
        {
            "id": "title",
            "headerName": "Title",
        },
        {
            "id": "description",
            "headerName": "Description",
        },
        {
            "id": "project_type",
            "headerName": "Type",
            "method": lambda r, h: r[h["id"]]["name"],
        },
        {
            "id": "sector",
            "headerName": "Sector",
            "method": lambda r, h: r[h["id"]]["name"],
        },
        {
            "id": "subsectors",
            "headerName": "Subsectors",
            "method": lambda r, h: ", ".join(ss["name"] for ss in r[h["id"]]),
            "column_width": WriteOnlyBase.COLUMN_WIDTH * 1.5,
        },
        {
            "id": "is_lvc",
            "headerName": "LVC/Non-LVC",
            "method": lambda r, h: r[h["id"]] and "LVC" or "Non-LVC",
            "column_width": WriteOnlyBase.COLUMN_WIDTH * 1.5,
        },
        {
            "id": "total_fund",
            "headerName": "Project funding",
        },
        {
            "id": "support_cost_psc",
            "headerName": "Project support cost",
        },
        {
            "id": "project_start_date",
            "headerName": "Project start date",
            "method": lambda r, h: format_iso_date(r[h["id"]]),
        },
        {
            "id": "project_end_date",
            "headerName": "Project end date",
            "method": lambda r, h: format_iso_date(r[h["id"]]),
        },
        {
            "id": "individual_consideration",
            "headerName": "Blanket consideration",
            "method": lambda r, h: r[h["id"]] and "No" or "Yes",
        },
    ]


def field_value(data, header):
    name = header["id"]
    field_data = data["field_data"]
    value = field_data.get(name, {}).get("value")
    return f"{value}" if value else "-"


def field_value_or_computed(data, header, is_date=False):
    name = header["id"]

    field_data = data["field_data"]
    computed_field_data = data["computed_field_data"]

    value = field_data.get(name, {}).get("value")

    is_computed = False

    if value is None:
        value = computed_field_data.get(name, None)
        is_computed = True

    if value and is_date:
        value = format_iso_date(value)

    value = value if value else "-"

    if is_computed:
        value = f"{value} (computed)"

    return value


def get_headers_metaproject():
    return [
        {
            "id": "project_funding",
            "headerName": "Project Funding (MYA)",
            "method": field_value_or_computed,
        },
        {
            "id": "support_cost",
            "headerName": "Support Cost (MYA)",
            "method": field_value_or_computed,
        },
        {
            "id": "start_date",
            "headerName": "Start date (MYA)",
            "method": partial(field_value_or_computed, is_date=True),
        },
        {
            "id": "end_date",
            "headerName": "End date (MYA)",
            "method": partial(field_value_or_computed, is_date=True),
        },
        {
            "id": "phase_out_odp",
            "headerName": "Phase out (ODP t) (MYA)",
            "method": field_value_or_computed,
        },
        {
            "id": "phase_out_mt",
            "headerName": "Phase out (Mt) (MYA)",
            "method": field_value_or_computed,
        },
        {
            "id": "targets",
            "headerName": "Targets",
            "method": field_value,
        },
        {
            "id": "starting_point",
            "headerName": "Starting point",
            "method": field_value,
        },
        {
            "id": "baseline",
            "headerName": "Baseline",
            "method": field_value,
        },
        {
            "id": "number_of_enterprises_assisted",
            "headerName": "Number of enterprises assisted",
            "method": field_value,
        },
        {
            "id": "number_of_enterprises",
            "headerName": "Number of enterprises",
            "method": field_value,
        },
        {
            "id": "aggregated_consumption",
            "headerName": "Aggregated consumption",
            "method": field_value,
        },
        {
            "id": "number_of_production_lines_assisted",
            "headerName": "Number of Production Lines assisted",
            "method": field_value,
        },
        {
            "id": "cost_effectiveness_kg",
            "headerName": "Cost effectiveness (US$/ Kg)",
            "method": field_value,
        },
        {
            "id": "cost_effectiveness_co2",
            "headerName": "Cost effectiveness (US$/ CO2-ep)",
            "method": field_value,
        },
    ]


def get_headers_specific_information(fields: Iterable[ProjectField]):
    result = []

    for field in fields:
        result.append({"id": field.read_field_name, "headerName": field.label})

    return result


def dict_as_obj(d):

    class Dummy:
        pass

    inst = Dummy()
    for key, value in d.items():
        if isinstance(value, dict):
            setattr(inst, key, dict_as_obj(value))
        else:
            setattr(inst, key, value)

    return inst


def get_activity_data_from_json(data):
    result = {}
    serializer = BPActivityExportSerializer()
    data_as_obj = dict_as_obj(data)
    for field, handler in serializer.get_fields().items():
        if field == "chemical_detail":
            value = "/".join(data.get("substances", []))
        elif isinstance(handler, serializers.ChoiceField):
            value = f"{data[field]}"
        elif isinstance(handler, serializers.SlugRelatedField):
            value = data[field][handler.slug_field] if data[field] else None
        elif isinstance(handler, serializers.SerializerMethodField):
            method_field = handler.method_name or f"get_{field}"
            value_getter = getattr(serializer, method_field, lambda x: x)
            value = value_getter(data_as_obj)
        else:
            value = data.get(field, None)
        result[field] = value
    return result


def get_activity_data_from_instance(data):
    result = None
    try:
        activity = BPActivity.objects.get(id=data["bp_activity"]["id"])
        result = BPActivityExportSerializer(activity).data
    except BPActivity.DoesNotExist:
        logger.warning(
            "Linked activity (%s) missing for project %s. Fallback to JSON.",
            data["bp_activity"]["id"],
            data["id"],
        )
    return result


def get_activity_data(data):
    """Serialize Activity if it exists, otherwise use the saved json."""
    result = None

    if data.get("bp_activity"):
        result = get_activity_data_from_instance(data)

    if not result:
        bp_data = data.get("bp_activity_json")
        if bp_data:
            result = get_activity_data_from_json(bp_data)

    return result


class ProjectsV2ProjectExport:
    wb: openpyxl.Workbook
    project: Project

    def __init__(self, project):
        self.project = project
        self.setup_workbook()

    def setup_workbook(self):
        wb = openpyxl.Workbook()
        # delete default sheet
        del wb[wb.sheetnames[0]]
        self.wb = wb

    def build_identifiers(self, data):
        sheet = self.add_sheet("Identifiers")
        WriteOnlyBase(sheet, get_headers_identifiers()).write([data])

    def build_bp(self, data):
        activity_data = get_activity_data(data)
        if activity_data:
            sheet = self.add_sheet("Identifiers - BP Activity")
            writer = BPActivitiesWriter(self.wb, min_year=None, max_year=None)
            # The BPActivitiesWriter creates a sheet that we don't need, replace it with our own
            del self.wb[writer.sheet.title]
            writer.sheet = sheet
            writer.write([activity_data])

    def build_cross_cutting(self, data):
        sheet = self.add_sheet("Cross-cutting")
        WriteOnlyBase(sheet, get_headers_cross_cutting()).write([data])

    def _write_project_specific_fields(
        self,
        fields_obj: ProjectSpecificFields,
        fields_section: str,
        sheet_name: str,
        data,
    ):
        fields = fields_obj.fields.filter(section__in=[fields_section])
        if fields:
            sheet = self.add_sheet(sheet_name)
            WriteOnlyBase(
                sheet,
                get_headers_specific_information(fields),
            ).write(data)

    def build_specific_information(self, data):
        project_specific_fields_obj = ProjectSpecificFields.objects.filter(
            cluster=self.project.cluster,
            type=self.project.project_type,
            sector=self.project.sector,
        ).first()

        if project_specific_fields_obj:
            self._write_project_specific_fields(
                project_specific_fields_obj,
                "Header",
                "Specific information - Overview",
                [data],
            )

            self._write_project_specific_fields(
                project_specific_fields_obj,
                "Substance Details",
                "Specific information - Substance details",
                data.get("ods_odp", []),
            )

            self._write_project_specific_fields(
                project_specific_fields_obj,
                "Impact",
                "Impact",
                [data],
            )

    def build_xls(self):
        serializer = ProjectDetailsV2Serializer(self.project)
        data = serializer.data
        self.build_identifiers(data)
        self.build_bp(data)
        self.build_cross_cutting(data)
        self.build_specific_information(data)

    def add_sheet(self, name):
        sheet = self.wb.create_sheet(name)
        configure_sheet_print(sheet, "landscape")
        return sheet

    def export_xls(self):
        self.build_xls()
        return workbook_response(f"Project {self.project.id}", self.wb)


def document_response(doc, filename):
    out = io.BytesIO()
    doc.save(out)
    out.seek(0)
    res = FileResponse(out, as_attachment=True, filename=filename)
    return res


class ProjectsV2ProjectExportDocx:
    user = User
    project: Project
    doc: docx.Document
    template_path = (
        pathlib.Path(__file__).parent.parent
        / "export"
        / "templates"
        / "Word Template for data entered into the system during project submission online.docx"
    )

    def __init__(self, project, user):
        self.user = user
        self.project = project
        with self.template_path.open("rb") as tpl:
            self.doc = docx.Document(tpl)

    def find_table(self, after_p_text=""):
        found = None

        if after_p_text:
            found_p = False
            for e in self.doc.element.body:
                if isinstance(e, CT_P) and after_p_text in e.text.strip():
                    found_p = True
                elif isinstance(e, CT_Tbl) and found_p:
                    found = Table(e, self.doc)
                    break

        return found

    def build_front_page(self, data):
        for p in self.doc.paragraphs:
            p_style = {
                "italic": False,
                "bold": False,
                "underline": False,
            }

            if p.runs:
                for k in p_style:
                    p_style[k] = getattr(p.runs[0], k, False)

            if p.text.startswith("Generated on"):
                now = datetime.datetime.utcnow().strftime("%d/%m/%Y")
                user = self.user.get_full_name() or self.user.username
                agency = self.user.agency.name if self.user.agency else ""
                p.text = f"Generated on {now} by {user}" + (
                    'of "{agency}"' if agency else ""
                )
            elif p.text.startswith("Project Title"):
                p.text = self.project.title
            elif p.text.startswith("Country:"):
                p.text = f"{p.text} {data['country']}"
            elif p.text.startswith("Agency:"):
                p.text = f"{p.text} {data['agency']}"
            elif p.text.startswith("Cluster:"):
                p.text = f"{p.text} {data['cluster']['name']}"
            elif p.text.startswith("Amount:"):
                p.text = f"{p.text} {data['total_fund']}"
            elif p.text.startswith("Project Description:"):
                p.text = f"{p.text} {data['description']}"

            if p.runs:
                for k, v in p_style.items():
                    setattr(p.runs[0], k, v)

    def _write_header_to_table(self, headers, table, data):
        for header in headers:
            row = table.add_row()
            for c_idx, cell in enumerate(row.cells):
                if c_idx == 0:
                    cell.text = header["headerName"]
                elif c_idx == 1 and header.get("method"):
                    cell.text = header["method"](data, header)
                elif c_idx == 1:
                    cell.text = str(data[header["id"]] or "")

    def _write_substance_table(self, _, table, data):
        for d in data:
            row = table.add_row()
            row_data = [
                "ods_display_name",
                "???",
                "ods_replacement",
                "phase_out_mt",
                "co2_mt",
                "odp",
            ]
            for c_idx, cell in enumerate(row.cells):
                cell.text = str(d.get(row_data[c_idx], "") or "")

    def _write_impact_target_actual(self, project, headers, table, data):
        planned_headers = {}
        actual_headers = {}
        for header in headers:
            if header["id"].endswith("_actual"):
                planned_name = header["id"].split("_actual")[0]
                actual_headers[planned_name] = header
            else:
                planned_headers[header["id"]] = header

        for field_id, header in planned_headers.items():
            row = table.add_row()
            row_data = [
                project.code,
                header["headerName"],
                data[field_id],
                data[actual_headers[field_id]["id"]],
            ]
            for c_idx, cell in enumerate(row.cells):
                cell.text = str(row_data[c_idx] or "")

    def build_related_tranches(self):
        table = self.find_table(
            "Related tranches (metacode and linked projects) Only MYA"
        )
        if table:
            related_projects = Project.objects.filter(
                meta_project__id=self.project.meta_project.id,
            )
            row = table.add_row()
            for project in related_projects:
                row_data = [
                    project.meta_project.new_code,
                    project.code,
                    project.status.name,
                ]
                for c_idx, cell in enumerate(row.cells):
                    cell.text = str(row_data[c_idx] or "")

    def build_cross_cutting(self, data):
        table = self.find_table("Cross-cutting fields")
        self._write_header_to_table(get_headers_cross_cutting(), table, data)

    def _get_fields_for_section(
        self,
        fields_obj: ProjectSpecificFields,
        section_name: str,
        **filters,
    ):
        return fields_obj.fields.filter(section__in=[section_name], **filters)

    def _write_project_specific_fields(
        self,
        table=None,
        fields=None,
        data=None,
        writer=None,
    ):
        writer = self._write_header_to_table if not writer else writer
        if data and table and fields:
            headers = get_headers_specific_information(fields)
            writer(headers, table, data)

    def _write_metaproject_fields(
        self,
        table=None,
        data=None,
    ):
        writer = self._write_header_to_table
        if data and table:
            headers = get_headers_metaproject()
            writer(headers, table, data)

    def build_specific_information(self, data):
        project_specific_fields_obj = ProjectSpecificFields.objects.filter(
            cluster=self.project.cluster,
            type=self.project.project_type,
            sector=self.project.sector,
        ).first()

        if project_specific_fields_obj:
            self._write_project_specific_fields(
                table=self.find_table("Project specific fields header"),
                fields=self._get_fields_for_section(
                    project_specific_fields_obj, "Header"
                ),
                data=data,
            )
            self._write_project_specific_fields(
                table=self.find_table("Substance details Page and tables"),
                fields=self._get_fields_for_section(
                    project_specific_fields_obj, "Substance Details"
                ),
                data=data.get("ods_odp", []),
                writer=self._write_substance_table,
            )
            self._write_project_specific_fields(
                table=self.find_table("Impact (tabular)"),
                fields=self._get_fields_for_section(
                    project_specific_fields_obj,
                    "Impact",
                    # is_actual=False,
                ),
                data=data,
            )

    def build_impact_previous_tranches(self):
        table = self.find_table("Impact (previous MYA tranches) If applicable")
        if table and self.project.tranche:
            related_projects = Project.objects.filter(
                meta_project__id=self.project.meta_project.id,
                # tranche__lt=self.project.tranche,
            )
            for project in related_projects:
                project_specific_fields_obj = ProjectSpecificFields.objects.filter(
                    cluster=project.cluster,
                    type=project.project_type,
                    sector=project.sector,
                ).first()

                fields = self._get_fields_for_section(
                    project_specific_fields_obj,
                    "Impact",
                )
                data = ProjectDetailsV2Serializer(project).data

                headers = get_headers_specific_information(fields)
                self._write_impact_target_actual(project, headers, table, data)

    def build_mya(self):
        metaproject = self.project.meta_project
        metaproject_data = {}

        if metaproject:
            metaproject_data = MetaProjecMyaDetailsSerializer(metaproject).data

        self._write_metaproject_fields(
            table=self.find_table("MYA (if applicable only new MYA)"),
            data=metaproject_data,
        )

    def build_document(self):
        serializer = ProjectDetailsV2Serializer(self.project)
        data = serializer.data

        self.build_front_page(data)
        self.build_related_tranches()
        self.build_cross_cutting(data)
        self.build_specific_information(data)
        self.build_impact_previous_tranches()
        self.build_mya()

    def export_docx(self):
        self.build_document()
        filename = self.project.code.replace("/", "_")
        return document_response(self.doc, filename=f"{filename}.docx")
