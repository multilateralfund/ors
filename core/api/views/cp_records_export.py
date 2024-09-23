import collections
import itertools
import openpyxl

from django.db.models import Prefetch
from django.db import models
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework import views
from rest_framework.exceptions import ValidationError


from core.api.export.cp_calculated_amount import CPCalculatedAmountWriter
from core.api.export.cp_data_extraction_all import (
    CPConsumptionODPWriter,
    CPDetailsExtractionWriter,
    CPHFCConsumptionMTCO2Writer,
    CPPricesExtractionWriter,
    HFC23EmissionWriter,
    HFC23GenerationWriter,
    MbrConsumptionWriter,
)
from core.api.export.cp_report_hfc_hcfc import CPReportHFCWriter, CPReportHCFCWriter
from core.api.export.cp_report_new import CPReportNewExporter
from core.api.export.cp_report_old import CPReportOldExporter
from core.api.export.cp_reports_list import CPReportListWriter
from core.api.permissions import IsSecretariat, IsViewer
from core.api.serializers import BlendSerializer
from core.api.serializers import SubstanceSerializer
from core.api.utils import workbook_pdf_response
from core.api.utils import workbook_response
from core.api.views.cp_records import CPRecordListView
from core.api.views.cp_report_empty_form import EmptyFormView
from core.api.views.utils import (
    SUBSTANCE_GROUP_ID_TO_CATEGORY,
    get_archive_reports_final_for_years,
    get_final_records_for_years,
    get_year_params_from_request,
)
from core.models import Blend
from core.models import ExcludedUsage
from core.models import Substance
from core.models.country_programme import (
    CPEmission,
    CPGeneration,
    CPPrices,
    CPRecord,
    CPReportFormatColumn,
    CPReportFormatRow,
    CPReport,
)
from core.models.country_programme_archive import (
    CPEmissionArchive,
    CPGenerationArchive,
    CPPricesArchive,
    CPRecordArchive,
)
from core.utils import IMPORT_DB_MAX_YEAR

# pylint: disable=C0302(too-many-lines)

EXCLUDE_FROM_CONSUMPTION = [
    "HBFC",
    "Other",
    "Legacy",
]


def get_record_chemical_category(record, set_hfc_preblended=True):
    """
    Returns the category of the chemical.
    If the chemical is a blend it will return hfc".

    @param set_hfc_preblended: If True, it will return "HFCs in Preblended Polyol"
        for pre-blended polyols.
    """

    if record.substance:
        substance_group = SUBSTANCE_GROUP_ID_TO_CATEGORY.get(
            record.substance.group.group_id
        )
        if not set_hfc_preblended:
            return substance_group

        # check if the substance is a pre-blended polyol
        substance_name = record.substance.name.lower()
        if "hfc" in substance_name and "pre-blended polyol" in substance_name:
            return "HFCs in Preblended Polyol"

        return substance_group

    if set_hfc_preblended and record.blend.is_related_preblended_polyol:
        return "HFCs in Preblended Polyol"

    # blends are considered as "HFC"
    return "HFC"


class CPRecordExportView(CPRecordListView):
    def get_usages(self, cp_report):
        empty_form = EmptyFormView.get_data(cp_report.year, cp_report)
        usages = empty_form.pop("usage_columns")
        return {
            **usages,
            **empty_form,
        }

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "cp_report_id",
                openapi.IN_QUERY,
                description="Country programme report id",
                type=openapi.TYPE_INTEGER,
            ),
        ],
    )
    def get(self, *args, **kwargs):
        convert_data = str(self.request.query_params.get("convert_data", None))
        cp_report = self._get_cp_report()
        if cp_report.year > IMPORT_DB_MAX_YEAR:
            exporter = CPReportNewExporter(cp_report)
        else:
            exporter = CPReportOldExporter(cp_report)

        wb = exporter.get_xlsx(
            self.get_data(cp_report), self.get_usages(cp_report), convert_data == "1"
        )
        return self.get_response(cp_report.name, wb)

    def get_response(self, name, wb):
        return workbook_response(name, wb)


class CPRecordPrintView(CPRecordExportView):
    def get_response(self, name, wb):
        return workbook_pdf_response(name, wb)


class CPEmptyExportView(CPRecordExportView):
    @swagger_auto_schema(
        required=["year"],
        manual_parameters=[
            openapi.Parameter(
                "year",
                openapi.IN_QUERY,
                description="What year to generate the empty report for",
                type=openapi.TYPE_INTEGER,
            ),
        ],
    )
    def get(self, *args, **kwargs):
        return super().get(*args, **kwargs)

    def _get_cp_report(self):
        try:
            year = int(self.request.query_params["year"])
        except (KeyError, TypeError, ValueError) as e:
            raise ValidationError({"year": "Invalid or missing parameter"}) from e
        return CPReport(year=year, name=f"Empty Country Programme {year}")

    def get_data(self, cp_report, full_history=False):
        displayed_chemicals = CPReportFormatRow.objects.get_for_year(cp_report.year)
        substances_ids = [x.substance_id for x in displayed_chemicals if x.substance_id]
        blends_ids = [x.blend_id for x in displayed_chemicals if x.blend_id]
        substances = SubstanceSerializer(
            Substance.objects.select_related("group")
            .prefetch_related(
                Prefetch(
                    "excluded_usages",
                    queryset=ExcludedUsage.objects.get_for_year(cp_report.year),
                ),
            )
            .filter(id__in=substances_ids)
            .order_by("group__name", "sort_order"),
            many=True,
            context={"with_usages": True},
        ).data
        blends = BlendSerializer(
            Blend.objects.prefetch_related(
                Prefetch(
                    "excluded_usages",
                    queryset=ExcludedUsage.objects.get_for_year(cp_report.year),
                ),
            )
            .filter(id__in=blends_ids)
            .order_by("sort_order"),
            many=True,
            context={"with_usages": True},
        ).data
        by_section = collections.defaultdict(list)

        for item in itertools.chain(substances, blends):
            for section in item["sections"]:
                item["display_name"] = item["name"]
                by_section["section_" + section.lower()].append(item)

        return {
            "cp_report": {
                "name": cp_report.name,
                "year": cp_report.year,
                "country": "XXXX",
            },
            **by_section,
        }


class CPReportListExportView(views.APIView):
    permission_classes = [IsSecretariat | IsViewer]

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "min_year",
                openapi.IN_QUERY,
                description="Minimum year",
                type=openapi.TYPE_INTEGER,
            ),
            openapi.Parameter(
                "max_year",
                openapi.IN_QUERY,
                description="Maximum year",
                type=openapi.TYPE_INTEGER,
            ),
        ],
    )
    def get(self, *args, **kwargs):
        wb = openpyxl.Workbook()
        exporter = CPReportListWriter(wb)
        min_year, max_year = get_year_params_from_request(self.request)
        data = (
            CPReport.objects.select_related("country", "created_by")
            .filter(year__gte=min_year, year__lte=max_year)
            .order_by("country__name", "created_at")
            .all()
        )
        exporter.write(data)

        # delete default sheet
        del wb[wb.sheetnames[0]]

        return workbook_response("CP Reports", wb)


class CPCalculatedAmountExportView(CPRecordListView):
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "cp_report_id",
                openapi.IN_QUERY,
                description="Country programme report id",
                type=openapi.TYPE_INTEGER,
            ),
        ],
    )
    def get(self, *args, **kwargs):
        cp_report = self._get_cp_report()

        wb = openpyxl.Workbook()
        exporter = CPCalculatedAmountWriter(wb, cp_report.year)
        data = self.get_data(cp_report)
        exporter.write(data)

        # delete default sheet
        del wb[wb.sheetnames[0]]

        return self.get_response(f"CalculatedAmount {cp_report.name}", wb)

    def get_data(self, cp_report, full_history=False):
        records = (
            CPRecord.objects.filter(country_programme_report_id=cp_report.id)
            .select_related("substance__group", "blend")
            .prefetch_related("record_usages")
            .all()
        )
        # set all consumption to 0
        data = {
            group: {"sectorial_total": 0, "consumption": 0}
            for group in SUBSTANCE_GROUP_ID_TO_CATEGORY.values()
            if group not in EXCLUDE_FROM_CONSUMPTION
        }
        data["HFCs in Preblended Polyol"] = {"sectorial_total": 0, "consumption": 0}
        data["HCFCs in Preblended Polyol"] = {"sectorial_total": 0, "consumption": 0}

        # calculate the consumption and sectorial total
        for record in records:
            # set the substance category
            substance_category = get_record_chemical_category(record)

            # set the substance category for HCFC pre-blended polyol
            substance_name = record.substance.name if record.substance else ""
            if "HCFC" in substance_name and "pre-blended polyol" in substance_name:
                substance_category = "HCFCs in Preblended Polyol"

            if substance_category in EXCLUDE_FROM_CONSUMPTION:
                continue

            # get the sectorial total
            sectorial_total = record.get_sectorial_total()
            consumption = record.get_consumption_value()

            # convert data
            if "HFC" in substance_category:
                # convert consumption value to CO₂ equivalent
                consumption *= record.get_chemical_gwp() or 0
                sectorial_total *= record.get_chemical_gwp() or 0
            else:
                # convert consumption value to ODP
                consumption *= record.get_chemical_odp() or 0
                sectorial_total *= record.get_chemical_odp() or 0

            data[substance_category]["sectorial_total"] += sectorial_total
            data[substance_category]["consumption"] += consumption

        # set the correct decimals number (for odp 2 decimals, for CO₂ 0 decimals)
        response_data = []
        for substance_category, values in data.items():
            if "HFC" in substance_category:
                values["consumption"] = round(values["consumption"], 2)
                values["sectorial_total"] = round(values["sectorial_total"], 2)
                values["unit"] = "CO₂-eq tonnes"
            else:
                values["consumption"] = round(values["consumption"], 2)
                values["sectorial_total"] = round(values["sectorial_total"], 2)
                values["unit"] = "ODP tonnes"

            substance_name = (
                substance_category if substance_category != "MBR" else "MB Non-QPS only"
            )
            response_data.append(
                {
                    "substance_name": substance_name,
                    **values,
                }
            )
        return response_data

    def get_response(self, name, wb):
        return workbook_response(name, wb)


class CPCalculatedAmountPrintView(CPCalculatedAmountExportView):
    def get_response(self, name, wb):
        return workbook_pdf_response(name, wb)


class CPHFCHCFCExportBaseView(views.APIView):
    """
    Base class for HCFC and HFC export views
    you need to implement get_usages and get methods
    - get usages method should return a list of usages for the given year
    - get method should return a response with the data for the given year
    """

    permission_classes = [IsSecretariat | IsViewer]

    def get_usages(self, year):
        raise NotImplementedError

    def get_data(self, year, section, filter_list=None):
        if not filter_list:
            filter_list = []
        filter_list.append(models.Q(section=section))

        return get_final_records_for_years(year, year, filter_list)

    def get_response(self, name, wb):
        return workbook_response(name, wb)

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "min_year",
                openapi.IN_QUERY,
                description="Minimum year",
                type=openapi.TYPE_INTEGER,
            ),
            openapi.Parameter(
                "max_year",
                openapi.IN_QUERY,
                description="Maximum year",
                type=openapi.TYPE_INTEGER,
            ),
        ],
    )
    def get(self, *args, **kwargs):
        raise NotImplementedError


class CPHCFCExportView(CPHFCHCFCExportBaseView):
    def get_usages(self, year):
        cp_report_formats = (
            CPReportFormatColumn.objects.get_for_year(year)
            .filter(section="A")
            .exclude(
                header_name__in=["Refrigeration", "Methyl bromide", "QPS", "Non-QPS"]
            )
            .select_related("usage")
            .order_by("sort_order")
        )
        usages = []
        for us_format in cp_report_formats:
            usages.append(
                {
                    "id": str(us_format.usage_id),
                    "headerName": (
                        us_format.usage.full_name
                        if "solvent" not in us_format.usage.full_name.lower()
                        else "Solvent"
                    ),
                    "columnCategory": "usage",
                    "convert_to_odp": True,
                    "align": "right",
                }
            )

        return usages

    def get_data(self, year, section, filter_list=None):
        if not filter_list:
            filter_list = []
        filter_list.append(models.Q(substance__name__icontains="HCFC"))
        return super().get_data(year, section, filter_list)

    def get(self, *args, **kwargs):
        min_year, max_year = get_year_params_from_request(self.request)
        wb = openpyxl.Workbook()

        for year in range(int(min_year), int(max_year) + 1):
            usages = self.get_usages(year)
            exporter = CPReportHCFCWriter(wb, usages, year)
            statistics = self.get_data(year, "A")
            exporter.write(statistics)

        # delete default sheet
        del wb[wb.sheetnames[0]]

        return self.get_response("CP Data Extraction-HCFC", wb)


class CPHFCExportView(CPHFCHCFCExportBaseView):
    def get_usages(self, year):
        cp_report_formats = (
            CPReportFormatColumn.objects.get_for_year(year)
            .filter(section="B")
            .select_related("usage")
            .order_by("sort_order")
        )
        usages = {
            "(MT)": [],
            "(CO₂-eq tonnes)": [],
        }
        for q_type in ["(MT)", "(CO₂-eq tonnes)"]:
            for us_format in cp_report_formats:
                usages[q_type].append(
                    {
                        "id": f"{us_format.usage_id} {q_type}",
                        "headerName": f"{us_format.usage.full_name} {q_type}",
                        "columnCategory": "usage",
                        "quantity_type": q_type,
                        "align": "right",
                    }
                )

        return usages["(MT)"], usages["(CO₂-eq tonnes)"]

    def get(self, *args, **kwargs):
        min_year, max_year = get_year_params_from_request(self.request)
        wb = openpyxl.Workbook()

        for year in range(int(min_year), int(max_year) + 1):
            usages_mt, usages_co2 = self.get_usages(year)
            exporter = CPReportHFCWriter(wb, usages_mt, usages_co2, year)
            statistics = self.get_data(year, "B")
            exporter.write(statistics)

        # delete default sheet
        del wb[wb.sheetnames[0]]

        return self.get_response("CP Data Extraction-HFC", wb)


class CPDataExtractionAllExport(views.APIView):
    permission_classes = [IsSecretariat | IsViewer]

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "min_year",
                openapi.IN_QUERY,
                description="Minimum year",
                type=openapi.TYPE_INTEGER,
            ),
            openapi.Parameter(
                "max_year",
                openapi.IN_QUERY,
                description="Maximum year",
                type=openapi.TYPE_INTEGER,
            ),
        ],
    )
    def get(self, *args, **kwargs):
        min_year, max_year = get_year_params_from_request(self.request)

        wb = openpyxl.Workbook()
        archive_reports = get_archive_reports_final_for_years(min_year, max_year)
        using_consumption_value_set = self.get_consumption_set(min_year, max_year)
        existent_reports = self.get_existent_reports(min_year, max_year)

        # ODS Price
        exporter = CPPricesExtractionWriter(wb, min_year, max_year)
        data = self.get_prices(min_year, max_year, archive_reports)
        exporter.write(data)

        # CP Details
        exporter = CPDetailsExtractionWriter(wb, min_year, max_year)
        data = self.get_cp_details(
            min_year, max_year, using_consumption_value_set, existent_reports
        )
        exporter.write(data)

        # CPConsumption(ODP)
        exporter = CPConsumptionODPWriter(wb, min_year, max_year)
        data = self._get_cp_consumption_data(
            min_year, max_year, using_consumption_value_set, existent_reports
        )
        exporter.write(data)

        # HFC-Consumption(MTvsCO₂)
        exporter = CPHFCConsumptionMTCO2Writer(wb, min_year, max_year)
        data = self._get_hfc_consumption_data(
            min_year, max_year, using_consumption_value_set, existent_reports
        )
        exporter.write(data)

        # HFC-23Generation
        exporter = HFC23GenerationWriter(wb)
        data = self._get_generations(min_year, max_year, archive_reports)
        exporter.write(data)

        # HFC23Emission
        exporter = HFC23EmissionWriter(wb)
        data = self._get_emissions(min_year, max_year, archive_reports)
        exporter.write(data)

        # MbrConsumption
        exporter = MbrConsumptionWriter(wb, min_year, max_year)
        data = self.get_mbr_consumption_data(min_year, max_year, archive_reports)
        exporter.write(data)

        # delete default sheet
        del wb[wb.sheetnames[0]]

        return workbook_response("CP Data Extraction-All", wb)

    def get_existent_reports(self, min_year, max_year):
        reports = (
            CPReport.objects.filter(year__gte=min_year, year__lte=max_year)
            .select_related("country")
            .all()
        )
        existent_reports = {}
        for report in reports:
            if report.country.name not in existent_reports:
                existent_reports[report.country.name] = []
            existent_reports[report.country.name].append(report.year)

        return existent_reports

    def get_consumption_set(self, min_year, max_year):
        """
        Get the set of country,year,section pairs for which
            the consumption value should be calculated
        For methyl bromide, the consumption value should be calculated using the sectorial total
        """
        records = get_final_records_for_years(min_year, max_year)
        return {
            (
                record.country_programme_report.country.name,
                record.country_programme_report.year,
                record.section,
            )
            for record in records
            if any([record.imports, record.exports, record.production])
        }

    def get_mbr_consumption_data(self, min_year, max_year, archive_reports):
        final_records = (
            CPRecord.objects.get_for_years(min_year, max_year)
            .filter(
                country_programme_report__status=CPReport.CPReportStatus.FINAL,
                substance__name__iexact="Methyl Bromide",
            )
            .annotate(
                country_name=models.F("country_programme_report__country__name"),
                year=models.F("country_programme_report__year"),
                methyl_bromide_qps=models.Sum(
                    "record_usages__quantity",
                    filter=models.Q(record_usages__usage__name__iexact="QPS"),
                    default=0,
                ),
                methyl_bromide_non_qps=models.Sum(
                    "record_usages__quantity",
                    filter=models.Q(record_usages__usage__name__iexact="Non-QPS"),
                    default=0,
                ),
                total=models.F("methyl_bromide_qps")
                + models.F("methyl_bromide_non_qps"),
            )
        ).values(
            "country_name",
            "year",
            "methyl_bromide_qps",
            "methyl_bromide_non_qps",
            "total",
        )

        if not archive_reports:
            mbr_list = list(final_records)
        else:
            archive_records = (
                CPRecordArchive.objects.get_for_years(min_year, max_year)
                .filter(
                    country_programme_report__status=CPReport.CPReportStatus.FINAL,
                    substance__name__iexact="Methyl Bromide",
                )
                .filter(
                    # get the records for the max version of the archive reports
                    *[
                        models.Q(
                            country_programme_report__country_id=c,
                            country_programme_report__year=y,
                            country_programme_report__version=v,
                        )
                        for c, y, v in archive_reports
                    ],
                    _connector=models.Q.OR,
                )
                .annotate(
                    country_name=models.F("country_programme_report__country__name"),
                    year=models.F("country_programme_report__year"),
                    methyl_bromide_qps=models.Sum(
                        "record_usages__quantity",
                        filter=models.Q(record_usages__usage__name__iexact="QPS"),
                        default=0,
                    ),
                    methyl_bromide_non_qps=models.Sum(
                        "record_usages__quantity",
                        filter=models.Q(record_usages__usage__name__iexact="Non-QPS"),
                        default=0,
                    ),
                    total=models.F("methyl_bromide_qps")
                    + models.F("methyl_bromide_non_qps"),
                )
            ).values(
                "country_name",
                "year",
                "methyl_bromide_qps",
                "methyl_bromide_non_qps",
                "total",
            )

            mbr_list = list(final_records) + list(archive_records)
        mbr_data = {}
        for mbr in mbr_list:
            if mbr["country_name"] not in mbr_data:
                mbr_data[mbr["country_name"]] = {
                    "country_name": mbr["country_name"],
                }
            mbr_data[mbr["country_name"]].update(
                {
                    f"methyl_bromide_qps_{mbr['year']}": mbr["methyl_bromide_qps"],
                    f"methyl_bromide_non_qps_{mbr['year']}": mbr[
                        "methyl_bromide_non_qps"
                    ],
                    f"total_{mbr['year']}": mbr["total"],
                }
            )
        mbr_data = dict(sorted(mbr_data.items(), key=lambda x: x[0]))
        return mbr_data.values()

    def get_prices(self, min_year, max_year, archive_reports):
        final_prices = (
            CPPrices.objects.select_related(
                "blend",
                "substance",
                "country_programme_report__country",
            )
            .prefetch_related("blend__components")
            .filter(
                country_programme_report__status=CPReport.CPReportStatus.FINAL,
                country_programme_report__year__gte=min_year,
                country_programme_report__year__lte=max_year,
            )
            .filter(
                models.Q(current_year_price__isnull=False)
                | models.Q(previous_year_price__isnull=False)
            )
            .order_by(
                "country_programme_report__year",
                "country_programme_report__country__name",
                "substance__sort_order",
                "blend__sort_order",
            )
            .all()
        )

        if not archive_reports:
            archive_prices = []
        else:
            archive_prices = (
                CPPricesArchive.objects.select_related(
                    "blend",
                    "substance",
                    "country_programme_report__country",
                )
                .prefetch_related("blend__components")
                .filter(
                    models.Q(current_year_price__isnull=False)
                    | models.Q(previous_year_price__isnull=False)
                )
                .filter(
                    # get the records for the max version of the archive reports
                    *[
                        models.Q(
                            country_programme_report__country_id=c,
                            country_programme_report__year=y,
                            country_programme_report__version=v,
                        )
                        for c, y, v in archive_reports
                    ],
                    _connector=models.Q.OR,
                )
                .order_by(
                    "country_programme_report__year",
                    "country_programme_report__country__name",
                    "substance__sort_order",
                    "blend__sort_order",
                )
                .all()
            )
        final_prices_dict = {}
        all_prices = list(final_prices) + list(archive_prices)
        for price in all_prices:
            key = (
                price.country_programme_report.country.name,
                price.get_chemical_display_name(),
            )
            year = price.country_programme_report.year
            if key not in final_prices_dict:
                final_prices_dict[key] = {}

            if not price.is_fob and not price.is_retail and year < 2023:
                is_fob = ""
                is_retail = ""
            else:
                is_fob = "Yes" if price.is_fob else "No"
                is_retail = "Yes" if price.is_retail else "No"

            final_prices_dict[key].update(
                {
                    f"price_{year}": price.current_year_price,
                    f"fob_{year}": is_fob,
                    f"retail_{year}": is_retail,
                    f"remarks_{year}": price.remarks,
                }
            )
        return dict(sorted(final_prices_dict.items(), key=lambda x: x[0]))

    def get_cp_details(
        self, min_year, max_year, using_consumption_value_set, existent_reports
    ):
        """
        Get CP details for the given year
        @param min_year: int
        @param max_year: int
        @param using_consumption_value_set: set
        @param existent_reports: dict
        @return: dict
        structure:
        {
            (country_name, chemical_name): {
                "substance_group": group_id,
                "substance_odp": value,
                "substance_gwp": value,
                "record_value_<year>": value,
            },
            ...
        }

        """
        records = get_final_records_for_years(min_year, max_year)
        cp_details = {}

        for record in records:
            country_name = record.country_programme_report.country.name
            year = record.country_programme_report.year
            chemical_name = record.get_chemical_display_name()

            # set consumption value
            # If Import, Export and Production are not provided for any substance in this report
            # it should be the TOTAL of Use by Sector
            cons_value = record.get_consumption_value(
                (country_name, year, record.section) in using_consumption_value_set
            )

            key = (country_name, chemical_name)

            if key not in cp_details:
                # initialize the row with default values
                cp_details[key] = {
                    "substance_group": (
                        record.substance.group.group_id if record.substance else "F"
                    ),
                    "substance_odp": record.get_chemical_odp(),
                    "substance_gwp": record.get_chemical_gwp(),
                }
                for data_year in existent_reports.get(country_name, []):
                    cp_details[key][f"record_value_{data_year}"] = 0

            cp_details[key][f"record_value_{year}"] = cons_value

        return dict(sorted(cp_details.items(), key=lambda x: x[0]))

    def _get_cp_consumption_data(
        self, min_year, max_year, using_consumption_value_set, existent_reports
    ):
        """
        Get CP consumption data for the given year

        @param year: int

        @return: dict
        structure:
        {
            (country_name, substance_group): {
                "record_value_<year>" : value
                ...
            }
            ...
        }
        """
        filters = [
            models.Q(substance__isnull=False),
            models.Q(section="A"),
        ]
        records = get_final_records_for_years(min_year, max_year, filters)
        country_records = {}
        for record in records:
            # set the group
            if (
                record.substance.name.lower()
                == "hcfc-141b in imported pre-blended polyol"
            ):
                group = "HCFC-141b Preblended Polyol"
            else:
                group = SUBSTANCE_GROUP_ID_TO_CATEGORY.get(
                    record.substance.group.group_id
                )
            if not group:
                continue

            country_name = record.country_programme_report.country.name
            year = record.country_programme_report.year

            key = (country_name, group)
            if key not in country_records:
                country_records[key] = {}
                for data_year in existent_reports.get(country_name, []):
                    country_records[key][f"record_value_{data_year}"] = 0

            # set consumption value
            consumption_value = record.get_consumption_value(
                (country_name, year, record.section) in using_consumption_value_set
            )
            consumption_value *= record.substance.odp

            country_records[key][f"record_value_{year}"] += consumption_value

        return dict(sorted(country_records.items(), key=lambda x: x[0]))

    def _get_hfc_consumption_data(
        self, min_year, max_year, using_consumption_value_set, existent_reports
    ):
        """
        Get HFC consumption data for the given year

        @param year: int
        @return: dict
        structure:
        {
            (country_name, substance_name) {
                    "substance_group": value,
                    "country_lvc": value,
                    "consumption_mt_<year>": value,
                    "consumption_co2_<year>": value,
                    "servicing_<year>": value,
                    "usages_total_<year>": value,
                },
                ...
            }
            ...
        }
        """
        records = get_final_records_for_years(
            min_year, max_year, [models.Q(section="B")]
        )
        country_records = {}
        for record in records:
            substance_name = get_record_chemical_category(record)
            if substance_name == "legacy" or (
                record.substance and "HFC" not in record.substance.name
            ):
                continue

            country = record.country_programme_report.country
            country_name = country.name
            year = record.country_programme_report.year

            key = (country_name, substance_name)
            if key not in country_records:
                country_records[key] = {
                    "country_lvc": "LVC" if country.is_lvc else "Non-LVC",
                    "substance_group": country.consumption_group,
                }
                for data_year in existent_reports.get(country_name, []):
                    country_records[key][f"consumption_mt_{data_year}"] = 0
                    country_records[key][f"consumption_co2_{data_year}"] = 0
                    country_records[key][f"servicing_{data_year}"] = 0
                    country_records[key][f"usages_total_{data_year}"] = 0

            # get consumption data
            consumption_value = record.get_consumption_value(
                (country_name, year, record.section) in using_consumption_value_set
            )
            country_records[key][f"consumption_mt_{year}"] += consumption_value

            # convert consumption value to CO₂ equivalent
            country_records[key][f"consumption_co2_{year}"] += (
                consumption_value * record.get_chemical_gwp()
            )

            for rec_us in record.record_usages.all():
                if "servicing" in rec_us.usage.full_name.lower():
                    country_records[key][f"servicing_{year}"] += rec_us.quantity
                country_records[key][f"usages_total_{year}"] += rec_us.quantity

        return dict(sorted(country_records.items(), key=lambda x: x[0]))

    def _get_generations(self, min_year, max_year, archive_reports):
        final_generations = (
            CPGeneration.objects.filter(
                country_programme_report__status=CPReport.CPReportStatus.FINAL,
                country_programme_report__year__gte=min_year,
                country_programme_report__year__lte=max_year,
            )
            .select_related("country_programme_report__country")
            .order_by(
                "country_programme_report__year",
                "country_programme_report__country__name",
            )
            .all()
        )

        if not archive_reports:
            return list(final_generations)

        archive_generations = (
            CPGenerationArchive.objects.filter(
                country_programme_report__status=CPReport.CPReportStatus.FINAL,
                country_programme_report__year__gte=min_year,
                country_programme_report__year__lte=max_year,
            )
            .filter(
                # get the records for the max version of the archive reports
                *[
                    models.Q(
                        country_programme_report__country_id=c,
                        country_programme_report__year=y,
                        country_programme_report__version=v,
                    )
                    for c, y, v in archive_reports
                ],
                _connector=models.Q.OR,
            )
            .select_related("country_programme_report__country")
            .order_by(
                "country_programme_report__year",
                "country_programme_report__country__name",
            )
            .all()
        )

        return list(final_generations) + list(archive_generations)

    def _get_emissions(self, min_year, max_year, archive_reports):
        final_emissions = (
            CPEmission.objects.filter(
                country_programme_report__status=CPReport.CPReportStatus.FINAL,
                country_programme_report__year__gte=min_year,
                country_programme_report__year__lte=max_year,
            )
            .select_related("country_programme_report__country")
            .order_by(
                "country_programme_report__year",
                "country_programme_report__country__name",
            )
            .all()
        )

        if not archive_reports:
            return list(final_emissions)

        archive_emissions = (
            CPEmissionArchive.objects.filter(
                country_programme_report__status=CPReport.CPReportStatus.FINAL,
                country_programme_report__year__gte=min_year,
                country_programme_report__year__lte=max_year,
            )
            .filter(
                # get the records for the max version of the archive reports
                *[
                    models.Q(
                        country_programme_report__country_id=c,
                        country_programme_report__year=y,
                        country_programme_report__version=v,
                    )
                    for c, y, v in archive_reports
                ],
                _connector=models.Q.OR,
            )
            .select_related("country_programme_report__country")
            .all()
        )

        return list(final_emissions) + list(archive_emissions)
