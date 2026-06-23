import pathlib
from copy import copy
from dataclasses import dataclass
from typing import Iterable
from typing import TypedDict

import openpyxl
from django.conf import settings
from django.db.models import F
from django.db.models import FloatField
from django.db.models import IntegerField
from django.db.models import QuerySet
from django.db.models import Sum
from django.db.models import TextField
from django.db.models import Value
from django.db.models.functions import Cast
from django.db.models.functions import Coalesce
from django.db.models.functions import Concat
from django.db.models.functions import NullIf
from django.db.models.functions import Round
from django.db.models.functions import Upper
from django.http import JsonResponse
from django.utils.html import strip_tags
from rest_framework import mixins
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from core.api.filters.blanket_approval_details import BlanketApprovalDetailsFilter
from core.api.permissions import HasProjectV2ViewAccess
from core.api.utils import workbook_response
from core.models import Project


def get_available_values(queryset: QuerySet[Project], field_name: str):
    rel_name = f"{field_name}__name"

    values = (
        queryset.order_by(rel_name).values_list(f"{field_name}_id", rel_name).distinct()
    )

    return [{"name": name, "id": pk} for pk, name in values if pk is not None]


class ProjectData(TypedDict):
    project_id: int
    project_title: str
    project_description: str
    agency_name: str
    country_pk: int
    country_name: str
    cluster_pk: int
    cluster_name: str
    project_type_pk: int
    project_type_name: str
    hcfc: float
    hfc: float
    project_funding: float
    project_support_cost: float
    total: float


class CountryData(TypedDict):
    cluster_id: int
    cluster_name: str
    project_type_id: int
    project_type_name: str
    projects: list[ProjectData]


class CountryTotal(TypedDict):
    hcfc: float
    hfc: float
    project_funding: float
    project_support_cost: float
    total: float


class CountryEntry(TypedDict):
    country_id: int
    country_name: str
    country_data: CountryData
    country_total: CountryTotal


@dataclass
class MergedValue:
    value: float | int | str
    size: int


class SheetWriter:
    def __init__(self, wb, last_row):
        self.wb = wb
        self.sheet = wb.active
        self.last_row = last_row

    def add_row(self, row, styles_from):
        cells = []

        to_merge = []

        idx = 0
        for val in row:
            src_cell = styles_from[idx]
            cell = copy(src_cell)
            for attr in ["font", "number_format", "alignment", "border"]:
                setattr(cell, attr, copy(getattr(src_cell, attr)))
            if idx == 0:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                cell.alignment = alignment
            if isinstance(val, MergedValue):
                to_merge.append(
                    (
                        self.last_row + 1,
                        idx + 1,
                        self.last_row + 1,
                        idx + 1 + val.size,
                    )
                )
                merged_value = val.value
                cell.value = None if merged_value == 0 else merged_value
                cells.append(cell)

                # Add placeholders for the extra merged columns
                for _ in range(val.size):
                    cells.append(None)

                idx += val.size + 1
            else:
                cell.value = None if val == 0 else val
                cells.append(cell)
                idx += 1

        self.sheet.append(cells)
        self.last_row += 1

        for start_row, start_column, end_row, end_column in to_merge:
            self.sheet.merge_cells(
                start_row=start_row,
                start_column=start_column,
                end_row=end_row,
                end_column=end_column,
            )


class BlanketApprovalDetailsViewset(
    viewsets.GenericViewSet,
    mixins.ListModelMixin,
):
    """ViewSet for blanket approval details."""

    filterset_class = BlanketApprovalDetailsFilter
    permission_classes = (HasProjectV2ViewAccess,)

    _template_path = (
        pathlib.Path(__file__).parent.parent
        / "export"
        / "templates"
        / "blanket_approval_details_template.xlsx"
    )

    def get_queryset(self):
        queryset = (
            Project.objects.really_all().filter(
                submission_status__name__in=[
                    "Recommended",
                    "Approved",
                ],
            )
            # Avoid duplication caused by project being in Approval (draft)
            # as it's submission_status is still Recommended until being approved.
            .exclude(submission_status__name="Recommended", version__gte=3)
        )

        # Requested in #35434.
        if self.request.user.has_perm("core.is_mlfs_user"):
            queryset = queryset.exclude(submission_status__name="Draft")

        return queryset

    def _extract_data(self):
        queryset: QuerySet[Project] = self.filter_queryset(self.get_queryset())

        per_country = {}
        total_projects = 0
        grand_total = {
            "hcfc": 0.0,
            "hfc": 0.0,
            "project_funding": 0.0,
            "project_support_cost": 0.0,
            "total": 0.0,
        }

        filtered_projects: Iterable[ProjectData] = queryset.values(  # type: ignore[assignment]
            project_id=F("id"),
            project_title=F("title"),
            project_description=Coalesce(
                NullIf(F("excom_provision"), Value("")),
                NullIf(
                    Concat(
                        F("decision__number"),
                        Value(": "),
                        F("decision__text"),
                        output_field=TextField(),
                    ),
                    Value(": "),
                ),
                output_field=TextField(),
            ),
            agency_name=F("agency__name"),
            country_pk=F("country"),
            country_name=Upper(F("country__name")),
            cluster_pk=F("cluster"),
            cluster_name=Upper(F("cluster__name")),
            project_type_pk=F("project_type"),
            project_type_name=F("project_type__name"),
            hcfc=Round(
                Coalesce(Sum("ods_odp__odp"), 0.0, output_field=FloatField()),
                precision=1,
            ),
            hfc=Cast(
                Round(
                    Coalesce(
                        Sum("ods_odp__co2_mt"),
                        0.0,
                        output_field=FloatField(),
                    )
                    / Value(1000.0),
                    precision=0,
                ),
                output_field=IntegerField(),
            ),
            project_funding=Coalesce(F("total_fund"), 0.0),
            project_support_cost=Coalesce(F("support_cost_psc"), 0.0),
            total=Coalesce(F("total_fund") + F("support_cost_psc"), 0.0),
        ).order_by(
            "country_name",
            "cluster_name",
            "project_type_name",
            "project_title",
            "project_id",
        )

        for project in filtered_projects:
            key = f"{project['project_type_pk']}, {project['cluster_pk']}"

            project["project_description"] = (
                strip_tags(project["project_description"])
                if project["project_description"]
                else ""
            )

            per_country.setdefault(
                project["country_pk"],
                {
                    "country_name": project["country_name"],
                    "country_id": project["country_pk"],
                },
            )
            per_country[project["country_pk"]].setdefault(
                key,
                {
                    "cluster_id": project["cluster_pk"],
                    "cluster_name": project["cluster_name"],
                    "project_type_id": project["project_type_pk"],
                    "project_type_name": project["project_type_name"],
                    "projects": [],
                },
            )

            result = per_country[project["country_pk"]][key]
            result["projects"].append(project)
            total_projects += 1

        result = []

        for country_data in per_country.values():
            country_id = country_data.pop("country_id")
            country_name = country_data.pop("country_name")
            country_total = {
                "hcfc": 0.0,
                "hfc": 0.0,
                "project_funding": 0.0,
                "project_support_cost": 0.0,
                "total": 0.0,
            }

            country_data = list(country_data.values())
            for clusters in country_data:
                for project in clusters["projects"]:
                    for key in country_total:
                        country_total[key] += project[key]

            result.append(
                {
                    "country_id": country_id,
                    "country_name": country_name,
                    "country_data": country_data,
                    "country_total": country_total,
                }
            )
            for key, v in country_total.items():
                grand_total[key] += v

        return total_projects, grand_total, result

    def list(self, request, *args, **kwargs):
        total_projects, grand_total, result = self._extract_data()
        return JsonResponse(
            {
                "total_projects": total_projects,
                "grand_total": grand_total,
                "result": result,
            }
        )

    @action(methods=["GET"], detail=False)
    def filters(self, request, *args, **kwargs):
        queryset: QuerySet[Project] = self.filter_queryset(self.get_queryset())

        result = {
            "country": get_available_values(queryset, "country"),
            "cluster": get_available_values(queryset, "cluster"),
            "project_type": get_available_values(queryset, "project_type"),
        }

        return Response(result)

    def _export_debug(self):
        result = []

        data = self._extract_data()
        result.append(
            {
                "result": data,
            }
        )

        return JsonResponse({"DEBUG": result})

    def _export_wb(self):
        wb = openpyxl.open(self._template_path)
        sheet = wb.active

        _, grand_total, data = self._extract_data()
        data: list[CountryEntry] = data

        i = 4
        row_country_name = sheet[i + 1]
        row_country_cluster = sheet[i + 2]
        row_country_type = sheet[i + 3]
        row_project_title = sheet[i + 4]
        row_project_description = sheet[i + 5]
        row_empty = sheet[i + 6]
        row_country_total = sheet[i + 7]

        # Remove the placeholder rows and their template merge before appending.
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.min_row >= 5 and merged_range.max_row <= 11:
                sheet.unmerge_cells(str(merged_range))

        sheet.delete_rows(5, 7)

        sw = SheetWriter(wb, sheet.max_row)

        for country in data:
            sw.add_row([country["country_name"]], row_country_name)
            for country_data in country["country_data"]:
                sw.add_row([country_data["cluster_name"]], row_country_cluster)
                sw.add_row([country_data["project_type_name"]], row_country_type)
                for project in country_data["projects"]:
                    sw.add_row(
                        [
                            project["project_title"],
                            project["agency_name"],
                            project["hcfc"],
                            project["hfc"],
                            project["project_funding"],
                            project["project_support_cost"],
                            project["total"],
                        ],
                        row_project_title,
                    )
                    sw.add_row(
                        [project["project_description"]], row_project_description
                    )
                    sw.add_row([""] * 7, row_empty)
            country_total: CountryTotal = country["country_total"]
            sw.add_row(
                [
                    MergedValue(
                        f"Total for {country['country_name'].title()}", size=+1
                    ),
                    country_total["hcfc"],
                    country_total["hfc"],
                    country_total["project_funding"],
                    country_total["project_support_cost"],
                    country_total["total"],
                ],
                row_country_total,
            )

        sw.add_row(
            [
                MergedValue("Grand total", size=+1),
                grand_total["hcfc"],
                grand_total["hfc"],
                grand_total["project_funding"],
                grand_total["project_support_cost"],
                grand_total["total"],
            ],
            row_country_total,
        )

        return workbook_response("List of projects and activities", wb)

    @action(methods=["GET"], detail=False)
    def export(self, request, *args, **kwargs):
        is_debug_request = settings.DEBUG and request.query_params.get("debug")

        if is_debug_request:
            return self._export_debug()

        return self._export_wb()
