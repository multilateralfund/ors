import typing
from functools import partial
from itertools import chain

import openpyxl
from dataclasses import dataclass
from dataclasses import field

from django.db.models import Model
from rest_framework.viewsets import GenericViewSet

from core.models.country import Country
from core.models.agency import Agency
from core.models.project import Project
from core.models.utils import SubstancesType
from core.models.substance import Substance
from core.models.blend import Blend

from core.models.project_metadata import ProjectCluster
from core.models.project_metadata import ProjectType
from core.models.project_metadata import ProjectSector
from core.models.project_metadata import ProjectSubSector
from core.models.project_metadata import ProjectStatus

from core.api.export.base import configure_sheet_print
from core.api.export.projects import ProjectWriter
from core.api.export.projects_v2 import ProjectV2Writer
from core.api.serializers.project import ProjectExportSerializer
from core.api.serializers.project_v2 import ProjectExportV2Serializer
from core.api.utils import workbook_response, workbook_pdf_response

from core.api.export.business_plan import ModelNameCodeWriter
from core.api.export.business_plan import ModelNameWriter


class ProjectsExport:
    view: GenericViewSet

    def __init__(self, view: GenericViewSet):
        self.view = view

    def export_xls(self):
        return self.get_wb(workbook_response)

    def export_pdf(self):
        return self.get_wb(workbook_pdf_response)

    def get_wb(self, method):
        queryset = self.view.filter_queryset(self.view.get_queryset())

        data = ProjectExportSerializer(queryset, many=True).data

        wb = openpyxl.Workbook(write_only=True)
        sheet = wb.create_sheet("Projects")
        configure_sheet_print(sheet, "landscape")

        ProjectWriter(sheet).write(data)

        name = "Projects"
        return method(name, wb)


@dataclass
class SheetDefinition:
    model: typing.Any | None
    data_getter: typing.Callable
    writer: type[ModelNameWriter | ModelNameCodeWriter]
    sheet_name: str
    validate_column: str
    enforce_validation: bool = True

    def get_data(self):
        return self.data_getter(self.model)

    def write_data(self, exporter: "ProjectsV2Export"):
        data = self.get_data()
        self.writer(exporter.wb, self.sheet_name).write(data)
        exporter.add_data_validation(
            self.validate_column,
            self.sheet_name,
            len(data),
            show_error=self.enforce_validation,
        )


@dataclass
class MultiModelSheetDefinition(SheetDefinition):
    models: typing.List[type[Model]] = field(kw_only=True, default_factory=list)

    def get_data(self):
        return list(chain(*(self.data_getter(model) for model in self.models)))


class ProjectsV2Export(ProjectsExport):
    wb: openpyxl.Workbook
    sheet: openpyxl.worksheet.worksheet.Worksheet

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setup_workbook()

    def setup_workbook(self):
        wb = openpyxl.Workbook()
        # delete default sheet
        del wb[wb.sheetnames[0]]
        sheet = wb.create_sheet("Projects")
        configure_sheet_print(sheet, "landscape")
        self.wb = wb
        self.sheet = sheet

    def get_name_and_codes(self, cls_name, code_name):
        queryset = cls_name.objects.values_list("name", code_name).order_by("name")
        return [{"name": name, "acronym": acronym} for name, acronym in queryset]

    def get_by_prop_name(self, cls_name, prop_name, filter_obsoletes=False):
        queryset = cls_name.objects.all()
        if filter_obsoletes:
            queryset = queryset.filter(obsolete=False)
        queryset = (
            queryset.values_list(prop_name, flat=True).distinct().order_by(prop_name)
        )
        return [{"name": name} for name in queryset]

    def get_names(self, cls_name, filter_obsoletes=False):
        return self.get_by_prop_name(
            cls_name, "name", filter_obsoletes=filter_obsoletes
        )

    def get_codes(self, cls_name):
        return self.get_by_prop_name(cls_name, "code")

    def get_enum_labels(self, enum):
        return [{"name": e.label} for e in enum]

    def get_wb(self, method):
        queryset = self.view.filter_queryset(self.view.get_queryset())
        data = ProjectExportV2Serializer(queryset, many=True).data
        ProjectV2Writer(self.sheet).write(data)

        data_sheets = [
            SheetDefinition(
                Project,
                self.get_codes,
                ModelNameWriter,
                "Codes",
                "A",
                enforce_validation=False,
            ),
            SheetDefinition(
                Project,
                partial(self.get_by_prop_name, prop_name="legacy_code"),
                ModelNameWriter,
                "Legacy codes",
                "B",
                enforce_validation=False,
            ),
            SheetDefinition(
                Project,
                partial(self.get_by_prop_name, prop_name="metacode"),
                ModelNameWriter,
                "Metacodes",
                "C",
                enforce_validation=False,
            ),
            SheetDefinition(
                ProjectCluster,
                partial(self.get_names, filter_obsoletes=True),
                ModelNameWriter,
                "Clusters",
                "D",
            ),
            SheetDefinition(
                Project.Category,
                self.get_enum_labels,
                ModelNameWriter,
                "Project categories",
                "E",
            ),
            SheetDefinition(
                ProjectType,
                partial(self.get_names, filter_obsoletes=True),
                ModelNameWriter,
                "Project types",
                "F",
            ),
            SheetDefinition(
                Project,
                partial(self.get_by_prop_name, prop_name="project_type_legacy"),
                ModelNameWriter,
                "Legacy project types",
                "G",
            ),
            SheetDefinition(Agency, self.get_names, ModelNameWriter, "Agencies", "H"),
            SheetDefinition(
                ProjectSector,
                partial(self.get_names, filter_obsoletes=True),
                ModelNameWriter,
                "Sectors",
                "I",
                enforce_validation=False,
            ),
            SheetDefinition(
                Project,
                partial(self.get_by_prop_name, prop_name="sector_legacy"),
                ModelNameWriter,
                "Legacy sectors",
                "J",
            ),
            SheetDefinition(
                ProjectSubSector,
                partial(self.get_names, filter_obsoletes=True),
                ModelNameWriter,
                "Subsectors",
                "K",
                enforce_validation=False,
            ),
            SheetDefinition(
                Project,
                partial(self.get_by_prop_name, prop_name="subsector_legacy"),
                ModelNameWriter,
                "Legacy subsectors",
                "L",
            ),
            SheetDefinition(
                SubstancesType,
                self.get_enum_labels,
                ModelNameWriter,
                "Substance types",
                "M",
            ),
            MultiModelSheetDefinition(
                None,
                self.get_names,
                ModelNameWriter,
                "Substances",
                "N",
                models=[Substance, Blend],
            ),
            SheetDefinition(
                ProjectStatus, self.get_names, ModelNameWriter, "Status", "O"
            ),
            SheetDefinition(
                Project,
                partial(self.get_by_prop_name, prop_name="serial_number"),
                ModelNameWriter,
                "Serial numbers",
                "P",
            ),
            SheetDefinition(
                Project,
                partial(self.get_by_prop_name, prop_name="serial_number_legacy"),
                ModelNameWriter,
                "Legacy serial numbers",
                "Q",
            ),
            SheetDefinition(
                Country,
                partial(self.get_name_and_codes, code_name="abbr"),
                ModelNameCodeWriter,
                "Countries",
                "R",
            ),
            SheetDefinition(
                Project,
                partial(self.get_by_prop_name, prop_name="title"),
                ModelNameWriter,
                "Titles",
                "S",
            ),
        ]

        for data_sheet in data_sheets:
            data_sheet.write_data(self)

        filename = "Projects"
        return method(filename, self.wb)

    def add_data_validation(
        self,
        column: str,
        validation_sheet: str,
        validation_range: int,
        allow_blank: bool = False,
        show_error: bool = False,
    ):
        validation_formula = f"'{validation_sheet}'!$A$2:$A${validation_range + 1}"
        data_validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=validation_formula,
            showDropDown=False,
            showErrorMessage=show_error,
            allow_blank=allow_blank,
        )
        data_validation.prompt = "Please select from the dropdown"
        data_validation.error = (
            "Invalid entry, please select a value from the dropdown list."
        )
        self.sheet.add_data_validation(data_validation)
        data_validation.add(f"{column}2:{column}1048576")
