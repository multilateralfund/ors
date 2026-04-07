from copy import copy
from typing import TYPE_CHECKING

import openpyxl
from django.db.models import F
from django.db.models import OuterRef
from django.db.models import Subquery
from openpyxl.utils import get_column_letter

from core.api.export.base import BaseWriter
from core.api.export.base import configure_sheet_print
from core.api.export.single_project_v2.helpers import format_iso_date
from core.api.utils import workbook_response
from core.models import MetaProject
from core.models import Project

if TYPE_CHECKING:
    from core.api.views import ProjectV2ViewSet
    from openpyxl.worksheet.worksheet import Worksheet


HEADERS = [
    {"headerName": "Metacode", "id": "umbrella_code"},
    {"headerName": "Country", "id": "country_name"},
    {"headerName": "Cluster", "id": "cluster_name"},
    {"headerName": "Lead agency", "id": "lead_agency_name"},
    {
        "headerName": "Start date (MYA)",
        "id": "start_date",
        "method": lambda r, h: format_iso_date(r[h["id"]]),
    },
    {
        "headerName": "End date (MYA)",
        "id": "end_date",
        "method": lambda r, h: format_iso_date(r[h["id"]]),
    },
    {
        "headerName": "Project duration (months)",
        "id": "project_duration",
        "in_grand_total": True,
    },
    {
        "headerName": "MYA Total agreed funding in principle (US $)",
        "id": "project_funding",
        "type": "number",
        "align": "right",
        "cell_format": "$###,###,##0.00#############",
        "in_grand_total": True,
    },
    {
        "headerName": "MYA Total support costs in principle (US $)",
        "id": "support_cost",
        "type": "number",
        "align": "right",
        "cell_format": "$###,###,##0.00#############",
        "in_grand_total": True,
    },
    {
        "headerName": "MYA Total agreed costs in principle (US $)",
        "id": "project_cost",
        "type": "number",
        "align": "right",
        "cell_format": "$###,###,##0.00#############",
        "in_grand_total": True,
    },
    {
        "headerName": "Phase-out (CO2-eq tonnes) (MYA)",
        "id": "phase_out_co2_eq_t",
        "in_grand_total": True,
    },
    {
        "headerName": "Phase-out (ODP tonnes) (MYA)",
        "id": "phase_out_odp",
        "in_grand_total": True,
    },
    {
        "headerName": "Phase-out (metric tonnes) (MYA)",
        "id": "phase_out_mt",
        "in_grand_total": True,
    },
    {
        "headerName": "Target in the last year (reduction in %)",
        "id": "target_reduction",
        "in_grand_total": True,
    },
    {
        "headerName": "Target in the last year (CO2-eq tonnes)",
        "id": "target_co2_eq_t",
        "in_grand_total": True,
    },
    {
        "headerName": "Target in the last year (ODP tonnes)",
        "id": "target_odp",
        "in_grand_total": True,
    },
    {
        "headerName": "Starting point for aggregate reductions in consumption or "
        "production (ODP tonnes)",
        "id": "starting_point_odp",
        "in_grand_total": True,
    },
    {
        "headerName": "Starting point for aggregate reductions in consumption or "
        "production (CO2-eq tonnes)",
        "id": "starting_point_co2_eq_t",
        "in_grand_total": True,
    },
    {
        "headerName": "Baseline (ODP tonnes)",
        "id": "baseline_odp",
        "in_grand_total": True,
    },
    {
        "headerName": "Baseline (CO2-eq tonnes)",
        "id": "baseline_co2_eq_t",
        "in_grand_total": True,
    },
    {
        "headerName": "Number of SMEs directly funded",
        "id": "number_of_smes_directly_funded",
        "in_grand_total": True,
    },
    {
        "headerName": "Number of non-SMEs directly funded",
        "id": "number_of_non_sme_directly_funded",
        "in_grand_total": True,
    },
    {
        "headerName": "Number of both SMEs and non-SMEs included in the project but "
        "not directly funded",
        "id": "number_of_both_sme_non_sme_not_directly_funded",
        "in_grand_total": True,
    },
    {
        "headerName": "Production sector: number of production lines assisted",
        "id": "number_of_production_lines_assisted",
        "in_grand_total": True,
    },
    {
        "headerName": "Cost effectiveness (US $/kg)",
        "id": "cost_effectiveness_kg",
        "type": "number",
        "align": "right",
        "cell_format": "$###,###,##0.00#############",
        "in_grand_total": True,
    },
    {
        "headerName": "Cost effectiveness (US $/CO2-eq tonnes) ",
        "id": "cost_effectiveness_co2",
        "type": "number",
        "align": "right",
        "cell_format": "$###,###,##0.00#############",
        "in_grand_total": True,
    },
]


class MyaWriter(BaseWriter):
    header_row_start_idx = 1

    def write(self, data):
        self.write_headers()
        self.write_data(data)
        self.write_totals_row()
        self.set_dimensions()
        self.sheet.freeze_panes = f"B{self.header_row_end_idx + 1}"
        last_col_letter = get_column_letter(self.max_column_idx)
        last_data_row = max(self.sheet.max_row - 1, self.header_row_end_idx)
        self.sheet.auto_filter.ref = (
            f"A{self.header_row_start_idx}:{last_col_letter}{last_data_row}"
        )

    def write_totals_row(self):
        row_idx = self.sheet.max_row + 1
        first_data_row = self.header_row_end_idx + 1
        has_data = row_idx > first_data_row

        for header_idx, header in enumerate(self.headers.values(), start=1):
            value = ""
            if header_idx == 1:
                value = "Grand total"
            elif header.get("in_grand_total"):
                col_letter = header["column_letter"]
                if has_data:
                    value = (
                        f"=SUM({col_letter}{first_data_row}:{col_letter}{row_idx - 1})"
                    )
                else:
                    value = 0
            cell = self._write_record_cell(
                row_idx,
                header["column"],
                value,
                align=header.get("align", "left"),
                cell_format=header.get("cell_format"),
            )
            font = copy(cell.font)
            font.bold = True
            cell.font = font
        self.sheet.row_dimensions[row_idx].height = self.ROW_HEIGHT


class MyaExport:
    wb: "openpyxl.Workbook"
    sheet: "Worksheet"
    view: "ProjectV2ViewSet"

    def __init__(self, view):
        self.view = view

    def setup_workbook(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "MYAs"
        configure_sheet_print(sheet, "landscape")
        self.wb = wb
        self.sheet = sheet

    def export_xls(self):
        self.setup_workbook()
        meta_project_ids = (
            self.view.filter_queryset(self.view.get_queryset())
            .exclude(meta_project__isnull=True)
            .order_by()
            .values_list("meta_project_id", flat=True)
            .distinct()
        )

        # self.sheet.append([h["headerName"] for h in HEADERS])

        first_project = Project.objects.filter(meta_project=OuterRef("pk")).order_by(
            "id"
        )[:1]

        meta_projects = MetaProject.objects.filter(id__in=meta_project_ids).values(
            *[h["id"] for h in HEADERS if h["id"]],
            country_name=F("country__name"),
            cluster_name=F("cluster__name"),
            lead_agency_name=Subquery(
                first_project.values("lead_agency__name")[:1],
            ),
        )

        writer = MyaWriter(self.sheet, HEADERS)
        writer.write(meta_projects)
        return workbook_response("Mya.xlsx", self.wb)
