# pylint: disable=too-many-locals,attribute-defined-outside-init,unused-argument
from collections import defaultdict
from itertools import chain
from typing import Sequence

import openpyxl
from openpyxl.styles import DEFAULT_FONT
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from core.api.export.base import configure_sheet_print
from core.api.export.single_project_v2.helpers import format_iso_date
from core.api.serializers.project_v2 import ProjectDetailsV2Serializer
from core.api.utils import workbook_response
from core.models import Project
from core.models import ProjectField
from core.models import ProjectSpecificFields

# pylint: disable=R0915

SECTIONS = (
    "Cross-Cutting",
    "Identifiers",
    "Header",
    "Impact",
    "Substance Details",
    "Approval",
    "MYA",
)

EXCLUDE_SECTIONS = (
    "Header",
    "Impact",
    "Substance Details",
    "MYA",
)

VARIANCE_FIELDS = (
    "total_fund",
    "support_cost_psc",
    "ods_odp.odp",
    "ods_odp.odp",
    "ods_odp.phase_out_mt",
    "ods_odp.co2_mt",
)


def serialize_project(p):
    serializer = ProjectDetailsV2Serializer(p)
    return serializer.data


def version_label(p):
    status = p.submission_status.name
    post_excom_meeting = p.post_excom_meeting.number if p.post_excom_meeting else ""
    meeting = p.meeting.number if p.meeting else ""
    if p.version > 3:
        return f"Version: {p.version}: {status} (ExCom {post_excom_meeting})"
    return f"Version: {p.version}: {status} (Meeting {meeting})"


def normalise_version(p):
    if p.version > 3:
        return 3
    return p.version


class CompareVersionsWriter:
    def __init__(self, sheet, candidates: Sequence[Project]):
        self.sheet = sheet
        self.candidates = candidates

    def write_headers(self, user, projects: Sequence[Project]):
        p1, p2 = list(projects)

        value_headers = self.get_other_headers(
            self.get_fields(
                user, versions=[normalise_version(p1), normalise_version(p2)]
            ),
            self.get_specific_information_fields(user),
        )

        per_section_headers = defaultdict(list)
        for header in value_headers:
            per_section_headers[header["section"]].append(header)

        per_section_merge_size = defaultdict(int)
        per_field_merge_size = defaultdict(int)

        h_sections = [None]
        h_fields = [None]
        v_fields = [None]
        for section in SECTIONS:
            members = per_section_headers[section]
            section_name = section if section != "Header" else ""
            for m in members:
                if m["id"] in VARIANCE_FIELDS:
                    h_sections.extend([section_name] * 3)
                    h_fields.extend([m["headerName"]] * 3)
                    v_fields.extend(
                        [
                            p1.submission_status.name,
                            p2.submission_status.name,
                            "Variance",
                        ]
                    )
                    per_section_merge_size[section_name] += 3
                    per_field_merge_size[m["id"]] = 3
                else:
                    h_sections.extend([section_name] * 2)
                    h_fields.extend([m["headerName"]] * 2)
                    v_fields.extend(
                        [p1.submission_status.name, p2.submission_status.name]
                    )
                    per_section_merge_size[section_name] += 2
                    per_field_merge_size[m["id"]] = 2

        self.sheet.append(h_sections)
        self.sheet.append(h_fields)
        self.sheet.append(v_fields)

        self.mark_header_row(1, len(h_sections))
        self.mark_header_row(2, len(h_fields))
        self.mark_header_row(3, len(v_fields))

        merge_start_idx = 2
        field_merge_start_idx = 2
        for section in SECTIONS:
            merge_size = per_section_merge_size[section]
            if merge_size:
                merge_end_idx = merge_start_idx + merge_size - 1
                self.sheet.merge_cells(
                    start_row=1,
                    start_column=merge_start_idx,
                    end_row=1,
                    end_column=merge_end_idx,
                )
                merge_start_idx = merge_end_idx + 1
            for m in per_section_headers[section]:
                field_merge_size = per_field_merge_size[m["id"]]
                if field_merge_size:
                    self.sheet.merge_cells(
                        start_row=2,
                        start_column=field_merge_start_idx,
                        end_row=2,
                        end_column=field_merge_start_idx + field_merge_size - 1,
                    )
                    field_merge_start_idx = field_merge_start_idx + field_merge_size

        self._h_fields = h_fields
        self._per_section_headers = per_section_headers

    def write(self, user, projects: Sequence[Project]):
        p1, p2 = list(projects)
        d1, d2 = serialize_project(p1), serialize_project(p2)

        data_row = [p1.final_version.id]

        data_remap = {
            "lead_agency": "lead_agency.name",
        }

        for section in SECTIONS:
            members = self._per_section_headers[section]
            for h in members:
                h_id = h["id"]
                if h_id_ovr := data_remap.get(h_id):
                    v1 = self.get_value(h_id_ovr, p1)
                    v2 = self.get_value(h_id_ovr, p2)
                else:
                    v1 = self.get_serialized_value(h_id, d1)
                    v2 = self.get_serialized_value(h_id, d2)

                variance = None
                if h_id in VARIANCE_FIELDS and v1 and v2:
                    try:
                        variance = abs(v2 - v1)
                    except TypeError:
                        pass

                if formatter := h.get("formatter"):
                    v1 = formatter(v1)
                    v2 = formatter(v2)

                if h_id in VARIANCE_FIELDS:
                    data_row.extend([v1, v2, variance])
                else:
                    data_row.extend([v1, v2])

        self.sheet.append(data_row)

        for i in range(1, len(self._h_fields) + 1):
            self.sheet.column_dimensions[get_column_letter(i)].width = 20

        # self.apply_number_format(self._per_section_headers)

    def apply_number_format(self, per_section_headers):
        for section in SECTIONS:
            members = per_section_headers[section]
            col = 1
            for h in members:
                col += 1
                if cell_format := h.get("cell_format"):
                    for row in range(3, 6):
                        self.sheet.cell(row, col).number_format = cell_format

    def mark_header_row(self, row, count):
        for i in range(1, count + 1):
            cell = self.sheet.cell(row, i)
            cell.font = Font(name=DEFAULT_FONT.name, bold=True, color=None)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    def mark_variance_row(self, row, count):
        for i in range(1, count + 1):
            cell = self.sheet.cell(row, i)
            cell.font = Font(name=DEFAULT_FONT.name, bold=False, color="FF0000")

    def get_fields(self, user, versions):
        result = []
        fields = (
            ProjectField.objects.get_visible_fields_for_user(user)
            .exclude(read_field_name="sort_order")
            .exclude(section__in=EXCLUDE_SECTIONS)
        )
        for f in fields:
            if set(f.get_visible_versions()).intersection(versions):
                result.append(f)
        return result

    def get_specific_information_fields(self, user):
        queryset = ProjectSpecificFields.objects.filter(
            cluster__in=[p.cluster for p in self.candidates],
            type__in=[p.project_type for p in self.candidates],
            sector__in=[p.sector for p in self.candidates],
        )
        return (
            queryset.first()
            .fields.get_visible_fields_for_user(user)
            .exclude(read_field_name="sort_order")
        )

    def get_other_headers(self, fields, specific_fields, exclude=None):
        result = []

        known_ids = []
        exclude = exclude if exclude else []

        all_fields = sorted(
            chain(fields, specific_fields), key=lambda x: (x.section, x.sort_order)
        )
        for f in all_fields:
            if f.read_field_name in exclude:
                continue
            if f.read_field_name in known_ids:
                continue
            path = (
                f"{f.table}.{f.read_field_name}"
                if f.table != "project"
                else f.read_field_name
            )
            h_def = {
                "id": path,
                "section": f.section,
                "headerName": f.label,
            }
            if f.data_type == "decimal-money":
                h_def["cell_format"] = "$###,###,##0.00#############"
            elif f.data_type == "decimal":
                h_def["cell_format"] = "###,###,##0.00#############"
            elif f.data_type == "date":
                h_def["formatter"] = format_iso_date
            result.append(h_def)
            known_ids.append(f.read_field_name)

        return result

    def get_value(self, name, project):
        last = None
        for n in name.split("."):
            if not last:
                last = getattr(project, n, None)
                continue

            if isinstance(last, (tuple, list, set)):
                last = [getattr(x, n) for x in last]
            elif isinstance(last, dict):
                last = last.get(n)
            else:
                last = getattr(last, n, None)

        return last

    def get_serialized_value(self, name, data):
        last = None
        for n in name.split("."):
            if not last:
                last = data.get(n)
                continue

            if isinstance(last, (tuple, list, set)):
                last = [x.get(n) for x in last]
            elif isinstance(last, dict):
                last = last.get(n)

        if isinstance(last, dict):
            last = last.get("name", last.get("number", str(last)))
        elif isinstance(last, (tuple, list, set)):
            last = ", ".join(map(str, last))

        elif isinstance(last, bool):
            last = "Yes" if last else "No"

        return last


class CompareVersionsProjectExport:
    wb: openpyxl.Workbook
    candidates: Sequence[Sequence[Project]]

    def __init__(self, user, candidates: Sequence[Sequence[Project]]):
        self.user = user
        self.candidates = candidates
        self.setup_workbook()

    def setup_workbook(self):
        wb = openpyxl.Workbook()
        # delete default sheet
        del wb[wb.sheetnames[0]]
        sheet = wb.create_sheet("Comparison of versions")
        configure_sheet_print(sheet, "landscape")
        self.wb = wb
        self.sheet = sheet

    def export(self, filename: str):
        flat_candidates = list(chain(*self.candidates))
        writer = CompareVersionsWriter(self.sheet, flat_candidates)
        writer.write_headers(self.user, self.candidates[0])
        for pair in self.candidates:
            writer.write(self.user, pair)
        return workbook_response(filename, self.wb)
