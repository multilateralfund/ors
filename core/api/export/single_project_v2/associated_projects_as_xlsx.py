import openpyxl
from django.db.models import Q
from django.db.models import QuerySet

from core.api.export.base import WriteOnlyBase
from core.api.export.base import configure_sheet_print
from core.api.export.single_project_v2.helpers import (
    format_iso_date,
    get_blanket_consideration_value,
    get_consumption_level_status_value,
)
from core.api.serializers.project_v2 import ProjectDetailsV2Serializer
from core.api.utils import workbook_response
from core.models import Project
from core.models import User


class ProjectWriter(WriteOnlyBase):
    header_row_start_idx = 1
    COLUMN_WIDTH = 20

    def __init__(self, sheet, user):
        headers = [
            {
                "id": "self_project_code",
                "headerName": "Associated project code",
                "column_width": self.COLUMN_WIDTH * 2,
            },
            {
                "id": "code",
                "headerName": "Project code",
                "column_width": self.COLUMN_WIDTH * 2,
            },
            {
                "id": "sector",
                "headerName": "Sector",
                "method": lambda r, h: r[h["id"]]["name"],
            },
            {
                "id": "subsectors",
                "headerName": "Sub-sector",
                "method": lambda r, h: ", ".join(ss["name"] for ss in r[h["id"]]),
                "column_width": WriteOnlyBase.COLUMN_WIDTH * 1.5,
            },
            {
                "id": "consumption_level_status",
                "headerName": "Consumption level status",
                "method": get_consumption_level_status_value,
                "column_width": WriteOnlyBase.COLUMN_WIDTH * 1.5,
            },
            {
                "id": "title",
                "headerName": "Title",
            },
            {
                "id": "description",
                "headerName": "Description",
            },
            {
                "id": "project_start_date",
                "headerName": "Project start date",
                "method": lambda r, h: format_iso_date(r[h["id"]]),
            },
            {
                "id": "project_end_date",
                "headerName": "Project end date",
                "method": lambda r, h: format_iso_date(r[h["id"]]),
            },
            {
                "id": "project_duration",
                "headerName": "Duration of project (months)",
            },
            {
                "id": "adjustment",
                "headerName": "Adjustment",
                "type": "boolean",
                "align": "center",
            },
            {
                "id": "total_fund",
                "headerName": "Project funding",
                "type": "number",
                "align": "right",
                "cell_format": "$###,###,##0.00#############",
            },
            {
                "id": "support_cost_psc",
                "headerName": "Project support cost",
                "type": "number",
                "align": "right",
                "cell_format": "$###,###,##0.00#############",
            },
            {
                "id": "blanket_or_individual_consideration",
                "headerName": "Blanket approval/Individual consideration",
                "method": get_blanket_consideration_value,
                "permission": "core.is_mlfs_user",
            },
            {
                "id": "version",
                "headerName": "Version",
            },
        ]

        allowed_headers = []

        for header in headers:
            perm = header.get("permission", None)
            if perm and user.has_perm(perm) or perm is None:
                allowed_headers.append(header)

        super().__init__(sheet, allowed_headers)


class ProjectsV2AssociatedProjectsExport:
    user = User
    wb: openpyxl.Workbook
    project: Project

    def __init__(self, project: Project, user: User):
        self.user = user
        self.project = project
        self.setup_workbook()

    def setup_workbook(self):
        wb = openpyxl.Workbook()
        # delete default sheet
        del wb[wb.sheetnames[0]]
        self.wb = wb

    def build(self, data):
        sheet = self.add_sheet("Associated projects")
        ProjectWriter(sheet, user=self.user).write(data)

    def get_associated_projects(self) -> QuerySet[Project]:
        result = Project.objects.none()

        project_filters = Q()
        if self.project.meta_project:
            project_filters &= Q(meta_project=self.project.meta_project)
        elif self.project.component is not None:
            project_filters &= Q(component=self.project.component)
        else:
            project_filters &= Q(id=self.project.id)

        result = Project.objects.filter(project_filters).exclude(id=self.project.id)

        # Requested in #35434.
        if self.user.has_perm("core.is_mlfs_user"):
            result = result.exclude(submission_status__name="Draft")

        return result

    def build_xls(self):
        data = []
        for project in self.get_associated_projects():
            serializer = ProjectDetailsV2Serializer(project)
            project_data = {"self_project_code": self.project.code}
            project_data.update(serializer.data)
            data.append(project_data)
        self.build(data)

    def add_sheet(self, name):
        sheet = self.wb.create_sheet(name)
        configure_sheet_print(sheet, "landscape")
        return sheet

    def export_xls(self):
        self.build_xls()
        return workbook_response(f"Associated projects for {self.project.id}", self.wb)
