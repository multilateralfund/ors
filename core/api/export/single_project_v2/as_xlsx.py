from typing import Annotated

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from core.api.export.base import BaseWriter, transpose_sheet
from core.api.export.base import configure_sheet_print
from core.api.export.business_plan import BPActivitiesWriter
from core.api.export.single_project_v2.xlsx_headers import get_headers_cross_cutting
from core.api.export.single_project_v2.xlsx_headers import get_headers_identifiers
from core.api.export.single_project_v2.xlsx_headers import (
    get_headers_specific_information,
)
from core.api.export.single_project_v2.helpers import get_activity_data
from core.api.serializers.project_v2 import ProjectDetailsV2Serializer
from core.api.utils import workbook_response
from core.models import Project, ProjectField
from core.models import ProjectSpecificFields
from core.models import User

# pylint: disable=R0913


class ProjectWriter(BaseWriter):
    header_row_start_idx = 1
    COLUMN_WIDTH = 20


class ProjectsV2ProjectExport:
    user = User
    wb: openpyxl.Workbook
    project: Project

    def __init__(self, project, user):
        self.user = user
        self.project = project
        self.setup_workbook()

    def setup_workbook(self):
        wb = openpyxl.Workbook()
        # delete default sheet
        del wb[wb.sheetnames[0]]
        self.wb = wb

    def build_identifiers(self, data):
        sheet = self.add_sheet("Identifiers")
        ProjectWriter(sheet, get_headers_identifiers()).write([data])

    def build_bp(self, data):
        activity_data = get_activity_data(data)
        if activity_data:
            sheet = self.add_sheet("Identifiers - BP Activity")
            writer = BPActivitiesWriter(self.wb, min_year=None, max_year=None)
            # The BPActivitiesWriter creates a sheet that we don't need, replace it with our own
            del self.wb[writer.sheet.title]
            writer.sheet = sheet
            writer.write([activity_data])

    def _write_transposed(self, sheet_name, headers, data):
        # Create a temporary horizontal sheet
        tmp_sheet = self.wb.create_sheet(f"_tmp_{sheet_name}")
        ProjectWriter(tmp_sheet, headers).write(data)

        # Create the sheet as the transposed horizontal one
        sheet = self.add_sheet(sheet_name, Worksheet.ORIENTATION_PORTRAIT)
        transpose_sheet(tmp_sheet, sheet)

        self.wb.remove(tmp_sheet)

    def _write_cross_cutting_fields(
        self,
        fields_section: str,
        sheet_name: Annotated[str, "max_length=31"],
        data,
    ):
        """
        Writes a new sheet.

        :param sheet_name: The sheet name should not exceed 31 characters, this is as Excel constraint.
        """
        fields = (
            ProjectField.objects.get_visible_fields_for_user(self.user)
            .filter(section__in=[fields_section])
            .exclude(read_field_name="sort_order")
        )
        if fields:
            self._write_transposed(sheet_name, get_headers_cross_cutting(fields), data)

    def build_cross_cutting(self, data):
        self._write_cross_cutting_fields(
            "Cross-Cutting",
            "Cross-cutting",
            [data],
        )

    def _write_project_specific_fields(
        self,
        fields_obj: ProjectSpecificFields,
        fields_section: str,
        sheet_name: Annotated[str, "max_length=31"],
        data,
        only_planned=False,
        transposed=False,
    ):
        """
        Writes a new sheet.

        :param sheet_name: The sheet name should not exceed 31 characters, this is an Excel constraint.
        """
        fields = (
            fields_obj.fields.get_visible_fields_for_user(self.user)
            .filter(section__in=[fields_section])
            .exclude(read_field_name="sort_order")
        )
        if only_planned:
            fields = fields.filter(is_actual=False)

        if fields:
            headers = get_headers_specific_information(fields)
            if transposed:
                self._write_transposed(sheet_name, headers, data)
            else:
                sheet = self.add_sheet(sheet_name)
                ProjectWriter(sheet, headers).write(data)

    def build_specific_information(self, data):
        project_specific_fields_obj = ProjectSpecificFields.objects.filter(
            cluster=self.project.cluster,
            type=self.project.project_type,
            sector=self.project.sector,
        ).first()

        if project_specific_fields_obj:
            self._write_project_specific_fields(
                project_specific_fields_obj,
                "Header",
                "SI - Overview",
                [data],
            )

            self._write_project_specific_fields(
                project_specific_fields_obj,
                "Substance Details",
                "SI - Substance details",
                [
                    {**d, "products_manufactured": data.get("products_manufactured")}
                    for d in data.get("ods_odp", [])
                ],
            )

            self._write_project_specific_fields(
                project_specific_fields_obj,
                "Impact",
                "Impact",
                [data],
                only_planned=True,
                transposed=True,
            )

    def build_xls(self):
        serializer = ProjectDetailsV2Serializer(self.project)
        data = serializer.data
        self.build_identifiers(data)
        self.build_bp(data)
        self.build_cross_cutting(data)
        self.build_specific_information(data)

    def add_sheet(self, name, orientation=Worksheet.ORIENTATION_LANDSCAPE):
        sheet = self.wb.create_sheet(name)
        configure_sheet_print(sheet, orientation)
        return sheet

    def export_xls(self):
        self.build_xls()
        return workbook_response(f"Project {self.project.id}", self.wb)
