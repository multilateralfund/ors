from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import DEFAULT_FONT
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.worksheet.datavalidation import DataValidation

from core.api.export.base import BaseWriter


class AdmWriter(BaseWriter):
    def __init__(self, sheet, headers, row_headers):
        super().__init__(sheet, headers)
        self.row_headers = row_headers
        self.questions_index_by_id = {}
        self.yes_no_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True,
            showErrorMessage=True,
        )
        self.sheet.add_data_validation(self.yes_no_validation)

    def write_headers(self):
        super().write_headers()
        self.write_row_headers()

    def write_row_headers(self):
        col_headers = {k: v for k, v in self.headers.items() if v.get("is_row_header")}

        for row_idx, row_header in enumerate(
            self.row_headers, start=self.header_row_end_idx + 1
        ):
            header_type = row_header.get("type")
            if header_type == "title":
                method = self._write_row_title_cell
            elif header_type == "subtitle":
                method = self._write_row_subtitle_cell
            else:
                method = self._write_record_cell
                item = {"row_idx": row_idx}
                self.questions_index_by_id[row_header["id"]] = item

                if row_header.get("choices"):
                    item["choices"] = {
                        choice["id"]: choice["value"]
                        for choice in row_header["choices"]
                    }
                    formula = ",".join(item["choices"].values())
                    item["data_validation"] = DataValidation(
                        type="list", formula1=f'"{formula}"', allow_blank=True
                    )
                    self.sheet.add_data_validation(item["data_validation"])

            col_idx = 1
            for key, col_header in col_headers.items():
                col_idx = col_header["column"]
                method(row_idx, col_idx, row_header[key])

            if header_type in ("title", "subtitle"):
                self.sheet.merge_cells(
                    start_row=row_idx,
                    start_column=col_idx,
                    end_row=row_idx,
                    end_column=self.max_column_idx,
                )

    def write_data(self, data):
        data_headers = {
            k: v for k, v in self.headers.items() if not v.get("is_row_header")
        }
        record_row_id = {record["row_id"]: record for record in data}
        for row_id, question in self.questions_index_by_id.items():
            row_idx = question["row_idx"]
            try:
                values_by_col_id = {
                    value["column_id"]: value
                    for value in record_row_id[row_id]["values"]
                }
            except KeyError:
                values_by_col_id = {}

            for header_id, header in data_headers.items():
                try:
                    value_obj = values_by_col_id[header_id]
                    choice_id = value_obj["value_choice_id"]
                    value_text = (
                        value_obj["value_text"] or question["choices"][choice_id]
                    )
                except KeyError:
                    value_text = ""

                cell = self._write_record_cell(row_idx, header["column"], value_text)
                if header.get("display_name", "").lower() == "yes / no":
                    self.yes_no_validation.add(cell)
                elif data_validation := question.get("data_validation"):
                    data_validation.add(cell)

    def _write_row_title_cell(self, row, column, value):
        cell = self.sheet.cell(row, column, value)
        cell.font = Font(name=DEFAULT_FONT.name, bold=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
        )
        return cell

    def _write_row_subtitle_cell(self, row, column, value):
        cell = self.sheet.cell(row, column, value)
        cell.font = Font(name=DEFAULT_FONT.name, bold=True, italic=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = PatternFill(
            start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
        )
        return cell
