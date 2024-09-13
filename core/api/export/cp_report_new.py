from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from core.api.export.base import CPReportBase
from core.api.export.section_export import SectionWriter

# pylint: disable=R0913


class CPReportNewExporter(CPReportBase):
    sections = (
        "section_a",
        "section_b",
        "section_c",
        "section_d",
        "section_e",
        "section_f",
    )

    def export_section_a(self, sheet, data, usages, convert_to="mt"):
        conversion = "METRIC TONNES" if convert_to == "mt" else "ODP TONNES"
        self._export_usage_section(
            sheet,
            data,
            usages,
            (
                "SECTION A. ANNEX A, ANNEX B, ANNEX C - GROUP I AND "
                f"ANNEX E - DATA ON CONTROLLED SUBSTANCES ({conversion})"
            ),
            manufacturing_blends=False,
            convert_to=convert_to,
        )

    def export_section_b(self, sheet, data, usages, convert_to="mt"):
        convertion = "METRIC TONNES" if convert_to == "mt" else "CO₂-eq tonnes"
        self._export_usage_section(
            sheet,
            data,
            usages,
            f"SECTION B. ANNEX F - DATA ON CONTROLLED SUBSTANCES ({convertion})",
            convert_to=convert_to,
        )

    def _export_usage_section(
        self, sheet, data, usages, title, manufacturing_blends=True, convert_to="mt"
    ):
        headers = [
            {
                "id": "sheet-title",
                "headerName": title,
                "children": [
                    {
                        "id": "display_name",
                        "headerName": "Substance",
                        "is_numeric": False,
                        "column_width": self.COLUMN_WIDTH * 2,
                    },
                    {
                        "id": "use-by-sector",
                        "headerName": "Use by Sector",
                        "children": [
                            *usages,
                            {
                                "id": "total",
                                "headerName": "TOTAL",
                                "is_sum_function": True,
                                "align": "right",
                            },
                        ],
                    },
                    {
                        "id": "imports",
                        "headerName": "Import",
                        "align": "right",
                    },
                    {
                        "id": "exports",
                        "headerName": "Export",
                        "align": "right",
                    },
                    {
                        "id": "production",
                        "headerName": "Production",
                        "align": "right",
                    },
                    *(
                        [
                            {
                                "id": "manufacturing_blends",
                                "headerName": "Manufacturing of Blends",
                            }
                        ]
                        if manufacturing_blends
                        else []
                    ),
                    {
                        "id": "import_quotas",
                        "headerName": "Import Quotas",
                        "align": "right",
                    },
                    {
                        "id": "banned_date",
                        "headerName": "Date ban commenced",
                        "type": "date",
                        "is_numeric": False,
                    },
                    {
                        "id": "remarks",
                        "headerName": "Remarks",
                        "is_numeric": False,
                        "column_width": self.COLUMN_WIDTH * 2,
                    },
                ],
            }
        ]
        SectionWriter(sheet, headers, convert_to).write(data)

    def export_section_c(self, sheet, data, *args):
        SectionWriter(
            sheet,
            [
                {
                    "id": "sheet-title",
                    "headerName": "SECTION C. AVERAGE ESTIMATED PRICE OF HCFCs, HFCs AND ALTERNATIVES (US $/kg)",
                    "children": [
                        {
                            "id": "display_name",
                            "headerName": "Substance",
                            "is_numeric": False,
                            "column_width": self.COLUMN_WIDTH * 2,
                        },
                        {
                            "id": "previous_year_price",
                            "headerName": "Previous year price",
                            "column_width": self.COLUMN_WIDTH * 2,
                            "align": "right",
                        },
                        {
                            "id": "current_year_price",
                            "headerName": "Current prices",
                            "column_width": self.COLUMN_WIDTH * 2,
                            "align": "right",
                        },
                        {
                            "id": "is_retail",
                            "headerName": "Is Retail price?",
                            "is_numeric": False,
                            "method": self.get_fob_retail_value,
                        },
                        {
                            "id": "is_fob",
                            "headerName": "Is FOB price?",
                            "is_numeric": False,
                            "method": self.get_fob_retail_value,
                        },
                        {
                            "id": "remarks",
                            "headerName": "Remarks",
                            "is_numeric": False,
                            "column_width": self.COLUMN_WIDTH * 2,
                        },
                    ],
                }
            ],
        ).write(data)

    def export_section_d(self, sheet, data, *args):
        SectionWriter(
            sheet,
            [
                {
                    "id": "sheet-title",
                    "headerName": "SECTION D. ANNEX F, GROUP II - DATA ON HFC-23 GENERATION (METRIC TONNES)",
                    "children": [
                        {
                            "id": "display_name",
                            "headerName": "Substance",
                            "is_numeric": False,
                            "column_width": self.COLUMN_WIDTH * 2,
                        },
                        {
                            "id": "all_uses",
                            "headerName": "Captured for all uses",
                            "column_width": self.COLUMN_WIDTH * 2,
                            "align": "right",
                        },
                        {
                            "id": "feedstock",
                            "headerName": "Captured for feedstock uses within your country",
                            "column_width": self.COLUMN_WIDTH * 2,
                            "align": "right",
                        },
                        {
                            "id": "destruction",
                            "headerName": "Captured for destruction",
                            "column_width": self.COLUMN_WIDTH * 2,
                            "align": "right",
                        },
                    ],
                }
            ],
        ).write(data)

    def export_section_e(self, sheet, data, *args):
        SectionWriter(
            sheet,
            [
                {
                    "id": "sheet-title",
                    "headerName": "SECTION E. ANNEX F, GROUP II - DATA ON HFC-23 EMISSIONS (METRIC TONNES)",
                    "children": [
                        {
                            "id": "facility",
                            "headerName": "Facility name or identifier",
                            "is_numeric": False,
                            "column_width": self.COLUMN_WIDTH * 2,
                        },
                        {
                            "id": "total",
                            "headerName": "Total amount generated",
                            "align": "right",
                        },
                        {
                            "id": "amount-generated",
                            "headerName": "Amount generated and captured",
                            "children": [
                                {
                                    "id": "all_uses",
                                    "headerName": "For all uses",
                                    "column_width": self.COLUMN_WIDTH * 2,
                                    "align": "right",
                                },
                                {
                                    "id": "feedstock_gc",
                                    "headerName": "For feedstock use in your country",
                                    "column_width": self.COLUMN_WIDTH * 2,
                                    "align": "right",
                                },
                                {
                                    "id": "destruction",
                                    "headerName": "For destruction",
                                    "column_width": self.COLUMN_WIDTH * 2,
                                    "align": "right",
                                },
                            ],
                        },
                        {
                            "id": "feedstock_wpc",
                            "headerName": "Amount used for feedstock without prior capture",
                            "column_width": self.COLUMN_WIDTH * 2,
                            "align": "right",
                        },
                        {
                            "id": " destruction_wpc",
                            "headerName": "Amount destroyed without prior capture",
                            "column_width": self.COLUMN_WIDTH * 2,
                            "align": "right",
                        },
                        {
                            "id": "generated_emissions",
                            "headerName": "Amount of generated emission",
                            "column_width": self.COLUMN_WIDTH * 2,
                            "align": "right",
                        },
                        {
                            "id": "remarks",
                            "headerName": "Remarks",
                            "is_numeric": False,
                            "column_width": self.COLUMN_WIDTH * 2,
                        },
                    ],
                }
            ],
        ).write(data)

    def export_section_f(self, sheet, data, *args):
        writer = SectionWriter(
            sheet,
            [
                {
                    "id": "sheet-title",
                    "headerName": "SECTION F. COMMENTS BY BILATERAL/IMPLEMENTING AGENCIES",
                    "children": [{"id": "remarks", "headerName": "Remarks"}],
                }
            ],
        )
        writer.write_headers()
        writer.set_dimensions()

        row_idx = writer.header_row_end_idx + 1
        col_idx = 1

        try:
            value = data["remarks"]
        except (KeyError, TypeError):
            # This may be missing when exporting an empty form
            value = ""

        cell = sheet.cell(row_idx, col_idx, value)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        sheet.column_dimensions[get_column_letter(col_idx)].width = (
            self.COLUMN_WIDTH * 4
        )
        sheet.row_dimensions[row_idx].height = self.ROW_HEIGHT * 16

    def get_fob_retail_value(self, cp_price, header):
        year = cp_price.country_programme_report.year
        if not cp_price["is_fob"] and not cp_price["is_retail"] and year < 2023:
            return ""
        return "Yes" if cp_price[header["id"]] else "No"
