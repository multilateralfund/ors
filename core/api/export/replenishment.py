# pylint: disable=C0302

from copy import copy
from datetime import datetime
from decimal import Decimal
from itertools import chain
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import DEFAULT_FONT, Font, Side, Border, Alignment
from openpyxl.utils import get_column_letter, range_boundaries

from django.db.models import F, Q

from core.api.export.base import configure_sheet_print, WriteOnlyBase
from core.models import (
    BilateralAssistance,
    ExternalIncomeAnnual,
    DisputedContribution,
    ExternalAllocation,
)

EMPTY_ROW = (None, None, None)
# pylint: disable=W0612
# pylint: disable=R0913


class DashboardWriter(WriteOnlyBase):
    COLUMN_WIDTH = 25

    def set_dimensions(self):
        self.sheet.column_dimensions["A"].width = self.COLUMN_WIDTH * 2
        self.sheet.column_dimensions["B"].width = self.COLUMN_WIDTH
        self.sheet.column_dimensions["C"].width = self.COLUMN_WIDTH

    def write_data(self, data):
        for key, cell_b, cell_c in data:
            record_row = []
            if cell_b is None and cell_c is None:
                record_row.append(self.write_header_cell(key))
            else:
                record_row.append(self.write_record_cell(key))

            record_row.append(self.write_record_cell(cell_b))
            record_row.append(self.write_record_cell(cell_c))

            self.sheet.append(record_row)
        # Merge cells for title, subtitle and date
        self.sheet.merge_cells("A4:C4")
        self.sheet.merge_cells("A6:C6")
        self.sheet.merge_cells("A5:C5")
        for row in range(4, 7):
            self.sheet.row_dimensions[row].height = 30
            self.sheet.cell(row=row, column=1).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    def write_header_cell(self, value, comment=None):
        cell = WriteOnlyCell(self.sheet, value=value)
        cell.font = Font(name="Times New Roman", bold=True)
        cell.border = Border(
            top=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return cell

    def write_record_cell(self, value, read_only=False):
        cell = WriteOnlyCell(self.sheet, value=value)
        cell.font = Font(name="Times New Roman")
        cell.border = Border(
            top=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if type(value) in (float, Decimal):
            cell.number_format = "###,###,##0.00###"

        return cell


class StatusOfContributionsWriter(WriteOnlyBase):
    BOLD_RECORD_CELLS = ["TOTAL", "SUB-TOTAL"]

    def __init__(self, sheet, period=None, extra_headers=None):
        headers = [
            {
                "id": "country",
                "headerName": "Party",
                "column_width": 25,
            },
            {
                "id": "agreed_contributions",
                "headerName": "Agreed contributions",
                "column_width": 25,
            },
            {
                "id": "cash_payments",
                "headerName": "Cash payments",
                "column_width": 25,
            },
            {
                "id": "bilateral_assistance",
                "headerName": "Bilateral assistance",
                "column_width": 25,
            },
            {
                "id": "promissory_notes",
                "headerName": "Promissory notes",
                "column_width": 25,
            },
            {
                "id": "outstanding_contributions",
                "headerName": "Outstanding contributions",
                "column_width": 25,
            },
        ]

        if extra_headers:
            headers.extend(extra_headers)
        self.period = period
        super().__init__(sheet, headers)

    def do_prewriting(self):
        # Merge cells for title, subtitle and date
        columns_number = len(self.headers)
        self.sheet.merge_cells(
            start_row=5, start_column=1, end_row=5, end_column=columns_number
        )
        self.sheet.merge_cells(
            start_row=6, start_column=1, end_row=6, end_column=columns_number
        )
        self.sheet.merge_cells(
            start_row=7, start_column=1, end_row=7, end_column=columns_number
        )
        self.sheet.merge_cells(
            start_row=8, start_column=1, end_row=8, end_column=columns_number
        )
        self.sheet.cell(row=5, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.sheet.cell(row=6, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.sheet.cell(row=7, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.sheet.cell(row=5, column=1).value = (
            "TRUST  FUND FOR THE  MULTILATERAL FUND FOR THE IMPLEMENTATION OF THE MONTREAL PROTOCOL"
        )
        self.sheet.cell(row=6, column=1).value = (
            f"Status of Contributions for {self.period} (US$)"
        )
        self.sheet.cell(row=7, column=1).value = "As at 24/05/2024"

    def write(self, data):
        self.do_prewriting()
        super().write(data)

    def write_header_cell(self, value, comment=None):
        cell = super().write_header_cell(value, comment)
        cell.font = Font(name="Times New Roman", bold=True)
        return cell

    def write_record_cell(self, value, read_only=False):
        cell = super().write_record_cell(value, read_only)
        if type(value) in (float, Decimal):
            cell.number_format = "###,###,##0.00###"

        if any(bold_string == value for bold_string in self.BOLD_RECORD_CELLS):
            cell.border = Border(
                top=Side(style="thick"),
                bottom=Side(style="thick"),
            )
            cell.font = Font(name="Times New Roman", bold=True)

        return cell


class StatisticsStatusOfContributionsWriter(WriteOnlyBase):
    BOLD_RECORD_CELLS = [
        "Total payments",
        "Interest earned",
        "Miscellaneous income",
        "TOTAL INCOME",
        "Accumulated figures",
    ]

    def __init__(self, sheet, headers, period=None):
        self.headers = headers
        self.period = period
        super().__init__(sheet, headers)

    def do_prewriting(self):
        # Merge cells for title, subtitle and date
        columns_number = len(self.headers)
        self.sheet.merge_cells(
            start_row=5, start_column=1, end_row=5, end_column=columns_number
        )
        self.sheet.merge_cells(
            start_row=6, start_column=1, end_row=6, end_column=columns_number
        )
        self.sheet.merge_cells(
            start_row=7, start_column=1, end_row=7, end_column=columns_number
        )
        self.sheet.merge_cells(
            start_row=8, start_column=1, end_row=8, end_column=columns_number
        )
        self.sheet.cell(row=5, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.sheet.cell(row=6, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.sheet.cell(row=7, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.sheet.cell(row=8, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.sheet.cell(row=5, column=1).value = (
            "TRUST  FUND FOR THE  MULTILATERAL FUND FOR THE IMPLEMENTATION OF THE MONTREAL PROTOCOL"
        )
        self.sheet.cell(row=6, column=1).value = (
            f"{self.period} SUMMARY STATUS OF CONTRIBUTIONS AND OTHER INCOME (US$)"
        )
        self.sheet.cell(row=7, column=1).value = (
            "BALANCE  AVAILABLE   FOR  NEW  ALLOCATIONS"
        )
        self.sheet.cell(row=8, column=1).value = "As at 24/05/2024"

    def write(self, data):
        self.do_prewriting()
        super().write(data)

    def write_header_cell(self, value, comment=None):
        cell = super().write_header_cell(value, comment)
        cell.font = Font(name="Times New Roman", bold=True)
        return cell

    def write_record_cell(self, value, read_only=False):
        cell = super().write_record_cell(value, read_only)
        if type(value) in (float, Decimal):
            cell.number_format = "###,###,##0.00###"

        if any(bold_string == value for bold_string in self.BOLD_RECORD_CELLS):
            cell.font = Font(name="Times New Roman", bold=True)

        return cell


class ScaleOfAssessmentWriter(WriteOnlyBase):
    # pylint: disable=too-many-arguments
    def __init__(
        self,
        sheet,
        past_un_period=None,
        past_period=None,
        current_period=None,
        ferm_year=None,
        comment=None,
    ):
        headers = [
            {
                "id": "no",
                "headerName": "No",
                "column_width": 5,
            },
            {
                "id": "country",
                "headerName": "Country",
                "column_width": 25,
            },
            {
                "id": "un_scale_of_assessment",
                "headerName": f"United Nations Scale of Assessment for the period {past_un_period}",
                "column_width": 25,
            },
            {
                "id": "adjusted_scale_of_assessment",
                "headerName": f"Adjusted UN Scale of Assessment using"
                f"{past_un_period} scale with no party "
                f"contributing more than 22%",
                "column_width": 25,
            },
            {
                "id": "yearly_amount",
                "headerName": f"Annual contributions for years {current_period} in (United States Dollar)",
                "column_width": 25,
            },
            {
                "id": "average_inflation_rate",
                "headerName": f"Average inflation rate for the period {past_period} (percent)",
                "column_width": 25,
            },
            {
                "id": "qualifies_for_fixed_rate_mechanism",
                "headerName": "Qualifying for fixed exchange rate mechanism",
                "column_width": 25,
            },
            {
                "id": "exchange_rate",
                "headerName": f"Fixed exchange rate mechanism users' "
                f"currencies rate of Exchange 01 Jan - 30 June "
                f"{ferm_year}",
                "column_width": 25,
            },
            {
                "id": "currency",
                "headerName": "Fixed exchange mechanism users national currencies",
                "column_width": 25,
            },
            {
                "id": "yearly_amount_local_currency",
                "headerName": "Fixed exchange mechanism users contribution "
                "amount in national currencies",
                "column_width": 25,
            },
        ]
        self.comment = comment
        super().__init__(sheet, headers)

    def write_header_cell(self, value, comment=None):
        cell = super().write_header_cell(value, comment)
        cell.font = Font(name="Times New Roman", bold=True)
        return cell

    def write_record_cell(self, value, read_only=False):
        cell = super().write_record_cell(value, read_only)
        if type(value) in (float, Decimal):
            cell.number_format = "###,###,##0.00###"
        cell.font = Font(name="Times New Roman")
        return cell

    def add_comments(self, merge_row):
        self.sheet.append([])
        self.sheet.merge_cells(
            start_row=merge_row,
            start_column=1,
            end_row=merge_row,
            end_column=10,
        )
        self.sheet.cell(row=merge_row, column=1).value = self.comment

    def write(self, data):
        super().write(data)
        # Data, header, total, empty row
        self.add_comments(len(data) + 3)


class BaseTemplateSheetWriter:
    """
    Base class for writing sheets based on an existing preloaded template sheet.

    No cell formatting, merging or table formatting are needed;
    just filling out values.

    Assumes a certain format of the template file, but easily configurable.
    """

    DATA_MAPPING = {}
    HEADERS_ROW = 1
    TEMPLATE_FIRST_DATA_ROW = 2
    TEMPLATE_LAST_DATA_ROW = 50

    CONVERT_BOOL_TO_NUMERIC = True

    # Distance between last data row and the totals row
    TOTALS_ROW_OFFSET = None

    def __init__(
        self,
        sheet,
        data,
        number_of_rows,
        start_year,
        as_of_date=None,
    ):
        self.sheet = sheet
        self.data = data
        self.number_of_rows = number_of_rows
        self.start_year = start_year
        self.as_of_date = as_of_date

    def write_headers(self):
        pass

    def _copy_row_format(self, source_row, new_row):

        for col in range(1, self.sheet.max_column + 1):

            source_cell = self.sheet.cell(row=source_row, column=col)
            new_cell = self.sheet.cell(row=new_row, column=col)

            new_cell.font = copy(source_cell.font)
            new_cell.border = copy(source_cell.border)
            new_cell.fill = copy(source_cell.fill)
            new_cell.alignment = copy(source_cell.alignment)
            new_cell.number_format = copy(source_cell.number_format)
            new_cell.protection = copy(source_cell.protection)

    def _delete_middle_rows(self, rows_number):
        start_row = self.TEMPLATE_FIRST_DATA_ROW + 1
        self.sheet.delete_rows(start_row, rows_number)

        # Shift merged cells to avoid footer breaking down
        merged_cells_range = self.sheet.merged_cells.ranges
        for merged_cell in merged_cells_range:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
            if min_row >= start_row:
                merged_cell.shift(0, -1 * rows_number)

    def _add_middle_rows(self, extra_rows):
        start_row = self.TEMPLATE_FIRST_DATA_ROW + 2

        # Shift merged cells to avoid footer breaking down
        merged_cells_range = self.sheet.merged_cells.ranges
        for merged_cell in merged_cells_range:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
            if min_row >= start_row:
                merged_cell.shift(0, extra_rows)

        self.sheet.insert_rows(start_row, extra_rows)

        # also copy row styles (based on row just above the inserted ones)
        for inserted_row in range(start_row, start_row + extra_rows):
            self._copy_row_format(start_row - 1, inserted_row)

    def write(self):
        """
        Writes the entire sheet.
        """
        # Overwrite headers
        self.write_headers()

        # Before writing, we need to:
        # - delete any extra rows that the template MIGHT have OR
        # - add any needed rows, copying an existing intermediate row's style
        template_data_rows_number = (
            self.TEMPLATE_LAST_DATA_ROW - self.TEMPLATE_FIRST_DATA_ROW + 1
        )
        data_rows_number = len(self.data)
        extra_rows = data_rows_number - template_data_rows_number
        if extra_rows < 0:
            self._delete_middle_rows(-1 * extra_rows)
        elif extra_rows > 0:
            self._add_middle_rows(extra_rows)

        # Write data
        for row_data in self.data:
            self.write_row(row_data, row_data["no"])

        # Write totals rows (existing formulas may need updating)
        self.write_totals()

    def write_row(self, row_data, row_index):
        """
        Writing row values for data rows.

        Totals row is calculated via formulas, leaving it alone.
        """
        row_to_overwrite = self.TEMPLATE_FIRST_DATA_ROW + row_index - 1
        for key in row_data.keys():
            if self.DATA_MAPPING.get(key) is None:
                continue
            column = self.DATA_MAPPING[key]["column"]
            cell = self.sheet.cell(row=row_to_overwrite, column=column)
            self.write_cell(
                cell,
                self.DATA_MAPPING[key].get("type"),
                self.DATA_MAPPING[key].get("format"),
                row_data[key],
            )

    def write_cell(self, cell, cell_type, cell_format, value):
        if value is None:
            value = "N/A"
            cell.alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=True
            )
            cell_format = None
        if self.CONVERT_BOOL_TO_NUMERIC:
            # Do not fall into the trap of using `if value`
            if value is True:
                value = 1
            if value is False:
                value = 0
        cell.value = value
        if cell_type in (Decimal, float) and cell_format is not None:
            cell.number_format = format

    def write_totals(self):
        """
        If the number of data rows changes, formulas need to be updated
        """
        if not self.TOTALS_ROW_OFFSET or not self.DATA_MAPPING:
            return
        # The last row of the actual output
        last_row = self.TEMPLATE_FIRST_DATA_ROW + len(self.data) - 1
        # The totals row of the actual output
        totals_row = last_row + self.TOTALS_ROW_OFFSET
        # Update sum values
        for key, item in self.DATA_MAPPING.items():
            cell = self.sheet.cell(totals_row, item["column"])
            value = cell.value
            if value and isinstance(value, str) and "=SUM" in value:
                cell.value = value.replace(
                    str(self.TEMPLATE_LAST_DATA_ROW), str(last_row)
                )


class ScaleOfAssessmentTemplateWriter(BaseTemplateSheetWriter):
    """
    Template sheet writer for SoA.
    """

    # Position and formatting for each filed from the serializer data rows
    DATA_MAPPING = {
        "no": {
            "column": 1,
            "type": int,
        },
        "country": {
            "column": 2,
            "type": str,
        },
        "un_scale_of_assessment": {
            "column": 3,
            "type": Decimal,
            "number_format": "###,###,##0.00###",
        },
        "adjusted_scale_of_assessment": {
            "column": 4,
            "type": Decimal,
            "number_format": "###,###,##0.00###",
        },
        "yearly_amount": {
            "column": 5,
            "type": Decimal,
            "number_format": "###,###,##0.00###",
        },
        "average_inflation_rate": {
            "column": 6,
            "type": Decimal,
            "number_format": "###,###,##0.00###",
        },
        "qualifies_for_fixed_rate_mechanism": {
            "column": 7,
            "type": bool,
        },
        "exchange_rate": {
            "column": 8,
            "type": Decimal,
            "number_format": "###,###,##0.00###",
        },
        # There is an empty column here, hence going from 8 to 10
        "currency": {
            "column": 10,
            "type": str,
        },
        # Another empty column here
        "yearly_amount_local_currency": {
            "column": 12,
            "type": Decimal,
            "number_format": "###,###,##0.00###",
        },
    }

    HEADERS_ROW = 1
    TEMPLATE_FIRST_DATA_ROW = 2
    TEMPLATE_LAST_DATA_ROW = 50

    # Totals row immediately under last data row
    TOTALS_ROW_OFFSET = 1

    def write_headers(self):
        for key, item in self.DATA_MAPPING.items():
            cell = self.sheet.cell(self.HEADERS_ROW, item["column"])
            value = cell.value
            if key == "un_scale_of_assessment":
                cell.value = value.replace(
                    "2022-24", f"{self.start_year - 2}-{self.start_year}"
                )
            if key == "adjusted_scale_of_assessment":
                cell.value = value.replace(
                    "2022-24", f"{self.start_year - 2}-{self.start_year - 2000}"
                )
            if key == "yearly_amount":
                cell.value = value.replace(
                    "2024, 2025 and 2026",
                    f"{self.start_year}, {self.start_year + 1} and {self.start_year + 2}",
                )
            if key == "average_inflation_rate":
                cell.value = value.replace(
                    "2021 -2023", f"{self.start_year - 3}-{self.start_year -  1}"
                )
            if key == "exchange_rate":
                cell.value = value.replace("(01 Jan - 30 June 2023)", "")


class StatusOfTheFundTemplateWriter(BaseTemplateSheetWriter):
    """
    Template sheet writer for Status of the Fund.
    """

    # Position and formatting for each filed from the serializer data rows
    DATA_MAPPING = {
        "3": {
            "column": 3,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "4": {
            "column": 4,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
    }

    TEMPLATE_FIRST_DATA_ROW = 12
    TEMPLATE_LAST_DATA_ROW = 42

    MEETING_ROW = 1
    MEETING_COLUMN = 2
    AS_OF_DATE_ROW = 8
    AS_OF_DATE_COLUMN = 2

    def write_headers(self):
        self.sheet.cell(column=self.MEETING_COLUMN, row=self.MEETING_ROW).value = (
            "UNEP/OzL.Pro/ExCom"
        )
        if self.as_of_date is not None:
            cell = self.sheet.cell(
                column=self.AS_OF_DATE_COLUMN, row=self.AS_OF_DATE_ROW
            )
            cell.value = f"As of {self.as_of_date.strftime('%d/%m/%Y')}"

    def write(self):
        self.write_headers()

        # Here data comes as tuples; no need for row shifting
        for index, row_data in enumerate(self.data):
            self.write_row(row_data, index)

    def write_row(self, row_data, row_index):
        """
        Writing row values for data rows.

        Here the values come as tuples and we need to do the mapping directly.
        """
        row_to_overwrite = self.TEMPLATE_FIRST_DATA_ROW + row_index

        # Skip empty rows
        if row_data is None:
            return
        if all(cell_data is None for cell_data in row_data):
            return

        for index, cell_data in enumerate(row_data[1:]):
            if cell_data is None:
                continue
            key = str(index + 3)
            column = self.DATA_MAPPING[key]["column"]
            cell = self.sheet.cell(row=row_to_overwrite, column=column)
            self.write_cell(
                cell,
                self.DATA_MAPPING[key].get("type"),
                self.DATA_MAPPING[key].get("format"),
                cell_data,
            )


class StatisticsTemplateWriter(BaseTemplateSheetWriter):
    """
    Template sheet writer for Statistics.
    """

    # All header cells are on column 1
    HEADER_COLUMN = 2
    TABLE_DESCRIPTION_ROW = 7

    AS_OF_DATE_ROW = 9
    AS_OF_DATE_COLUMN = HEADER_COLUMN

    MEETING_ROW = 1
    MEETING_COLUMN = 14

    TEMPLATE_FIRST_DATA_ROW = 11
    TEMPLATE_LAST_DATA_ROW = 35
    TEMPLATE_FIRST_DATA_COLUMN = 3

    def write_headers(self):
        description_cell = self.sheet.cell(
            row=self.TABLE_DESCRIPTION_ROW, column=self.HEADER_COLUMN
        )
        value = description_cell.value
        current_year = datetime.now().year
        description_cell.value = value.replace("2024", str(current_year))

        self.sheet.cell(column=self.MEETING_COLUMN, row=self.MEETING_ROW).value = (
            "UNEP/OzL.Pro/ExCom"
        )
        if self.as_of_date is not None:
            cell = self.sheet.cell(
                column=self.AS_OF_DATE_COLUMN, row=self.AS_OF_DATE_ROW
            )
            cell.value = f"As of {self.as_of_date.strftime('%d/%m/%Y')}"

    def write(self):
        self.write_headers()
        for column_index, triennial_data in enumerate(self.data):
            for row_key in triennial_data.keys():
                row_to_overwrite = self.TEMPLATE_FIRST_DATA_ROW + row_key
                column_to_overwrite = self.TEMPLATE_FIRST_DATA_COLUMN + column_index
                cell = self.sheet.cell(row=row_to_overwrite, column=column_to_overwrite)

                value = triennial_data[row_key] or Decimal(0.0)
                cell.value = value


class StatusOfContributionsSummaryTemplateWriter(BaseTemplateSheetWriter):
    # TODO: HEADERS_ROW could be a list
    HEADERS_ROW = 7
    TEMPLATE_FIRST_DATA_ROW = 11
    TEMPLATE_LAST_DATA_ROW = 65

    # Distance between last "normal" data row and the disputed contributions row
    DISPUTED_CONTRIBUTIONS_ROW_OFFSET = 2
    DISPUTED_CONTRIBUTIONS_COLUMN = 3

    # Position and formatting for each filed from the serializer data rows
    DATA_MAPPING = {
        "country": {
            "column": 2,
            "type": str,
        },
        "agreed_contributions": {
            "column": 3,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "cash_payments": {
            "column": 4,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "bilateral_assistance": {
            "column": 5,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "promissory_notes": {
            "column": 6,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "outstanding_contributions": {
            "column": 7,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "gain_loss": {
            "column": 8,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
    }

    def __init__(
        self,
        sheet,
        data,
        number_of_rows,
        start_year,
        as_of_date=None,
        disputed_contributions=None,
        ceit_data=None,
    ):
        self.disputed_contributions = disputed_contributions
        self.ceit_data = ceit_data
        super().__init__(sheet, data, number_of_rows, start_year, as_of_date)

    MEETING_ROW = 1
    MEETING_COLUMN = 2
    AS_OF_DATE_ROW = 8
    AS_OF_DATE_COLUMN = 2

    def write_headers(self):
        self.sheet.cell(column=self.MEETING_COLUMN, row=self.MEETING_ROW).value = (
            "UNEP/OzL.Pro/ExCom"
        )
        if self.as_of_date is not None:
            cell = self.sheet.cell(
                column=self.AS_OF_DATE_COLUMN, row=self.AS_OF_DATE_ROW
            )
            cell.value = f"As of {self.as_of_date.strftime('%d/%m/%Y')}"

    def write(self):
        """
        Overwriting base write() method to also write the disputed contributions row
        """
        super().write()

        if (
            self.disputed_contributions is not None
            and hasattr(self, "DISPUTED_CONTRIBUTIONS_ROW_OFFSET")
            and hasattr(self, "DISPUTED_CONTRIBUTIONS_COLUMN")
        ):
            # The last row of the actual output
            last_row = self.TEMPLATE_FIRST_DATA_ROW + len(self.data) - 1
            # The contributions row of the actual output
            disputed_row = last_row + self.DISPUTED_CONTRIBUTIONS_ROW_OFFSET

            cell = self.sheet.cell(
                row=disputed_row, column=self.DISPUTED_CONTRIBUTIONS_COLUMN
            )
            cell.value = self.disputed_contributions


class StatusOfContributionsTriennialTemplateWriter(
    StatusOfContributionsSummaryTemplateWriter
):
    # Only overwriting what's needed
    TEMPLATE_LAST_DATA_ROW = 59

    MEETING_ROW = 1
    MEETING_COLUMN = 6

    # Relative to the last data row
    CEIT_ROW_OFFSET = 5

    # Position and formatting for each filed from the serializer data rows
    DATA_MAPPING = {
        "country": {
            "column": 2,
            "type": str,
        },
        "agreed_contributions": {
            "column": 3,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "cash_payments": {
            "column": 4,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "bilateral_assistance": {
            "column": 5,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "promissory_notes": {
            "column": 6,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
        "outstanding_contributions": {
            "column": 7,
            "type": Decimal,
            "number_format": "###,###,##0",
        },
    }

    PERIOD_ROW = 7
    PERIOD_COLUMN = 2

    def write_headers(self):
        super().write_headers()

        cell = self.sheet.cell(row=self.PERIOD_ROW, column=self.PERIOD_COLUMN)
        value = cell.value
        cell.value = value.replace(
            "2024-2026", f"{self.start_year}-{self.start_year + 2}"
        )

    def write_ceit_row(self):
        if self.ceit_data is not None and hasattr(self, "CEIT_ROW_OFFSET"):
            # The last row of the actual output
            last_row = self.TEMPLATE_FIRST_DATA_ROW + len(self.data) - 1
            # The CEIT row of the actual output
            ceit_row = last_row + self.CEIT_ROW_OFFSET

            for key in self.ceit_data.keys():
                if self.DATA_MAPPING.get(key) is None:
                    continue
                column = self.DATA_MAPPING[key]["column"]
                cell = self.sheet.cell(row=ceit_row, column=column)
                self.write_cell(
                    cell,
                    self.DATA_MAPPING[key].get("type"),
                    self.DATA_MAPPING[key].get("format"),
                    self.ceit_data[key],
                )

    def write(self):
        super().write()

        self.write_ceit_row()


class StatusOfContributionsAnnualTemplateWriter(
    StatusOfContributionsTriennialTemplateWriter
):
    def write_headers(self):
        super().write_headers()

        cell = self.sheet.cell(row=self.PERIOD_ROW, column=self.PERIOD_COLUMN)
        value = cell.value
        cell.value = value.replace(
            f"{self.start_year}-{self.start_year + 2}", f"{self.start_year}"
        )


class ConsolidatedInputDataWriter:
    COLUMN_WIDTH = 15

    def __init__(self, wb):
        self.wb = wb

    def write_headers(self, ws, labels):
        for index, name in enumerate(labels):
            cell = ws.cell(row=1, column=index + 1)
            cell.value = name
            cell.font = Font(name=DEFAULT_FONT.name, bold=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            column_letter = get_column_letter(index + 1)
            ws.column_dimensions[column_letter].width = self.COLUMN_WIDTH

    def write_row(self, ws, row_index, row_data):
        for index, value in enumerate(row_data):
            cell = ws.cell(row=row_index, column=index + 1)
            cell.value = value
            cell.font = Font(name="Times New Roman")
            cell.border = Border(
                top=Side(style="thin"),
                left=Side(style="thin"),
                right=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            if type(value) in (float, Decimal):
                cell.number_format = "###,###,##0.00###"

    def write_data(self, ws, data):
        for index, row_data in enumerate(data):
            self.write_row(ws, index + 2, row_data)

    def write_agency_data(self, ws, data, agency, start_row):
        for index, row_data in enumerate(data):
            enhanced_row_data = (agency,) + row_data
            self.write_row(ws, start_row + index, enhanced_row_data)

    def export_bilateral(self, ws):
        columns = [
            "Country",
            "Year",
            "Meeting",
            "Decision",
            "Amount",
            "Comment",
        ]
        expressions = [
            "country__name",
            "year",
            "meeting__number",
            "decision_number",
            "amount",
            "comment",
        ]

        data = BilateralAssistance.objects.filter(
            Q(amount__gte=Decimal(5)) | Q(amount__lte=Decimal(-5))
        ).values_list(*expressions)
        self.write_headers(ws, columns)
        self.write_data(ws, data)

    def export_interest(self, ws):
        columns = [
            "Agency",
            "Year",
            "Meeting",
            "Quarter",
            "Amount",
            "Comment",
        ]
        expressions = [
            "agency_name",
            "year",
            "meeting__number",
            "quarter",
            "interest_earned",
            "comment",
        ]

        data = ExternalIncomeAnnual.objects.filter(
            Q(interest_earned__gte=Decimal(5)) | Q(interest_earned__lte=Decimal(-5)),
            year__isnull=False,
        ).values_list(*expressions)
        self.write_headers(ws, columns)
        self.write_data(ws, data)

    def export_miscellaneous_income(self, ws):
        columns = ["Year", "Meeting", "Amount", "Comment"]
        expressions_annual = [
            "year",
            "meeting__number",
            "miscellaneous_income",
            "comment",
        ]
        queryset_annual = ExternalIncomeAnnual.objects.filter(
            Q(miscellaneous_income__gte=Decimal(5))
            | Q(miscellaneous_income__lte=Decimal(-5)),
            year__isnull=False,
        )

        expressions_triennial = [
            "triennial_start_year",
            "meeting__number",
            "miscellaneous_income",
            "comment",
        ]
        queryset_triennial = ExternalIncomeAnnual.objects.filter(
            Q(miscellaneous_income__gte=Decimal(5))
            | Q(miscellaneous_income__lte=Decimal(-5)),
            triennial_start_year__isnull=False,
        )
        triennial_data = [
            (f"{data[0]}-{data[0] + 2}", *data[1:])
            for data in queryset_triennial.values_list(*expressions_triennial)
        ]

        data = chain(
            queryset_annual.values_list(*expressions_annual),
            triennial_data,
        )

        self.write_headers(ws, columns)
        self.write_data(ws, data)

    def export_disputed_contributions(self, ws):
        columns = ["Country", "Year", "Meeting", "Decision", "Amount", "Comment"]
        expressions = [
            "country__name",
            "year",
            "meeting",
            "decision_number",
            "amount",
            "comment",
        ]
        data = DisputedContribution.objects.filter(
            Q(amount__gte=Decimal(5)) | Q(amount__lte=Decimal(-5))
        ).values_list(*expressions)

        self.write_headers(ws, columns)
        self.write_data(ws, data)

    def export_allocations(self, ws):
        columns = ["Agency", "Year", "Meeting", "Amount", "Comment"]
        # Agency is inferred from another field
        expressions_first = ["year", "meeting__number"]
        expressions_second = ["comment"]
        agencies_mapping = [
            ("UNDP", "undp"),
            ("UNEP", "unep"),
            ("UNIDO", "unido"),
            ("World Bank", "world_bank"),
        ]
        self.write_headers(ws, columns)
        start_row = 2
        for agency_name, amount_field in agencies_mapping:
            agency_expressions = expressions_first + [amount_field] + expressions_second
            queryset = ExternalAllocation.objects.annotate(
                amount=F(amount_field)
            ).exclude(amount=Decimal(0))
            data = queryset.values_list(*agency_expressions)
            self.write_agency_data(
                ws,
                data,
                agency=agency_name,
                start_row=start_row,
            )
            start_row += queryset.count()

    def export_secretariat_budget(self, ws):
        columns = ["Meeting", "Decision", "Year", "Amount", "Comment"]
        expressions = [
            "meeting__number",
            "decision_number",
            "year",
            "staff_contracts",
            "comment",
        ]
        data = ExternalAllocation.objects.exclude(
            staff_contracts=Decimal(0)
        ).values_list(*expressions)

        self.write_headers(ws, columns)
        self.write_data(ws, data)

    def export_treasurer_budget(self, ws):
        columns = ["Meeting", "Decision", "Year", "Amount", "Comment"]
        expressions = [
            "meeting__number",
            "decision_number",
            "year",
            "treasury_fees",
            "comment",
        ]
        data = ExternalAllocation.objects.exclude(treasury_fees=Decimal(0)).values_list(
            *expressions
        )

        self.write_headers(ws, columns)
        self.write_data(ws, data)

    def export_evaluation_budget(self, ws):
        columns = ["Meeting", "Decision", "Year", "Amount", "Comment"]
        expressions = [
            "meeting__number",
            "decision_number",
            "year",
            "monitoring_fees",
            "comment",
        ]
        data = ExternalAllocation.objects.exclude(
            monitoring_fees=Decimal(0)
        ).values_list(*expressions)

        self.write_headers(ws, columns)
        self.write_data(ws, data)

    EXPORTED_SHEETS = [
        ("Bilateral", export_bilateral),
        ("Interest", export_interest),
        ("Misc Income", export_miscellaneous_income),
        ("Disputed Contributions", export_disputed_contributions),
        ("Allocations", export_allocations),
        ("Secretariat Budget", export_secretariat_budget),
        ("Treasurer Budget", export_treasurer_budget),
        ("Evaluation Budget", export_evaluation_budget),
    ]

    def write(self):
        for sheet_name, export_method in self.EXPORTED_SHEETS:
            ws = self.wb.create_sheet(sheet_name)
            configure_sheet_print(ws, "landscape")
            # pylint: : disable=E1121
            export_method(self, ws)
