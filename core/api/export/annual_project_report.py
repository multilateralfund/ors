from copy import copy
from collections import defaultdict
from datetime import datetime
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from core.api.serializers.annual_project_report import AnnualProjectReportReadSerializer
from core.models import AnnualProjectReport, ProjectStatus

# pylint: disable=R0902,R0911,R0913,R0914,R0915,R1705,W0212,C0302


class APRExportWriter:
    """
    Generates Excel export for APR using Annex I template.
    """

    TEMPLATE_PATH = (
        settings.ROOT_DIR / "api" / "export" / "templates" / "APRAnnexI.xlsx"
    )

    # Caching the raw template bytes so disk I/O only happens once per process.
    _template_bytes = None

    @classmethod
    def _get_template_workbook(cls):
        if cls._template_bytes is None:
            cls._template_bytes = cls.TEMPLATE_PATH.read_bytes()

        return load_workbook(BytesIO(cls._template_bytes))

    SHEET_NAME = "Annex I APR report "
    STATUS_SHEET_NAME = "Status Values"
    # Last header row (only one now present in the template)
    HEADER_ROW = 1
    FIRST_DATA_ROW = 2

    DATE_FORMAT = "dd/mm/yyyy"

    DATE_FIELDS = {
        "date_approved",
        "date_completion_proposal",
        "date_first_disbursement",
        "date_planned_completion",
        "date_actual_completion",
        "date_financial_completion",
        "date_of_completion_per_agreement_or_decisions",
    }

    NUMERIC_FIELDS = {
        "consumption_phased_out_odp_proposal",
        "consumption_phased_out_mt_proposal",
        "consumption_phased_out_co2_proposal",
        "production_phased_out_odp_proposal",
        "production_phased_out_mt_proposal",
        "production_phased_out_co2_proposal",
        "consumption_phased_out_odp",
        "consumption_phased_out_mt",
        "consumption_phased_out_co2",
        "production_phased_out_odp",
        "production_phased_out_mt",
        "production_phased_out_co2",
        "approved_funding",
        "adjustment",
        "approved_funding_plus_adjustment",
        "per_cent_funds_disbursed",
        "balance",
        "support_cost_approved",
        "support_cost_adjustment",
        "support_cost_approved_plus_adjustment",
        "support_cost_balance",
        "funds_disbursed",
        "funds_committed",
        "estimated_disbursement_current_year",
        "support_cost_disbursed",
        "support_cost_committed",
        "disbursements_made_to_final_beneficiaries",
        "funds_advanced",
    }

    PERCENTAGE_FIELDS = {
        "per_cent_funds_disbursed",
    }

    BOOLEAN_FIELDS = {
        "pcr_due",
        "gender_policy",
    }

    @classmethod
    def build_column_mapping(cls):
        """
        Generate column index-to-field mapping from serializer's excel_fields.
        This ensures consistency between serializer and Excel export.

        All fields — including any excluded ones — are mapped to their natural
        (1-indexed) column positions so that the output always stays aligned
        with the template headers. Excluded fields are written as None by
        _write_row_data rather than being removed from the mapping.
        """
        excel_fields = AnnualProjectReportReadSerializer.Meta.excel_fields
        # Map fields to column numbers (1-indexed)
        return {field: idx + 1 for idx, field in enumerate(excel_fields)}

    def __init__(
        self,
        year=None,
        agency_name=None,
        project_reports_data=None,
        exclude_fields=None,
        progress_callback=None,
    ):
        """
        If agency_name is None, the report includes all agencies.
        If year is None, it's a cumulative report for all years.
        If exclude_fields is provided, those columns are omitted from the export.
        If progress_callback is provided, it is called with (rows_written, total)
        every 100 rows while writing data.
        """
        self.year = year
        self.agency_name = agency_name
        self.project_reports_data = project_reports_data or []
        self.exclude_fields = exclude_fields
        self.progress_callback = progress_callback
        self.workbook = None
        self.worksheet = None
        self.status_worksheet = None
        self.column_mapping = self.build_column_mapping()
        self.status_column_idx = None

        # Find the status column index
        for field, idx in self.column_mapping.items():
            if field == "status":
                self.status_column_idx = idx
                break

    def _build_workbook(self):
        """Build the workbook with all data, formatting, and validation."""
        self.workbook = self._get_template_workbook()
        self.worksheet = self.workbook[self.SHEET_NAME]

        # Remove hidden sheets & extra columns; create status reference sheet
        self._remove_hidden_sheets()
        self._remove_extra_columns()
        self._create_status_sheet()

        # Write data rows
        self._write_data_rows()

        # Apply cell formatting (dates, numbers, booleans)
        self._apply_cell_formatting()

        # Apply data validation (status dropdown)
        self._apply_data_validation()

    def generate(self):
        self._build_workbook()
        return self._create_response()

    def generate_to_file(self, filepath):
        """Generate the Excel file and save it to the given file path."""
        self._build_workbook()
        self.workbook.save(filepath)

    def _remove_hidden_sheets(self):
        """
        Removes all hidden sheets from the template - except the ones we want to keep.
        """
        sheets_to_keep = {self.SHEET_NAME, self.STATUS_SHEET_NAME}
        sheets_to_remove = []

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            if sheet.sheet_state == "hidden" and sheet_name not in sheets_to_keep:
                sheets_to_remove.append(sheet_name)

        for sheet_name in sheets_to_remove:
            del self.workbook[sheet_name]

    def _remove_extra_columns(self):
        """
        Remove any extra columns - ensures only our data columns are exported.
        """
        max_needed_column = max(self.column_mapping.values())

        if self.worksheet.max_column > max_needed_column:
            self.worksheet.delete_cols(
                max_needed_column + 1, self.worksheet.max_column - max_needed_column
            )

    def _create_status_sheet(self):
        """
        Creates a sheet with all valid project status names from the DB.
        This will be used as the source for the status dropdown validation.
        """
        if self.STATUS_SHEET_NAME in self.workbook.sheetnames:
            self.status_worksheet = self.workbook[self.STATUS_SHEET_NAME]
        else:
            self.status_worksheet = self.workbook.create_sheet(self.STATUS_SHEET_NAME)

        statuses = ProjectStatus.objects.all().order_by("name")

        # Write the status names to the sheet (one per row, in column A)
        for idx, status in enumerate(statuses, start=1):
            self.status_worksheet.cell(row=idx, column=1, value=status.name)

    def _write_data_rows(self):
        """Write all project report data to the worksheet."""
        # The template's first data row is used as a formatting source for all other rows
        template_row = self.FIRST_DATA_ROW

        # Delete any extra template rows (if present)
        # We'll keep just the first template data row and duplicate it...
        if self.worksheet.max_row > template_row:
            self.worksheet.delete_rows(
                template_row + 1, self.worksheet.max_row - template_row
            )

        total = len(self.project_reports_data)

        # Insert all extra rows at once for speed optimization.
        # Eeach insert_rows() call shifts every row below it and would lead to O(n*n).
        if total > 1:
            self.worksheet.insert_rows(template_row + 1, total - 1)

        # Read the template's cell styles once, then reuse them for every inserted row.
        template_styles = self._read_row_styles(template_row)

        for idx, report_data in enumerate(self.project_reports_data):
            current_row = self.FIRST_DATA_ROW + idx
            if idx > 0:
                self._apply_cached_row_styles(template_styles, current_row)

            self._write_row_data(current_row, report_data)

            rows_written = idx + 1
            if self.progress_callback and rows_written % 100 == 0:
                self.progress_callback(rows_written, total)

    def _read_row_styles(self, source_row):
        """Reads all cell styles from a row into a column-index-keyed dictionary."""
        styles = {}
        for col_idx in range(1, len(self.column_mapping) + 1):
            src = self.worksheet.cell(source_row, col_idx)
            if src.has_style:
                styles[col_idx] = {
                    "font": copy(src.font),
                    "border": copy(src.border),
                    "fill": copy(src.fill),
                    "number_format": src.number_format,
                    "protection": copy(src.protection),
                    "alignment": copy(src.alignment),
                }
        return styles

    def _apply_cached_row_styles(self, styles, target_row):
        """Apply pre-read template cell styles to all cells of a target row."""
        for col_idx, style in styles.items():
            tgt = self.worksheet.cell(target_row, col_idx)
            tgt.font = copy(style["font"])
            tgt.border = copy(style["border"])
            tgt.fill = copy(style["fill"])
            tgt.number_format = style["number_format"]
            tgt.protection = copy(style["protection"])
            tgt.alignment = copy(style["alignment"])

    def _write_row_data(self, row_number, report_data):
        """Writes a single project report's data to the row identified by row_number"""
        excluded_fields = self.exclude_fields or set()
        for field_name, col_idx in self.column_mapping.items():
            cell = self.worksheet.cell(row_number, col_idx)
            value = (
                None
                if field_name in excluded_fields
                else self._format_field_value(field_name, report_data)
            )
            cell.value = value

    def _format_field_value(self, field_name, report_data):
        """
        Format field value appropriately for Excel.
        Dates are converted to date objects, numbers remain as numbers.
        """
        value = report_data.get(field_name)
        if value is None:
            return None

        # Handle numeric fields
        if field_name in self.NUMERIC_FIELDS:
            if isinstance(value, (int, float, Decimal)):
                return float(value)
            return None

        # Handle boolean fields - convert to Yes/No
        if field_name in self.BOOLEAN_FIELDS or isinstance(value, bool):
            if value in (True, 1, "1"):
                return "Yes"
            if value in (False, 0, "0"):
                return "No"
            return ""

        # Handle date fields - convert to date object for Excel
        if field_name in self.DATE_FIELDS:
            result = None
            if isinstance(value, str):
                try:
                    result = datetime.fromisoformat(value.replace("Z", "+00:00")).date()
                except (ValueError, AttributeError):
                    pass
            return result

        # All other fields as string
        return str(value) if value else ""

    def _apply_cell_formatting(self):
        """
        Applies proper formats to date, numeric, and boolean columns.
        """
        if not self.project_reports_data:
            return

        first_row = self.FIRST_DATA_ROW
        last_row = self.FIRST_DATA_ROW + len(self.project_reports_data) - 1

        for field_name, col_idx in self.column_mapping.items():
            if field_name in self.DATE_FIELDS:
                for row in range(first_row, last_row + 1):
                    cell = self.worksheet.cell(row, col_idx)
                    cell.number_format = self.DATE_FORMAT

            elif field_name in self.PERCENTAGE_FIELDS:
                for row in range(first_row, last_row + 1):
                    cell = self.worksheet.cell(row, col_idx)
                    cell.number_format = "0.00%"

            elif field_name in self.NUMERIC_FIELDS:
                for row in range(first_row, last_row + 1):
                    cell = self.worksheet.cell(row, col_idx)
                    # Format with thousands separator and 2 decimal places
                    cell.number_format = "#,##0.00"

            elif field_name in self.BOOLEAN_FIELDS:
                for row in range(first_row, last_row + 1):
                    cell = self.worksheet.cell(row, col_idx)
                    # Text format ensures Yes/No displays properly
                    cell.number_format = "@"

    def _apply_data_validation(self):
        """
        Apply data validation to the status column.
        Users can only select values from the Status Values sheet.
        """
        if not self.status_column_idx or not self.project_reports_data:
            return

        status_count = ProjectStatus.objects.count()
        if status_count == 0:
            return

        # Create data validation referencing the status sheet
        # Formula references the range in the Status Values sheet
        status_validation = DataValidation(
            type="list",
            formula1=f"'{self.STATUS_SHEET_NAME}'!$A$1:$A${status_count}",
            allow_blank=True,
            showErrorMessage=True,
            error="Please select a valid status from the dropdown list.",
            errorTitle="Invalid Status",
        )

        self.worksheet.add_data_validation(status_validation)

        # Apply validation to all status cells in data rows
        first_row = self.FIRST_DATA_ROW
        last_row = self.FIRST_DATA_ROW + len(self.project_reports_data) - 1
        col_letter = get_column_letter(self.status_column_idx)

        # Add the range to the validation
        status_validation.add(f"{col_letter}{first_row}:{col_letter}{last_row}")

    def _create_response(self):
        if self.agency_name:
            safe_agency_name = "".join(
                c for c in self.agency_name if c.isalnum() or c in (" ", "-", "_")
            ).strip()
            if self.year:
                filename = f"APR_{self.year}_{safe_agency_name}.xlsx"
            else:
                filename = "APR_Cumulative_{safe_agency_name}.xlsx"
        elif self.year:
            # This is a multi-agency report (for MLFS to edit)
            filename = f"APR_{self.year}_All_Agencies.xlsx"
        else:
            filename = "APR_Cumulative_All_Agencies.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        self.workbook.save(response)
        return response


class APRSummaryTablesExportWriter:
    """
    Generates multi-sheet Excel export with summary tables for APR.
    Unlike APRExportWriter, this includes *all* data regardless of UI filters.
    """

    TEMPLATE_PATH = (
        settings.ROOT_DIR / "api" / "export" / "templates" / "APRSummaryTables.xlsx"
    )

    # Caching the raw template bytes so disk I/O only happens once per process.
    _template_bytes = None

    @classmethod
    def _get_template_workbook(cls):
        if cls._template_bytes is None:
            cls._template_bytes = cls.TEMPLATE_PATH.read_bytes()

        return load_workbook(BytesIO(cls._template_bytes))

    @staticmethod
    def _get_field_value(obj, field_path):
        """Utility to traverse a double-underscore field path on a Django ORM object"""
        for part in field_path.split("__"):
            obj = getattr(obj, part, None)
            if obj is None:
                return None
        return obj

    # Sheet name constants
    SHEET_SUMMARY = "I.1 Summary Data "
    SHEET_SUMMARY_CLUSTER = "I.2 Summary data by cluster"
    SHEET_COMPLETION_YEAR = "Project completion_Rep year"
    SHEET_ANNUAL = "Annex I (a)"
    SHEET_INVESTMENT = "Annex I (b)"
    SHEET_NON_INVESTMENT = "Annex I (c)"
    SHEET_PREPARATION = "Annex I (d)"
    SHEET_ONGOING_INVESTMENT = "Annex I (e)"
    SHEET_ONGOING_NON_INVESTMENT = "Annex I (f)"
    SHEET_ONGOING_PREPARATION = "Annex I (g)"

    # Row position constants - shared by sheets (a) through (g)
    HEADER_ROW = 3
    DATA_START_ROW = 4

    # I.1 Summary Data fixed row positions
    SUMMARY_DATA_START_ROW = 4

    # I.2 and Project completion_Rep year data start (after merged header rows 3-4)
    CLUSTER_DATA_START_ROW = 5
    # Number of template placeholder data rows in the cluster sheets (before Total)
    CLUSTER_TEMPLATE_DATA_ROWS = 5

    @classmethod
    def build_annual_column_mapping(cls):
        """Column mapping for Annual summary sheet (a)"""
        return {
            "approval_year": 1,
            "num_approvals": 2,
            "num_completed": 3,
            "pct_completed": 4,
            "approved_funding": 5,
            "funds_disbursed": 6,
            "balance": 7,
            "pct_funds_disbursed": 8,
        }

    def __init__(self, agency=None, year=None):
        self.agency = agency
        self.year = year
        self.workbook = None
        self.annual_column_mapping = self.build_annual_column_mapping()

        # Get all APR data
        queryset = AnnualProjectReport.objects.all().select_related(
            "project",
            "project__agency",
            "project__cluster",
            "project__country",
            "project__country__parent",
            "project__sector",
            "project__status",
            "project__project_type",
            "report",
            "report__progress_report",
        )

        status_names = list(
            ProjectStatus.objects.filter(
                code__in=["COM", "FIN", "ONG"]
            ).values_list("name", flat=True)
        )
        queryset = queryset.filter(status__in=status_names)
        if self.agency:
            queryset = queryset.filter(project__agency=self.agency)
        self.queryset = queryset
        # Materialize the queryset once so we don't keep re-querying the DB for each tab
        self.records = list(queryset)

        self.serialized_data = AnnualProjectReportReadSerializer(
            self.records, many=True
        ).data

    def generate(self):
        self.workbook = self._get_template_workbook()

        # Remove hidden sheets from the template
        self._remove_hidden_sheets()

        # I.1: Overall summary data
        self._write_summary_data_sheet()

        # I.2: Summary by cluster
        self._write_summary_by_cluster_sheet()

        # Project completion in the reporting year
        self._write_project_completion_year_sheet()

        # (a) Annual summary by approval year
        self._write_annual_summary_sheet()

        # (b) Completed investment projects
        self._write_investment_projects_sheet()

        # (c) Completed non-investment projects
        self._write_non_investment_projects_sheet()

        # (d) Completed preparation activities
        self._write_preparation_activities_sheet()

        # (e) Ongoing investment projects
        self._write_ongoing_investment_sheet()

        # (f) Ongoing non-investment projects
        self._write_ongoing_non_investment_sheet()

        # (g) Ongoing preparation activities
        self._write_ongoing_preparation_sheet()

        return self._create_response()

    def _remove_hidden_sheets(self):
        """Remove all hidden sheets from the template."""
        sheets_to_remove = [
            name
            for name in self.workbook.sheetnames
            if self.workbook[name].sheet_state == "hidden"
        ]
        for name in sheets_to_remove:
            del self.workbook[name]

    def _clear_template_data_rows(self, ws, start_row, end_row=None):
        """
        Clear any template/sample data from the worksheet.
        Does *not* delete rows as this would shift row numbers; just clears contents.
        """
        if end_row is None:
            end_row = ws.max_row

        for row in range(start_row, end_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).value = None

    def _find_section_row(self, ws, header_text, search_from):
        """Return the row number where col-1 matches header_text (case-insensitive)."""
        needle = header_text.strip().lower()
        for r in range(search_from, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and str(val).strip().lower() == needle:
                return r
        return None

    def _copy_row_style_summary(self, ws, source_row, target_row):
        """Copy all cell styles from source_row to target_row."""
        for col in range(1, ws.max_column + 1):
            src = ws.cell(source_row, col)
            tgt = ws.cell(target_row, col)
            if src.has_style:
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.number_format = src.number_format
                tgt.protection = copy(src.protection)
                tgt.alignment = copy(src.alignment)

    def _adjust_section_rows(self, ws, section_start, template_count, needed_count):
        """
        Resize a contiguous data section from template_count rows to needed_count rows
        by inserting or deleting rows (and carrying styles along when inserting).
        """
        if needed_count == template_count:
            return
        if needed_count > template_count:
            n_insert = needed_count - template_count
            insert_at = section_start + template_count
            ws.insert_rows(insert_at, n_insert)
            style_source = (
                section_start + template_count - 1
                if template_count > 0
                else section_start
            )
            for i in range(n_insert):
                self._copy_row_style_summary(ws, style_source, insert_at + i)
        else:
            ws.delete_rows(section_start + needed_count, template_count - needed_count)

    def _write_summary_data_sheet(self):
        """I.1: Overall summary data"""
        ws = self.workbook[self.SHEET_SUMMARY]

        num_approvals = len(self.records)
        num_completed = sum(
            1
            for apr in self.records
            if apr.project.status and apr.project.status.code == "COM"
        )

        total_funds_approved = sum(
            apr.approved_funding_plus_adjustment_denorm or 0 for apr in self.records
        )
        total_funds_disbursed = sum(apr.funds_disbursed or 0 for apr in self.records)

        # Approved phase-out = sum of proposal/denorm values (consumption + production)
        total_approved_odp = sum(
            (apr.consumption_phased_out_odp_proposal_denorm or 0)
            + (apr.production_phased_out_odp_proposal_denorm or 0)
            for apr in self.records
        )
        total_actual_odp = sum(
            (apr.consumption_phased_out_odp or 0) + (apr.production_phased_out_odp or 0)
            for apr in self.records
        )
        total_approved_mt = sum(
            (apr.consumption_phased_out_mt_proposal_denorm or 0)
            + (apr.production_phased_out_mt_proposal_denorm or 0)
            for apr in self.records
        )
        total_actual_mt = sum(
            (apr.consumption_phased_out_mt or 0) + (apr.production_phased_out_mt or 0)
            for apr in self.records
        )
        total_approved_co2 = sum(
            (apr.consumption_phased_out_co2_proposal_denorm or 0)
            + (apr.production_phased_out_co2_proposal_denorm or 0)
            for apr in self.records
        )
        total_actual_co2 = sum(
            (apr.consumption_phased_out_co2 or 0) + (apr.production_phased_out_co2 or 0)
            for apr in self.records
        )

        # Write values to column B, rows 4-13
        values = [
            num_approvals,
            num_completed,
            total_funds_approved,
            total_funds_disbursed,
            total_approved_odp,
            total_actual_odp,
            total_approved_mt,
            total_actual_mt,
            total_approved_co2,
            total_actual_co2,
        ]

        for idx, value in enumerate(values):
            cell = ws.cell(self.SUMMARY_DATA_START_ROW + idx, 2, value)
            cell.number_format = "#,##0"

    def _write_summary_by_cluster_sheet(self):
        """I.2: Summary data by cluster"""
        ws = self.workbook[self.SHEET_SUMMARY_CLUSTER]

        # Group the in-memory objects list by cluster, sorted by (sort_order, name)
        cluster_buckets = {}
        for apr in self.records:
            cluster = apr.project.cluster
            if cluster is None:
                continue
            key = (getattr(cluster, "sort_order", 0) or 0, cluster.name)
            if key not in cluster_buckets:
                cluster_buckets[key] = []
            cluster_buckets[key].append(apr)

        # Pre-compute all rows before touching the sheet structure
        rows_data = []
        totals = {
            "num_approved": 0,
            "num_completed": 0,
            "approved_funding": 0,
            "disbursed": 0,
            "balance": 0,
        }
        for (_, cluster_name), cluster_records in sorted(cluster_buckets.items()):
            num_approved = len(cluster_records)
            num_completed = sum(
                1
                for apr in cluster_records
                if apr.project.status and apr.project.status.code == "COM"
            )
            pct_completed = (
                round(num_completed / num_approved * 100) if num_approved else 0
            )
            approved_funding = sum(
                apr.approved_funding_plus_adjustment_denorm or 0
                for apr in cluster_records
            )
            disbursed = sum(apr.funds_disbursed or 0 for apr in cluster_records)
            balance = approved_funding - disbursed
            pct_disbursed = (
                round(disbursed / approved_funding * 100) if approved_funding else 0
            )
            rows_data.append(
                (
                    cluster_name,
                    num_approved,
                    num_completed,
                    pct_completed,
                    approved_funding,
                    disbursed,
                    balance,
                    pct_disbursed,
                )
            )
            totals["num_approved"] += num_approved
            totals["num_completed"] += num_completed
            totals["approved_funding"] += approved_funding
            totals["disbursed"] += disbursed
            totals["balance"] += balance

        n_rows = len(rows_data)

        # Resize data section; Total row (already labelled in template) shifts with it
        self._adjust_section_rows(
            ws, self.CLUSTER_DATA_START_ROW, self.CLUSTER_TEMPLATE_DATA_ROWS, n_rows
        )

        # Clear data rows; preserve the Total row label (col 1) from the template
        total_row = self.CLUSTER_DATA_START_ROW + n_rows
        self._clear_template_data_rows(ws, self.CLUSTER_DATA_START_ROW, total_row - 1)
        for col in range(2, ws.max_column + 1):
            ws.cell(total_row, col).value = None

        # Write data rows
        row = self.CLUSTER_DATA_START_ROW
        for (
            cluster_name,
            num_approved,
            num_completed,
            pct_completed,
            approved_funding,
            disbursed,
            balance,
            pct_disbursed,
        ) in rows_data:
            ws.cell(row, 1, cluster_name)
            ws.cell(row, 2, num_approved)
            ws.cell(row, 3, num_completed)
            ws.cell(row, 4, pct_completed)
            ws.cell(row, 5, approved_funding)
            ws.cell(row, 6, disbursed)
            ws.cell(row, 7, balance)
            ws.cell(row, 8, pct_disbursed)
            for col in range(2, 9):
                ws.cell(row, col).number_format = "#,##0"
            row += 1

        # Total row: label already in template, write only the values
        total_pct_completed = (
            round(totals["num_completed"] / totals["num_approved"] * 100)
            if totals["num_approved"]
            else 0
        )
        total_pct_disbursed = (
            round(totals["disbursed"] / totals["approved_funding"] * 100)
            if totals["approved_funding"]
            else 0
        )
        ws.cell(total_row, 2, totals["num_approved"])
        ws.cell(total_row, 3, totals["num_completed"])
        ws.cell(total_row, 4, total_pct_completed)
        ws.cell(total_row, 5, totals["approved_funding"])
        ws.cell(total_row, 6, totals["disbursed"])
        ws.cell(total_row, 7, totals["balance"])
        ws.cell(total_row, 8, total_pct_disbursed)
        for col in range(2, 9):
            ws.cell(total_row, col).number_format = "#,##0"

    def _write_project_completion_year_sheet(self):
        """Project completion in the reporting year, grouped by cluster"""
        ws = self.workbook[self.SHEET_COMPLETION_YEAR]

        # Filter and group by cluster in-memory, to avoid repeated DB hits
        if self.year:
            completed_records = [
                apr
                for apr in self.records
                if apr.project.status
                and apr.project.status.code == "COM"
                and apr.report
                and apr.report.progress_report
                and apr.report.progress_report.year == self.year
            ]
        else:
            completed_records = [
                apr
                for apr in self.records
                if apr.project.status and apr.project.status.code == "COM"
            ]

        cluster_buckets = {}
        for apr in completed_records:
            cluster = apr.project.cluster
            if cluster is None:
                continue
            key = (getattr(cluster, "sort_order", 0) or 0, cluster.name)
            if key not in cluster_buckets:
                cluster_buckets[key] = []
            cluster_buckets[key].append(apr)

        # Pre-compute all rows before touching the sheet structure
        rows_data = []
        totals = {"num_completed": 0, "odp": 0, "mt": 0, "co2": 0}
        for (_, cluster_name), cluster_records in sorted(cluster_buckets.items()):
            num_completed = len(cluster_records)
            total_odp = sum(
                (apr.consumption_phased_out_odp or 0)
                + (apr.production_phased_out_odp or 0)
                for apr in cluster_records
            )
            total_mt = sum(
                (apr.consumption_phased_out_mt or 0)
                + (apr.production_phased_out_mt or 0)
                for apr in cluster_records
            )
            total_co2 = sum(
                (apr.consumption_phased_out_co2 or 0)
                + (apr.production_phased_out_co2 or 0)
                for apr in cluster_records
            )
            rows_data.append(
                (cluster_name, num_completed, total_odp, total_mt, total_co2)
            )
            totals["num_completed"] += num_completed
            totals["odp"] += total_odp
            totals["mt"] += total_mt
            totals["co2"] += total_co2

        n_rows = len(rows_data)

        # Resize data section; Total row (already labelled in template) shifts with it
        self._adjust_section_rows(
            ws, self.CLUSTER_DATA_START_ROW, self.CLUSTER_TEMPLATE_DATA_ROWS, n_rows
        )

        # Clear data rows; preserve the Total row label (col 1) from the template
        total_row = self.CLUSTER_DATA_START_ROW + n_rows
        self._clear_template_data_rows(ws, self.CLUSTER_DATA_START_ROW, total_row - 1)
        for col in range(2, ws.max_column + 1):
            ws.cell(total_row, col).value = None

        # Write data rows
        row = self.CLUSTER_DATA_START_ROW
        for cluster_name, num_completed, total_odp, total_mt, total_co2 in rows_data:
            ws.cell(row, 1, cluster_name)
            ws.cell(row, 2, num_completed)
            ws.cell(row, 3, total_odp)
            ws.cell(row, 4, total_mt)
            ws.cell(row, 5, total_co2)
            for col in range(2, 6):
                ws.cell(row, col).number_format = "#,##0"
            row += 1

        # Total row: label already in template, write only the values
        ws.cell(total_row, 2, totals["num_completed"])
        ws.cell(total_row, 3, totals["odp"])
        ws.cell(total_row, 4, totals["mt"])
        ws.cell(total_row, 5, totals["co2"])
        for col in range(2, 6):
            ws.cell(total_row, col).number_format = "#,##0"

    def _write_annual_row(self, ws, row, col_map, row_data, bold=False):
        """Helper for writing a single data row in the annual summary sheet."""
        label_cell = ws.cell(row, col_map["approval_year"], row_data["label"])
        if bold:
            label_cell.font = Font(bold=True)
        ws.cell(row, col_map["num_approvals"], row_data["num_approvals"])
        ws.cell(row, col_map["num_completed"], row_data["num_completed"])
        ws.cell(row, col_map["pct_completed"], row_data["pct_completed"])
        ws.cell(row, col_map["pct_completed"]).number_format = "0%"
        ws.cell(row, col_map["approved_funding"], row_data["approved_funding"])
        ws.cell(row, col_map["funds_disbursed"], row_data["funds_disbursed"])
        ws.cell(row, col_map["balance"], row_data["balance"])
        ws.cell(row, col_map["pct_funds_disbursed"], row_data["pct_funds_disbursed"])

        for col_name in ["approved_funding", "funds_disbursed", "balance"]:
            ws.cell(row, col_map[col_name]).number_format = "#,##0"
        ws.cell(row, col_map["pct_funds_disbursed"]).number_format = "0"

    def _write_annual_summary_sheet(self):
        """Sheet (a): Annual summary data by approval year"""
        ws = self.workbook[self.SHEET_ANNUAL]

        self._clear_template_data_rows(ws, self.DATA_START_ROW)

        # Group serialized data by approval year
        year_data = defaultdict(list)
        for item in self.serialized_data:
            date_approved = item.get("date_approved")
            if date_approved:
                if isinstance(date_approved, str):
                    try:
                        year = datetime.fromisoformat(
                            date_approved.replace("Z", "+00:00")
                        ).year
                    except (ValueError, AttributeError):
                        continue
                else:
                    continue
                year_data[year].append(item)

        # Process each year
        col_map = self.annual_column_mapping
        row = self.DATA_START_ROW

        # Running totals for the Total row
        grand_totals = {
            "num_approvals": 0,
            "num_completed": 0,
            "approved_funding": 0,
            "funds_disbursed": 0,
            "balance": 0,
        }

        for approval_year in sorted(year_data.keys()):
            year_projects = year_data[approval_year]

            num_approvals = len(year_projects)
            num_completed = sum(1 for p in year_projects if p.get("status") == "COM")

            total_approved_funding = sum(
                p.get("approved_funding_plus_adjustment") or 0 for p in year_projects
            )
            total_balance = sum(p.get("balance") or 0 for p in year_projects)
            total_funds_disbursed = sum(
                p.get("funds_disbursed") or 0 for p in year_projects
            )

            avg_pct_disbursed = (
                total_funds_disbursed / total_approved_funding * 100
                if total_approved_funding
                else 0
            )

            pct_completed = (num_completed / num_approvals) if num_approvals else 0

            self._write_annual_row(
                ws,
                row,
                col_map,
                {
                    "label": approval_year,
                    "num_approvals": num_approvals,
                    "num_completed": num_completed,
                    "pct_completed": pct_completed,
                    "approved_funding": total_approved_funding,
                    "funds_disbursed": total_funds_disbursed,
                    "balance": total_balance,
                    "pct_funds_disbursed": avg_pct_disbursed,
                },
            )

            # Accumulate grand totals
            grand_totals["num_approvals"] += num_approvals
            grand_totals["num_completed"] += num_completed
            grand_totals["approved_funding"] += total_approved_funding
            grand_totals["funds_disbursed"] += total_funds_disbursed
            grand_totals["balance"] += total_balance

            row += 1

        # Write Total row
        gt = grand_totals
        pct_completed_total = (
            (gt["num_completed"] / gt["num_approvals"]) if gt["num_approvals"] else 0
        )
        avg_pct_total = (
            gt["funds_disbursed"] / gt["approved_funding"] * 100
            if gt["approved_funding"]
            else 0
        )

        self._write_annual_row(
            ws,
            row,
            col_map,
            {
                "label": "Total",
                "num_approvals": gt["num_approvals"],
                "num_completed": gt["num_completed"],
                "pct_completed": pct_completed_total,
                "approved_funding": gt["approved_funding"],
                "funds_disbursed": gt["funds_disbursed"],
                "balance": gt["balance"],
                "pct_funds_disbursed": avg_pct_total,
            },
            bold=True,
        )

    def _filter_records(self, status_code=None, type_code=None, exclude_type_codes=()):
        """Returns a filtered sub-list of self.records, using in-memory filtering"""
        result = []
        for apr in self.records:
            if status_code and not (
                apr.project.status and apr.project.status.code == status_code
            ):
                continue
            pt_code = (
                apr.project.project_type.code if apr.project.project_type else None
            )
            if type_code and pt_code != type_code:
                continue
            if exclude_type_codes and pt_code in exclude_type_codes:
                continue
            result.append(apr)
        return result

    def _write_investment_projects_sheet(self):
        """Sheet (b): Cumulative completed investment projects"""
        ws = self.workbook[self.SHEET_INVESTMENT]
        self._write_flat_aggregation_sheet(
            ws,
            self._filter_records(status_code="COM", type_code="INV"),
            include_odp_co2=True,
            sheet_type="cumulative",
        )

    def _write_non_investment_projects_sheet(self):
        """Sheet (c): Cumulative completed non-investment projects"""
        ws = self.workbook[self.SHEET_NON_INVESTMENT]
        self._write_flat_aggregation_sheet(
            ws,
            self._filter_records(status_code="COM", exclude_type_codes=("INV", "PRP")),
            include_odp_co2=False,
            sheet_type="cumulative",
        )

    def _write_preparation_activities_sheet(self):
        """Sheet (d): Cumulative completed project preparation activities"""
        ws = self.workbook[self.SHEET_PREPARATION]
        self._write_flat_aggregation_sheet(
            ws,
            self._filter_records(status_code="COM", type_code="PRP"),
            include_odp_co2=False,
            sheet_type="cumulative",
        )

    def _write_ongoing_investment_sheet(self):
        """Sheet (e): Cumulative ongoing investment projects"""
        ws = self.workbook[self.SHEET_ONGOING_INVESTMENT]
        self._write_flat_aggregation_sheet(
            ws,
            self._filter_records(status_code="ONG", type_code="INV"),
            include_odp_co2=True,
            sheet_type="ongoing_investment",
        )

    def _write_ongoing_non_investment_sheet(self):
        """Sheet (f): Cumulative ongoing non-investment projects"""
        ws = self.workbook[self.SHEET_ONGOING_NON_INVESTMENT]
        self._write_flat_aggregation_sheet(
            ws,
            self._filter_records(status_code="ONG", exclude_type_codes=("INV", "PRP")),
            include_odp_co2=False,
            sheet_type="ongoing_non_investment",
        )

    def _write_ongoing_preparation_sheet(self):
        """Sheet (g): Cumulative ongoing preparation activities"""
        ws = self.workbook[self.SHEET_ONGOING_PREPARATION]
        self._write_flat_aggregation_sheet(
            ws,
            self._filter_records(status_code="ONG", type_code="PRP"),
            include_odp_co2=False,
            sheet_type="ongoing_preparation",
        )

    def _write_flat_aggregation_sheet(self, ws, records, include_odp_co2, sheet_type):
        """
        Write flat aggregation layout, as used by sheets (b)-(g).
        Structure: Total row, "Region" label, region data, "Sector" label, sector data.
        Inserts or deletes rows as needed so the section sizes match the actual data.
        """
        column_specs = self._get_column_specs(sheet_type, include_odp_co2)

        # Pre-compute all data before touching the sheet structure
        total_data = self._compute_group_data(records, include_odp_co2, sheet_type)
        region_items = list(
            self._compute_grouped_data(
                records,
                "project__country__parent__name",
                include_odp_co2,
                sheet_type,
            )
        )
        sector_items = list(
            self._compute_grouped_data(
                records,
                "project__sector__code",
                include_odp_co2,
                sheet_type,
            )
        )
        n_regions = len(region_items)
        n_sectors = len(sector_items)

        # Locate the "Sector" section header in the template to measure section sizes.
        # The template layout (by rows) looks like this:
        # DATA_START_ROW            -> Total
        # +1                        -> Region header
        # +2...sector_header-1      -> region rows
        # sector_header             -> Sector header
        # sector_header+1...max_row -> sector rows
        sector_header_row = self._find_section_row(
            ws, "Sector", self.DATA_START_ROW + 1
        )
        if sector_header_row is None:
            sector_header_row = ws.max_row + 1  # degenerate fallback

        template_n_regions = sector_header_row - self.DATA_START_ROW - 2
        template_n_sectors = ws.max_row - sector_header_row

        # Resize the region data section
        region_data_start = self.DATA_START_ROW + 2
        self._adjust_section_rows(ws, region_data_start, template_n_regions, n_regions)

        # After adjustment the Sector header has moved to:
        new_sector_header_row = region_data_start + n_regions

        # Resize the sector data section
        sector_data_start = new_sector_header_row + 1
        self._adjust_section_rows(ws, sector_data_start, template_n_sectors, n_sectors)

        # Clear only data rows — section headers ("Region", "Sector") come from the template
        self._clear_template_data_rows(ws, self.DATA_START_ROW, self.DATA_START_ROW)
        self._clear_template_data_rows(
            ws, region_data_start, region_data_start + n_regions - 1
        )
        self._clear_template_data_rows(
            ws, sector_data_start, sector_data_start + n_sectors - 1
        )

        # Write Total row (label also in template, but overwriting is harmless)
        self._write_aggregation_row(
            ws, self.DATA_START_ROW, "Total", total_data, column_specs, bold=True
        )

        # Region data rows (header already in template at DATA_START_ROW + 1)
        for row, (group_name, group_data) in enumerate(
            region_items, start=region_data_start
        ):
            self._write_aggregation_row(ws, row, group_name, group_data, column_specs)

        # Sector data rows (header already in template at new_sector_header_row)
        for row, (group_name, group_data) in enumerate(
            sector_items, start=sector_data_start
        ):
            self._write_aggregation_row(ws, row, group_name, group_data, column_specs)

    def _write_aggregation_row(self, ws, row, label, data, column_specs, bold=False):
        """Write a single aggregation data row to the worksheet."""
        cell = ws.cell(row, 1, label)
        if bold:
            cell.font = Font(bold=True)

        col = 2
        for field_name, number_format in column_specs:
            cell = ws.cell(row, col, data.get(field_name) or 0)
            if number_format:
                cell.number_format = number_format
            col += 1

    def _get_column_specs(self, sheet_type, include_odp_co2):
        """Returns list of (field_name, number_format) tuples for data columns"""
        if sheet_type == "ongoing_investment":
            return [
                ("num_projects", "#,##0"),
                ("total_approved_funding", "#,##0"),
                ("total_funds_disbursed", "#,##0"),
                ("avg_pct_disbursed", "#,##0"),
                ("num_disbursing", "#,##0"),
                ("pct_disbursing", "0%"),
                ("avg_months_to_disbursement", "#,##0"),
                ("avg_months_to_completion", "#,##0"),
                ("avg_delay", "#,##0"),
                ("total_phaseout_combined_kg", "#,##0"),
                ("cost_effectiveness", "#,##0.00"),
            ]
        elif sheet_type == "ongoing_non_investment":
            return [
                ("num_projects", "#,##0"),
                ("total_approved_funding", "#,##0"),
                ("total_funds_disbursed", "#,##0"),
                ("avg_pct_disbursed", "#,##0"),
                ("num_disbursing", "#,##0"),
                ("pct_disbursing", "0%"),
                ("avg_months_to_disbursement", "#,##0"),
                ("avg_months_to_completion", "#,##0"),
                ("avg_delay", "#,##0"),
            ]
        elif sheet_type == "ongoing_preparation":
            return [
                ("num_projects", "#,##0"),
                ("total_approved_funding", "#,##0"),
                ("total_funds_disbursed", "#,##0"),
                ("avg_pct_disbursed", "#,##0"),
                ("avg_months_to_disbursement", "#,##0"),
                ("avg_months_to_completion", "#,##0"),
                ("avg_delay", "#,##0"),
            ]
        else:
            # cumulative completed sheets (b), (c), (d)
            specs = [
                ("num_completed", "#,##0"),
                ("total_approved_funding", "#,##0"),
                ("avg_pct_disbursed", "#,##0"),
            ]
            if include_odp_co2:
                specs.extend(
                    [
                        ("total_consumption_odp", "#,##0"),
                        ("total_production_odp", "#,##0"),
                        ("total_consumption_mt", "#,##0"),
                        ("total_production_mt", "#,##0"),
                        ("total_consumption_co2", "#,##0"),
                        ("total_production_co2", "#,##0"),
                    ]
                )
            specs.extend(
                [
                    ("avg_months_to_disbursement", "#,##0"),
                    ("avg_months_to_completion", "#,##0"),
                ]
            )
            if include_odp_co2:
                specs.append(("cost_effectiveness", "#,##0.00"))
            return specs

    def _compute_group_data(self, records, include_odp_co2, sheet_type):
        """
        Helper for computing aggregation data for a list of records;
        used for Total and per-group.
        """
        data = {}
        count = len(records)

        is_ongoing = sheet_type in [
            "ongoing_investment",
            "ongoing_non_investment",
            "ongoing_preparation",
        ]

        if is_ongoing:
            data["num_projects"] = count
        else:
            data["num_completed"] = count

        data["total_approved_funding"] = sum(
            apr.approved_funding_plus_adjustment_denorm or 0 for apr in records
        )
        data["total_funds_disbursed"] = sum(apr.funds_disbursed or 0 for apr in records)

        data["avg_pct_disbursed"] = (
            data["total_funds_disbursed"] / data["total_approved_funding"] * 100
            if data["total_approved_funding"]
            else 0
        )

        data["avg_months_to_disbursement"] = self._calculate_avg_months(
            records, "project__date_approved", "date_first_disbursement"
        )

        if is_ongoing:
            data["avg_months_to_completion"] = self._calculate_avg_months(
                records, "project__date_approved", "date_planned_completion"
            )
            data["avg_delay"] = self._calculate_avg_months(
                records,
                "date_of_completion_per_agreement_or_decisions_denorm",
                "date_planned_completion",
            )
            num_disbursing = sum(
                1 for apr in records if apr.funds_disbursed and apr.funds_disbursed > 0
            )
            data["num_disbursing"] = num_disbursing
            data["pct_disbursing"] = (num_disbursing / count) if count > 0 else 0
        else:
            data["avg_months_to_completion"] = self._calculate_avg_months(
                records, "project__date_approved", "date_actual_completion"
            )

        if include_odp_co2:
            total_consumption_odp = sum(
                apr.consumption_phased_out_odp or 0 for apr in records
            )
            total_production_odp = sum(
                apr.production_phased_out_odp or 0 for apr in records
            )
            data["total_consumption_odp"] = total_consumption_odp
            data["total_production_odp"] = total_production_odp

            data["total_consumption_mt"] = sum(
                apr.consumption_phased_out_mt or 0 for apr in records
            )
            data["total_production_mt"] = sum(
                apr.production_phased_out_mt or 0 for apr in records
            )

            data["total_consumption_co2"] = sum(
                apr.consumption_phased_out_co2 or 0 for apr in records
            )
            data["total_production_co2"] = sum(
                apr.production_phased_out_co2 or 0 for apr in records
            )

            total_odp = total_consumption_odp + total_production_odp
            data["cost_effectiveness"] = (
                data["total_approved_funding"] / (total_odp * 1000) if total_odp else 0
            )

            if sheet_type == "ongoing_investment":
                # Sum of phaseout combined (proposal values) in kg
                total_phaseout_combined = sum(
                    (apr.consumption_phased_out_odp_proposal_denorm or 0)
                    + (apr.production_phased_out_odp_proposal_denorm or 0)
                    for apr in records
                )
                data["total_phaseout_combined_kg"] = total_phaseout_combined * 1000
                data["cost_effectiveness"] = (
                    data["total_approved_funding"] / data["total_phaseout_combined_kg"]
                    if data["total_phaseout_combined_kg"]
                    else 0
                )

        return data

    def _compute_grouped_data(self, records, group_field, include_odp_co2, sheet_type):
        """
        Helper for computing aggregation data grouped by a field.
        Returns list of (name, data) tuples, sorted by group value.
        """
        buckets = {}
        for apr in records:
            val = self._get_field_value(apr, group_field)
            key = val if val is not None else ""
            if key not in buckets:
                buckets[key] = []
            buckets[key].append(apr)

        result = []
        for group_val in sorted(buckets.keys(), key=lambda x: (x == "", str(x))):
            group_name = group_val or "Unknown"
            group_data = self._compute_group_data(
                buckets[group_val], include_odp_co2, sheet_type
            )
            result.append((group_name, group_data))

        return result

    def _calculate_avg_months(self, records, start_date_field, end_date_field):
        """Helper for all month averaging performed throughout the file"""
        total_months = 0
        count = 0

        for apr in records:
            start_date = self._get_field_value(apr, start_date_field)
            end_date = self._get_field_value(apr, end_date_field)

            if start_date and end_date:
                delta = relativedelta(end_date, start_date)
                months = delta.years * 12 + delta.months
                total_months += months
                count += 1

        return total_months / count if count > 0 else 0

    def _create_response(self):
        """Create HTTP response with the workbook"""
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        filename = "APR_Summary_Tables_Cumulative"
        if self.agency:
            filename += f"_{self.agency.name.replace(' ', '_')}"
        filename += ".xlsx"

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
