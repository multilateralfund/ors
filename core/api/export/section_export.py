from openpyxl.utils import get_column_letter
from dateutil.parser import parse

from core.api.export.base import BaseWriter


class SectionWriter(BaseWriter):

    def __init__(self, sheet, headers, convert_to="mt"):
        self.convert_to = convert_to
        super().__init__(sheet, headers)

    def write_data(self, data):
        """Write data row by row:

        - Write group names if groups are present in the records.
        - Write sub-totals for each column of each group.
        - Write a final total row for each column.
        """
        if not data:
            # If there is no data, add an empty row so the formulas actually work
            data = [{}]

        row_idx = self.header_row_end_idx + 1
        group_ranges = []
        current_group, group_row_start_idx = None, None
        for record in data:
            new_group = record.get("group")
            if new_group and new_group != current_group:
                if group_row_start_idx:
                    group_ranges.append((group_row_start_idx, row_idx - 1))
                    self._write_subtotal_row(group_row_start_idx, row_idx)
                    row_idx += 1

                self._write_header_cell(row_idx, 1, new_group)
                self.sheet.merge_cells(
                    start_row=row_idx,
                    start_column=1,
                    end_row=row_idx,
                    end_column=self.max_column_idx,
                )

                row_idx += 1
                current_group = new_group
                group_row_start_idx = row_idx

            self._write_record_row(row_idx, record)
            row_idx += 1

        if group_row_start_idx:
            group_ranges.append((group_row_start_idx, row_idx - 1))
            self._write_subtotal_row(group_row_start_idx, row_idx)
            row_idx += 1
        else:
            # No groups in this section, make one big group instead
            group_ranges.append((self.header_row_end_idx + 1, row_idx - 1))

        self._write_total_row(row_idx, group_ranges)

    def _write_subtotal_row(self, group_row_start_idx, row_idx):
        self._write_record_cell(row_idx, 1, "Sub-total", read_only=True)
        for header in list(self.headers.values())[1:]:
            col = header["column_letter"]
            col_idx = header["column"]

            value = ""
            if header.get("is_numeric", True):
                value = f"=SUM({col}{group_row_start_idx}:{col}{row_idx-1})"

            self._write_record_cell(
                row_idx,
                col_idx,
                value,
                read_only=True,
                align=header.get("align", "left"),
            )

    def _write_record_row(self, row_idx, record):
        excluded = set(record.get("excluded_usages", []))
        quantity_key = (
            "quantity" if self.convert_to == "mt" else f"quantity_{self.convert_to}"
        )
        by_usage_id = {
            item["usage_id"]: item[quantity_key]
            for item in record.get("record_usages", [])
        }
        for header_id, header in self.headers.items():
            is_oversized = False
            if header.get("columnCategory") == "usage":
                value = by_usage_id.get(header_id)
            else:
                value = record.get(header_id)

            read_only = False
            if header.get("is_sum_function"):
                col_letter = get_column_letter(header["column"] - 1)
                value = f"=SUM(B{row_idx}:{col_letter}{row_idx})"
                read_only = True
            elif header_id in excluded:
                value = ""
                read_only = True
            elif header.get("type", None) == "boolean":
                value = "Yes" if value else "No"
            elif header.get("is_numeric", True):
                value = float(value or 0)
                value = round(value, 2)
            elif header.get("type", None) == "date":
                # DD/MM/YYYY
                value = parse(value).strftime("%d/%m/%Y") if value else ""
            else:
                value = value or ""

            self._write_record_cell(
                row_idx,
                header["column"],
                value,
                read_only=read_only,
                align=header.get("align", "left"),
                can_be_clipped=header.get("can_be_clipped", False),
            )

    def _write_total_row(self, row_idx, group_ranges):
        self._write_header_cell(row_idx, 1, "TOTAL")
        for header in list(self.headers.values())[1:]:
            col_idx = header["column"]
            col = header["column_letter"]
            value = ""
            if header.get("is_numeric", True):
                ranges = ",".join(
                    [f"{col}{start}:{col}{end}" for start, end in group_ranges]
                )
                value = f"=SUM({ranges})"
            self._write_record_cell(
                row_idx,
                col_idx,
                value,
                read_only=True,
                align=header.get("align", "left"),
            )
