from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side

from core.api.export.base import WriteOnlyBase


class SummaryOfProjectsWriter(WriteOnlyBase):
    ROW_HEIGHT = 35
    COLUMN_WIDTH = 20
    header_row_start_idx = 2
    FONT_NAME = "Times New Roman"
    TOTAL_VISIBLE_FIELDS = {"projects_count", "amounts_recommended"}
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, sheet, is_blanket_approval=False):
        headers = [
            {
                "id": "text",
                "headerName": "Projects and activities",
                "column_width": 36,
            },
            {
                "id": "countries_count",
                "headerName": "No. of countries",
                "align": "center",
                "cell_format": "0;-0;;@",
                "column_width": 8,
            },
            {
                "id": "projects_count",
                "headerName": "No. of funding requests",
                "align": "center",
                "cell_format": "0;-0;;@",
                "column_width": 8,
            },
            {
                "id": "amounts_recommended",
                "headerName": "Amounts recommended\n(US$)",
                "type": "number",
                "align": "right",
                "cell_format": "#,##0.00#############;-#,##0.00#############;;@",
                "column_width": 14,
            },
        ]

        if not is_blanket_approval:
            headers.append(
                {
                    "id": "amounts_in_principle",
                    "headerName": "Amounts in principle\n(US$)",
                    "type": "number",
                    "align": "right",
                    "cell_format": "#,##0.00#############;-#,##0.00#############;;@",
                    "column_width": 14,
                }
            )

        super().__init__(sheet, headers)

    def write(self, data):
        self.sheet.append([])
        super().write(data)
        if data:
            self.format_total_row(self.header_row_start_idx + len(data))

    def format_total_row(self, row_idx):
        self.sheet.merge_cells(
            start_row=row_idx,
            start_column=1,
            end_row=row_idx,
            end_column=2,
        )

        for column_idx in range(1, self.max_column_idx + 1):
            cell = self.sheet.cell(row=row_idx, column=column_idx)
            cell.font = Font(name=self.FONT_NAME, bold=True, sz=10)
            cell.border = self.BORDER

    def write_header_cell(self, value, comment=None):
        cell = super().write_header_cell(value, comment)
        cell.font = Font(name=self.FONT_NAME, bold=True, sz=10)
        cell.border = self.BORDER
        cell.fill = PatternFill()
        return cell

    def write_record_cell(self, value, read_only=False, align="left", cell_format=None):
        cell = super().write_record_cell(value, read_only, align, cell_format)
        cell.font = Font(name=self.FONT_NAME, sz=10)
        cell.border = self.BORDER
        return cell
