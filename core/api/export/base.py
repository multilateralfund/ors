import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.styles import Font
from openpyxl.styles import DEFAULT_FONT
from openpyxl.utils import get_column_letter

# pylint: disable=R0913


class BaseWriter:
    ROW_HEIGHT = 30
    COLUMN_WIDTH = 15
    OVERSIZED_CELL_THRESHOLD = 25
    header_row_start_idx = 2

    def __init__(self, sheet, headers):
        self.sheet = sheet
        self.headers = {}
        self.all_headers = {}

        next_column_idx = self._compute_header_positions(
            headers, row=self.header_row_start_idx
        )
        self.header_row_end_idx = max(h["row"] for h in self.headers.values())
        self.max_column_idx = next_column_idx - 1

    def write(self, data):
        self.write_headers()
        self.write_data(data)
        self.set_dimensions()
        # Freeze so top and side headers always stay visible.
        # Help with scrolling in the very large sections.
        self.sheet.freeze_panes = f"B{self.header_row_end_idx + 1}"

    def write_headers(self):
        """Write the entire header structure and merge cells."""
        # Write empty values in the entire header range to ensure everything
        # has the same style.
        for i in range(self.header_row_start_idx, self.header_row_end_idx + 1):
            for j in range(1, self.max_column_idx + 1):
                self._write_header_cell(i, j, "")

        for parsed_header in self.all_headers.values():
            comment = None
            name = parsed_header.get("headerName") or parsed_header.get("display_name")
            if parsed_header.get("type") == "date":
                comment = name
                name = "Date"

            row_idx, col_idx = parsed_header["row"], parsed_header["column"]
            end_row_idx, end_col_idx = row_idx, parsed_header["end_column"]
            if not parsed_header.get("children"):
                # Merge to the last header row if this is a leaf node.
                end_row_idx = self.header_row_end_idx

            self._write_header_cell(
                row_idx,
                col_idx,
                name,
                comment=comment,
            )
            self.sheet.merge_cells(
                start_row=row_idx,
                start_column=col_idx,
                end_row=end_row_idx,
                end_column=end_col_idx,
            )

    def write_data(self, data):
        for row_idx, record in enumerate(data, start=self.header_row_end_idx + 1):
            for header_id, header in self.headers.items():
                header_type = header.get("type")
                if method := header.get("method"):
                    value = method(record, header)
                else:
                    value = record.get(header_id)

                if header_type == "number":
                    value = float(value or 0)
                elif header_type == "int":
                    value = int(value or 0)
                elif header_type == "bool":
                    value = "Yes" if value else "No"
                else:
                    value = value or ""

                self._write_record_cell(
                    row_idx,
                    header["column"],
                    value,
                    read_only=header.get("read_only", False),
                    align=header.get("align", "left"),
                    cell_format=header.get("cell_format"),
                )

    def set_dimensions(self):
        for header in self.headers.values():
            self.sheet.column_dimensions[header["column_letter"]].width = header.get(
                "column_width", self.COLUMN_WIDTH
            )

        for row in range(self.header_row_start_idx, self.sheet.max_row + 1):
            # For "oversized" cell contents, the height can be set in the previous step,
            # so we need to avoid overwriting it.
            if not self.sheet.row_dimensions[row].height:
                self.sheet.row_dimensions[row].height = self.ROW_HEIGHT

    def _compute_header_positions(self, items, column=1, row=3):
        """
        Compute the positions of the headers depending on their hierarchy.
        And transform headers list in a dictionary where the key is the header id.
        """
        for item in items:
            # Keep track of the start column, as the header might span over many
            # columns if it has more than one child; for those cases, the cells need
            # to be merged horizontally.
            start_column = column
            children = item.get("children")
            if children:
                column = self._compute_header_positions(
                    children, column=column, row=row + 1
                )
            else:
                column += 1

            header = {
                **item,
                "row": row,
                "column_letter": get_column_letter(start_column),
                "column": start_column,
                "end_column": column - 1,
            }
            self.all_headers[item["id"]] = header
            # If header has no children, keep track of them separately
            # as these are the places where values are actually added.
            if not children:
                self.headers[item["id"]] = header
        return column

    def _write_header_cell(self, row, column, value, comment=None):
        cell = self.sheet.cell(row, column, value)
        cell.font = Font(name=DEFAULT_FONT.name, bold=True, color=None)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        if comment:
            cell.comment = Comment(comment, "")
        # Hack to avoid clipped header in exported PDF.
        # TODO: find a way (?) to improve this in the future.
        if value == "Other unidentified manufacturing":
            self.sheet.row_dimensions[row].height = self.ROW_HEIGHT * 4 / 3
        return cell

    def _write_record_cell(
        self,
        row,
        column,
        value,
        read_only=False,
        align="left",
        can_be_clipped=False,
        cell_format=None,
    ):
        cell = self.sheet.cell(row, column, value)
        cell.font = Font(color=None)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="hair"),
            right=Side(style="hair"),
            top=Side(style="hair"),
            bottom=Side(style="hair"),
        )
        if can_be_clipped:
            # We assume that only string values can be clipped.
            # Adding 2 chars to the length just in case.
            needed_rows = (len(value) + 2) / self.OVERSIZED_CELL_THRESHOLD
            # Assuming (naively) that 50 characters fit into "normal" 2-row cells.
            if needed_rows > 2:
                # Add 1/2 * ROW_HEIGHT for each extra needed row
                self.sheet.row_dimensions[row].height = (
                    self.ROW_HEIGHT * needed_rows / 2
                )
        if cell_format:
            cell.number_format = cell_format
        elif align == "right":
            # this cell will contain a number
            cell.number_format = "###,###,##0.00#############"
        if read_only:
            cell.fill = PatternFill(
                start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
            )
        return cell


class CPReportBase:
    ROW_HEIGHT = 30
    COLUMN_WIDTH = 15
    sections = ()

    def __init__(self, cp_report):
        self.cp_report = cp_report

    def get_xlsx(self, data, usages, convert_data=False):
        cp_report = data["cp_report"]
        wb = openpyxl.Workbook()
        for section in self.sections:
            name = section.replace("_", " ").title()
            sheet = wb.create_sheet(name)
            configure_sheet_print(sheet, sheet.ORIENTATION_PORTRAIT)
            title_cell = sheet.cell(
                1, 1, "Country: %(country)s Year: %(year)s" % cp_report
            )
            title_cell.font = Font(bold=True, size=20, color=None)

            args = {}
            args["usages"] = usages.get(section, [])
            if section in "section_a,section_b" and convert_data:
                args["convert_to"] = "odp" if section == "section_a" else "gwp"
            elif section == "section_c":
                args["year"] = cp_report["year"]

            getattr(self, f"export_{section}")(sheet, data.get(section, []), **args)

        # Remove the default sheet before saving
        del wb[wb.sheetnames[0]]
        return wb


class WriteOnlyBase:
    ROW_HEIGHT = 30
    COLUMN_WIDTH = 15
    header_row_start_idx = 1

    def __init__(self, sheet, headers):
        self.sheet = sheet
        self.headers = headers
        self.header_row_end_idx = 1
        self.max_column_idx = len(self.headers)

    def write(self, data):
        self.set_dimensions()
        self.write_headers()
        self.write_data(data)
        # # Freeze so top and side headers always stay visible.
        # # Help with scrolling in the very large sections.
        # self.sheet.freeze_panes = f"B{self.header_row_end_idx + 1}"

    def write_headers(self):
        header_row = []
        for parsed_header in self.headers:
            comment = None
            name = parsed_header.get("headerName") or parsed_header.get("display_name")
            if parsed_header.get("type") == "date":
                comment = name
                name = "Date"
            header_row.append(
                self.write_header_cell(
                    name,
                    comment=comment,
                )
            )
        self.sheet.append(header_row)

    def write_data(self, data):
        for record in data:
            record_row = []
            for header in self.headers:
                header_id = header.get("id")
                header_type = header.get("type")
                if method := header.get("method"):
                    value = method(record, header)
                else:
                    value = record.get(header_id)

                if header_type == "number":
                    value = float(value or 0)
                elif header_type == "int":
                    value = int(value or 0)
                elif header_type == "bool":
                    value = "Yes" if value else "No"
                else:
                    value = value or ""

                record_row.append(self.write_record_cell(value))
            self.sheet.append(record_row)

    def set_dimensions(self):
        for i, header in enumerate(self.headers):
            column_letter = get_column_letter(i + 1)
            self.sheet.column_dimensions[column_letter].width = header.get(
                "column_width", self.COLUMN_WIDTH
            )

    def write_header_cell(self, value, comment=None):
        cell = WriteOnlyCell(self.sheet, value=value)
        cell.font = Font(name=DEFAULT_FONT.name, bold=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        if comment:
            cell.comment = Comment(comment, "")
        return cell

    def write_record_cell(self, value, read_only=False):
        cell = WriteOnlyCell(self.sheet, value=value)
        cell.font = Font(color=None)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="hair"),
            right=Side(style="hair"),
            top=Side(style="hair"),
            bottom=Side(style="hair"),
        )
        if read_only:
            cell.fill = PatternFill(
                start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
            )
        return cell


class CPDataHFCHCFCWriterBase(BaseWriter):
    """
    Base class for writing CP data for HCFC and HFC sections.
    You need to implement get_value_for_header method
    You may need to implement __init__ method
    - get_value_for_header method should return the value for the given header
    - __init__ method should call super().__init__ and set the headers
    """

    OVERSIZED_CELL_THRESHOLD = 12

    def write_data(self, data):
        row_idx = self.header_row_end_idx + 1
        for record in data:
            self._write_record_row(row_idx, record)
            row_idx += 1

    def get_value_for_header(self, header_id, header, record, by_usage_id):
        raise NotImplementedError

    def _write_record_row(self, row_idx, record):
        by_usage_id = {
            str(item.usage_id): item.quantity for item in record.record_usages.all()
        }
        first_column_sum = None
        for header_id, header in self.headers.items():
            # delete quantity type from header_id
            quantity_type = header.get("quantity_type")
            if quantity_type:
                header_id = header_id.replace(quantity_type, "").strip()

            value = self.get_value_for_header(header_id, header, record, by_usage_id)

            # set first column sum
            if header.get("columnCategory") == "usage" and not first_column_sum:
                first_column_sum = get_column_letter(header["column"])

            # set value for sum columns
            if header.get("is_sum_function"):
                col_letter = get_column_letter(header["column"] - 1)
                value = f"=SUM({first_column_sum}{row_idx}:{col_letter}{row_idx})"
                first_column_sum = None

            self._write_record_cell(
                row_idx,
                header["column"],
                value,
                align=header.get("align", "left"),
                can_be_clipped=header.get("can_be_clipped", False),
            )


def configure_sheet_print(sheet, orientation):
    sheet.page_setup.orientation = orientation
    sheet.page_setup.paperSize = getattr(sheet.page_setup, "PAPERSIZE_A4", 9)
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.sheet_properties.pageSetUpPr.autoPageBreaks = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = False
    sheet.page_margins.top = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.left = 0.25
