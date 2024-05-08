from openpyxl.utils import get_column_letter
from core.api.export.base import BaseWriter
from core.api.export.base import COLUMN_WIDTH


class CPReportHFCWriter(BaseWriter):
    header_row_start_idx = 1

    def __init__(self, wb, usages_mt, usages_co2, year):
        usages_headers = []
        for q_type, usages in [("(MT)", usages_mt), ("(CO2)", usages_co2)]:
            usages_headers.extend(
                [
                    *usages,
                    {
                        "id": f"total {q_type}",
                        "headerName": f"Total {q_type}",
                        "is_sum_function": True,
                        "quantity_type": q_type,
                    },
                    {
                        "id": f"imports {q_type}",
                        "headerName": f"Import {q_type}",
                        "quantity_type": q_type,
                    },
                    {
                        "id": f"exports {q_type}",
                        "headerName": f"Export{q_type}",
                        "quantity_type": q_type,
                    },
                    {
                        "id": f"production {q_type}",
                        "headerName": f"Production{q_type}",
                        "quantity_type": q_type,
                    },
                ]
            )
        headers = [
            {
                "id": "country_name",
                "headerName": "Country",
            },
            {
                "id": "country_is_lvc",
                "headerName": "Status",
                "columnCategory": "lvc_status",
            },
            {
                "id": "chemical_name",
                "headerName": "Chemical",
            },
            {
                "id": "chemical_gwp",
                "headerName": "GWP",
            },
            {
                "id": "year",
                "headerName": "Year",
            },
            *usages_headers,
            {
                "id": "remarks",
                "headerName": "Notes",
            },
        ]
        sheet = wb.create_sheet(str(year))
        super().__init__(sheet, headers)

    def write_data(self, records):
        row_idx = self.header_row_end_idx + 1
        for record in records:
            self._write_record_row(row_idx, record)
            row_idx += 1

    def _write_record_row(self, row_idx, record):
        by_usage_id = {
            str(item.usage_id): item.quantity for item in record.record_usages.all()
        }
        gwp = record.get_chemical_gwp()
        first_column_sum = None
        for header_id, header in self.headers.items():
            # delete quantity type from header_id
            quantity_type = header.get("quantity_type")
            if quantity_type:
                header_id = header_id.replace(quantity_type, "").strip()

            # set value for custom columns
            if header_id == "country_name":
                value = record.country_programme_report.country.name
            elif header_id == "country_is_lvc":
                lvc_status = record.country_programme_report.country.is_lvc
                value = "LVC" if lvc_status else "Non-LVC"
            elif header_id == "chemical_name":
                value = record.get_chemical_display_name()
            elif header_id == "chemical_gwp":
                value = gwp
            elif header.get("columnCategory") == "usage":
                value = by_usage_id.get(header_id) or 0
            elif header_id == "year":
                value = record.country_programme_report.year
            else:
                value = getattr(record, header_id, None)

            # convert value to co2 equivalent if needed
            if quantity_type:
                value = value or 0
                if quantity_type == "(CO2)":
                    value *= gwp

            # set first column sum
            if header.get("columnCategory") == "usage" and not first_column_sum:
                first_column_sum = get_column_letter(header["column"])

            # set value for sum columns
            if header.get("is_sum_function"):
                col_letter = get_column_letter(header["column"] - 1)
                value = f"=SUM({first_column_sum}{row_idx}:{col_letter}{row_idx})"
                first_column_sum = None

            self._write_record_cell(
                row_idx,
                header["column"],
                value,
            )
