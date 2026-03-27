"""
Tests for the export_all_projects_as_apr management command.
"""

import os
import tempfile
from datetime import date

import pytest
from django.core.management import call_command
from openpyxl import load_workbook

from core.api.export.annual_project_report import APRExportWriter
from core.api.serializers.annual_project_report import (
    AnnualProjectReportReadSerializer,
)
from core.api.tests.factories import (
    ProjectFactory,
    ProjectStatusFactory,
)
from core.management.commands.export_all_projects_as_apr import (
    COMPUTED_FIELDS,
    DENORM_FIELD_MAP,
    _build_project_dict,
    _generate_project_dicts,
)
from core.models import AnnualProjectReport

# pylint: disable=W0613,W0621,R0914


@pytest.fixture
def ongoing_status():
    return ProjectStatusFactory.create(name="Ongoing", code="ONG")


@pytest.fixture
def project_v3(ongoing_status):
    """A self-contained version-3 project with funding and support cost set."""
    return ProjectFactory(
        code="TST/RO1/001/INV/01",
        version=3,
        latest_project=None,
        total_fund=10000.0,
        support_cost_psc=1000.0,
        date_approved=date(2020, 6, 15),
        status=ongoing_status,
    )


@pytest.fixture
def second_project_v3(ongoing_status):
    return ProjectFactory(
        code="TST/RO1/002/INV/01",
        version=3,
        latest_project=None,
        total_fund=5000.0,
        support_cost_psc=500.0,
        status=ongoing_status,
    )


def _make_populated_apr(project, year=2024):
    """
    Build an in-memory AnnualProjectReport for a project the same way
    _generate_project_dicts does: inject latest_project_version_for_year,
    then call populate_derived_fields().
    """
    apr = AnnualProjectReport(project=project)
    apr.project_id = project.id
    apr.status = project.status.name if project.status else ""
    apr.__dict__["latest_project_version_for_year"] = project
    apr.__dict__["report_year"] = year
    apr.populate_derived_fields()
    return apr


@pytest.mark.django_db
class TestBuildProjectDict:
    def test_all_excel_fields_present(self, project_v3):
        apr = _make_populated_apr(project_v3)
        data = _build_project_dict(apr)

        for field in AnnualProjectReportReadSerializer.Meta.excel_fields:
            assert (
                field in data
            ), f"Field '{field}' missing from _build_project_dict output"

    def test_denorm_fields_come_from_apr_attributes(self, project_v3):
        apr = _make_populated_apr(project_v3)
        data = _build_project_dict(apr)

        for field_name, attr_name in DENORM_FIELD_MAP.items():
            raw = getattr(apr, attr_name, None)
            # Dates are converted to ISO strings
            expected = raw.isoformat() if hasattr(raw, "isoformat") else raw
            assert (
                data[field_name] == expected
            ), f"Field '{field_name}' expected {expected!r}, got {data[field_name]!r}"

    def test_status_field_is_set(self, project_v3):
        apr = _make_populated_apr(project_v3)
        data = _build_project_dict(apr)
        assert data["status"] == project_v3.status.name

    def test_apr_input_fields_are_none(self, project_v3):
        apr = _make_populated_apr(project_v3)
        data = _build_project_dict(apr)

        input_fields = [
            "funds_disbursed",
            "funds_committed",
            "support_cost_disbursed",
            "support_cost_committed",
            "last_year_remarks",
            "current_year_remarks",
            "gender_policy",
            "date_first_disbursement",
            "date_actual_completion",
            "estimated_disbursement_current_year",
        ]
        for field in input_fields:
            assert (
                data[field] is None
            ), f"Input field '{field}' should be None, got {data[field]!r}"

    def test_all_computed_fields_present(self, project_v3):
        apr = _make_populated_apr(project_v3)
        data = _build_project_dict(apr)

        for field in COMPUTED_FIELDS:
            assert field in data, f"Computed field '{field}' missing from dict"

    def test_calculated_fields(self, project_v3):
        apr = _make_populated_apr(project_v3)
        data = _build_project_dict(apr)

        assert data["balance"] == project_v3.total_fund
        assert data["per_cent_funds_disbursed"] is None
        assert data["support_cost_balance"] == project_v3.support_cost_psc

    def test_date_fields_serialized_as_iso_strings(self, project_v3):
        apr = _make_populated_apr(project_v3)
        data = _build_project_dict(apr)

        # project_v3 has date_approved set, so date_approved_denorm should be a string
        assert isinstance(data["date_approved"], str)
        assert data["date_approved"] == project_v3.date_approved.isoformat()


@pytest.mark.django_db
class TestGenerateProjectDicts:
    def test_yields_one_dict_per_project(self, project_v3, second_project_v3):
        dicts = list(_generate_project_dicts([project_v3, second_project_v3], 2024))
        assert len(dicts) == 2

    def test_empty_input_yields_empty_output(self):
        dicts = list(_generate_project_dicts([], 2024))
        assert not dicts

    def test_project_code_matches(self, project_v3):
        dicts = list(_generate_project_dicts([project_v3], 2024))
        assert dicts[0]["project_code"] == project_v3.code

    def test_balance_is_approved_funding_plus_adjustment(self, ongoing_status):
        v3 = ProjectFactory(
            code="TST/ADJ/003/INV/01",
            version=3,
            latest_project=None,
            total_fund=10000.0,
            status=ongoing_status,
        )
        v4 = ProjectFactory(
            code="TST/ADJ/003/INV/01",
            version=4,
            latest_project=v3,
            total_fund=12000.0,
            status=ongoing_status,
        )
        # Inject v4 as latest version so adjustment = 12000 - 10000 = 2000
        apr = AnnualProjectReport(project=v3)
        apr.project_id = v3.id
        apr.status = v3.status.name
        apr.__dict__["latest_project_version_for_year"] = v4
        apr.__dict__["report_year"] = 2024
        apr.populate_derived_fields()
        data = _build_project_dict(apr)

        # balance must use approved_funding_plus_adjustment (12000), not approved_funding (10000)
        assert data["approved_funding"] == 10000.0
        assert data["approved_funding_plus_adjustment"] == 12000.0
        assert data["balance"] == 12000.0

    def test_per_cent_funds_disbursed_none_without_disbursement(self, project_v3):
        dicts = list(_generate_project_dicts([project_v3], 2024))
        assert dicts[0]["per_cent_funds_disbursed"] is None

    def test_support_cost_balance_populated(self, project_v3):
        dicts = list(_generate_project_dicts([project_v3], 2024))
        assert dicts[0]["support_cost_balance"] == project_v3.support_cost_psc

    def test_apr_input_fields_are_none(self, project_v3):
        dicts = list(_generate_project_dicts([project_v3], 2024))
        for field in ["funds_disbursed", "last_year_remarks", "gender_policy"]:
            assert dicts[0][field] is None


@pytest.mark.django_db
class TestExportAllProjectsAsAprCommand:
    def _run_command(self):
        """Run the command into a temp file, return the workbook."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            call_command("export_all_projects_as_apr", output=path)
            return load_workbook(path), path
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_xlsx_output_structure_and_field_values(self, project_v3):
        """
        Consolidates 4 single-project export checks: correct sheet name,
        balance/approved_funding values, per_cent_funds_disbursed=None without
        disbursements, and APR input columns are empty.
        """
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            call_command("export_all_projects_as_apr", output=path)
            assert os.path.exists(path)
            wb = load_workbook(path)
            ws = wb[APRExportWriter.SHEET_NAME]
            assert APRExportWriter.SHEET_NAME in wb.sheetnames

            col_map = APRExportWriter.build_column_mapping()
            row = APRExportWriter.FIRST_DATA_ROW

            # balance and approved_funding should equal total_fund (no disbursements)
            assert (
                ws.cell(row, col_map["approved_funding"]).value == project_v3.total_fund
            )
            assert ws.cell(row, col_map["balance"]).value == project_v3.total_fund

            # per_cent_funds_disbursed should be None without disbursements
            assert ws.cell(row, col_map["per_cent_funds_disbursed"]).value is None

            # APR input columns should be empty
            for field in [
                "funds_disbursed",
                "funds_committed",
                "support_cost_disbursed",
            ]:
                val = ws.cell(row, col_map[field]).value
                assert val is None, f"Column '{field}' should be empty, got {val!r}"
        finally:
            os.unlink(path)

    def test_writes_one_row_per_project(self, project_v3, second_project_v3):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            call_command("export_all_projects_as_apr", output=path)
            wb = load_workbook(path)
            ws = wb[APRExportWriter.SHEET_NAME]
            data_rows = ws.max_row - (APRExportWriter.FIRST_DATA_ROW - 1)
            assert data_rows == 2
        finally:
            os.unlink(path)

    def test_empty_db_still_produces_file(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            call_command("export_all_projects_as_apr", output=path)
            assert os.path.exists(path)
        finally:
            os.unlink(path)
