import io

import openpyxl
import pytest
from django.urls import reverse

from core.api.tests.base import BaseTest
from core.api.tests.conftest import pdf_text

pytestmark = pytest.mark.django_db
# pylint: disable=C8008, W0221


class TestCPExportXLSX(BaseTest):
    url = reverse("country-programme-export")

    def test_get_cp_export_anon(self, cp_report_2019):
        response = self.client.get(self.url, {"cp_report_id": cp_report_2019.id})
        assert response.status_code == 403

    def test_get_cp_export_new(
        self, secretariat_user, cp_report_2019, _setup_new_cp_report
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"cp_report_id": cp_report_2019.id})
        assert response.status_code == 200
        assert response.filename == cp_report_2019.name + ".xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == [
            "Section A",
            "Section B",
            "Section C",
            "Section D",
            "Section E",
            "Section F",
        ]
        assert wb["Section A"]["A1"].value == "Country: Romania Year: 2019"

    def test_get_cp_export_new_converted(
        self, secretariat_user, cp_report_2019, _setup_new_cp_report
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(
            self.url, {"cp_report_id": cp_report_2019.id, "convert_data": 1}
        )
        assert response.status_code == 200
        assert response.filename == cp_report_2019.name + ".xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == [
            "Section A",
            "Section B",
            "Section C",
            "Section D",
            "Section E",
            "Section F",
        ]
        assert wb["Section A"]["A1"].value == "Country: Romania Year: 2019"
        assert "ODP" in wb["Section A"]["A2"].value
        assert "CO₂-eq tonnes" in wb["Section B"]["A2"].value

    def test_get_cp_export_old(
        self, secretariat_user, cp_report_2005, _setup_old_cp_report
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"cp_report_id": cp_report_2005.id})
        assert response.status_code == 200
        assert response.filename == cp_report_2005.name + ".xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == [
            "Section A",
            "Adm B",
            "Adm C",
            "Section C",
            "Adm D",
        ]
        assert wb["Section A"]["A1"].value == "Country: Romania Year: 2005"


class TestCPExportPDF(BaseTest):
    url = reverse("country-programme-print")

    def test_get_cp_export_anon(self, cp_report_2019):
        response = self.client.get(self.url, {"cp_report_id": cp_report_2019.id})
        assert response.status_code == 403

    def test_get_cp_export_new(
        self, secretariat_user, cp_report_2019, _setup_new_cp_report
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"cp_report_id": cp_report_2019.id})
        assert response.status_code == 200
        assert response.filename == cp_report_2019.name + ".pdf"

        text = pdf_text(io.BytesIO(response.getvalue()))
        assert "Country: Romania Year: 2019" in text
        for name in [
            "Section A",
            "Section B",
            "Section C",
            "Section D",
            "Section E",
            "Section F",
        ]:
            assert name.upper() in text

    def test_get_cp_export_old(
        self, secretariat_user, cp_report_2005, _setup_old_cp_report
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"cp_report_id": cp_report_2005.id})
        assert response.status_code == 200
        assert response.filename == cp_report_2005.name + ".pdf"

        text = pdf_text(io.BytesIO(response.getvalue()))
        assert "Country: Romania Year: 2005" in text

        for name in [
            "Section A",
            "Adm B",
            "Adm C",
            "Section C",
            "Adm D",
        ]:
            assert name.upper() in text


class TestCPExportEmpty(BaseTest):
    url = reverse("country-programme-export-empty")

    def test_get_cp_export_anon(self, cp_report_2019):
        response = self.client.get(self.url, {"cp_report_id": cp_report_2019.id})
        assert response.status_code == 403

    def test_get_cp_export_new(self, secretariat_user):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"year": "2019"})
        assert response.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == [
            "Section A",
            "Section B",
            "Section C",
            "Section D",
            "Section E",
            "Section F",
        ]
        assert wb["Section A"]["A1"].value == "Country: XXXX Year: 2019"

    def test_get_cp_export_old(self, secretariat_user):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"year": 2005})
        assert response.status_code == 200

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == [
            "Section A",
            "Adm B",
            "Adm C",
            "Section C",
            "Adm D",
        ]
        assert wb["Section A"]["A1"].value == "Country: XXXX Year: 2005"


class TestCPArchiveExportXLSX(BaseTest):
    url = reverse("country-programme-archive-export")

    def test_get_cp_export_anon(self, cp_report_2019):
        response = self.client.get(self.url, {"cp_report_id": cp_report_2019.id})
        assert response.status_code == 403

    def test_get_cp_archive_export_new(self, secretariat_user, _setup_old_version_2019):
        self.client.force_authenticate(user=secretariat_user)

        cp_ar = _setup_old_version_2019

        response = self.client.get(self.url, {"cp_report_id": cp_ar.id})
        assert response.status_code == 200
        assert response.filename == cp_ar.name + ".xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == [
            "Section A",
            "Section B",
            "Section C",
            "Section D",
            "Section E",
            "Section F",
        ]
        assert wb["Section A"]["A1"].value == "Country: Romania Year: 2019"

    def test_get_cp_export_old(self, secretariat_user, _setup_old_version_2005):
        self.client.force_authenticate(user=secretariat_user)

        cp_ar = _setup_old_version_2005

        response = self.client.get(self.url, {"cp_report_id": cp_ar.id})
        assert response.status_code == 200
        assert response.filename == cp_ar.name + ".xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == [
            "Section A",
            "Adm B",
            "Adm C",
            "Section C",
            "Adm D",
        ]
        assert wb["Section A"]["A1"].value == "Country: Romania Year: 2005"


class TestCPHCFCExtractionExport(BaseTest):
    url = reverse("country-programme-hcfc-export")

    def test_get_cp_export(
        self, secretariat_user, _setup_new_cp_report, _cp_report_format
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"min_year": 2019, "max_year": 2020})
        assert response.status_code == 200
        assert response.filename == "CP Data Extraction-HCFC.xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == ["2019", "2020"]
        # check number of rows
        assert wb["2019"].max_row == 5
        assert wb["2020"].max_row == 1

        # check number of columns
        # country, substnace, categ, year, usage, total, import, export, production, notes
        assert wb["2019"].max_column == 10
        assert wb["2020"].max_column == 10


class TestCPHFCExtractionExport(BaseTest):
    url = reverse("country-programme-hfc-export")

    def test_get_cp_export(
        self, secretariat_user, _setup_new_cp_report, _cp_report_format
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"min_year": 2019, "max_year": 2020})
        assert response.status_code == 200
        assert response.filename == "CP Data Extraction-HFC.xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == ["2019", "2020"]
        # check number of rows
        assert wb["2019"].max_row == 9
        assert wb["2020"].max_row == 1

        # check number of columns
        # country, status, chemical, gwp, year,
        # usage(mt), total(mt), import(mt), export(mt), production(mt),
        # usage(CO₂), total(CO₂), import(CO₂), export(CO₂), production(CO₂), notes
        assert wb["2019"].max_column == 16
        assert wb["2020"].max_column == 16


class TestCPExtractionALLExport(BaseTest):
    url = reverse("country-programme-extraction-all-export")

    def test_get_cp_export(
        self, secretariat_user, _setup_new_cp_report, _cp_report_format
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"min_year": 2019, "max_year": 2019})
        assert response.status_code == 200
        assert response.filename == "CP Data Extraction-All.xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == [
            "ODSPrice",
            "CP-Details",
            "CPConsumption(ODP)",
            "HFC-Consumption(MTvsCO₂Equi)",
            "HFC-23Generation",
            "HFC-23Emission",
            "MbrConsumption",
        ]
        # check number of rows
        assert wb["ODSPrice"].max_row == 8
        assert wb["CP-Details"].max_row == 13
        assert wb["CPConsumption(ODP)"].max_row == 8
        assert wb["HFC-Consumption(MTvsCO₂Equi)"].max_row == 2
        assert wb["HFC-23Generation"].max_row == 2
        assert wb["HFC-23Emission"].max_row == 3
        assert wb["MbrConsumption"].max_row == 2

        # check number of columns
        # country, substance, price,fob, retail_proce, remarks, notes
        assert wb["ODSPrice"].max_column == 7
        # country, annex_group, substance, odp_value, gwp_value, 2019, notes
        assert wb["CP-Details"].max_column == 7
        # country, substance, value, notes
        assert wb["CPConsumption(ODP)"].max_column == 4
        # country, status, chemical, group,
        # 2019 MT, 2019 CO₂, 2019 Servicing, 2019 Usage Total, notes
        assert wb["HFC-Consumption(MTvsCO₂Equi)"].max_column == 9
        # country, year, substance, all_uses, feedstock, destruction, notes
        assert wb["HFC-23Generation"].max_column == 7
        # country, year, facility, total, all_uses, feedstock_gc, destruction
        # feedstock_wpc, destruction_wpc, generated_emissions, remarks, notes
        assert wb["HFC-23Emission"].max_column == 12
        # country, qps, non-qps, total
        assert wb["MbrConsumption"].max_column == 4


class TestCPCalculatedAmountExport(BaseTest):
    url = reverse("country-programme-calculated-amount-export")

    def test_get_cp_export(
        self, secretariat_user, cp_report_2019, _setup_new_cp_report
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"cp_report_id": cp_report_2019.id})
        assert response.status_code == 200
        assert response.filename == f"CalculatedAmount {cp_report_2019.name}.xlsx"

        wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
        assert wb.sheetnames == ["Calculated Amount"]
        assert wb["Calculated Amount"].max_row == 10
        assert wb["Calculated Amount"].max_column == 4


class TestCPCalculatedAmountExportPDF(BaseTest):
    url = reverse("country-programme-calculated-amount-print")

    def test_get_cp_export(
        self, secretariat_user, cp_report_2019, _setup_new_cp_report
    ):
        self.client.force_authenticate(user=secretariat_user)

        response = self.client.get(self.url, {"cp_report_id": cp_report_2019.id})
        assert response.status_code == 200
        assert response.filename == f"CalculatedAmount {cp_report_2019.name}.pdf"

        text = pdf_text(io.BytesIO(response.getvalue()))
        for name in [
            "Substances",
            "Unit",
            "Sectoral",
            "Consumption",
        ]:
            assert name in text

        for name in [
            "CFC",
            "Halon",
            "CTC",
            "TCA",
            "HCFC",
            "MB Non-QPS only",
            "HFC",
        ]:
            assert name in text
