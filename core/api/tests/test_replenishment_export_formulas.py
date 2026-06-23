"""
Regression tests for the dynamic totals/aggregation formulas in the
replenishment (status of contributions / scale of assessment) Excel exports.

The template sheets carry SUM and other aggregation formulas with absolute cell
references sized for a fixed number of rows. When the actual number of donor
countries differs, rows are inserted/deleted and the formulas must be rewritten
so the column totals keep tracking the real data range (openpyxl does not do this
on its own). See ``BaseTemplateSheetWriter._translate_formulas``.
"""

from decimal import Decimal

import openpyxl

from core.api.export.replenishment import (
    ScaleOfAssessmentTemplateWriter,
    StatusOfContributionsSummaryTemplateWriter,
    StatusOfContributionsTriennialTemplateWriter,
    translate_formula_rows,
)
from core.api.views.replenishment import EXPORT_RESOURCES_DIR

CONTRIBUTIONS_TEMPLATE = EXPORT_RESOURCES_DIR / "ContributionsFormatted.xlsx"
SOA_TEMPLATE = EXPORT_RESOURCES_DIR / "Scale of Assesement.xlsx"

SUMMARY_SHEET = "Summary Status of Contributions"
TRIENNIAL_SHEET = "2024-26 Contributions"


def _load(template):
    wb = openpyxl.load_workbook(filename=template)
    # pylint: disable=protected-access
    wb._external_links = []
    return wb


def _soc_data(count):
    return [
        {
            "no": i + 1,
            "country": f"Country {i + 1}",
            "agreed_contributions": Decimal(100),
            "cash_payments": Decimal(90),
            "bilateral_assistance": Decimal(5),
            "promissory_notes": Decimal(0),
            "outstanding_contributions": Decimal(5),
            "gain_loss": Decimal(1),
        }
        for i in range(count)
    ]


def _soa_data(count):
    return [
        {
            "no": i + 1,
            "country": f"Country {i + 1}",
            "un_scale_of_assessment": Decimal(1),
            "adjusted_scale_of_assessment": Decimal(1),
            "yearly_amount": Decimal(1),
            "average_inflation_rate": Decimal(1),
            "qualifies_for_fixed_rate_mechanism": False,
            "exchange_rate": Decimal(1),
            "currency": "USD",
            "yearly_amount_local_currency": Decimal(1),
        }
        for i in range(count)
    ]


# ---------------------------------------------------------------------------
# Unit tests for the formula translation helper
# ---------------------------------------------------------------------------


class TestTranslateFormulaRows:
    def test_insert_expands_enclosing_range(self):
        # Inserting rows inside a SUM range extends its lower bound.
        result = translate_formula_rows(
            "=SUM(C11:C65)", position=13, count=5, deleting=False
        )
        assert result == "=SUM(C11:C70)"

    def test_delete_shrinks_enclosing_range(self):
        result = translate_formula_rows(
            "=SUM(C11:C65)", position=12, count=35, deleting=True
        )
        assert result == "=SUM(C11:C30)"

    def test_insert_shifts_references_below(self):
        result = translate_formula_rows(
            "=C66+C67", position=13, count=5, deleting=False
        )
        assert result == "=C71+C72"

    def test_delete_shifts_references_below(self):
        result = translate_formula_rows(
            "=C66+C67", position=12, count=35, deleting=True
        )
        assert result == "=C31+C32"

    def test_absolute_reference_row_is_translated_dollar_kept(self):
        result = translate_formula_rows(
            "=(C2/46.037)*$C$51+C2", position=3, count=5, deleting=False
        )
        assert result == "=(C2/46.037)*$C$56+C2"

    def test_reference_above_change_is_untouched(self):
        result = translate_formula_rows(
            "=SUM(C11:C65)", position=70, count=5, deleting=False
        )
        assert result == "=SUM(C11:C65)"

    def test_cross_sheet_reference_is_not_translated(self):
        result = translate_formula_rows(
            "='YR2024'!F59", position=12, count=35, deleting=True
        )
        assert result == "='YR2024'!F59"

    def test_external_reference_is_not_translated(self):
        result = translate_formula_rows(
            "=[3]Summary!$D$100", position=12, count=35, deleting=True
        )
        assert result == "=[3]Summary!$D$100"

    def test_numeric_constants_are_not_treated_as_references(self):
        result = translate_formula_rows(
            "=C2/46.037*22", position=2, count=5, deleting=False
        )
        assert result == "=C7/46.037*22"


# ---------------------------------------------------------------------------
# Writer-level tests against the real template files
# ---------------------------------------------------------------------------


class TestSummaryStatusOfContributionsFormulas:
    FIRST_DATA_ROW = 11

    def _write(self, count):
        wb = _load(CONTRIBUTIONS_TEMPLATE)
        ws = wb[SUMMARY_SHEET]
        StatusOfContributionsSummaryTemplateWriter(
            ws,
            _soc_data(count),
            count,
            None,
            disputed_contributions=Decimal(7),
        ).write()
        return ws

    def _assert_consistent(self, count):
        ws = self._write(count)
        last_data_row = self.FIRST_DATA_ROW + count - 1
        subtotal_row = last_data_row + 1
        disputed_row = subtotal_row + 1
        total_row = disputed_row + 1

        # Column totals span exactly the data rows (no spill, no exclusion).
        for col in "CDEFGH":
            assert (
                ws[f"{col}{subtotal_row}"].value
                == f"=SUM({col}{self.FIRST_DATA_ROW}:{col}{last_data_row})"
            )
        # Grand total references the (moved) sub-total and disputed rows.
        for col in "CDEF":
            assert (
                ws[f"{col}{total_row}"].value
                == f"={col}{subtotal_row}+{col}{disputed_row}"
            )

    def test_fewer_countries_than_template(self):
        self._assert_consistent(20)

    def test_template_sized(self):
        # 55 data rows == template; formulas must be untouched.
        self._assert_consistent(55)

    def test_more_countries_than_template(self):
        # The reviewer's concern: an added donor row must be in the totals.
        self._assert_consistent(60)


class TestTriennialStatusOfContributionsFormulas:
    FIRST_DATA_ROW = 11

    def _write(self, count):
        wb = _load(CONTRIBUTIONS_TEMPLATE)
        ws = wb[TRIENNIAL_SHEET]
        StatusOfContributionsTriennialTemplateWriter(
            ws,
            _soc_data(count),
            count,
            2024,
            disputed_contributions=Decimal(7),
            ceit_data={
                "country": "CEIT",
                "agreed_contributions": Decimal(3),
                "cash_payments": Decimal(2),
                "bilateral_assistance": Decimal(1),
                "promissory_notes": Decimal(0),
                "outstanding_contributions": Decimal(1),
            },
        ).write()
        return ws

    def _assert_consistent(self, count):
        ws = self._write(count)
        last_data_row = self.FIRST_DATA_ROW + count - 1
        total_row = last_data_row + 1
        disputed_row = total_row + 1
        grand_total_row = disputed_row + 1

        for col in "CDEG":
            assert (
                ws[f"{col}{total_row}"].value
                == f"=SUM({col}{self.FIRST_DATA_ROW}:{col}{last_data_row})"
            )
        for col in "CDEG":
            assert (
                ws[f"{col}{grand_total_row}"].value
                == f"={col}{total_row}+{col}{disputed_row}"
            )

    def test_fewer_countries_than_template(self):
        self._assert_consistent(20)

    def test_template_sized(self):
        self._assert_consistent(49)

    def test_more_countries_than_template(self):
        self._assert_consistent(60)


class TestScaleOfAssessmentFormulas:
    FIRST_DATA_ROW = 2

    def _write(self, count):
        wb = _load(SOA_TEMPLATE)
        ws = wb.active
        ScaleOfAssessmentTemplateWriter(ws, _soa_data(count), count, 2024).write()
        return ws

    def _assert_consistent(self, count):
        ws = self._write(count)
        last_data_row = self.FIRST_DATA_ROW + count - 1
        totals_row = last_data_row + 1
        for col in "DE":
            assert (
                ws[f"{col}{totals_row}"].value
                == f"=SUM({col}{self.FIRST_DATA_ROW}:{col}{last_data_row})"
            )

    def test_fewer_countries_than_template(self):
        self._assert_consistent(20)

    def test_template_sized(self):
        self._assert_consistent(49)

    def test_more_countries_than_template(self):
        self._assert_consistent(60)
