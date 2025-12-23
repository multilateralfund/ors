from datetime import datetime, date
from io import BytesIO

from constance import config
from openpyxl import load_workbook
from zipfile import ZipFile

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.utils import timezone
from rest_framework import status

from core.api.export.annual_project_report import APRExportWriter
from core.api.tests.base import BaseTest
from core.models import (
    AnnualProgressReport,
    AnnualAgencyProjectReport,
    AnnualProjectReport,
    AnnualProjectReportFile,
)
from core.api.tests.factories import (
    AgencyFactory,
    AnnualProgressReportFactory,
    AnnualProjectReportFactory,
    AnnualAgencyProjectReportFactory,
    AnnualProjectReportFileFactory,
    ProjectFactory,
    MetaProjectFactory,
    ProjectClusterFactory,
)

# pylint: disable=W0221,W0613,C0302,R0913


@pytest.mark.django_db
class TestAPRCurrentYearView(BaseTest):
    def test_without_login(self):
        self.client.force_authenticate(user=None)
        url = reverse("apr-current-year")
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_requires_apr_view_permission(self, user):
        self.client.force_authenticate(user=user)
        url = reverse("apr-current-year")
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_returns_unendorsed_year_when_exists(
        self, apr_agency_viewer_user, annual_progress_report
    ):
        AnnualProgressReportFactory(year=2022, endorsed=True)

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-current-year")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["current_year"] == annual_progress_report.year
        assert response.data["endorsed"] is False
        assert len(response.data["apr_list"]) == 2
        # Find the unendorsed APR and check it (should be the param fixture)
        unendorsed = next(a for a in response.data["apr_list"] if not a["endorsed"])
        assert unendorsed["year"] == annual_progress_report.year

    def test_returns_latest_unendorsed_when_multiple_exist(
        self, apr_agency_viewer_user
    ):
        AnnualProgressReportFactory(year=2023, endorsed=False)
        AnnualProgressReportFactory(year=2024, endorsed=False)
        AnnualProgressReportFactory(year=2022, endorsed=True)

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-current-year")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["current_year"] == 2024
        assert response.data["endorsed"] is False
        assert len(response.data["apr_list"]) == 3

    def test_returns_latest_endorsed_when_no_unendorsed(
        self, apr_agency_viewer_user, annual_progress_report_endorsed
    ):
        AnnualProgressReportFactory(year=2022, endorsed=True)

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-current-year")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["current_year"] == annual_progress_report_endorsed.year
        assert response.data["endorsed"] is True
        assert len(response.data["apr_list"]) == 2

    def test_returns_latest_endorsed_among_multiple(self, apr_agency_viewer_user):
        AnnualProgressReportFactory(year=2021, endorsed=True)
        AnnualProgressReportFactory(year=2023, endorsed=True)
        AnnualProgressReportFactory(year=2022, endorsed=True)

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-current-year")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["current_year"] == 2023
        assert response.data["endorsed"] is True
        assert len(response.data["apr_list"]) == 3

    def test_no_reports_exist(self, apr_agency_viewer_user):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-current-year")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["current_year"] is None
        assert response.data["endorsed"] is False
        assert len(response.data["apr_list"]) == 0

    def test_uses_unendorsed_over_endorsed(self, apr_agency_viewer_user):
        AnnualProgressReportFactory(year=2023, endorsed=False)
        AnnualProgressReportFactory(year=2024, endorsed=True)

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-current-year")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["current_year"] == 2023
        assert response.data["endorsed"] is False
        assert len(response.data["apr_list"]) == 2

    def test_mlfs_user_can_access(
        self, secretariat_viewer_user, annual_progress_report
    ):
        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse("apr-current-year")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["current_year"] == annual_progress_report.year
        assert response.data["endorsed"] is False
        assert len(response.data["apr_list"]) == 1

    def test_agency_user_can_access(
        self, apr_agency_viewer_user, annual_progress_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-current-year")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["current_year"] == annual_progress_report.year
        assert response.data["endorsed"] is False
        assert len(response.data["apr_list"]) == 1


@pytest.mark.django_db
class TestAPRWorkspaceView(BaseTest):
    def test_without_login(self, apr_year):
        self.client.force_authenticate(user=None)
        url = reverse("apr-workspace", kwargs={"year": apr_year})
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_workspace_creates_agency_report(
        self, apr_agency_viewer_user, apr_year, agency, annual_progress_report
    ):
        """
        An agency report is created (if not existing) upon accessing the workspace view.
        Test that this is happening.
        """

        self.client.force_authenticate(user=apr_agency_viewer_user)

        url = reverse("apr-workspace", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["agency_id"] == agency.id
        assert response.data["progress_report_year"] == apr_year
        assert (
            response.data["status"] == AnnualAgencyProjectReport.SubmissionStatus.DRAFT
        )

        # Check that a database record was created
        assert AnnualAgencyProjectReport.objects.filter(
            agency=agency, progress_report__year=apr_year
        ).exists()

    def test_get_workspace_creates_project_reports(
        self,
        apr_agency_viewer_user,
        apr_year,
        multiple_projects_for_apr,
        annual_progress_report,
    ):
        """
        Project reports are created (if not existing) upon accessing the workspace view.
        Test that this is happening.
        """
        self.client.force_authenticate(user=apr_agency_viewer_user)

        url = reverse("apr-workspace", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        # Call should return 5 project reports (3 ongoing, 2 completed)
        assert response.data["total_projects"] == len(multiple_projects_for_apr)
        assert len(response.data["project_reports"]) == 5

    def test_get_workspace_requires_apr_view_permission(self, user, apr_year):
        self.client.force_authenticate(user=user)

        url = reverse("apr-workspace", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_workspace_filters_by_status(
        self,
        apr_agency_viewer_user,
        apr_year,
        annual_progress_report,
        multiple_projects_for_apr,
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)

        # Only filter by completed projects
        url = reverse("apr-workspace", kwargs={"year": apr_year})
        response = self.client.get(url, {"status": "COM"})

        assert response.status_code == status.HTTP_200_OK
        # Response should also include ONG projects, even if not selected
        assert response.data["total_projects"] == 5

    def test_get_workspace_idempotent(
        self,
        apr_agency_viewer_user,
        apr_year,
        annual_progress_report,
        multiple_projects_for_apr,
    ):
        """Test that multiple calls to workspace don't create duplicates."""
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-workspace", kwargs={"year": apr_year})

        # Call twice
        response1 = self.client.get(url)
        response2 = self.client.get(url)

        assert response1.status_code == status.HTTP_200_OK
        assert response2.status_code == status.HTTP_200_OK

        # both calls should return the same report, with no extra project report created
        assert response1.data["id"] == response2.data["id"]
        assert response1.data["total_projects"] == response2.data["total_projects"]
        assert (
            AnnualProjectReport.objects.filter(
                report__progress_report__year=apr_year
            ).count()
            == 5
        )

    def test_get_agency_report_nested_data(
        self,
        apr_agency_viewer_user,
        annual_progress_report,
        annual_agency_report,
        annual_project_report,
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)

        AnnualProjectReportFileFactory(report=annual_agency_report)

        url = reverse(
            "apr-workspace",
            kwargs={
                "year": annual_agency_report.progress_report.year,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert "project_reports" in response.data
        assert len(response.data["project_reports"]) == 1
        assert "files" in response.data
        assert len(response.data["files"]) == 1

    def test_get_workspace_only_own_agency(
        self, apr_agency_viewer_user, annual_progress_report
    ):
        # Create report for a different agency to check it can't be accessed
        other_agency = AgencyFactory()
        other_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=other_agency,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)

        url = reverse(
            "apr-workspace",
            kwargs={
                "year": other_report.progress_report.year,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert "project_reports" in response.data
        assert len(response.data["project_reports"]) == 0

    def test_support_cost_derived_fields(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        multiple_project_versions_for_apr,
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
            support_cost_disbursed=8000.0,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == latest_version.code
            ),
            None,
        )
        assert project_data is not None

        assert project_data["support_cost_approved"] == version3.support_cost_psc

        expected_adjustment = (
            latest_version.support_cost_psc - version3.support_cost_psc
        )
        assert project_data["support_cost_adjustment"] == expected_adjustment

        assert (
            project_data["support_cost_approved_plus_adjustment"]
            == latest_version.support_cost_psc
        )
        expected_balance = latest_version.support_cost_psc - 8000.0
        assert project_data["support_cost_balance"] == expected_balance

    def test_support_cost_derived_fields_single_version(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        agency,
        project_ongoing_status,
    ):
        # This project only has version 3
        single_version = ProjectFactory(
            agency=agency,
            status=project_ongoing_status,
            version=3,
            latest_project=None,
            support_cost_psc=15000.0,
            date_approved=date(
                annual_agency_report.progress_report.year - 1, 6, 15
            ),
        )

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=single_version,
            support_cost_disbursed=5000.0,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == single_version.code
            ),
            None,
        )
        assert project_data is not None

        assert project_data["support_cost_approved"] == 15000.0
        assert project_data["support_cost_adjustment"] == None
        assert project_data["support_cost_approved_plus_adjustment"] == 15000.0
        assert project_data["support_cost_balance"] == 10000.0


@pytest.mark.django_db
class TestAPRBulkUpdateView(BaseTest):
    def test_without_login(self, annual_agency_report):
        self.client.force_authenticate(user=None)
        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_bulk_update_project_reports(
        self,
        apr_agency_inputter_user,
        annual_agency_report,
        multiple_projects_for_apr,
    ):
        for project in multiple_projects_for_apr[:3]:
            AnnualProjectReportFactory(report=annual_agency_report, project=project)

        self.client.force_authenticate(user=apr_agency_inputter_user)

        update_data = {
            "project_reports": [
                {
                    "project_code": multiple_projects_for_apr[0].code,
                    "funds_disbursed": 50000.0,
                    "date_first_disbursement": "2024-01-15",
                },
                {
                    "project_code": multiple_projects_for_apr[1].code,
                    "consumption_phased_out_odp": 25.5,
                    "last_year_remarks": "On track",
                },
            ]
        }

        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, update_data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["updated_count"] == 2
        assert response.data["error_count"] == 0

    def test_bulk_update_requires_edit_permission(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)

        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"project_reports": []}, format="json")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_bulk_update_submitted_reports(
        self,
        apr_agency_inputter_user,
        mlfs_admin_user,
        annual_agency_report,
        annual_project_report,
    ):
        annual_agency_report.status = annual_agency_report.SubmissionStatus.SUBMITTED
        annual_agency_report.save()

        update_data = {
            "project_reports": [
                {
                    "project_code": annual_project_report.project.code,
                    "funds_disbursed": 50000.0,
                }
            ]
        }
        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )

        # Agency users should not be able to update, but mlfs ones should
        self.client.force_authenticate(user=apr_agency_inputter_user)
        response = self.client.post(url, update_data, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST

        self.client.force_authenticate(user=mlfs_admin_user)
        response = self.client.post(url, update_data, format="json")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["updated_count"] == 1

    def test_bulk_update_unlocked_reports(
        self,
        apr_agency_inputter_user,
        annual_agency_report,
        annual_project_report,
    ):
        annual_agency_report.status = annual_agency_report.SubmissionStatus.SUBMITTED
        annual_agency_report.is_unlocked = True
        annual_agency_report.save()

        self.client.force_authenticate(user=apr_agency_inputter_user)

        update_data = {
            "project_reports": [
                {
                    "project_code": annual_project_report.project.code,
                    "funds_disbursed": 50000.0,
                }
            ]
        }
        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, update_data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["updated_count"] == 1

    def test_bulk_update_endorsed_reports(
        self,
        apr_agency_inputter_user,
        mlfs_admin_user,
        annual_agency_report,
        annual_project_report,
    ):
        annual_agency_report.progress_report.endorsed = True
        annual_agency_report.progress_report.save()

        update_data = {
            "project_reports": [
                {
                    "project_code": annual_project_report.project.code,
                    "funds_disbursed": 50000.0,
                }
            ]
        }
        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )

        self.client.force_authenticate(user=apr_agency_inputter_user)
        response = self.client.post(url, update_data, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST

        self.client.force_authenticate(user=mlfs_admin_user)
        response = self.client.post(url, update_data, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_bulk_update_duplicate_codes(
        self,
        apr_agency_inputter_user,
        annual_agency_report,
        annual_project_report,
    ):
        self.client.force_authenticate(user=apr_agency_inputter_user)

        update_data = {
            "project_reports": [
                {
                    "project_code": annual_project_report.project.code,
                    "funds_disbursed": 50000.0,
                },
                {
                    "project_code": annual_project_report.project.code,
                    "funds_committed": 30000.0,
                },
            ]
        }

        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, update_data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Duplicate project codes" in str(response.data)

    def test_bulk_update_invalid_project_code(
        self,
        apr_agency_inputter_user,
        annual_agency_report,
        annual_project_report,
    ):
        self.client.force_authenticate(user=apr_agency_inputter_user)

        update_data = {
            "project_reports": [
                {
                    "project_code": "INVALID/CODE/99/XXX/99",
                    "funds_disbursed": 50000.0,
                },
                {
                    "project_code": annual_project_report.project.code,
                    "funds_disbursed": 30000.0,
                },
            ]
        }

        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, update_data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["updated_count"] == 1
        assert response.data["error_count"] == 1
        assert "errors" in response.data

    def test_bulk_update_partial_fields(
        self,
        apr_agency_inputter_user,
        annual_agency_report,
        annual_project_report,
    ):
        self.client.force_authenticate(user=apr_agency_inputter_user)

        # First update
        update_data = {
            "project_reports": [
                {
                    "project_code": annual_project_report.project.code,
                    "funds_disbursed": 50000.0,
                }
            ]
        }

        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, update_data, format="json")
        assert response.status_code == status.HTTP_200_OK

        # Second update, but affecting a different field
        del update_data["project_reports"][0]["funds_disbursed"]
        update_data["project_reports"][0]["consumption_phased_out_odp"] = 25.5

        response = self.client.post(url, update_data, format="json")
        assert response.status_code == status.HTTP_200_OK

        # Check that both fields are set
        annual_project_report.refresh_from_db()
        assert annual_project_report.funds_disbursed == 50000.0
        assert annual_project_report.consumption_phased_out_odp == 25.5


@pytest.mark.django_db
class TestAPRFileUploadView(BaseTest):
    def test_without_login(self, annual_agency_report):
        self.client.force_authenticate(user=None)
        test_file = SimpleUploadedFile(
            "test_report.pdf", b"file_content", content_type="application/pdf"
        )
        data = {
            "file": test_file,
            "file_name": "Annual Progress Report 2024",
            "file_type": AnnualProjectReportFile.FileType.ANNUAL_PROGRESS_FINANCIAL_REPORT,
        }

        url = reverse(
            "apr-file-upload",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, data, format="multipart")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_upload_file(self, apr_agency_inputter_user, annual_agency_report):
        self.client.force_authenticate(user=apr_agency_inputter_user)

        test_financial_file = SimpleUploadedFile(
            "test_report.pdf", b"file_content", content_type="application/pdf"
        )
        test_supporting_file = SimpleUploadedFile(
            "test_support.pdf",
            b"supporting_file_content",
            content_type="application/pdf",
        )
        data = {
            "financial_file": test_financial_file,
            "supporting_files": [test_supporting_file],
        }

        url = reverse(
            "apr-file-upload",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, data, format="multipart")

        assert response.status_code == status.HTTP_201_CREATED
        assert "files" in response.data
        assert len(response.data["files"]) == 2
        assert response.data["message"] == "Files uploaded successfully."

        assert AnnualProjectReportFile.objects.filter(
            report=annual_agency_report
        ).exists()

    def test_upload_file_permissions(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)

        url = reverse(
            "apr-file-upload",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="multipart")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_upload_submitted_report(
        self, apr_agency_inputter_user, annual_agency_report
    ):
        annual_agency_report.status = annual_agency_report.SubmissionStatus.SUBMITTED
        annual_agency_report.save()

        self.client.force_authenticate(user=apr_agency_inputter_user)

        test_file = SimpleUploadedFile(
            "test.pdf", b"content", content_type="application/pdf"
        )

        url = reverse(
            "apr-file-upload",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"file": test_file}, format="multipart")

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_upload_file_replaces_existing_annual_report(
        self, apr_agency_inputter_user, annual_agency_report
    ):
        # First add "existing" file to the annual agency report
        old_file = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_type=AnnualProjectReportFile.FileType.ANNUAL_PROGRESS_FINANCIAL_REPORT,
        )

        self.client.force_authenticate(user=apr_agency_inputter_user)

        # Then upload new file to check that it overwrites the old one
        new_file = SimpleUploadedFile(
            "new_report.pdf", b"new_content", content_type="application/pdf"
        )
        data = {
            "financial_file": new_file,
        }
        url = reverse(
            "apr-file-upload",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, data, format="multipart")

        assert response.status_code == status.HTTP_201_CREATED

        # Check that only one file exists and the old one was deleted
        assert (
            AnnualProjectReportFile.objects.filter(
                report=annual_agency_report,
                file_type=AnnualProjectReportFile.FileType.ANNUAL_PROGRESS_FINANCIAL_REPORT,
            ).count()
            == 1
        )
        assert not AnnualProjectReportFile.objects.filter(id=old_file.id).exists()

    def test_upload_unlocked_report(
        self, apr_agency_inputter_user, annual_agency_report
    ):
        annual_agency_report.status = annual_agency_report.SubmissionStatus.SUBMITTED
        annual_agency_report.is_unlocked = True
        annual_agency_report.save()

        self.client.force_authenticate(user=apr_agency_inputter_user)

        test_file = SimpleUploadedFile(
            "test.pdf", b"content", content_type="application/pdf"
        )
        data = {
            "financial_file": test_file,
        }
        url = reverse(
            "apr-file-upload",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, data, format="multipart")

        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["message"] == "Files uploaded successfully."

    def test_upload_file_to_endorsed_report(
        self, apr_agency_inputter_user, mlfs_admin_user, annual_agency_report
    ):
        annual_agency_report.progress_report.endorsed = True
        annual_agency_report.progress_report.save()

        test_file = SimpleUploadedFile(
            "test.pdf", b"content", content_type="application/pdf"
        )
        data = {"financial_file": test_file}

        url = reverse(
            "apr-file-upload",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )

        # Agency user cannot upload
        self.client.force_authenticate(user=apr_agency_inputter_user)
        response = self.client.post(url, data, format="multipart")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "endorsed" in str(response.data).lower()

        # MLFS admin also cannot upload
        self.client.force_authenticate(user=mlfs_admin_user)
        response = self.client.post(url, data, format="multipart")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "endorsed" in str(response.data).lower()

    def test_delete_file_from_endorsed_report(
        self, apr_agency_inputter_user, mlfs_admin_user, annual_agency_report
    ):
        file_obj = AnnualProjectReportFileFactory(report=annual_agency_report)

        annual_agency_report.progress_report.endorsed = True
        annual_agency_report.progress_report.save()

        url = reverse(
            "apr-file-delete",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
                "pk": file_obj.pk,
            },
        )

        self.client.force_authenticate(user=apr_agency_inputter_user)
        response = self.client.delete(url)
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "endorsed" in str(response.data).lower()

        self.client.force_authenticate(user=mlfs_admin_user)
        response = self.client.delete(url)
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "endorsed" in str(response.data).lower()


@pytest.mark.django_db
class TestAPRFileDownloadView(BaseTest):
    def test_without_login(self, annual_agency_report):
        file_obj = AnnualProjectReportFileFactory(report=annual_agency_report)

        self.client.force_authenticate(user=None)
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
                "pk": file_obj.pk,
            },
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_can_download_own_file(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        file_obj = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_name="test_report.pdf",
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
                "pk": file_obj.pk,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Disposition"] == 'attachment; filename="test_report.pdf"'
        )
        assert "Content-Type" in response

    def test_agency_user_cannot_download_other_agency_file(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        other_agency = AgencyFactory()
        other_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_agency_report.progress_report,
            agency=other_agency,
            is_unlocked=False,
        )
        file_obj = AnnualProjectReportFileFactory(report=other_report)

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": other_report.progress_report.year,
                "agency_id": other_report.agency.id,
                "pk": file_obj.pk,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_user_can_download_any_file(
        self, secretariat_viewer_user, annual_agency_report
    ):
        file_obj = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_name="mlfs_download_test.pdf",
        )

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
                "pk": file_obj.pk,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Disposition"]
            == 'attachment; filename="mlfs_download_test.pdf"'
        )

    def test_download_nonexistent_file(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
                "pk": 99999,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_download_file_with_wrong_year_or_agency(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        file_obj = AnnualProjectReportFileFactory(report=annual_agency_report)

        self.client.force_authenticate(user=apr_agency_viewer_user)

        # Wrong year
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": 9999,
                "agency_id": annual_agency_report.agency.id,
                "pk": file_obj.pk,
            },
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_404_NOT_FOUND

        # Wrong agency
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": 9999,
                "pk": file_obj.pk,
            },
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_download_financial_report_file(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        file_obj = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_type=AnnualProjectReportFile.FileType.ANNUAL_PROGRESS_FINANCIAL_REPORT,
            file_name="financial_report_2024.pdf",
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
                "pk": file_obj.pk,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            'attachment; filename="financial_report_2024.pdf"'
            in response["Content-Disposition"]
        )

    def test_download_supporting_document_file(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        file_obj = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_type=AnnualProjectReportFile.FileType.OTHER_SUPPORTING_DOCUMENT,
            file_name="supporting_doc.docx",
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-file-download",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
                "pk": file_obj.pk,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            'attachment; filename="supporting_doc.docx"'
            in response["Content-Disposition"]
        )

    def test_serializer_file_url(self, apr_agency_viewer_user, annual_agency_report):
        file_obj = AnnualProjectReportFileFactory(report=annual_agency_report)

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data["files"]) == 1

        file_url = response.data["files"][0]["file_url"]
        expected_file_url = (
            f"/api/annual-project-report/{annual_agency_report.progress_report.year}"
            f"/agency/{annual_agency_report.agency.id}/files/{file_obj.pk}/download/"
        )
        assert expected_file_url in file_url
        assert "/media/" not in file_url


@pytest.mark.django_db
class TestAPRGlobalViewSet(BaseTest):
    def test_without_login(self, apr_year):
        self.client.force_authenticate(user=None)
        url = reverse("apr-mlfs-list", kwargs={"year": apr_year})
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_cannot_access(self, apr_agency_viewer_user, apr_year):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-mlfs-list", kwargs={"year": apr_year})
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_viewer_can_access_reports(
        self, secretariat_viewer_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.DRAFT,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse("apr-mlfs-list", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data) == 1
        assert (
            response.data[0]["status"]
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )

    def test_mlfs_full_access_can_access_reports(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.DRAFT,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-list", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        # MLFS "full access" users cannot access draft reports
        assert len(response.data) == 1
        assert response.data[0]["agency_id"] == agency1.id

    def test_filter_by_agency(self, mlfs_admin_user, apr_year, annual_progress_report):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-list", kwargs={"year": apr_year})
        response = self.client.get(url, {"agency": agency1.id})

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data) == 1
        assert response.data[0]["agency_id"] == agency1.id

    def test_filter_by_status(
        self,
        mlfs_admin_user,
        apr_year,
        annual_progress_report,
        project_ongoing_status,
        project_closed_status,
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        report1 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        report2 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        project_ong = ProjectFactory(
            agency=agency1,
            status=project_ongoing_status,
            version=3,
            latest_project=None,
            date_approved=date(apr_year - 1, 1, 1),
        )
        project_clo = ProjectFactory(
            agency=agency2,
            status=project_closed_status,
            version=3,
            latest_project=None,
            date_approved=date(apr_year - 1, 1, 1),
        )

        AnnualProjectReportFactory(
            report=report1, project=project_ong, status=project_ong.status.name
        )
        AnnualProjectReportFactory(
            report=report2, project=project_clo, status=project_clo.status.name
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-list", kwargs={"year": apr_year})

        response = self.client.get(url, {"status": "CLO"})

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data) == 2

        # Check that them nested project reports are filtered correctly
        agency1_data = next(r for r in response.data if r["agency_id"] == agency1.id)
        assert len(agency1_data["project_reports"]) == 1
        assert agency1_data["project_reports"][0]["status"] == "Ongoing"

        agency2_data = next(r for r in response.data if r["agency_id"] == agency2.id)
        assert len(agency2_data["project_reports"]) == 1
        assert agency2_data["project_reports"][0]["status"] == "Closed"

    def test_filter_by_country(
        self, mlfs_admin_user, apr_year, annual_progress_report, country_ro, new_country
    ):
        agency = AgencyFactory()
        annual_agency_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        project = ProjectFactory(country=country_ro, agency=agency)
        project_new = ProjectFactory(country=new_country, agency=agency)

        AnnualProjectReportFactory(report=annual_agency_report, project=project)
        AnnualProjectReportFactory(report=annual_agency_report, project=project_new)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-list", kwargs={"year": apr_year})
        response = self.client.get(url, {"country": country_ro.name})

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data) == 1
        # Check that only the filtered country's project is returned
        assert len(response.data[0]["project_reports"]) == 1
        assert response.data[0]["project_reports"][0]["country_name"] == country_ro.name

    def test_unlocked_flag_list(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency = AgencyFactory()
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=True,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-list", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert len(response.data) == 0

    def test_get_report_detail(
        self,
        mlfs_admin_user,
        annual_agency_report,
        annual_project_report,
    ):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-detail",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["id"] == annual_agency_report.id
        assert response.data["agency_id"] == annual_agency_report.agency.id
        assert len(response.data["project_reports"]) == 1
        assert response.data["project_reports"][0]["id"] == annual_project_report.id

    def test_unlocked_flag_detail(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency = AgencyFactory()
        # It's unlocked, so it should not be accessible!
        annual_agency_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=True,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-detail",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_get_report_nested_data(
        self, mlfs_admin_user, annual_agency_report, annual_project_report
    ):
        self.client.force_authenticate(user=mlfs_admin_user)

        AnnualProjectReportFileFactory(report=annual_agency_report)
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )

        annual_agency_report.save()

        url = reverse(
            "apr-mlfs-detail",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert "project_reports" in response.data
        assert "files" in response.data
        assert len(response.data["files"]) == 1


@pytest.mark.django_db
class TestAPRToggleLockView(BaseTest):
    def test_without_login(self, annual_agency_report):
        self.client.force_authenticate(user=None)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"is_unlocked": True}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_cannot_toggle_lock(
        self, apr_agency_inputter_user, annual_agency_report
    ):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=apr_agency_inputter_user)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"is_unlocked": True}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_viewer_cannot_toggle_lock(
        self, secretariat_viewer_user, annual_agency_report
    ):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"is_unlocked": True}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_full_access_can_unlock(self, mlfs_admin_user, annual_agency_report):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.is_unlocked = False
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"is_unlocked": True}, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["is_unlocked"] is True
        assert "unlocked successfully" in response.data["message"]

        # Verify database was updated
        annual_agency_report.refresh_from_db()
        assert annual_agency_report.is_unlocked is True

    def test_mlfs_full_access_can_lock(self, mlfs_admin_user, annual_agency_report):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.is_unlocked = True
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"is_unlocked": False}, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["is_unlocked"] is False
        assert "locked successfully" in response.data["message"]

        # Verify database was updated
        annual_agency_report.refresh_from_db()
        assert annual_agency_report.is_unlocked is False

    def test_cannot_toggle_draft_report(self, mlfs_admin_user, annual_agency_report):
        annual_agency_report.status = AnnualAgencyProjectReport.SubmissionStatus.DRAFT
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"is_unlocked": True}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_cannot_toggle_endorsed_report(self, mlfs_admin_user, annual_agency_report):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.progress_report.endorsed = True
        annual_agency_report.progress_report.save()
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {"is_unlocked": True}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_missing_is_unlocked_parameter(self, mlfs_admin_user, annual_agency_report):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST


@pytest.mark.django_db
class TestAPREndorseView(BaseTest):
    def test_without_login(self, apr_year):
        self.client.force_authenticate(user=None)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_cannot_endorse(
        self, apr_agency_inputter_user, apr_year, annual_progress_report
    ):
        self.client.force_authenticate(user=apr_agency_inputter_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_viewer_cannot_endorse(
        self, secretariat_viewer_user, apr_year, annual_progress_report
    ):
        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_full_access_can_endorse(
        self, mlfs_admin_user, apr_year, meeting_apr_same_year, annual_progress_report
    ):
        # Create some submitted reports to allow endorsing
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        data = {
            "date_endorsed": timezone.now().date().isoformat(),
            "meeting_endorsed": meeting_apr_same_year.id,
            "remarks_endorsed": "Endorsed with all reports submitted.",
        }
        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert "endorsed successfully" in response.data["message"]
        assert response.data["year"] == apr_year
        assert response.data["total_agencies"] == 2

        annual_progress_report.refresh_from_db()
        assert annual_progress_report.endorsed is True

    def test_endorse_missing_required_fields(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency = AgencyFactory()
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})

        # Missing all required fields
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "date_endorsed" in response.data
        assert "meeting_endorsed" in response.data

    def test_cannot_endorse_if_draft_reports_exist(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.DRAFT,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_cannot_endorse_if_unlocked_reports_exist(
        self, mlfs_admin_user, apr_year, annual_progress_report, meeting_apr_same_year
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        # Fixturing two submitted reports - one locked, one unlocked
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        unlocked_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=True,
        )

        self.client.force_authenticate(user=mlfs_admin_user)

        # GET should show report as not endorsable
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["is_endorsable"] is False
        assert response.data["total_agencies"] == 2
        assert response.data["submitted_agencies"] == 1
        assert response.data["draft_agencies"] == 1
        assert unlocked_report.agency.name in response.data["draft_agency_names"]

        # POST should fail to endorse
        data = {
            "date_endorsed": timezone.now().date().isoformat(),
            "meeting_endorsed": meeting_apr_same_year.id,
            "remarks_endorsed": "Should not work with unlocked reports.",
        }
        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "unlocked" in str(response.data).lower()
        assert unlocked_report.agency.name in str(response.data)

    def test_cannot_endorse_if_no_submitted_reports_exist(
        self, mlfs_admin_user, apr_year, annual_progress_report, meeting_apr_same_year
    ):
        self.client.force_authenticate(user=mlfs_admin_user)

        # GET should show report as not endorsable
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["is_endorsable"] is False
        assert response.data["total_agencies"] == 0
        assert response.data["submitted_agencies"] == 0
        assert response.data["draft_agencies"] == 0

        # POST should fail to endorse
        data = {
            "date_endorsed": timezone.now().date().isoformat(),
            "meeting_endorsed": meeting_apr_same_year.id,
            "remarks_endorsed": "Should not work with unlocked reports.",
        }
        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "no agencies have submitted reports" in str(response.data).lower()

    def test_endorse_already_endorsed(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        annual_progress_report.endorsed = True
        annual_progress_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_endorse_nonexistent_year(self, mlfs_admin_user):
        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-endorse", kwargs={"year": 9999})
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_get_endorsement_status_not_endorsed(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["year"] == apr_year
        assert response.data["endorsed"] is False
        assert response.data["is_endorsable"] is True
        assert response.data["total_agencies"] == 2
        assert response.data["submitted_agencies"] == 2
        assert response.data["draft_agencies"] == 0
        assert response.data["draft_agency_names"] == []

    def test_get_endorsement_status_with_draft_reports(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        draft_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.DRAFT,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["is_endorsable"] is False
        assert response.data["total_agencies"] == 2
        assert response.data["submitted_agencies"] == 1
        assert response.data["draft_agencies"] == 1
        assert draft_report.agency.name in response.data["draft_agency_names"]

    def test_get_endorsement_status_already_endorsed(
        self,
        mlfs_admin_user,
        apr_year,
        annual_progress_report_endorsed,
        meeting_apr_same_year,
    ):
        agency = AgencyFactory()
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report_endorsed,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["endorsed"] is True
        assert response.data["is_endorsable"] is False
        assert (
            response.data["date_endorsed"]
            == annual_progress_report_endorsed.date_endorsed.isoformat()
        )
        assert response.data["meeting_endorsed"] == meeting_apr_same_year.id
        assert response.data["remarks_endorsed"] == "Test endorsement"

    def test_get_endorsement_status_requires_mlfs_access(
        self, apr_agency_viewer_user, apr_year, annual_progress_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-endorse", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
class TestAPRExportView(BaseTest):
    def test_without_login(self, apr_year, agency):
        self.client.force_authenticate(user=None)
        url = reverse(
            "apr-export",
            kwargs={"year": apr_year, "agency_id": agency.id},
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_can_export_own_report(
        self, apr_agency_viewer_user, annual_agency_report, annual_project_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response["Content-Disposition"]
        assert (
            f"APR_{annual_agency_report.progress_report.year}"
            in response["Content-Disposition"]
        )
        assert (
            annual_agency_report.agency.name[:10] in response["Content-Disposition"]
        )  # Partial name match

        assert len(response.content) > 0

    def test_agency_user_cannot_export_other_agency_report(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        other_agency = AgencyFactory()
        other_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_agency_report.progress_report,
            agency=other_agency,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={
                "year": other_report.progress_report.year,
                "agency_id": other_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_user_can_export_any_agency_report(
        self, secretariat_viewer_user, annual_agency_report, annual_project_report
    ):
        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert len(response.content) > 0

    def test_export_with_multiple_projects(
        self, apr_agency_viewer_user, annual_agency_report, multiple_projects_for_apr
    ):
        for project in multiple_projects_for_apr[:3]:
            AnnualProjectReportFactory(
                report=annual_agency_report,
                project=project,
                funds_disbursed=10000.0,
            )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert len(response.content) > 0

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        # Ensure that all 3 rows are exported, as there are no filters
        data_rows = worksheet.max_row - APRExportWriter.HEADER_ROW
        assert data_rows == 3

    def test_export_with_status_filter(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        annual_project_report,
        project_ongoing_status,
    ):
        annual_project_report.project.status = project_ongoing_status
        annual_project_report.project.save(update_fields=["status"])

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url, {"status": "ONG"})

        assert response.status_code == status.HTTP_200_OK
        assert len(response.content) > 0

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.HEADER_ROW
        assert data_rows == 1

    def test_export_empty_report(self, apr_agency_viewer_user, annual_agency_report):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_200_OK
        assert len(response.content) > 0

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.HEADER_ROW
        # For empty reports, we still export one empty data row
        assert data_rows == 1

        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW + 1
        row_values = [
            worksheet.cell(first_data_row, col).value
            for col in range(1, len(columns) + 1)
        ]
        assert all(value is None or value == "" for value in row_values)

    def test_export_nonexistent_report(self, apr_agency_viewer_user):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={"year": 9999, "agency_id": 9999},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_export_data_correctness(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        project_ongoing_status,
    ):
        project = ProjectFactory(
            code="TEST/CODE/2024/001",
            title="Test Project Title",
            agency=annual_agency_report.agency,
            status=project_ongoing_status,
        )

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project,
            funds_disbursed=123456.78,
            last_year_remarks="Test remarkss",
            consumption_phased_out_odp=45.67,
            date_first_disbursement="2024-03-15",
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        # All project data gets written to first data row; check it matches
        project_code_col = columns["project_code"]
        assert (
            worksheet.cell(first_data_row, project_code_col).value
            == "TEST/CODE/2024/001"
        )

        project_title_col = columns["project_title"]
        assert (
            worksheet.cell(first_data_row, project_title_col).value
            == "Test Project Title"
        )

        funds_disbursed_col = columns["funds_disbursed"]
        assert worksheet.cell(first_data_row, funds_disbursed_col).value == 123456.78

        odp_col = columns["consumption_phased_out_odp"]
        assert worksheet.cell(first_data_row, odp_col).value == 45.67

        date_col = columns["date_first_disbursement"]
        cell_value = worksheet.cell(first_data_row, date_col).value
        assert cell_value == datetime(2024, 3, 15, 0, 0)

        remarks_col = columns["last_year_remarks"]
        assert worksheet.cell(first_data_row, remarks_col).value == "Test remarkss"

        agency_col = columns["agency_name"]
        assert worksheet.cell(first_data_row, agency_col).value == project.agency.name

        country_col = columns["country_name"]
        assert worksheet.cell(first_data_row, country_col).value == project.country.name

    def test_export_null_values(
        self, apr_agency_viewer_user, annual_agency_report, project_ongoing_status
    ):
        project = ProjectFactory(
            code="TEST/CODE/2024/001",
            title="Test Project Title",
            agency=annual_agency_report.agency,
            status=project_ongoing_status,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project,
            funds_disbursed=None,
            consumption_phased_out_odp=45.67,
            last_year_remarks="",
            date_first_disbursement=None,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-export",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        funds_col = columns["funds_disbursed"]
        assert worksheet.cell(first_data_row, funds_col).value is None

        date_col = columns["date_first_disbursement"]
        assert worksheet.cell(first_data_row, date_col).value is None

        remarks_col = columns["last_year_remarks"]
        cell_value = worksheet.cell(first_data_row, remarks_col).value
        assert cell_value is None


@pytest.mark.django_db
class TestAnnualProjectReportDerivedProperties(BaseTest):
    def test_without_login(self, **kwargs):
        pass

    def test_derived_properties_with_multiple_versions(
        self, annual_agency_report, multiple_project_versions_for_apr
    ):
        version3 = multiple_project_versions_for_apr[0]
        # just a sanity check
        assert version3.version == 3

        middle_version = multiple_project_versions_for_apr[1]
        # just a sanity check
        assert middle_version.total_fund == 125000.0

        latest_version = multiple_project_versions_for_apr[2]

        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version3,
            funds_disbursed=80000.0,
            support_cost_disbursed=8000.0,
        )

        assert (
            annual_report.adjustment == latest_version.total_fund - version3.total_fund
        )
        assert (
            annual_report.approved_funding_plus_adjustment == latest_version.total_fund
        )
        assert (
            annual_report.per_cent_funds_disbursed
            == 80000.0 / latest_version.total_fund
        )
        assert annual_report.support_cost_adjustment == 5000.0
        assert annual_report.support_cost_approved_plus_adjustment == 15000.0
        assert annual_report.support_cost_balance == 15000.0 - 8000.0

    def test_derived_properties_with_later_latest_version(
        self,
        annual_agency_report,
        late_post_excom_versions_for_apr,
    ):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()
        # Below, we are also setting the project to the *latest* (next-year) version,
        # to check the logic still behaves correctly.
        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=late_post_excom_versions_for_apr[1],
            funds_disbursed=80000.0,
            support_cost_disbursed=8000.0,
        )

        # All properties depending on latest_version_for_year should return None
        assert annual_report.adjustment is None
        assert annual_report.support_cost_adjustment is None

        # And the addition/substraction-based ones should return initial values
        assert annual_report.approved_funding_plus_adjustment == 100000.0
        assert annual_report.support_cost_approved_plus_adjustment == 10000.0
        assert annual_report.support_cost_balance == 10000.0 - 8000.0
        assert annual_report.per_cent_funds_disbursed == 0.8

    def test_derived_properties_with_no_latest_version(
        self,
        annual_agency_report,
        initial_project_version_for_apr,
    ):
        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=initial_project_version_for_apr,
            funds_disbursed=80000.0,
            support_cost_disbursed=8000.0,
        )

        # All properties depending on latest_version_for_year should return None
        assert annual_report.adjustment is None
        assert annual_report.support_cost_adjustment is None

        # And the addition/substraction-based ones should return initial values
        assert annual_report.approved_funding_plus_adjustment == 100000.0
        assert annual_report.support_cost_approved_plus_adjustment == 10000.0
        assert annual_report.support_cost_balance == 10000.0 - 8000.0
        assert annual_report.per_cent_funds_disbursed == 0.8

    def test_pcr_due_when_ong_to_com_in_same_year(
        self,
        annual_agency_report,
        agency,
        country_ro,
        sector,
        project_ongoing_status,
        project_completed_status,
        decision_apr_same_year,
    ):
        version4 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_completed_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/ONG/01",
            version=4,
            post_excom_decision=decision_apr_same_year,
            total_fund=100000.0,
        )
        version3 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/ONG/01",
            version=3,
            post_excom_decision=None,
            latest_project=version4,
            total_fund=100000.0,
        )

        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version3,
        )

        assert annual_report.pcr_due is True

    def test_pcr_due_when_ong_previous_year_to_com_current_year(
        self,
        annual_agency_report,
        agency,
        country_ro,
        sector,
        project_ongoing_status,
        project_completed_status,
        decision_apr_previous_year,
        decision_apr_same_year,
    ):
        version4 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_completed_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/PREV/01",
            version=4,
            post_excom_decision=decision_apr_same_year,
            total_fund=100000.0,
        )
        version3 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/PREV/01",
            version=3,
            post_excom_decision=decision_apr_previous_year,
            latest_project=version4,
            total_fund=100000.0,
        )

        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version3,
        )

        assert annual_report.pcr_due is True

    def test_pcr_due_false_when_already_completed_previous_year(
        self,
        annual_agency_report,
        agency,
        country_ro,
        sector,
        project_completed_status,
        decision_apr_previous_year,
    ):
        version3 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_completed_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/OLD/01",
            version=3,
            post_excom_decision=decision_apr_previous_year,
            total_fund=100000.0,
        )

        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version3,
        )
        assert annual_report.pcr_due is False

    def test_pcr_due_false_when_still_ongoing(
        self,
        annual_agency_report,
        agency,
        country_ro,
        sector,
        project_ongoing_status,
        decision_apr_same_year,
    ):
        version3 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/STILL/01",
            version=3,
            post_excom_decision=decision_apr_same_year,
            total_fund=100000.0,
        )

        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version3,
        )

        assert annual_report.pcr_due is False

    def test_pcr_due_with_clo_status(
        self,
        annual_agency_report,
        agency,
        country_ro,
        sector,
        project_ongoing_status,
        project_closed_status,
        decision_apr_same_year,
    ):
        version4 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_closed_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/CLO/01",
            version=4,
            post_excom_decision=decision_apr_same_year,
            total_fund=100000.0,
        )
        version3 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/CLO/01",
            version=3,
            post_excom_decision=None,
            latest_project=version4,
            total_fund=100000.0,
        )

        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version3,
        )
        assert annual_report.pcr_due is False

    def test_pcr_due_false_with_no_versions_in_year(
        self,
        annual_agency_report,
        agency,
        country_ro,
        sector,
        project_ongoing_status,
    ):
        version3 = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            date_approved=date(2021, 6, 1),
            code="TEST/PCR/NONE/01",
            version=3,
            post_excom_decision=None,
            total_fund=100000.0,
        )

        annual_report = AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version3,
        )

        assert annual_report.pcr_due is False


@pytest.mark.django_db
class TestAPRMLFSBulkUpdateView(BaseTest):
    def test_without_login(self, apr_year):
        self.client.force_authenticate(user=None)
        url = reverse("apr-mlfs-bulk-update", kwargs={"year": apr_year})
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_can_bulk_update_across_agencies(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        report1 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        report2 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        project1 = ProjectFactory(
            code="TEST/CODE/2024/001", agency=agency1, version=3, latest_project=None
        )
        project2 = ProjectFactory(
            code="TEST/CODE/2024/002", agency=agency2, version=3, latest_project=None
        )

        pr1 = AnnualProjectReportFactory(report=report1, project=project1)
        pr2 = AnnualProjectReportFactory(report=report2, project=project2)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-bulk-update", kwargs={"year": apr_year})

        data = {
            "project_reports": [
                {
                    "id": pr1.id,
                    "project_code": pr1.project.code,
                    "funds_disbursed": 100000,
                },
                {
                    "id": pr2.id,
                    "project_code": pr2.project.code,
                    "funds_disbursed": 200000,
                },
            ]
        }

        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["updated_count"] == 2
        assert response.data["error_count"] == 0

        pr1.refresh_from_db()
        pr2.refresh_from_db()
        assert pr1.funds_disbursed == 100000
        assert pr2.funds_disbursed == 200000

    def test_mlfs_bulk_update_without_id(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-bulk-update", kwargs={"year": apr_year})

        data = {
            "project_reports": [{"project_code": "12345", "funds_disbursed": 100000}]
        }

        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["updated_count"] == 0
        assert response.data["error_count"] == 1
        assert "Missing 'id' field" in response.data["errors"][0]["error"]

    def test_mlfs_bulk_update_nonexistent_id(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-bulk-update", kwargs={"year": apr_year})

        data = {
            "project_reports": [
                {"id": 99999, "project_code": "12345", "funds_disbursed": 100000}
            ]
        }

        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["updated_count"] == 0
        assert response.data["error_count"] == 1
        assert "not found" in response.data["errors"][0]["error"]

    def test_mlfs_bulk_update_duplicate_ids(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency = AgencyFactory()
        report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        project = ProjectFactory(
            agency=agency, code="12345", version=3, latest_project=None
        )
        pr = AnnualProjectReportFactory(report=report, project=project)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-bulk-update", kwargs={"year": apr_year})

        data = {
            "project_reports": [
                # we're using the same id for both reports
                {"id": pr.id, "project_code": project.code, "funds_disbursed": 100000},
                {"id": pr.id, "project_code": project.code, "funds_disbursed": 200000},
            ]
        }

        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Duplicate project report IDs" in str(response.data)

    def test_agency_user_cannot_access_mlfs_bulk_update(
        self, apr_agency_inputter_user, apr_year, annual_progress_report
    ):
        self.client.force_authenticate(user=apr_agency_inputter_user)
        url = reverse("apr-mlfs-bulk-update", kwargs={"year": apr_year})

        data = {"project_reports": [{"id": 1, "funds_disbursed": 100000}]}

        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_cannot_update_endorsed_reports(
        self, mlfs_admin_user, apr_year, annual_progress_report_endorsed
    ):
        agency = AgencyFactory()
        report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report_endorsed,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        project = ProjectFactory(agency=agency, version=3, latest_project=None)
        project_report = AnnualProjectReportFactory(report=report, project=project)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-bulk-update", kwargs={"year": apr_year})

        data = {
            "project_reports": [{"id": project_report.id, "funds_disbursed": 500000}]
        }

        response = self.client.post(url, data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "endorsed" in str(response.data).lower()


@pytest.mark.django_db
class TestAPRKickStartView(BaseTest):
    def test_without_login(self):
        self.client.force_authenticate(user=None)
        url = reverse("apr-kick-start")

        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_cannot_kick_start(self, apr_agency_viewer_user):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-kick-start")

        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_viewer_cannot_kick_start(
        self, secretariat_viewer_user, annual_progress_report_endorsed
    ):
        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse("apr-kick-start")

        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_status_with_no_endorsed_years(self, mlfs_admin_user):
        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["can_kick_start"] is False
        assert response.data["latest_endorsed_year"] is None
        assert response.data["next_year"] is None
        assert "No endorsed APRs exist" in response.data["message"]

    def test_get_status_when_can_kick_start(
        self, mlfs_admin_user, annual_progress_report_endorsed
    ):
        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["can_kick_start"] is True
        assert (
            response.data["latest_endorsed_year"]
            == annual_progress_report_endorsed.year
        )
        assert response.data["next_year"] == annual_progress_report_endorsed.year + 1
        assert response.data["unendorsed_years"] == []

    def test_get_status_when_unendorsed_year_exists(
        self, mlfs_admin_user, annual_progress_report_endorsed
    ):
        # Create an unendorsed APR for next year
        next_year = annual_progress_report_endorsed.year + 1
        AnnualProgressReportFactory(year=next_year, endorsed=False)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["can_kick_start"] is False
        assert response.data["next_year"] is None
        assert next_year in response.data["unendorsed_years"]
        assert "must be endorsed before creating" in response.data["message"]

    def test_kick_start_with_no_endorsed_years(self, mlfs_admin_user):
        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "No endorsed APRs exist" in str(response.data)

    def test_kick_start_when_unendorsed_year_exists(
        self, mlfs_admin_user, annual_progress_report_endorsed
    ):
        # Create an "existing" unendorsed APR, so that another one can't be created
        next_year = annual_progress_report_endorsed.year + 1
        AnnualProgressReportFactory(year=next_year, endorsed=False)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert f"APR for {next_year} already exists" in str(response.data)

    def test_mlfs_admin_can_kick_start_new_year(
        self, mlfs_admin_user, annual_progress_report_endorsed
    ):
        next_year = annual_progress_report_endorsed.year + 1

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["year"] == next_year
        assert response.data["previous_year"] == annual_progress_report_endorsed.year
        assert "initialized successfully" in response.data["message"]

        new_apr = AnnualProgressReport.objects.get(year=next_year)
        assert new_apr.endorsed is False
        assert new_apr.created_by == mlfs_admin_user
        assert new_apr.created_at is not None

    def test_kick_start_creates_unendorsed_apr(
        self, mlfs_admin_user, annual_progress_report_endorsed
    ):
        next_year = annual_progress_report_endorsed.year + 1

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_201_CREATED

        new_apr = AnnualProgressReport.objects.get(year=next_year)
        assert new_apr.endorsed is False
        assert new_apr.meeting_endorsed is None
        assert new_apr.date_endorsed is None
        assert new_apr.remarks_endorsed == ""
        assert new_apr.created_by == mlfs_admin_user

    def test_kick_start_uses_latest_endorsed_year(self, mlfs_admin_user):
        AnnualProgressReportFactory(year=2022, endorsed=True)
        AnnualProgressReportFactory(year=2024, endorsed=True)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["year"] == 2025
        assert response.data["previous_year"] == 2024

    def test_kick_start_idempotent(
        self, mlfs_admin_user, annual_progress_report_endorsed
    ):
        # Cannot kick-start the same year twice.
        next_year = annual_progress_report_endorsed.year + 1

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-kick-start")

        response1 = self.client.post(url, {}, format="json")
        assert response1.status_code == status.HTTP_201_CREATED

        response2 = self.client.post(url, {}, format="json")
        assert response2.status_code == status.HTTP_400_BAD_REQUEST
        assert f"APR for {next_year} already exists" in str(response2.data)


@pytest.mark.django_db
class TestAPRWorkspaceAccessControl(BaseTest):
    """Tests that workspace access & initialization requires an initial kick-start"""

    def test_without_login(self, **kwargs):
        pass

    def test_workspace_access_without_kickstart(self, apr_agency_viewer_user, agency):
        non_existent_year = 2030

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-workspace", kwargs={"year": non_existent_year})

        response = self.client.get(url)

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "not been started" in str(response.data)

    def test_workspace_access_after_kickstart(
        self, apr_agency_viewer_user, agency, mlfs_admin_user
    ):
        AnnualProgressReportFactory(year=2023, endorsed=True)

        self.client.force_authenticate(user=mlfs_admin_user)
        kickstart_url = reverse("apr-kick-start")
        response = self.client.post(kickstart_url, {}, format="json")
        assert response.status_code == status.HTTP_201_CREATED

        self.client.force_authenticate(user=apr_agency_viewer_user)
        workspace_url = reverse("apr-workspace", kwargs={"year": 2024})

        response = self.client.get(workspace_url)

        assert response.status_code == status.HTTP_200_OK
        assert response.data["progress_report_year"] == 2024

    def test_workspace_creates_agency_report_after_kickstart(
        self, apr_agency_viewer_user, agency, annual_progress_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-workspace", kwargs={"year": annual_progress_report.year})

        assert not AnnualAgencyProjectReport.objects.filter(
            progress_report=annual_progress_report, agency=agency
        ).exists()

        response = self.client.get(url)
        assert response.status_code == status.HTTP_200_OK

        agency_report = AnnualAgencyProjectReport.objects.get(
            progress_report=annual_progress_report, agency=agency
        )
        assert agency_report.status == AnnualAgencyProjectReport.SubmissionStatus.DRAFT
        assert agency_report.created_by == apr_agency_viewer_user

    def test_workspace_prepopulates_from_previous_year(
        self,
        apr_agency_viewer_user,
        agency,
        mlfs_admin_user,
        annual_progress_report_endorsed,
        multiple_projects_for_apr,
    ):
        previous_year = annual_progress_report_endorsed.year
        previous_agency_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report_endorsed,
            agency=agency,
        )

        project = multiple_projects_for_apr[0]
        AnnualProjectReportFactory(
            report=previous_agency_report,
            project=project,
            status="ONG",
            funds_disbursed=100000.0,
            funds_committed=50000.0,
            last_year_remarks="Previous year remarks",
            gender_policy=True,
        )

        # Kick-start next year
        self.client.force_authenticate(user=apr_agency_viewer_user)
        next_year = previous_year + 1

        self.client.force_authenticate(user=mlfs_admin_user)
        kickstart_url = reverse("apr-kick-start")
        self.client.post(kickstart_url, {}, format="json")

        # Now access workspace as agency; data should be pre-populated
        self.client.force_authenticate(user=apr_agency_viewer_user)
        workspace_url = reverse("apr-workspace", kwargs={"year": next_year})

        response = self.client.get(workspace_url)
        assert response.status_code == status.HTTP_200_OK

        new_agency_report = AnnualAgencyProjectReport.objects.get(
            progress_report__year=next_year, agency=agency
        )
        new_project_report = new_agency_report.project_reports.get(project=project)

        assert new_project_report.status == "ONG"
        assert new_project_report.funds_disbursed == 100000.0
        assert new_project_report.funds_committed == 50000.0
        assert new_project_report.last_year_remarks == "Previous year remarks"
        assert new_project_report.gender_policy is True

    def test_workspace_creates_new_projects_not_in_previous_year(
        self,
        apr_agency_viewer_user,
        agency,
        mlfs_admin_user,
        country_ro,
        sector,
        project_ongoing_status,
        annual_progress_report_endorsed,
    ):
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report_endorsed,
            agency=agency,
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        kickstart_url = reverse("apr-kick-start")
        next_year_apr = self.client.post(kickstart_url, {}, format="json")
        next_year = next_year_apr.data["year"]

        # Create a new project to check its report is not pre-populated
        new_project = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            code="NEW/PROJECT/2024/INV/01",
            version=3,
            latest_project=None,
            date_approved=date(next_year - 1, 1, 1),
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        workspace_url = reverse("apr-workspace", kwargs={"year": next_year})

        response = self.client.get(workspace_url)
        assert response.status_code == status.HTTP_200_OK

        # Check that data for the new project report was created with default values
        # (not pre-populated)
        new_agency_report = AnnualAgencyProjectReport.objects.get(
            progress_report__year=next_year, agency=agency
        )
        new_project_report = new_agency_report.project_reports.get(project=new_project)

        assert new_project_report.funds_disbursed is None
        assert new_project_report.last_year_remarks == ""

    def test_workspace_matches_by_project_code_and_agency(
        self,
        apr_agency_viewer_user,
        agency,
        mlfs_admin_user,
        country_ro,
        sector,
        project_ongoing_status,
        annual_progress_report_endorsed,
    ):
        previous_year = annual_progress_report_endorsed.year
        previous_agency_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report_endorsed,
            agency=agency,
        )

        project = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            code="UNIQUE/CODE/2023/INV/01",
            version=3,
            latest_project=None,
            date_approved=date(previous_year - 1, 1, 1),
        )

        AnnualProjectReportFactory(
            report=previous_agency_report,
            project=project,
            funds_disbursed=250000.0,
            current_year_remarks="Match by code test",
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        kickstart_url = reverse("apr-kick-start")
        response = self.client.post(kickstart_url, {}, format="json")
        next_year = response.data["year"]

        self.client.force_authenticate(user=apr_agency_viewer_user)
        workspace_url = reverse("apr-workspace", kwargs={"year": next_year})

        response = self.client.get(workspace_url)
        assert response.status_code == status.HTTP_200_OK

        new_agency_report = AnnualAgencyProjectReport.objects.get(
            progress_report__year=next_year, agency=agency
        )
        new_project_report = new_agency_report.project_reports.get(project=project)

        assert new_project_report.funds_disbursed == 250000.0
        assert new_project_report.current_year_remarks == "Match by code test"

    def test_workspace_no_prepopulation_if_no_previous_year(
        self,
        apr_agency_viewer_user,
        agency,
        multiple_projects_for_apr,
    ):
        # Create "first ever" APR (no previous APR year)
        first_apr = AnnualProgressReportFactory(year=2025, endorsed=False)

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-workspace", kwargs={"year": 2025})

        response = self.client.get(url)
        assert response.status_code == status.HTTP_200_OK

        # Check that all project reports were created with default values
        # (this means that status was inherited from project)
        agency_report = AnnualAgencyProjectReport.objects.get(
            progress_report=first_apr, agency=agency
        )
        assert agency_report.project_reports.exists() is True
        for project_report in agency_report.project_reports.all():
            assert project_report.funds_disbursed is None
            assert project_report.last_year_remarks == ""

            # Check that status was carried over from the project
            project = next(
                p
                for p in multiple_projects_for_apr
                if p.code == project_report.project.code
            )
            assert project.status.name == project_report.status

    def test_workspace_uses_previous_year_status_over_project_status(
        self,
        apr_agency_viewer_user,
        agency,
        mlfs_admin_user,
        country_ro,
        sector,
        project_ongoing_status,
        project_completed_status,
        annual_progress_report_endorsed,
    ):
        previous_year = annual_progress_report_endorsed.year
        project = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            version=3,
            date_approved=date(previous_year - 1, 6, 1),
        )
        previous_agency_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report_endorsed,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
        )

        # This has a different status from the project's status
        AnnualProjectReportFactory(
            report=previous_agency_report,
            project=project,
            status=project_completed_status.name,
            funds_disbursed=50000.0,
            last_year_remarks="Previous year data",
        )

        # Kick-start new year & then access agency workspace
        self.client.force_authenticate(user=mlfs_admin_user)
        kickstart_url = reverse("apr-kick-start")
        kickstart_response = self.client.post(kickstart_url)
        assert kickstart_response.status_code == status.HTTP_201_CREATED

        new_year = previous_year + 1
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-workspace", kwargs={"year": new_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        project_reports = response.data["project_reports"]
        project_report = next(
            (pr for pr in project_reports if pr["project_code"] == project.code),
            None,
        )
        assert project_report is not None

        # Status should be taken from the APR, not the project
        assert project_report["status"] == project_completed_status.name
        assert project_report["status"] != project.status.name
        assert project_report["funds_disbursed"] == 50000.0
        assert project_report["last_year_remarks"] == "Previous year data"

    def test_workspace_only_prepopulates_matching_projects(
        self,
        apr_agency_viewer_user,
        agency,
        mlfs_admin_user,
        country_ro,
        sector,
        project_ongoing_status,
        annual_progress_report_endorsed,
    ):
        previous_year = annual_progress_report_endorsed.year
        previous_agency_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report_endorsed,
            agency=agency,
        )

        old_project = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            code="OLD/PROJECT/2023/INV/01",
            version=3,
            latest_project=None,
            date_approved=date(previous_year - 1, 1, 1),
        )

        # Only create an existing report for the old project
        AnnualProjectReportFactory(
            report=previous_agency_report,
            project=old_project,
            funds_disbursed=100000.0,
        )

        # Now kick-start a new APR, then create a new project (which did not exist)
        self.client.force_authenticate(user=mlfs_admin_user)
        kickstart_url = reverse("apr-kick-start")
        response = self.client.post(kickstart_url, {}, format="json")
        next_year = response.data["year"]

        new_project = ProjectFactory(
            agency=agency,
            country=country_ro,
            sector=sector,
            status=project_ongoing_status,
            code="NEW/PROJECT/2024/INV/01",
            version=3,
            latest_project=None,
            date_approved=date(next_year - 1, 1, 1),
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        workspace_url = reverse("apr-workspace", kwargs={"year": next_year})

        response = self.client.get(workspace_url)
        assert response.status_code == status.HTTP_200_OK

        new_agency_report = AnnualAgencyProjectReport.objects.get(
            progress_report__year=next_year, agency=agency
        )

        old_project_report = new_agency_report.project_reports.get(project=old_project)
        assert old_project_report.funds_disbursed == 100000.0

        new_project_report = new_agency_report.project_reports.get(project=new_project)
        assert new_project_report.funds_disbursed is None


@pytest.mark.django_db
class TestAPRStatusView(BaseTest):
    def test_without_login(self, annual_agency_report):
        self.client.force_authenticate(user=None)
        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_viewer_cannot_submit(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_inputter_cannot_submit(
        self, apr_agency_inputter_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_inputter_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_submitter_can_submit_draft_report(
        self, apr_agency_submitter_user, annual_agency_report
    ):
        assert (
            annual_agency_report.status
            == AnnualAgencyProjectReport.SubmissionStatus.DRAFT
        )
        assert annual_agency_report.submitted_at is None
        assert annual_agency_report.submitted_by is None
        assert annual_agency_report.is_unlocked is False

        self.client.force_authenticate(user=apr_agency_submitter_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["message"] == "Report submitted successfully."
        assert (
            response.data["status"]
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        assert response.data["submitted_at"] is not None
        assert response.data["submitted_by"] == apr_agency_submitter_user.username

        annual_agency_report.refresh_from_db()
        assert (
            annual_agency_report.status
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        assert annual_agency_report.submitted_at is not None
        assert annual_agency_report.submitted_by == apr_agency_submitter_user
        assert annual_agency_report.is_unlocked is False

    def test_mail_after_agency_submission(
        self,
        apr_agency_submitter_user,
        annual_agency_report,
        mock_send_agency_submission_notification,
    ):
        config.APR_AGENCY_SUBMISSION_NOTIFICATIONS_ENABLED = True
        config.APR_AGENCY_SUBMISSION_NOTIFICATIONS_EMAILS = (
            apr_agency_submitter_user.email
        )

        self.client.force_authenticate(user=apr_agency_submitter_user)
        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["message"] == "Report submitted successfully."
        assert (
            response.data["status"]
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        assert response.data["submitted_at"] is not None
        assert response.data["submitted_by"] == apr_agency_submitter_user.username

        mock_send_agency_submission_notification.assert_called()
        assert mock_send_agency_submission_notification.call_count == 1

    def test_mail_not_sent_after_agency_submission_if_disabled(
        self,
        apr_agency_submitter_user,
        annual_agency_report,
        mock_send_agency_submission_notification,
    ):
        config.APR_AGENCY_SUBMISSION_NOTIFICATIONS_ENABLED = False
        config.APR_AGENCY_SUBMISSION_NOTIFICATIONS_EMAILS = (
            apr_agency_submitter_user.email
        )

        self.client.force_authenticate(user=apr_agency_submitter_user)
        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert response.data["message"] == "Report submitted successfully."
        assert (
            response.data["status"]
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        assert response.data["submitted_at"] is not None
        assert response.data["submitted_by"] == apr_agency_submitter_user.username

        mock_send_agency_submission_notification.assert_not_called()

    def test_mlfs_can_submit_draft_report(self, mlfs_admin_user, annual_agency_report):
        self.client.force_authenticate(user=mlfs_admin_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert (
            response.data["status"]
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )

    def test_cannot_submit_already_submitted_locked_report(
        self, apr_agency_submitter_user, annual_agency_report
    ):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.submitted_at = timezone.now()
        annual_agency_report.submitted_by = apr_agency_submitter_user
        annual_agency_report.save()

        self.client.force_authenticate(user=apr_agency_submitter_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Cannot transition" in str(response.data)

    def test_submit_unlocked_report_locks_it_again(
        self, apr_agency_submitter_user, annual_agency_report
    ):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.submitted_at = timezone.now()
        annual_agency_report.submitted_by = apr_agency_submitter_user
        annual_agency_report.is_unlocked = True
        annual_agency_report.save()

        self.client.force_authenticate(user=apr_agency_submitter_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")

        assert response.status_code == status.HTTP_200_OK
        assert (
            response.data["status"]
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )

        annual_agency_report.refresh_from_db()
        assert annual_agency_report.is_unlocked is False
        assert (
            annual_agency_report.status
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )

    def test_explicit_status_in_request_body(
        self, apr_agency_submitter_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_submitter_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.post(
            url,
            {"status": AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED},
            format="json",
        )

        assert response.status_code == status.HTTP_200_OK
        assert (
            response.data["status"]
            == AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )

    def test_cannot_submit_other_agency_report(
        self, apr_agency_submitter_user, annual_agency_report
    ):
        other_agency = AgencyFactory()
        other_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_agency_report.progress_report, agency=other_agency
        )

        self.client.force_authenticate(user=apr_agency_submitter_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": other_report.progress_report.year,
                "agency_id": other_report.agency.id,
            },
        )
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_submit_nonexistent_report(self, apr_agency_submitter_user, apr_year):
        self.client.force_authenticate(user=apr_agency_submitter_user)

        url = reverse(
            "apr-status",
            kwargs={
                "year": apr_year,
                "agency_id": 99999,
            },
        )
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_404_NOT_FOUND


@pytest.mark.django_db
class TestAPRMLFSExportView(BaseTest):
    def test_without_login(self, apr_year):
        self.client.force_authenticate(user=None)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_cannot_access(self, apr_agency_viewer_user, apr_year):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_viewer_can_export_all_agencies(
        self, secretariat_viewer_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory(name="A")
        agency2 = AgencyFactory(name="B")

        report1 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        report2 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        project1 = ProjectFactory(
            code="AG1/TEST/2024/001",
            agency=agency1,
            version=3,
            latest_project=None,
        )
        project2 = ProjectFactory(
            code="AG2/TEST/2024/001",
            agency=agency2,
            version=3,
            latest_project=None,
        )

        AnnualProjectReportFactory(report=report1, project=project1)
        AnnualProjectReportFactory(report=report2, project=project2)

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert f"APR_{apr_year}_All_Agencies.xlsx" in response["Content-Disposition"]

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.FIRST_DATA_ROW + 1
        assert data_rows == 2

        first_data_row = APRExportWriter.FIRST_DATA_ROW
        columns = APRExportWriter.build_column_mapping()
        project_col = columns["project_code"]

        project_1_value = worksheet.cell(first_data_row, project_col).value
        assert project_1_value == project1.code
        project_2_value = worksheet.cell(first_data_row + 1, project_col).value
        assert project_2_value == project2.code

    def test_only_submitted_and_locked_reports_exported(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()
        agency3 = AgencyFactory()

        # Submitted and locked - should be included
        report1 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        # Draft - should be excluded
        report2 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.DRAFT,
            is_unlocked=False,
        )

        # Submitted but unlocked - should be excluded
        report3 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency3,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=True,
        )

        project1 = ProjectFactory(agency=agency1, version=3, latest_project=None)
        project2 = ProjectFactory(agency=agency2, version=3, latest_project=None)
        project3 = ProjectFactory(agency=agency3, version=3, latest_project=None)

        AnnualProjectReportFactory(report=report1, project=project1)
        AnnualProjectReportFactory(report=report2, project=project2)
        AnnualProjectReportFactory(report=report3, project=project3)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.FIRST_DATA_ROW + 1
        assert data_rows == 1

        project_col = APRExportWriter.build_column_mapping()["project_code"]
        project_1_value = worksheet.cell(
            APRExportWriter.FIRST_DATA_ROW, project_col
        ).value
        assert project_1_value == project1.code

    def test_filter_by_agency(self, mlfs_admin_user, apr_year, annual_progress_report):
        agency1 = AgencyFactory(name="Agency One")
        agency2 = AgencyFactory(name="Agency Two")

        report1 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        report2 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        project1 = ProjectFactory(agency=agency1, version=3, latest_project=None)
        project2 = ProjectFactory(agency=agency2, version=3, latest_project=None)

        AnnualProjectReportFactory(report=report1, project=project1)
        AnnualProjectReportFactory(report=report2, project=project2)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url, {"agency": f"{agency1.id}"})

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.FIRST_DATA_ROW + 1
        assert data_rows == 1

        first_data_row = APRExportWriter.FIRST_DATA_ROW
        columns = APRExportWriter.build_column_mapping()
        agency_col = columns["agency_name"]
        agency_value = worksheet.cell(first_data_row, agency_col).value
        assert agency_value == "Agency One"

    def test_filter_by_country(
        self, mlfs_admin_user, apr_year, annual_progress_report, country_ro, new_country
    ):
        agency1 = AgencyFactory()
        agency2 = AgencyFactory()

        report1 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        report2 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        project1 = ProjectFactory(
            agency=agency1, country=country_ro, version=3, latest_project=None
        )
        project2 = ProjectFactory(
            agency=agency2, country=new_country, version=3, latest_project=None
        )

        AnnualProjectReportFactory(report=report1, project=project1)
        AnnualProjectReportFactory(report=report2, project=project2)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url, {"country": country_ro.name})

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.FIRST_DATA_ROW + 1
        assert data_rows == 1

        country_column = APRExportWriter.build_column_mapping()["country_name"]
        country = worksheet.cell(APRExportWriter.FIRST_DATA_ROW, country_column).value
        assert country == country_ro.name

    def test_filter_by_project_status(
        self,
        mlfs_admin_user,
        apr_year,
        annual_progress_report,
        project_ongoing_status,
        project_completed_status,
        project_closed_status,
    ):
        agency = AgencyFactory()

        report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        project1 = ProjectFactory(
            code="TEST/ONG/INV/01",
            agency=agency,
            status=project_ongoing_status,
            version=3,
            latest_project=None,
        )
        project2 = ProjectFactory(
            code="TEST/COM/INV/02",
            agency=agency,
            status=project_completed_status,
            version=3,
            latest_project=None,
        )
        project3 = ProjectFactory(
            code="TEST/CLO/INV/03s",
            agency=agency,
            status=project_closed_status,
            version=3,
            latest_project=None,
        )

        AnnualProjectReportFactory(report=report, project=project1)
        AnnualProjectReportFactory(report=report, project=project2)
        AnnualProjectReportFactory(report=report, project=project3)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        # This will actually include both ONG and COM, but not CLO
        response = self.client.get(url, {"status": "ONG"})

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.FIRST_DATA_ROW + 1
        assert data_rows == 2

        code_column = APRExportWriter.build_column_mapping()["project_code"]
        first_code = worksheet.cell(APRExportWriter.FIRST_DATA_ROW, code_column).value
        second_code = worksheet.cell(
            APRExportWriter.FIRST_DATA_ROW + 1, code_column
        ).value
        assert {first_code, second_code} == {project1.code, project2.code}

    def test_empty_export_with_headers(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]

        # Should have headers but no data rows (or one empty template row)
        data_rows = worksheet.max_row - APRExportWriter.FIRST_DATA_ROW + 1
        # This is just an empty "template" data row
        assert data_rows == 1
        first_data_row = APRExportWriter.FIRST_DATA_ROW + 1
        columns = APRExportWriter.build_column_mapping()
        row_values = [
            worksheet.cell(first_data_row, col).value
            for col in range(1, len(columns) + 1)
        ]
        assert all(value is None or value == "" for value in row_values)

    def test_duplicate_project_codes_across_agencies(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency1 = AgencyFactory(name="Agency One")
        agency2 = AgencyFactory(name="Agency Two")

        report1 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency1,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )
        report2 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency2,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        # Same project code for both agencies, both should be in the export
        project1 = ProjectFactory(
            code="DUPLICATE/CODE/001",
            agency=agency1,
            version=3,
            latest_project=None,
        )
        project2 = ProjectFactory(
            code="DUPLICATE/CODE/001",
            agency=agency2,
            version=3,
            latest_project=None,
        )

        AnnualProjectReportFactory(report=report1, project=project1)
        AnnualProjectReportFactory(report=report2, project=project2)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.FIRST_DATA_ROW + 1
        assert data_rows == 2

    def test_agencies_ordered_by_name(
        self, mlfs_admin_user, apr_year, annual_progress_report
    ):
        agency_z = AgencyFactory(name="Zero")
        agency_a = AgencyFactory(name="Alpha")
        agency_o = AgencyFactory(name="Omega")

        for agency in [agency_z, agency_a, agency_o]:
            report = AnnualAgencyProjectReportFactory(
                progress_report=annual_progress_report,
                agency=agency,
                status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
                is_unlocked=False,
            )
            project = ProjectFactory(agency=agency, version=3, latest_project=None)
            AnnualProjectReportFactory(report=report, project=project)

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-export", kwargs={"year": apr_year})
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        agency_col = columns["agency_name"]

        first_row = APRExportWriter.FIRST_DATA_ROW
        assert worksheet.cell(first_row, agency_col).value == "Alpha"
        assert worksheet.cell(first_row + 1, agency_col).value == "Omega"
        assert worksheet.cell(first_row + 2, agency_col).value == "Zero"

    def test_empty_export_nonexistent_year(self, mlfs_admin_user):
        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse("apr-mlfs-export", kwargs={"year": 9999})
        response = self.client.get(url)

        # Should return empty export (no reports for that year)
        assert response.status_code == status.HTTP_200_OK
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        data_rows = worksheet.max_row - APRExportWriter.HEADER_ROW

        # This is just an empty "template" data row
        assert data_rows == 1
        first_data_row = APRExportWriter.FIRST_DATA_ROW + 1
        columns = APRExportWriter.build_column_mapping()
        row_values = [
            worksheet.cell(first_data_row, col).value
            for col in range(1, len(columns) + 1)
        ]
        assert all(value is None or value == "" for value in row_values)

    def test_derived_fields_with_multiple_versions(
        self, mlfs_admin_user, annual_agency_report, multiple_project_versions_for_apr
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
            funds_disbursed=80000.0,
            support_cost_disbursed=8000.0,
        )

        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-export",
            kwargs={"year": annual_agency_report.progress_report.year},
        )

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        assert (
            worksheet.cell(first_data_row, columns["approved_funding"]).value
            == version3.total_fund
        )

        expected_adjustment = latest_version.total_fund - version3.total_fund
        assert (
            worksheet.cell(first_data_row, columns["adjustment"]).value
            == expected_adjustment
        )

        assert (
            worksheet.cell(
                first_data_row, columns["approved_funding_plus_adjustment"]
            ).value
            == latest_version.total_fund
        )

        assert (
            worksheet.cell(first_data_row, columns["support_cost_approved"]).value
            == version3.support_cost_psc
        )

        expected_sc_adjustment = (
            latest_version.support_cost_psc - version3.support_cost_psc
        )
        assert (
            worksheet.cell(first_data_row, columns["support_cost_adjustment"]).value
            == expected_sc_adjustment
        )

    def test_derived_fields_with_archive_version_apr(
        self, mlfs_admin_user, annual_agency_report, multiple_project_versions_for_apr
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        # Create an APR pointing to version3 instead of the latest
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version3,
            funds_disbursed=75000.0,
        )

        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-export",
            kwargs={"year": annual_agency_report.progress_report.year},
        )

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        assert (
            worksheet.cell(first_data_row, columns["approved_funding"]).value
            == version3.total_fund
        )
        expected_adjustment = latest_version.total_fund - version3.total_fund
        assert (
            worksheet.cell(first_data_row, columns["adjustment"]).value
            == expected_adjustment
        )

    def test_derived_fields_with_no_version_3(
        self,
        mlfs_admin_user,
        annual_agency_report,
        initial_project_version_2_for_apr,
    ):
        # Derived fields should return None when there's no version 3.
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=initial_project_version_2_for_apr,
        )

        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-export",
            kwargs={"year": annual_agency_report.progress_report.year},
        )

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        assert worksheet.cell(first_data_row, columns["approved_funding"]).value is None
        assert (
            worksheet.cell(first_data_row, columns["support_cost_approved"]).value
            is None
        )

    def test_derived_fields_with_no_latest_version_in_year(
        self,
        mlfs_admin_user,
        annual_agency_report,
        late_post_excom_versions_for_apr,
    ):
        version3 = late_post_excom_versions_for_apr[0]
        later_version = late_post_excom_versions_for_apr[1]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=later_version,
        )

        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-export",
            kwargs={"year": annual_agency_report.progress_report.year},
        )

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        # adjustment should be None (no version in the current year)
        assert worksheet.cell(first_data_row, columns["adjustment"]).value is None

        # approved_funding_plus_adjustment should fall back to version3
        assert (
            worksheet.cell(
                first_data_row, columns["approved_funding_plus_adjustment"]
            ).value
            == version3.total_fund
        )

    def test_phaseout_fields_from_version3(
        self, mlfs_admin_user, annual_agency_report, multiple_project_versions_for_apr
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
        )

        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-export",
            kwargs={"year": annual_agency_report.progress_report.year},
        )

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        # All phase-out proposal fields should come from version 3
        assert (
            worksheet.cell(
                first_data_row, columns["consumption_phased_out_odp_proposal"]
            ).value
            == version3.consumption_phase_out_odp
        )

        assert (
            worksheet.cell(
                first_data_row, columns["consumption_phased_out_co2_proposal"]
            ).value
            == version3.consumption_phase_out_co2
        )

        assert (
            worksheet.cell(
                first_data_row, columns["production_phased_out_odp_proposal"]
            ).value
            == version3.production_phase_out_odp
        )

        assert (
            worksheet.cell(
                first_data_row, columns["production_phased_out_co2_proposal"]
            ).value
            == version3.production_phase_out_co2
        )

    def test_date_fields_from_different_versions(
        self, mlfs_admin_user, annual_agency_report, multiple_project_versions_for_apr
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
        )

        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-export",
            kwargs={"year": annual_agency_report.progress_report.year},
        )

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        date_approved_cell = worksheet.cell(
            first_data_row, columns["date_approved"]
        ).value
        assert date_approved_cell.date() == version3.date_approved

        date_completion_proposal_cell = worksheet.cell(
            first_data_row, columns["date_completion_proposal"]
        ).value
        assert date_completion_proposal_cell.date() == version3.date_completion

        date_completion_agreement_cell = worksheet.cell(
            first_data_row,
            columns["date_of_completion_per_agreement_or_decisions"],
        ).value
        assert date_completion_agreement_cell.date() == latest_version.date_completion


@pytest.mark.django_db
class TestAPRFilesDownloadAllView(BaseTest):
    def test_without_login(self, apr_year, agency):
        self.client.force_authenticate(user=None)
        url = reverse(
            "apr-files-download-all",
            kwargs={"year": apr_year, "agency_id": agency.id},
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_cannot_access(
        self, apr_agency_viewer_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-files-download-all",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_mlfs_viewer_can_download_all_files(
        self, secretariat_viewer_user, annual_agency_report
    ):
        file1 = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_name="report1.pdf",
            file_type=AnnualProjectReportFile.FileType.ANNUAL_PROGRESS_FINANCIAL_REPORT,
        )
        file2 = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_name="support_doc.docx",
            file_type=AnnualProjectReportFile.FileType.OTHER_SUPPORTING_DOCUMENT,
        )

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-files-download-all",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == "application/zip"
        assert "attachment" in response["Content-Disposition"]
        assert (
            f"APR_{annual_agency_report.progress_report.year}"
            in response["Content-Disposition"]
        )
        assert annual_agency_report.agency.name[:10] in response["Content-Disposition"]

        zip_content = BytesIO(response.content)
        with ZipFile(zip_content, "r") as zipf:
            file_list = zipf.namelist()
            assert len(file_list) == 2
            assert file1.file_name in file_list
            assert file2.file_name in file_list

    def test_mlfs_admin_can_download_all_files(
        self, mlfs_admin_user, annual_agency_report
    ):
        AnnualProjectReportFileFactory(
            report=annual_agency_report, file_name="test_file.pdf"
        )

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-files-download-all",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == "application/zip"

    def test_download_all_with_no_files(
        self, secretariat_viewer_user, annual_agency_report
    ):
        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-files-download-all",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == "application/zip"

        zip_content = BytesIO(response.content)
        # This should be an empty zip, no files were created for the agency report
        with ZipFile(zip_content, "r") as zipf:
            assert len(zipf.namelist()) == 0

    def test_download_all_nonexistent_report(self, secretariat_viewer_user):
        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-files-download-all",
            kwargs={"year": 9999, "agency_id": 9999},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_download_all_with_multiple_file_types(
        self, secretariat_viewer_user, annual_agency_report
    ):
        financial_file = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_name="financial_2024.pdf",
            file_type=AnnualProjectReportFile.FileType.ANNUAL_PROGRESS_FINANCIAL_REPORT,
        )
        support_file1 = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_name="support_doc_1.docx",
            file_type=AnnualProjectReportFile.FileType.OTHER_SUPPORTING_DOCUMENT,
        )
        support_file2 = AnnualProjectReportFileFactory(
            report=annual_agency_report,
            file_name="support_doc_2.xlsx",
            file_type=AnnualProjectReportFile.FileType.OTHER_SUPPORTING_DOCUMENT,
        )

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-files-download-all",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        zip_content = BytesIO(response.content)
        with ZipFile(zip_content, "r") as zipf:
            file_list = zipf.namelist()
            assert len(file_list) == 3
            assert financial_file.file_name in file_list
            assert support_file1.file_name in file_list
            assert support_file2.file_name in file_list

    def test_zip_filename_format(self, secretariat_viewer_user, annual_agency_report):
        AnnualProjectReportFileFactory(report=annual_agency_report)

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-files-download-all",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        content_disposition = response["Content-Disposition"]

        assert f"APR_{annual_agency_report.progress_report.year}" in content_disposition
        assert "_Files.zip" in content_disposition
        assert "attachment" in content_disposition

    def test_handles_special_characters_in_agency_name(
        self, secretariat_viewer_user, apr_year, annual_progress_report
    ):
        agency = AgencyFactory(name="Test Agency / Special & Co.")
        report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency,
        )
        AnnualProjectReportFileFactory(report=report)

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse(
            "apr-files-download-all",
            kwargs={
                "year": apr_year,
                "agency_id": agency.id,
            },
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        content_disposition = response["Content-Disposition"]

        assert "Test_Agency_Special_Co" in content_disposition
        assert "/" not in content_disposition
        assert "&" not in content_disposition


@pytest.mark.django_db
class TestAPRDerivedFieldsAPI(BaseTest):
    """
    Test class to check that *all* derived fields for AnnualProjectReport
    are correctly calculated and returned via API based on existing projects.
    """

    def test_without_login(self):
        """No need for this here, it's tested elsewhere"""

    def test_project_identification_derived_fields(
        self, apr_agency_viewer_user, annual_agency_report, country_ro, country_europe
    ):
        country_ro.parent = country_europe
        country_ro.save()

        project = ProjectFactory(
            code="ROM/FOA/80/TAS/123",
            legacy_code="ROM/80/TAS/123",
            title="Test Project for Derived Fields",
            agency=annual_agency_report.agency,
            country=country_ro,
        )

        project.meta_project = MetaProjectFactory(code="ROM/FOA/80/TAS/123")
        project.metacode = "ROM/FOA/80/TAS/123/1"
        project.cluster = ProjectClusterFactory.create(
            name="KPP1", code="KPP1", sort_order=1
        )
        project.save()

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == project.code
            ),
            None,
        )
        assert project_data is not None

        assert project_data["meta_code"] == "ROM/FOA/80/TAS/123/1"
        assert project_data["project_code"] == "ROM/FOA/80/TAS/123"
        assert project_data["legacy_code"] == "ROM/80/TAS/123"
        assert project_data["agency_name"] == project.agency.name
        assert project_data["cluster_name"] == project.cluster.name
        assert project_data["region_name"] == project.country.parent.name
        assert project_data["country_name"] == project.country.name
        assert project_data["type_code"] == project.project_type.code
        assert project_data["sector_code"] == project.sector.code
        assert project_data["project_title"] == "Test Project for Derived Fields"

    def test_date_derived_fields(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        multiple_project_versions_for_apr,
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == latest_version.code
            ),
            None,
        )
        assert project_data is not None

        # date_approved should come from version 3
        assert project_data["date_approved"] == version3.date_approved.isoformat()

        # date_completion_proposal should come from version 3
        assert (
            project_data["date_completion_proposal"]
            == version3.date_completion.isoformat()
        )

        # date_of_completion_per_agreement_or_decisions comes from latest version >= 3
        assert (
            project_data["date_of_completion_per_agreement_or_decisions"]
            == latest_version.date_completion.isoformat()
        )

    def test_phaseout_proposal_derived_fields(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        multiple_project_versions_for_apr,
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == version3.code
            ),
            None,
        )
        assert project_data is not None

        # All phaseout proposal fields should come from version 3
        assert (
            project_data["consumption_phased_out_odp_proposal"]
            == version3.consumption_phase_out_odp
        )
        assert (
            project_data["consumption_phased_out_co2_proposal"]
            == version3.consumption_phase_out_co2
        )
        assert (
            project_data["production_phased_out_odp_proposal"]
            == version3.production_phase_out_odp
        )
        assert (
            project_data["production_phased_out_co2_proposal"]
            == version3.production_phase_out_co2
        )

    def test_financial_approved_and_adjustment_fields(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        multiple_project_versions_for_apr,
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
            funds_disbursed=80000.0,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == latest_version.code
            ),
            None,
        )
        assert project_data is not None

        assert project_data["approved_funding"] == version3.total_fund

        expected_adjustment = latest_version.total_fund - version3.total_fund
        assert project_data["adjustment"] == expected_adjustment

        assert (
            project_data["approved_funding_plus_adjustment"]
            == latest_version.total_fund
        )

    def test_financial_calculated_fields(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        multiple_project_versions_for_apr,
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        funds_disbursed = 80000.0
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
            funds_disbursed=funds_disbursed,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == latest_version.code
            ),
            None,
        )
        assert project_data is not None

        expected_balance = version3.total_fund - funds_disbursed
        assert project_data["balance"] == expected_balance

        expected_percentage = funds_disbursed / latest_version.total_fund
        assert project_data["per_cent_funds_disbursed"] == expected_percentage

    def test_support_cost_derived_fields(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        multiple_project_versions_for_apr,
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
            support_cost_disbursed=8000.0,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == latest_version.code
            ),
            None,
        )
        assert project_data is not None

        # support_cost_approved comes from version 3
        assert project_data["support_cost_approved"] == version3.support_cost_psc

        # support_cost_adjustment is latest - approved
        expected_adjustment = (
            latest_version.support_cost_psc - version3.support_cost_psc
        )
        assert project_data["support_cost_adjustment"] == expected_adjustment

        # support_cost_approved_plus_adjustment
        expected_total = version3.support_cost_psc + expected_adjustment
        assert project_data["support_cost_approved_plus_adjustment"] == expected_total

    def test_support_cost_balance(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        multiple_project_versions_for_apr,
    ):
        latest_version = multiple_project_versions_for_apr[2]

        support_cost_disbursed = 8000.0
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
            support_cost_disbursed=support_cost_disbursed,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == latest_version.code
            ),
            None,
        )
        assert project_data is not None

        expected_balance = latest_version.support_cost_psc - support_cost_disbursed
        assert project_data["support_cost_balance"] == expected_balance

    def test_implementation_delays_field(
        self, apr_agency_viewer_user, annual_agency_report, annual_project_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == annual_project_report.project.code
            ),
            None,
        )
        assert project_data is not None

        # This field is currently hardcoded to return an empty string
        assert project_data["implementation_delays_status_report_decisions"] == ""

    def test_derived_fields_with_no_version_3(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        initial_project_version_2_for_apr,
    ):
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=initial_project_version_2_for_apr,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == initial_project_version_2_for_apr.code
            ),
            None,
        )
        assert project_data is not None

        # Fields that depend on version 3 should be None
        assert project_data["consumption_phased_out_odp_proposal"] is None
        assert project_data["consumption_phased_out_co2_proposal"] is None
        assert project_data["production_phased_out_odp_proposal"] is None
        assert project_data["production_phased_out_co2_proposal"] is None
        assert project_data["approved_funding"] is None
        assert project_data["support_cost_approved"] is None

    def test_derived_fields_with_no_latest_version(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        late_post_excom_versions_for_apr,
    ):
        version4 = late_post_excom_versions_for_apr[1]
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=version4,
            funds_disbursed=80000.0,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == version4.code
            ),
            None,
        )
        assert project_data is not None

        # Fields that depend on latest_version_for_year should be == None
        assert project_data["adjustment"] is None
        assert project_data["support_cost_adjustment"] is None

        # But approved_funding_plus_adjustment should use version 3 as fallback
        version3 = late_post_excom_versions_for_apr[0]
        assert project_data["approved_funding_plus_adjustment"] == version3.total_fund

    def test_null_handling_in_calculations(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        multiple_project_versions_for_apr,
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
            funds_disbursed=None,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        project_data = next(
            (
                p
                for p in response.data["project_reports"]
                if p["project_code"] == latest_version.code
            ),
            None,
        )
        assert project_data is not None

        # per_cent_funds_disbursed should be None when funds_disbursed is None
        assert project_data["per_cent_funds_disbursed"] is None

        # balance should treat None as 0
        assert project_data["balance"] == version3.total_fund

    def test_all_derived_fields_via_mlfs_export(
        self, mlfs_admin_user, annual_agency_report, multiple_project_versions_for_apr
    ):
        version3 = multiple_project_versions_for_apr[0]
        latest_version = multiple_project_versions_for_apr[2]

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=latest_version,
            funds_disbursed=80000.0,
            support_cost_disbursed=8000.0,
        )
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=mlfs_admin_user)
        url = reverse(
            "apr-mlfs-export",
            kwargs={"year": annual_agency_report.progress_report.year},
        )
        url += f"?agency={annual_agency_report.agency.id}"

        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook[APRExportWriter.SHEET_NAME]
        columns = APRExportWriter.build_column_mapping()
        first_data_row = APRExportWriter.FIRST_DATA_ROW

        assert (
            worksheet.cell(first_data_row, columns["approved_funding"]).value
            == version3.total_fund
        )
        assert worksheet.cell(first_data_row, columns["adjustment"]).value == (
            latest_version.total_fund - version3.total_fund
        )
        assert worksheet.cell(
            first_data_row, columns["approved_funding_plus_adjustment"]
        ).value == (latest_version.total_fund)


@pytest.mark.django_db
class TestAPRPermissions(BaseTest):
    def test_without_login(self, **kwargs):
        pass

    def test_apr_agency_viewer_can_view(
        self, apr_agency_viewer_user, annual_progress_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-workspace",
            kwargs={"year": annual_progress_report.year},
        )
        response = self.client.get(url)
        assert response.status_code == status.HTTP_200_OK

    def test_apr_agency_viewer_cannot_edit(
        self, apr_agency_viewer_user, annual_agency_report, annual_project_report
    ):
        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency_id,
            },
        )
        response = self.client.post(
            url,
            {"project_reports": []},
            format="json",
        )
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_apr_agency_inputter_can_edit(
        self, apr_agency_inputter_user, annual_agency_report, annual_project_report
    ):
        self.client.force_authenticate(user=apr_agency_inputter_user)
        url = reverse(
            "apr-update",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency_id,
            },
        )
        response = self.client.post(
            url,
            {
                "project_reports": [
                    {
                        "project_code": annual_project_report.project.code,
                        "funds_disbursed": 50000.0,
                    }
                ]
            },
            format="json",
        )
        assert response.status_code == status.HTTP_200_OK

    def test_apr_agency_inputter_cannot_submit(
        self, apr_agency_inputter_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_inputter_user)
        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency_id,
            },
        )
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_apr_agency_submitter_can_submit(
        self, apr_agency_submitter_user, annual_agency_report
    ):
        self.client.force_authenticate(user=apr_agency_submitter_user)
        url = reverse(
            "apr-status",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency_id,
            },
        )
        response = self.client.post(url, {}, format="json")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["status"] == "submitted"

    def test_apr_mlfs_full_access_can_unlock(
        self, apr_mlfs_full_access_user, annual_agency_report
    ):
        annual_agency_report.status = (
            AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED
        )
        annual_agency_report.save()

        self.client.force_authenticate(user=apr_mlfs_full_access_user)
        url = reverse(
            "apr-toggle-lock",
            kwargs={
                "year": annual_agency_report.progress_report.year,
                "agency_id": annual_agency_report.agency_id,
            },
        )
        response = self.client.post(url, {"is_unlocked": True}, format="json")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["is_unlocked"] is True

    def test_apr_mlfs_full_access_can_endorse(
        self, apr_mlfs_full_access_user, annual_progress_report, meeting_apr_same_year
    ):
        agency = AgencyFactory()
        AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
            agency=agency,
            status=AnnualAgencyProjectReport.SubmissionStatus.SUBMITTED,
            is_unlocked=False,
        )

        self.client.force_authenticate(user=apr_mlfs_full_access_user)
        url = reverse("apr-endorse", kwargs={"year": annual_progress_report.year})
        response = self.client.post(
            url,
            {
                "meeting_endorsed": meeting_apr_same_year.id,
                "date_endorsed": "2024-12-01",
                "remarks_endorsed": "Test endorsement",
            },
            format="json",
        )
        assert response.status_code == status.HTTP_200_OK

    def test_separate_apr_and_projects_permissions(
        self, agency, viewer_user, apr_agency_viewer_user
    ):
        # User with only Projects (not APR) permissions
        assert not viewer_user.has_perm("core.has_apr_view_access")

        # And the reverse (only APR permissions)
        assert apr_agency_viewer_user.has_perm("core.has_apr_view_access")
        assert not apr_agency_viewer_user.has_perm("core.has_project_v2_edit_access")
