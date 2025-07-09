import io
from http import HTTPStatus

from django.urls import reverse

from core.api.tests.base import BaseTest

from core.models.project import Project

import openpyxl
import pytest


pytestmark = pytest.mark.django_db


def validate_single_project_export(project: Project, response):
    assert response.filename == f"Project {project.id}.xlsx", response.filename

    wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
    assert set(
        [
            "Identifiers",
            "Identifiers - BP Activity",
            "Cross-cutting",
            "Specific information - Overview",
            "Specific information - Substance details",
            "Impact",
        ]
    ).intersection(wb.sheetnames)
    sheet = wb["Identifiers"]
    assert sheet["A1"].value == "Country", sheet["A1"].value
    assert sheet["A2"].value == project.country.name, sheet["A1"].value


def validate_projects_export(project: Project, response):
    assert response.filename == "Projects.xlsx", response.filename

    wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
    assert set(
        [
            "Projects",
            "Codes",
            "Legacy codes",
            "Metaproject codes",
            "Clusters",
            "Metaproject categories",
            "Project types",
            "Legacy project types",
            "Sectors",
            "Legacy sectors",
            "Subsectors",
            "Legacy subsectors",
            "Substance types",
            "Substances",
            "Status",
            "Serial numbers",
            "Legacy serial numbers",
            "Countries",
            "Titles",
        ]
    ).intersection(wb.sheetnames)
    sheet = wb["Projects"]
    assert sheet["A1"].value == "Code", sheet["A1"].value
    assert sheet["A2"].value == project.code, sheet["A2"].value


class TestProjectV2ExportXLSX(BaseTest):
    url = reverse("project-v2-export")

    def test_export_project_anon(self, project):
        self.client.force_authenticate(user=None)
        response = self.client.get(self.url, {"project_id": project.id})
        assert response.status_code == HTTPStatus.FORBIDDEN, response.data

    def test_export_project_agency_submitter(self, project, agency_inputter_user):
        self.client.force_authenticate(user=agency_inputter_user)
        response = self.client.get(self.url, {"project_id": project.id})
        assert response.status_code == HTTPStatus.OK, response.data
        validate_single_project_export(project, response)

    def test_export_project_secretariat(self, project, secretariat_viewer_user):
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url, {"project_id": project.id})
        assert response.status_code == HTTPStatus.OK, response.data
        validate_single_project_export(project, response)

    def test_export_projects_anon(self, project):
        self.client.force_authenticate(user=None)
        response = self.client.get(self.url, {"project_id": project.id})
        assert response.status_code == HTTPStatus.FORBIDDEN, response.data

    def test_export_projects_agency_submitter(self, project, agency_inputter_user):
        self.client.force_authenticate(user=agency_inputter_user)
        response = self.client.get(self.url)
        assert response.status_code == HTTPStatus.OK, response.data
        validate_projects_export(project, response)

    def test_export_projects_secretariat(self, project, secretariat_viewer_user):
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url)
        assert response.status_code == HTTPStatus.OK, response.data
        validate_projects_export(project, response)
