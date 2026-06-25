# pylint: disable=W0613,R0904
import io
from decimal import Decimal
from http import HTTPStatus

import openpyxl
import pytest
from constance import config
from django.urls import reverse
from openpyxl.utils import get_column_letter

from core.api.tests.base import BaseTest
from core.api.tests.factories import (
    CountryFactory,
    MetaProjectFactory,
    ProjectFactory,
    ProjectTypeFactory,
)
from core.api.views.mya_export import HEADERS as MYA_HEADERS
from core.models.country import Country
from core.models.project import MetaProject, Project, ProjectOdsOdp
from core.models.project_metadata import ProjectField

pytestmark = pytest.mark.django_db


def get_sheet_headers(sheet):
    return {
        cell.value: get_column_letter(cell.column)
        for cell in sheet[1]
        if cell.value is not None
    }


def get_project_row(sheet, project_id):
    return next(
        (
            idx
            for idx in range(2, sheet.max_row + 1)
            if sheet[f"A{idx}"].value == project_id
        ),
        None,
    )


def load_projects_sheet(response):
    wb = openpyxl.load_workbook(io.BytesIO(response.getvalue()))
    return wb["Projects"]


def load_workbook(response):
    return openpyxl.load_workbook(io.BytesIO(response.getvalue()))


class TestProjectsDashboardExport(BaseTest):
    url = reverse("project-v2-dashboards-all")

    # Auth
    def test_unauthenticated_returns_403(self):
        self.client.force_authenticate(user=None)
        response = self.client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    # Basic structure
    def test_returns_xlsx_with_expected_sheets(
        self, secretariat_viewer_user, approved_project
    ):
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url, {"mock_data": "false"})
        assert response.status_code == HTTPStatus.OK
        wb = load_workbook(response)
        assert set(wb.sheetnames) == {
            "Projects",
            "Substances",
            "Funds",
            "Global fields",
            "MetaProjects",
        }

    def test_global_fields_sheet_lists_constant_values(
        self, secretariat_viewer_user, approved_project
    ):
        config.TOTAL_SAVINGS_TO_SOCIETY_IN_US_DOLLAR = Decimal("123.45")

        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url, {"mock_data": "false"})
        sheet = load_workbook(response)["Global fields"]

        assert [c.value for c in sheet[1]] == ["Field", "Value"]

        rows = {
            sheet[f"A{r}"].value: sheet[f"B{r}"].value
            for r in range(2, sheet.max_row + 1)
        }
        # Decimal value is written as a number, and the empty GLOBAL_FIELD_*
        # placeholders are excluded.
        assert rows["Total savings to society in US $"] == 123.45
        assert "Global field 1" not in rows

    def test_projects_sheet_has_expected_headers(
        self, secretariat_viewer_user, approved_project
    ):
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url, {"mock_data": "false"})
        headers = get_sheet_headers(load_projects_sheet(response))
        for expected in (
            "country_iso",
            "Region",
            "Sub-Region",
            "Type Simple",
            "Is latest version",
            "Actual funds",
            "Actual PSC",
        ):
            assert expected in headers, f"Missing header: {expected}"

    def test_actual_funds_and_psc_appear_exactly_once(
        self, secretariat_viewer_user, approved_project
    ):
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url, {"mock_data": "false"})
        header_values = [cell.value for cell in load_projects_sheet(response)[1]]
        assert header_values.count("Actual funds") == 1
        assert header_values.count("Actual PSC") == 1

    # latest_only
    def test_latest_only_true_excludes_older_versions(
        self, secretariat_viewer_user, approved_project
    ):
        older = ProjectFactory.create(
            latest_project_id=approved_project.id,
            meta_project=approved_project.meta_project,
            status=approved_project.status,
            submission_status=approved_project.submission_status,
        )
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url, {"latest_only": "true", "mock_data": "false"}
        )
        sheet = load_projects_sheet(response)
        ids = [sheet[f"A{r}"].value for r in range(2, sheet.max_row + 1)]
        assert approved_project.id in ids
        assert older.id not in ids

    def test_latest_only_false_includes_all_versions(
        self, secretariat_viewer_user, approved_project
    ):
        older = ProjectFactory.create(
            latest_project_id=approved_project.id,
            meta_project=approved_project.meta_project,
            status=approved_project.status,
            submission_status=approved_project.submission_status,
        )
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url, {"latest_only": "false", "mock_data": "false"}
        )
        sheet = load_projects_sheet(response)
        ids = [sheet[f"A{r}"].value for r in range(2, sheet.max_row + 1)]
        assert approved_project.id in ids
        assert older.id in ids

    # exclude_production
    def test_exclude_production_true_omits_production_projects(
        self, secretariat_viewer_user, approved_project
    ):
        prod = ProjectFactory.create(
            production=True,
            status=approved_project.status,
            submission_status=approved_project.submission_status,
        )
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url, {"exclude_production": "true", "mock_data": "false"}
        )
        sheet = load_projects_sheet(response)
        ids = [sheet[f"A{r}"].value for r in range(2, sheet.max_row + 1)]
        assert prod.id not in ids

    def test_exclude_production_false_includes_production_projects(
        self, secretariat_viewer_user, approved_project
    ):
        prod = ProjectFactory.create(
            production=True,
            status=approved_project.status,
            submission_status=approved_project.submission_status,
        )
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url, {"exclude_production": "false", "mock_data": "false"}
        )
        sheet = load_projects_sheet(response)
        ids = [sheet[f"A{r}"].value for r in range(2, sheet.max_row + 1)]
        assert prod.id in ids

    # fill_substance_type
    def test_fill_substance_type_false_leaves_null(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = None
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "fill_substance_type": "false",
                "latest_only": "false",
                "mock_data": "false",
            },
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert sheet[f"{headers['substance_type']}{row}"].value is None

    def test_fill_substance_type_true_fills_with_hfc(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = None
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "fill_substance_type": "true",
                "latest_only": "false",
                "mock_data": "false",
            },
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert sheet[f"{headers['substance_type']}{row}"].value == "HFC"

    # merge_methyl_bromide
    def test_merge_methyl_bromide_false_keeps_methyl_bromide(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = "Methyl Bromide"
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "merge_methyl_bromide": "false",
                "latest_only": "false",
                "mock_data": "false",
            },
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert sheet[f"{headers['substance_type']}{row}"].value == "Methyl Bromide"

    def test_merge_methyl_bromide_true_reclassifies_as_cfc(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = "Methyl Bromide"
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "merge_methyl_bromide": "true",
                "latest_only": "false",
                "mock_data": "false",
            },
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert sheet[f"{headers['substance_type']}{row}"].value == "CFC"

    def test_fill_substance_type_takes_precedence_over_merge_methyl_bromide_when_null(
        self, secretariat_viewer_user, approved_project
    ):
        # Both flags true, substance_type is None: fill_substance_type fires (sets "HFC"),
        # merge_methyl_bromide is elif so it does not fire.
        approved_project.substance_type = None
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "fill_substance_type": "true",
                "merge_methyl_bromide": "true",
                "latest_only": "false",
                "mock_data": "false",
            },
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert sheet[f"{headers['substance_type']}{row}"].value == "HFC"

    # Mock Data
    def test_mock_data_false_leaves_impact_metrics_empty(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = "HCFC"
        approved_project.total_fund = 1_000_000
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {"mock_data": "false", "latest_only": "false"},
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"

        def cell(field):
            return sheet[f"{headers[field]}{row}"].value

        assert cell("total_number_of_technicians_trained_actual") is None
        # New fields (float + bool) are likewise empty when mock_data is off.
        assert cell("energy_savings_actual") is None
        assert cell("quantity_controlled_substances_destroyed_mt_actual") is None

    def test_mock_data_true_fills_impact_metrics(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = "HCFC"
        approved_project.total_fund = 1_000_000
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "mock_data": "true",
                "mock_types": "hcfc",
                "latest_only": "false",
            },
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"

        def cell(field):
            return sheet[f"{headers[field]}{row}"].value

        assert cell("total_number_of_technicians_trained_actual") > 0
        # NOU personnel (Poisson) and controlled-substance destruction
        # (funding-scaled float) are non-investment metrics too.
        assert cell("total_number_of_nou_personnel_supported_actual") >= 0
        assert cell("number_of_female_nou_personnel_supported_actual") >= 0
        assert cell("quantity_controlled_substances_destroyed_mt_actual") > 0
        assert cell("quantity_controlled_substances_destroyed_co2_eq_t_actual") > 0
        # Energy-efficiency / MEPS booleans render as Yes/No.
        for boolean_field in (
            "ee_demonstration_project_actual",
            "meps_developed_domestic_refrigeration_actual",
            "meps_developed_commercial_refrigeration_actual",
            "meps_developed_residential_ac_actual",
            "meps_developed_commercial_ac_actual",
        ):
            assert cell(boolean_field) in ("Yes", "No")

    def test_mock_types_cfc_catches_methyl_bromide(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = "Methyl Bromide"
        approved_project.total_fund = 1_000_000
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "mock_data": "true",
                "mock_types": "cfc",
                "mock_seed": "42",
                "latest_only": "false",
            },
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        value = sheet[
            f"{headers['total_number_of_technicians_trained_actual']}{row}"
        ].value
        assert value is not None and value > 0

    def test_mock_types_excludes_unmatched_substance_types(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = "CFC"
        approved_project.total_fund = 1_000_000
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "mock_data": "true",
                "mock_types": "hcfc",
                "latest_only": "false",
            },
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert (
            sheet[f"{headers['total_number_of_technicians_trained_actual']}{row}"].value
            is None
        )

    def _mock_cells(self, response, project_id, field_names):
        """Read several columns for one project from a single response.

        The response body is a stream that can only be consumed once, so load
        the workbook here and resolve all requested fields from it.
        """
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, project_id)
        assert row is not None, "Project not found in export"
        return {f: sheet[f"{headers[f]}{row}"].value for f in field_names}

    def test_mock_data_fills_energy_savings_float(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.project_type = ProjectTypeFactory.create(code="INV")
        approved_project.substance_type = "HCFC"
        approved_project.total_fund = 1_000_000
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {"mock_data": "true", "mock_types": "hcfc", "latest_only": "false"},
        )
        cells = self._mock_cells(
            response, approved_project.id, ["energy_savings_actual"]
        )
        value = cells["energy_savings_actual"]
        assert value is not None and value > 0
        # Funding-scaled magnitude, not a tiny integer count.
        assert value > 1000

    def test_mock_data_fills_hfc23_production_fields(
        self, secretariat_viewer_user, approved_project
    ):
        prod = ProjectFactory.create(
            production=True,
            substance_type="HCFC",
            total_fund=1_000_000,
            status=approved_project.status,
            submission_status=approved_project.submission_status,
        )
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "mock_data": "true",
                "mock_types": "hcfc",
                "exclude_production": "false",
                "latest_only": "false",
            },
        )
        quantity_fields = [
            "quantity_hfc_23_by_product_generated_actual",
            "quantity_hfc_23_by_product_destroyed_actual",
            "quantity_hfc_23_by_product_emitted_actual",
        ]
        cells = self._mock_cells(
            response,
            prod.id,
            quantity_fields + ["hfc_23_by_product_generation_rate_actual"],
        )
        for field in quantity_fields:
            assert cells[field] is not None and cells[field] > 0
        assert 1.5 <= cells["hfc_23_by_product_generation_rate_actual"] <= 4.0

    def test_mock_data_hfc23_rate_does_not_scale_with_funding(
        self, secretariat_viewer_user, approved_project
    ):
        prod = ProjectFactory.create(
            production=True,
            substance_type="HCFC",
            total_fund=100_000_000,
            status=approved_project.status,
            submission_status=approved_project.submission_status,
        )
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url,
            {
                "mock_data": "true",
                "mock_types": "hcfc",
                "exclude_production": "false",
                "latest_only": "false",
            },
        )
        cells = self._mock_cells(
            response, prod.id, ["hfc_23_by_product_generation_rate_actual"]
        )
        # Bounded regardless of (large) funding.
        assert 1.5 <= cells["hfc_23_by_product_generation_rate_actual"] <= 4.0

    def test_mock_data_reproducible_with_seed(
        self, secretariat_viewer_user, approved_project
    ):
        approved_project.substance_type = "HCFC"
        approved_project.total_fund = 1_000_000
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        params = {
            "mock_data": "true",
            "mock_types": "hcfc",
            "latest_only": "false",
            "mock_seed": "7",
        }
        field = "quantity_controlled_substances_destroyed_mt_actual"
        first = self._mock_cells(
            self.client.get(self.url, params), approved_project.id, [field]
        )[field]
        same = self._mock_cells(
            self.client.get(self.url, params), approved_project.id, [field]
        )[field]
        different = self._mock_cells(
            self.client.get(self.url, {**params, "mock_seed": "8"}),
            approved_project.id,
            [field],
        )[field]
        assert first == same
        assert first != different

    # Geo / type columns
    def test_geo_columns_populated(self, secretariat_viewer_user, approved_project):
        region = CountryFactory.create(
            name="Test Region", location_type=Country.LocationType.REGION, iso3="RGN"
        )
        subregion = CountryFactory.create(
            name="Test Subregion",
            location_type=Country.LocationType.SUBREGION,
            iso3="SRG",
            parent=region,
        )
        country = CountryFactory.create(
            name="Test Country",
            location_type=Country.LocationType.COUNTRY,
            iso3="TCY",
            parent=subregion,
        )
        approved_project.country = country
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url, {"latest_only": "false", "mock_data": "false"}
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert sheet[f"{headers['country_iso']}{row}"].value == "TCY"
        assert sheet[f"{headers['Region']}{row}"].value == "Test Region"
        assert sheet[f"{headers['Sub-Region']}{row}"].value == "Test Subregion"

    def test_type_simple_investment(self, secretariat_viewer_user, approved_project):
        inv_type = ProjectTypeFactory.create(code="INV")
        approved_project.project_type = inv_type
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url, {"latest_only": "false", "mock_data": "false"}
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert sheet[f"{headers['Type Simple']}{row}"].value == "Investment"

    def test_type_simple_non_investment(
        self, secretariat_viewer_user, approved_project
    ):
        non_inv_type = ProjectTypeFactory.create(code="TAS")
        approved_project.project_type = non_inv_type
        approved_project.save()
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url, {"latest_only": "false", "mock_data": "false"}
        )
        sheet = load_projects_sheet(response)
        headers = get_sheet_headers(sheet)
        row = get_project_row(sheet, approved_project.id)
        assert row is not None, "Project not found in export"
        assert sheet[f"{headers['Type Simple']}{row}"].value == "Non-Investment"

    # Substances sheet
    def test_substances_sheet_includes_ods_odp_rows(
        self, secretariat_viewer_user, approved_project, substance
    ):
        ProjectField.objects.create(
            import_name="ods_display_name",
            label="Substance",
            read_field_name="ods_display_name",
            write_field_name="ods_display_name",
            table="ods_odp",
            data_type="text",
            section="Substance Details",
        )
        ProjectOdsOdp.objects.create(
            project=approved_project,
            ods_substance=substance,
            odp=1,
            co2_mt=1,
            sort_order=1,
        )
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(
            self.url, {"latest_only": "false", "mock_data": "false"}
        )
        sheet = load_workbook(response)["Substances"]
        assert sheet.max_row > 1

    # MetaProjects sheet — reuses MyaExport, so these only cover the
    # dashboard-specific bits: header parity, TEMP exclusion, fallback flow.
    def test_metaprojects_sheet_headers_match_mya_export(
        self, secretariat_viewer_user, approved_project
    ):
        # Fails loudly if MyaExport restructures HEADERS.
        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url, {"mock_data": "false"})
        sheet = load_workbook(response)["MetaProjects"]
        header_row = [c.value for c in sheet[1]]
        assert header_row == [h["headerName"] for h in MYA_HEADERS]

    def test_metaprojects_sheet_includes_non_draft_excludes_temp(
        self, secretariat_viewer_user, project_approved_status
    ):
        # Permanent umbrella_code should be included; project_cost is
        # hand-entered so the stored value must appear.
        included_mp = MetaProjectFactory.create(
            type=MetaProject.MetaProjectType.MYA,
            is_draft=False,
            umbrella_code="BTN/CPOP/32",
            project_cost=Decimal("123456.50"),
        )
        ProjectFactory.create(
            meta_project=included_mp,
            category=Project.Category.MYA,
            submission_status=project_approved_status,
        )

        # TEMP segment in umbrella_code (not yet ready) → excluded.
        temp_mp = MetaProjectFactory.create(
            type=MetaProject.MetaProjectType.MYA,
            is_draft=False,
            umbrella_code="KNA/HPMP2/TEMP/1",
            project_cost=Decimal("999.00"),
        )
        ProjectFactory.create(
            meta_project=temp_mp,
            category=Project.Category.MYA,
            submission_status=project_approved_status,
        )

        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url, {"mock_data": "false"})
        sheet = load_workbook(response)["MetaProjects"]
        headers = get_sheet_headers(sheet)

        metacode_col = headers["Metacode"]
        cost_col = headers["MYA Total agreed costs in principle (US $)"]
        rows = {
            sheet[f"{metacode_col}{r}"].value: sheet[f"{cost_col}{r}"].value
            for r in range(2, sheet.max_row + 1)
        }
        assert rows["BTN/CPOP/32"] == 123456.50
        assert "KNA/HPMP2/TEMP/1" not in rows

    def test_metaprojects_sheet_uses_computed_fallback(
        self, secretariat_viewer_user, project_approved_status
    ):
        # Null funding + approved sub-projects → MyaExport's fallback sums
        # total_fund. Proves the shared calc flows through (not raw nulls).
        mp = MetaProjectFactory.create(
            type=MetaProject.MetaProjectType.MYA,
            is_draft=False,
            umbrella_code="BTN/CPOP/40",
            project_funding=None,
        )
        ProjectFactory.create(
            meta_project=mp,
            category=Project.Category.MYA,
            submission_status=project_approved_status,
            total_fund=100,
        )
        ProjectFactory.create(
            meta_project=mp,
            category=Project.Category.MYA,
            submission_status=project_approved_status,
            total_fund=250,
        )

        self.client.force_authenticate(user=secretariat_viewer_user)
        response = self.client.get(self.url, {"mock_data": "false"})
        sheet = load_workbook(response)["MetaProjects"]
        headers = get_sheet_headers(sheet)

        metacode_col = headers["Metacode"]
        funding_col = headers["MYA Total agreed funding in principle (US $)"]
        rows = {
            sheet[f"{metacode_col}{r}"].value: sheet[f"{funding_col}{r}"].value
            for r in range(2, sheet.max_row + 1)
        }
        assert rows["BTN/CPOP/40"] == 350.0
