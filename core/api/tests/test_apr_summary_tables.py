"""
Tests for APR Summary Tables Export functionality.
"""

from datetime import date
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status

from core.api.export.annual_project_report import APRSummaryTablesExportWriter
from core.api.tests.base import BaseTest
from core.api.tests.factories import (
    AnnualAgencyProjectReportFactory,
    AnnualProjectReportFactory,
    CountryFactory,
    ProjectFactory,
    ProjectSectorFactory,
    ProjectTypeFactory,
)
from core.models import Country
from core.models.project_metadata import ProjectCluster, ProjectType

# pylint: disable=W0221,W0613,R0913,R0914,R0904


@pytest.mark.django_db
class TestAPRSummaryTablesExport(BaseTest):
    """Test the summary tables export functionality"""

    def test_without_login(self):
        self.client.force_authenticate(user=None)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_requires_apr_view_permission(self, user):
        self.client.force_authenticate(user=user)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_agency_user_sees_only_own_projects(
        self,
        apr_agency_viewer_user,
        annual_progress_report,
        annual_agency_report,
    ):
        own_project = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            code=f"TEST_OWN_{apr_agency_viewer_user.agency.id}",
            date_approved=date(2023, 1, 15),
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=own_project,
        )

        other_agency_report = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
        )
        other_project = ProjectFactory(
            agency=other_agency_report.agency,
            code=f"TEST_OTHER_{other_agency_report.agency.id}",
            date_approved=date(2023, 2, 20),
        )
        AnnualProjectReportFactory(
            report=other_agency_report,
            project=other_project,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Check I.1 summary only counts the agency's own project
        workbook = load_workbook(BytesIO(response.content))
        summary_sheet = workbook[APRSummaryTablesExportWriter.SHEET_SUMMARY]

        # B4 = Number of Approvals — should be 1 (only own project)
        assert summary_sheet.cell(4, 2).value == 1

    def test_mlfs_user_sees_all_projects(
        self,
        secretariat_viewer_user,
        annual_progress_report,
    ):
        agency_report1 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
        )
        project1 = ProjectFactory(
            agency=agency_report1.agency,
            code="AGENCY1/001",
            date_approved=date(2023, 1, 15),
        )
        AnnualProjectReportFactory(
            report=agency_report1,
            project=project1,
        )

        agency_report2 = AnnualAgencyProjectReportFactory(
            progress_report=annual_progress_report,
        )
        project2 = ProjectFactory(
            agency=agency_report2.agency,
            code="AGENCY2/001",
            date_approved=date(2023, 2, 20),
        )
        AnnualProjectReportFactory(
            report=agency_report2,
            project=project2,
        )

        self.client.force_authenticate(user=secretariat_viewer_user)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK

        # Check I.1 summary counts both projects
        workbook = load_workbook(BytesIO(response.content))
        summary_sheet = workbook[APRSummaryTablesExportWriter.SHEET_SUMMARY]

        # B4 = Number of Approvals — should be 2 (both projects)
        assert summary_sheet.cell(4, 2).value == 2

    def test_export_structure_and_metadata(
        self,
        apr_agency_viewer_user,
        annual_progress_report,
        annual_agency_report,
    ):
        """
        Consolidates 6 structural/metadata checks into a single export call.
        Covers: sheet names, hidden-sheet removal, Annex I (a)/(b)/(c) headers,
        flat-layout structure, and Content-Disposition filename.
        """
        project = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            date_approved=date(2023, 1, 15),
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Content-Disposition includes agency name, year, and extension
        content_disposition = response["Content-Disposition"]
        assert "APR_Summary_Tables_Cumulative" in content_disposition
        assert (
            apr_agency_viewer_user.agency.name.replace(" ", "_") in content_disposition
        )
        assert ".xlsx" in content_disposition

        workbook = load_workbook(BytesIO(response.content))
        sheet_names = workbook.sheetnames

        # All expected sheets are present
        assert APRSummaryTablesExportWriter.SHEET_SUMMARY in sheet_names
        assert APRSummaryTablesExportWriter.SHEET_SUMMARY_CLUSTER in sheet_names
        assert APRSummaryTablesExportWriter.SHEET_COMPLETION_YEAR in sheet_names
        assert "Annex I (a)" in sheet_names
        assert "Annex I (b)" in sheet_names
        assert "Annex I (c)" in sheet_names
        assert "Annex I (d)" in sheet_names
        assert "Annex I (e)" in sheet_names
        assert "Annex I (f)" in sheet_names
        assert "Annex I (g)" in sheet_names

        # No hidden sheets
        for sheet_name in sheet_names:
            assert workbook[sheet_name].sheet_state != "hidden"

        # Annex I (a): header row structure
        sheet_a = workbook["Annex I (a)"]
        assert sheet_a["A1"].value == "(a) Annual summary data"
        assert "Year approval" in str(sheet_a["A3"].value)
        assert "Number of Approvals" in str(sheet_a["B3"].value)
        assert "Number of completed" in str(sheet_a["C3"].value)
        assert "Per cent completed" in str(sheet_a["D3"].value)
        assert "Approved funding" in str(sheet_a["E3"].value)
        assert "Funds disbursed" in str(sheet_a["F3"].value)
        assert "Balance" in str(sheet_a["G3"].value)

        # Annex I (b): flat layout structure
        sheet_b = workbook["Annex I (b)"]
        assert sheet_b["A1"].value == "(b) Cumulative completed investment projects"
        assert "Item" in str(sheet_b.cell(3, 1).value)
        assert sheet_b.cell(4, 1).value == "Total"
        assert any(
            "Region" in str(sheet_b.cell(row, 1).value or "")
            for row in range(5, sheet_b.max_row + 1)
        )
        assert any(
            "Sector" in str(sheet_b.cell(row, 1).value or "")
            for row in range(5, sheet_b.max_row + 1)
        )

        # Annex I (c): flat layout structure
        sheet_c = workbook["Annex I (c)"]
        assert sheet_c["A1"].value == "(c) Cumulative completed non-investment projects"
        assert sheet_c.cell(4, 1).value == "Total"

    def test_annual_summary_aggregation_by_year(
        self,
        apr_agency_viewer_user,
        annual_progress_report,
        annual_agency_report,
        project_ongoing_status,
        project_completed_status,
    ):
        project_2022 = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            date_approved=date(2022, 3, 15),
            status=project_completed_status,
            version=3,
            total_fund=100000,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project_2022,
            status="COM",
            funds_disbursed=100000,
        )

        project_2023_1 = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            date_approved=date(2023, 1, 20),
            status=project_ongoing_status,
            version=3,
            total_fund=50000,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project_2023_1,
            funds_disbursed=25000,
        )

        project_2023_2 = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            date_approved=date(2023, 6, 10),
            status=project_completed_status,
            version=3,
            total_fund=75000,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project_2023_2,
            status="COM",
            funds_disbursed=75000,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["Annex I (a)"]

        # Find the 2022 and 2023 data rows (data starts at row 4)
        year_2022_row = None
        year_2023_row = None
        col_map = APRSummaryTablesExportWriter.build_annual_column_mapping()
        for row in range(
            APRSummaryTablesExportWriter.DATA_START_ROW, sheet.max_row + 1
        ):
            year_val = sheet.cell(row, col_map["approval_year"]).value
            if year_val == 2022:
                year_2022_row = row
            elif year_val == 2023:
                year_2023_row = row

        assert year_2022_row is not None
        assert year_2023_row is not None

        # Number of approvals, number completed, approved funding
        assert sheet.cell(year_2022_row, 2).value == 1
        assert sheet.cell(year_2022_row, 3).value == 1
        assert sheet.cell(year_2022_row, 5).value == 100000

        assert sheet.cell(year_2023_row, 2).value == 2
        assert sheet.cell(year_2023_row, 3).value == 1
        assert sheet.cell(year_2023_row, 5).value == 125000
        assert sheet.cell(year_2023_row, 6).value == 100000

    def test_completed_project_summary_and_totals(
        self,
        apr_agency_viewer_user,
        annual_progress_report,
        annual_agency_report,
        project_completed_status,
    ):
        """
        Testing summary and totals for completed projects in a single export.
        """
        project = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            date_approved=date(2023, 1, 15),
            status=project_completed_status,
            version=3,
            total_fund=100000,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project,
            status="COM",
            funds_disbursed=80000,
            consumption_phased_out_odp=10,
            production_phased_out_odp=5,
            consumption_phased_out_mt=100,
            production_phased_out_mt=50,
            consumption_phased_out_co2=1000,
            production_phased_out_co2=500,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)

        workbook = load_workbook(BytesIO(response.content))

        # I.1 Summary Data sheet structure
        summary_sheet = workbook[APRSummaryTablesExportWriter.SHEET_SUMMARY]
        assert "Summary data" in str(summary_sheet["A1"].value)
        # Number of Approvals
        assert summary_sheet.cell(4, 2).value == 1
        # Number of completed
        assert summary_sheet.cell(5, 2).value == 1
        # Total funds disbursed
        assert summary_sheet.cell(7, 2).value == 80000
        # ODP = 10 + 5
        assert summary_sheet.cell(9, 2).value == 15
        # MT = 100 + 50
        assert summary_sheet.cell(11, 2).value == 150
        # CO2 = 1000 + 500
        assert summary_sheet.cell(13, 2).value == 1500

        # Annex I (a): Total row
        sheet_a = workbook["Annex I (a)"]
        total_row = None
        for row in range(
            APRSummaryTablesExportWriter.DATA_START_ROW, sheet_a.max_row + 1
        ):
            if sheet_a.cell(row, 1).value == "Total":
                total_row = row
                break
        assert total_row is not None
        # Number of approvals & completed
        assert sheet_a.cell(total_row, 2).value == 1
        assert sheet_a.cell(total_row, 3).value == 1

    def test_project_type_sheet_routing(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        project_completed_status,
        project_ongoing_status,
    ):
        """
        Testing that various sheets correctly filter which types of projects they include
        """
        inv_type = ProjectType.objects.create(
            name="Investment", code="INV", sort_order=1
        )
        prp_type = ProjectType.objects.create(
            name="Preparation", code="PRP", sort_order=2
        )
        oth_type = ProjectType.objects.create(name="Other", code="OTH", sort_order=3)

        # COM+INV -> sheet (b): completed investment
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                status=project_completed_status,
                project_type=inv_type,
            ),
        )
        # ONG+INV -> sheet (e): ongoing investment, excluded from (b)
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                status=project_ongoing_status,
                project_type=inv_type,
            ),
        )
        # COM+PRP -> sheet (d): completed preparation, excluded from (b) and (c)
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                status=project_completed_status,
                project_type=prp_type,
            ),
        )
        # COM+OTH -> sheet (c): completed non-investment, excluded from (b) and (d)
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                status=project_completed_status,
                project_type=oth_type,
            ),
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        wb = load_workbook(
            BytesIO(self.client.get(reverse("apr-summary-tables-export")).content)
        )
        data_row = APRSummaryTablesExportWriter.DATA_START_ROW

        # Sheet (b): only COM+INV are included
        assert (
            wb[APRSummaryTablesExportWriter.SHEET_INVESTMENT].cell(data_row, 2).value
            == 1
        )
        # Sheet (c): only COM+OTH are included
        assert (
            wb[APRSummaryTablesExportWriter.SHEET_NON_INVESTMENT]
            .cell(data_row, 2)
            .value
            == 1
        )
        # Sheet (d): only COM+PRP are included
        assert (
            wb[APRSummaryTablesExportWriter.SHEET_PREPARATION].cell(data_row, 2).value
            == 1
        )
        # Sheet (e): only ONG+INV are included
        assert (
            wb[APRSummaryTablesExportWriter.SHEET_ONGOING_INVESTMENT]
            .cell(data_row, 2)
            .value
            == 1
        )

    def test_includes_all_statuses_in_summary(
        self,
        apr_agency_viewer_user,
        annual_progress_report,
        annual_agency_report,
        project_ongoing_status,
        project_completed_status,
        project_closed_status,
    ):
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                date_approved=date(2023, 1, 15),
                status=project_ongoing_status,
            ),
        )

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                date_approved=date(2023, 2, 15),
                status=project_completed_status,
            ),
        )

        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                date_approved=date(2023, 3, 15),
                status=project_closed_status,
            ),
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)

        workbook = load_workbook(BytesIO(response.content))
        summary_sheet = workbook[APRSummaryTablesExportWriter.SHEET_SUMMARY]

        # B4 = Number of Approvals — should include all statuses
        num_approvals = summary_sheet.cell(4, 2).value
        assert num_approvals == 3

    def test_summary_by_cluster_sheet(
        self,
        apr_agency_viewer_user,
        annual_progress_report,
        annual_agency_report,
        project_completed_status,
        project_ongoing_status,
    ):
        cluster1 = ProjectCluster.objects.create(name="HPMP stage I", sort_order=1)
        cluster2 = ProjectCluster.objects.create(name="KIP stage I", sort_order=2)

        project1 = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            date_approved=date(2023, 1, 15),
            status=project_completed_status,
            cluster=cluster1,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project1,
            status="COM",
            funds_disbursed=100000,
        )

        project2 = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            date_approved=date(2023, 2, 15),
            status=project_ongoing_status,
            cluster=cluster2,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project2,
            funds_disbursed=50000,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        url = reverse("apr-summary-tables-export")
        response = self.client.get(url)

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook[APRSummaryTablesExportWriter.SHEET_SUMMARY_CLUSTER]

        assert "Summary data by cluster" in str(sheet["A1"].value)

        cluster_names = []
        for row in range(
            APRSummaryTablesExportWriter.CLUSTER_DATA_START_ROW, sheet.max_row + 1
        ):
            val = sheet.cell(row, 1).value
            if val and val != "Total":
                cluster_names.append(val)
            elif val == "Total":
                break

        assert "HPMP stage I" in cluster_names
        assert "KIP stage I" in cluster_names

    def test_project_completion_year_sheet(
        self,
        apr_agency_viewer_user,
        annual_progress_report,
        annual_agency_report,
        project_completed_status,
    ):
        cluster = ProjectCluster.objects.create(name="HPMP stage I", sort_order=1)

        project = ProjectFactory(
            agency=apr_agency_viewer_user.agency,
            date_approved=date(2023, 1, 15),
            status=project_completed_status,
            cluster=cluster,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=project,
            status="COM",
            consumption_phased_out_odp=10,
            production_phased_out_odp=5,
            consumption_phased_out_mt=100,
            production_phased_out_mt=50,
            consumption_phased_out_co2=1000,
            production_phased_out_co2=500,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        report_year = annual_progress_report.year
        url = reverse("apr-summary-tables-export") + f"?year={report_year}"
        response = self.client.get(url)

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook[APRSummaryTablesExportWriter.SHEET_COMPLETION_YEAR]

        assert "Project completion" in str(sheet["A1"].value)

        found_cluster = False
        for row in range(
            APRSummaryTablesExportWriter.CLUSTER_DATA_START_ROW, sheet.max_row + 1
        ):
            if sheet.cell(row, 1).value == "HPMP stage I":
                found_cluster = True
                assert sheet.cell(row, 2).value == 1
                # ODP = consumption(10) + production(5)
                assert sheet.cell(row, 3).value == 15
                # MT = consumption(100) + production(50)
                assert sheet.cell(row, 4).value == 150
                # CO2 = consumption(1000) + production(500)
                assert sheet.cell(row, 5).value == 1500
                break

        assert found_cluster

    def test_sector_header_row_when_more_regions_than_template(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        project_ongoing_status,
    ):
        inv_type = ProjectTypeFactory(code="INV", sort_order=1)
        sector = ProjectSectorFactory()
        for i in range(6):
            region = CountryFactory(
                name=f"ExtraRegion{i}",
                location_type=Country.LocationType.REGION,
            )
            country = CountryFactory(name=f"ExtraCountry{i}", parent=region)
            AnnualProjectReportFactory(
                report=annual_agency_report,
                project=ProjectFactory(
                    agency=apr_agency_viewer_user.agency,
                    status=project_ongoing_status,
                    project_type=inv_type,
                    sector=sector,
                    country=country,
                ),
            )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        response = self.client.get(reverse("apr-summary-tables-export"))
        ws = load_workbook(BytesIO(response.content))[
            APRSummaryTablesExportWriter.SHEET_ONGOING_INVESTMENT
        ]

        n_regions = 6
        # Layout: Totals row -> "Region" header -> n_regions data rows -> "Sector" header
        expected_sector_row = (
            APRSummaryTablesExportWriter.DATA_START_ROW + 1 + n_regions + 1
        )
        assert (
            str(ws.cell(expected_sector_row, 1).value or "").strip().lower() == "sector"
        )
        assert (
            str(ws.cell(expected_sector_row - 1, 1).value or "").strip().lower()
            != "sector"
        )

    def test_sector_header_row_when_fewer_regions_than_template(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        project_ongoing_status,
    ):
        inv_type = ProjectTypeFactory(code="INV", sort_order=1)
        sector = ProjectSectorFactory()
        region = CountryFactory(
            name="SoleRegion", location_type=Country.LocationType.REGION
        )
        country = CountryFactory(name="SoleCountry", parent=region)
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                status=project_ongoing_status,
                project_type=inv_type,
                sector=sector,
                country=country,
            ),
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        response = self.client.get(reverse("apr-summary-tables-export"))
        ws = load_workbook(BytesIO(response.content))[
            APRSummaryTablesExportWriter.SHEET_ONGOING_INVESTMENT
        ]

        n_regions = 1
        # Layout: Totals row -> "Region" header -> n_regions data rows -> "Sector" header
        expected_sector_row = (
            APRSummaryTablesExportWriter.DATA_START_ROW + 1 + n_regions + 1
        )
        assert (
            str(ws.cell(expected_sector_row, 1).value or "").strip().lower() == "sector"
        )
        assert (
            str(ws.cell(expected_sector_row - 1, 1).value or "").strip() == "SoleRegion"
        )

    def test_cluster_total_row_position_when_count_exceeds_template(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        project_completed_status,
    ):
        num_clusters = 8
        for i in range(num_clusters):
            cluster = ProjectCluster.objects.create(name=f"XCluster{i}", sort_order=i)
            AnnualProjectReportFactory(
                report=annual_agency_report,
                project=ProjectFactory(
                    agency=apr_agency_viewer_user.agency,
                    status=project_completed_status,
                    cluster=cluster,
                ),
            )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        response = self.client.get(reverse("apr-summary-tables-export"))
        ws = load_workbook(BytesIO(response.content))[
            APRSummaryTablesExportWriter.SHEET_SUMMARY_CLUSTER
        ]

        expected_total_row = (
            APRSummaryTablesExportWriter.CLUSTER_DATA_START_ROW + num_clusters
        )
        assert str(ws.cell(expected_total_row, 1).value or "").strip() == "Total"
        assert ws.cell(expected_total_row, 2).value == num_clusters

    def test_pct_disbursed_uses_fund_totals_not_per_project_average(
        self,
        apr_agency_viewer_user,
        annual_agency_report,
        project_completed_status,
    ):
        """
        Regression: pct_funds_disbursed must be total_disbursed / total_approved_funding,
        not an average of individual per-project percentages.

        With two unequally-sized projects:
          Project A: approved=3000, disbursed=3000 → 100%
          Project B: approved=1000, disbursed=0     →   0%
          Correct  (new): 3000 / 4000 * 100 = 75
          Wrong  (old avg): (100 + 0) / 2   = 50

        Covers: _write_annual_summary_sheet (per-year + Total rows)
        and _compute_group_data (flat-sheet Total row).
        """
        inv_type = ProjectType.objects.create(
            name="Investment", code="INV", sort_order=1
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                date_approved=date(2023, 1, 15),
                status=project_completed_status,
                project_type=inv_type,
                version=3,
                total_fund=3000,
            ),
            status="COM",
            funds_disbursed=3000,
        )
        AnnualProjectReportFactory(
            report=annual_agency_report,
            project=ProjectFactory(
                agency=apr_agency_viewer_user.agency,
                date_approved=date(2023, 6, 1),
                status=project_completed_status,
                project_type=inv_type,
                version=3,
                total_fund=1000,
            ),
            status="COM",
            funds_disbursed=0,
        )

        self.client.force_authenticate(user=apr_agency_viewer_user)
        wb = load_workbook(
            BytesIO(self.client.get(reverse("apr-summary-tables-export")).content)
        )

        col_map = APRSummaryTablesExportWriter.build_annual_column_mapping()
        sheet_a = wb[APRSummaryTablesExportWriter.SHEET_ANNUAL]

        # Per-year row for 2023
        year_2023_row = next(
            r
            for r in range(
                APRSummaryTablesExportWriter.DATA_START_ROW, sheet_a.max_row + 1
            )
            if sheet_a.cell(r, col_map["approval_year"]).value == 2023
        )
        assert sheet_a.cell(year_2023_row, col_map["pct_funds_disbursed"]).value == 75.0

        # Total row
        total_row = next(
            r
            for r in range(
                APRSummaryTablesExportWriter.DATA_START_ROW, sheet_a.max_row + 1
            )
            if sheet_a.cell(r, col_map["approval_year"]).value == "Total"
        )
        assert sheet_a.cell(total_row, col_map["pct_funds_disbursed"]).value == 75.0

        # Sheet (b) Total row: avg_pct_disbursed is at col 4
        # (col 1=label, col 2=num_completed, col 3=total_approved_funding, col 4=avg_pct_disbursed)
        sheet_b = wb[APRSummaryTablesExportWriter.SHEET_INVESTMENT]
        assert (
            sheet_b.cell(APRSummaryTablesExportWriter.DATA_START_ROW, 4).value == 75.0
        )
