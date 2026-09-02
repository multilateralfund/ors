from pathlib import Path
from unittest.mock import ANY
from unittest.mock import Mock

import pytest
from django.core.exceptions import ImproperlyConfigured
from django.test import override_settings
from openpyxl import Workbook

from core.api import utils


def response_content(response):
    try:
        return b"".join(response.streaming_content)
    finally:
        response.close()


def test_workbook_pdf_response_uses_local_libreoffice(monkeypatch):
    check_call = Mock()

    def create_pdf(command, cwd, shell):
        check_call(command, cwd=cwd, shell=shell)
        (Path(cwd) / "report.pdf").write_bytes(b"local PDF")

    monkeypatch.setattr(utils.shutil, "which", lambda _executable: "/usr/bin/soffice")
    monkeypatch.setattr(utils.subprocess, "check_call", create_pdf)
    unoserver_client = Mock()
    monkeypatch.setattr(utils, "UnoClient", unoserver_client)

    response = utils.workbook_pdf_response("report", Workbook())

    assert response_content(response) == b"local PDF"
    assert check_call.call_args.args[0][0] == "/usr/bin/soffice"
    unoserver_client.assert_not_called()


@override_settings(UNOSERVER_HOST="libreoffice:2003")
def test_workbook_pdf_response_uses_unoserver_when_libreoffice_is_missing(
    monkeypatch,
):
    monkeypatch.setattr(utils.shutil, "which", lambda _executable: None)
    converter = Mock()
    converter.convert.return_value = b"remote PDF"
    unoserver_client = Mock(return_value=converter)
    monkeypatch.setattr(utils, "UnoClient", unoserver_client)

    response = utils.workbook_pdf_response("report", Workbook())

    assert response_content(response) == b"remote PDF"
    unoserver_client.assert_called_once_with(
        server="libreoffice",
        port="2003",
        host_location="remote",
    )
    converter.convert.assert_called_once_with(indata=ANY, convert_to="pdf")
    assert converter.convert.call_args.kwargs["indata"].startswith(b"PK")


@pytest.mark.parametrize(
    "unoserver_host",
    ["", "libreoffice", ":2003", "libreoffice:not-a-port", "libreoffice:65536"],
)
def test_workbook_pdf_response_requires_valid_unoserver_host(
    monkeypatch,
    settings,
    unoserver_host,
):
    settings.UNOSERVER_HOST = unoserver_host
    monkeypatch.setattr(utils.shutil, "which", lambda _executable: None)

    with pytest.raises(ImproperlyConfigured, match="host:port"):
        utils.workbook_pdf_response("report", Workbook())
